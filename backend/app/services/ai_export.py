"""
Export Excel untuk hasil Asisten AI — dibangun RAPI & profesional (bukan dump polos).

Saat ini: perbandingan PART dua unit (banding_rangka) → workbook berwarna dengan
sheet Ringkasan + sheet part-beda per sisi + sheet part-sama. Sumber sama dengan
tool banding_rangka (EPC Loading List per-VIN), tapi TANPA cap 30 baris — Excel
memuat SELURUH part yang berbeda.
"""
from __future__ import annotations

import io
import re
import shutil
import tempfile
import threading
import time
import uuid
import zipfile
from concurrent.futures import ThreadPoolExecutor
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import catalog_bom, epc_bom, part_index

# ── Palet warna (samakan dengan UI brand) ──
_BRAND = "028912"
_BRAND_DK = "026A0E"
_HEAD_FILL = PatternFill("solid", fgColor=_BRAND)
_SUB1_FILL = PatternFill("solid", fgColor="EAF6EC")   # hijau muda
_SUB2_FILL = PatternFill("solid", fgColor="FFF3E2")   # oranye muda (sisi 2)
_ZEBRA = PatternFill("solid", fgColor="F8F9F7")
# Warna STATUS baris (builder sheet_status): hijau=ready, merah=kosong, kuning=perlu perhatian.
_ROW_HIJAU = PatternFill("solid", fgColor="EAF6EC")
_ROW_MERAH = PatternFill("solid", fgColor="FDECEA")
_ROW_KUNING = PatternFill("solid", fgColor="FFF6DC")
_STATUS_FILL = {"hijau": _ROW_HIJAU, "merah": _ROW_MERAH, "kuning": _ROW_KUNING}
_WHITE = Font(color="FFFFFF", bold=True, size=11)
_TITLE_FONT = Font(color="FFFFFF", bold=True, size=15)
_BOLD = Font(bold=True, color="1B211D")
_INK = Font(color="1B211D")
_MONO = Font(name="Consolas", color="0F1411")
_THIN = Side(style="thin", color="E1E4E1")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center")   # sel angka

# Cegah CSV/Excel FORMULA INJECTION: sel teks yang diawali = + - @ (atau kendali
# tab/CR/LF) diperlakukan Excel sebagai FORMULA/DDE. Nama part/tabel di file ini
# berasal dari EPC/SIMS/keluaran model (tak tepercaya) → escape dgn prefiks '
# agar ditulis sebagai TEKS. Angka & non-string dibiarkan apa adanya.
_FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r", "\n")


def _safe(v):
    if isinstance(v, str):
        s = v.lstrip("﻿")  # buang BOM di depan, jangan tertipu
        if s[:1] in _FORMULA_LEAD:
            return "'" + v
    return v


# ── Perbandingan PENUH dua VIN (tanpa cap) ──
def compare_rangka(rangka_1: str, rangka_2: str, kategori: str = "") -> dict:
    """Bandingkan SET PART dua unit dari EPC Loading List. Return:
      {ok:True, frame_1, frame_2, kat_nama, total_1, total_2,
       same:[{pn,nama,qty}], only1:[...], only2:[...]}  atau  {ok:False, error}."""
    r1 = (rangka_1 or "").strip()
    r2 = (rangka_2 or "").strip()
    if not r1 or not r2:
        return {"ok": False, "error": "Butuh dua nomor rangka."}

    with ThreadPoolExecutor(max_workers=2) as ex:
        f1 = ex.submit(epc_bom.loading_list, r1)
        f2 = ex.submit(epc_bom.loading_list, r2)
        ll1, ll2 = f1.result(), f2.result()
    for ll in (ll1, ll2):
        if not ll.get("found"):
            return {"ok": False, "error": "Salah satu unit tak ditemukan / token EPC bermasalah."}
        if ll.get("partial"):
            return {"ok": False, "error": "Data EPC salah satu unit terbaca tidak lengkap — coba lagi."}

    code = None
    kat_nama = "SEMUA part"
    if kategori:
        code = catalog_bom.resolve_kategori(kategori) if catalog_bom.available() else None
        if not code:
            return {"ok": False, "error": f"Kategori '{kategori}' tak dikenal."}
        kat_nama = catalog_bom.KATEGORI_NAMA.get(code, kategori)

    pncat = catalog_bom.pn_category_map() if catalog_bom.available() else {}

    def _cat(pn: str) -> str:
        return (pncat.get(catalog_bom._norm(pn)) or {}).get("kategori") or "00"

    def _set(ll: dict) -> dict:
        out = {}
        for p in ll.get("parts", []):
            pn = p.get("pn")
            if not pn or (code and _cat(pn) != code):
                continue
            out[pn] = p
        return out

    A, B = _set(ll1), _set(ll2)
    sa, sb = set(A), set(B)
    only1, only2, same = sorted(sa - sb), sorted(sb - sa), sorted(sa & sb)

    # Nama Inggris lokal untuk SEMUA PN (sekali query), fallback ke kamus EPC.
    localn: dict[str, str] = {}
    for r in part_index.search_exact_pns(list(sa | sb)):
        pn = (r.get("part_number") or "").upper()
        if pn and pn not in localn:
            localn[pn] = r.get("part_name") or ""

    def _row(pn: str, src: dict) -> dict:
        p = src.get(pn, {})
        en = localn.get(pn) or epc_bom.translate_cn(p.get("nama_cn"))
        return {"pn": pn, "nama": " ".join((en or p.get("nama_cn") or "").split()),
                "nama_china": " ".join((p.get("nama_cn") or "").split()),
                "qty": p.get("qty")}

    return {
        "ok": True,
        "frame_1": ll1.get("frame_number"), "frame_2": ll2.get("frame_number"),
        "kat_nama": kat_nama,
        "total_1": len(A), "total_2": len(B),
        "same": [_row(pn, A) for pn in same],
        "only1": [_row(pn, A) for pn in only1],
        "only2": [_row(pn, B) for pn in only2],
    }


# ── Pembangun Excel ber-styling ──
def _style_table(ws, start_row: int, headers: list[str], rows: list[list],
                 head_fill: PatternFill, widths: list[int]) -> int:
    """Tulis satu tabel bergaya mulai dari start_row. Return baris SETELAH tabel."""
    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.fill = head_fill
        c.font = _WHITE
        c.alignment = _CENTER
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(j)].width = w
    r = start_row + 1
    for i, row in enumerate(rows):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=_safe(val))
            c.border = _BORDER
            c.alignment = _CENTER if j in (1, len(headers)) else _LEFT
            if j == 2:  # kolom PN → mono
                c.font = _MONO
            else:
                c.font = _INK
            if i % 2:
                c.fill = _ZEBRA
        r += 1
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return r


def _title(ws, text: str, sub: str, ncol: int) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = ws.cell(row=1, column=1, value=text)
    t.fill = _HEAD_FILL
    t.font = _TITLE_FONT
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    s = ws.cell(row=2, column=2 if False else 1, value=sub)
    s.font = Font(color="535B56", size=10, italic=True)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return 4  # baris mulai konten


# .xlsx itu arsip ZIP, dan setiap ENTRI-nya membawa stempel waktu sendiri yang
# diambil dari jam dinding saat penulisan. Jadi memasang wb.properties saja TIDAK
# cukup: dua build data identik yang kebetulan jatuh di detik berbeda tetap
# menghasilkan bytes berbeda. Cacat ini laten — test byte-stable lolos selama
# kedua pemanggilan mendarat di detik yang sama, lalu gagal acak begitu mesin
# sedang sibuk (mis. saat suite penuh berjalan). Stempel dipin ke batas bawah
# yang diizinkan format ZIP (1980-01-01).
_ZIP_EPOCH = (1980, 1, 1, 0, 0, 0)
_CORE_XML = "docProps/core.xml"
# openpyxl MENIMPA properties.modified dengan waktu simpan di dalam save(), jadi
# memasangnya sebelum save() (lihat _save_stable) hanya mengunci `created`.
# Nilainya dinormalkan di sini, setelah workbook jadi.
_MODIFIED_RE = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")
_MODIFIED_TETAP = b"2000-01-01T00:00:00Z"


def _zip_stabil(raw: bytes, modified: bytes = _MODIFIED_TETAP) -> bytes:
    """Tulis ulang arsip xlsx menjadi byte-stabil: stempel waktu tiap entri ZIP
    dipin, dan `dcterms:modified` di core.xml dinormalkan. Urutan entri & metode
    kompresi dipertahankan agar isinya identik bagi Excel.

    `modified` bisa diisi nilai ASLI file user (lihat `_save_asli`) supaya properti
    dokumen miliknya tidak ikut berubah."""
    src = zipfile.ZipFile(io.BytesIO(raw))
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w") as dst:
        for info in src.infolist():
            data = src.read(info.filename)
            if info.filename == _CORE_XML:
                data = _MODIFIED_RE.sub(rb"\g<1>" + modified + rb"\g<2>", data)
            zi = zipfile.ZipInfo(info.filename, date_time=_ZIP_EPOCH)
            zi.compress_type = info.compress_type
            zi.external_attr = info.external_attr
            zi.internal_attr = info.internal_attr
            zi.create_system = info.create_system
            dst.writestr(zi, data)
    return out.getvalue()


def _save_stable(wb) -> bytes:
    """Simpan workbook → bytes DETERMINISTIK. openpyxl cap created/modified dgn
    datetime.now() → dua build data sama menghasilkan bytes berbeda; pin ke tanggal
    tetap supaya byte-stable (aman cache & diff)."""
    from datetime import datetime
    fixed = datetime(2000, 1, 1)
    try:
        wb.properties.created = fixed
        wb.properties.modified = fixed
    except Exception:  # pragma: no cover
        pass
    buf = io.BytesIO()
    wb.save(buf)
    try:
        return _zip_stabil(buf.getvalue())
    except Exception:  # pragma: no cover — file tetap sah walau tak ter-normalisasi
        return buf.getvalue()


def banding_rangka_excel(rangka_1: str, rangka_2: str, kategori: str = "") -> tuple[bytes | None, str]:
    """Bangun workbook perbandingan. Return (bytes, filename) atau (None, pesan_error)."""
    d = compare_rangka(rangka_1, rangka_2, kategori)
    if not d.get("ok"):
        return None, d.get("error") or "Gagal membandingkan."

    f1, f2 = d["frame_1"], d["frame_2"]
    kat = d["kat_nama"]
    identik = not d["only1"] and not d["only2"]
    wb = Workbook()

    # ── Sheet 1: Ringkasan ──
    ws = wb.active
    ws.title = "Ringkasan"
    ws.sheet_view.showGridLines = False
    _title(ws, f"Perbandingan Part {kat} — {f1} vs {f2}",
           "Sumber: EPC Loading List per-VIN (Sinotruk) · MASPART Asisten AI", 4)
    rows = [
        ["Total part", d["total_1"], d["total_2"], ""],
        ["Part SAMA", len(d["same"]), len(d["same"]), ""],
        ["Hanya di unit ini", len(d["only1"]), len(d["only2"]), ""],
        ["Total BEDA", len(d["only1"]) + len(d["only2"]),
         "", "identik" if identik else "berbeda"],
    ]
    after = _style_table(
        ws, 4, ["Item", f"Rangka 1 ({f1})", f"Rangka 2 ({f2})", "Keterangan"],
        rows, _HEAD_FILL, [26, 22, 22, 16])
    # Kesimpulan
    verdict = ("✔ KEDUA UNIT SAMA PERSIS (semua part identik pada kategori ini)."
               if identik else
               f"✘ TIDAK SAMA PERSIS — {len(d['only1']) + len(d['only2'])} part berbeda "
               f"({len(d['only1'])} hanya di {f1}, {len(d['only2'])} hanya di {f2}).")
    ws.merge_cells(start_row=after + 1, start_column=1, end_row=after + 1, end_column=4)
    cc = ws.cell(row=after + 1, column=1, value=verdict)
    cc.font = Font(bold=True, color=(_BRAND_DK if identik else "B35C00"), size=11)
    cc.fill = _SUB1_FILL if identik else _SUB2_FILL
    cc.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[after + 1].height = 26

    # ── Sheet 2 & 3: part beda per sisi ──
    def _diff_sheet(name: str, frame: str, items: list[dict], fill: PatternFill):
        w = wb.create_sheet(name)
        w.sheet_view.showGridLines = False
        _title(w, f"Part hanya di {frame} ({kat})",
               f"{len(items)} part — tidak ada di unit pembanding", 4)
        data = [[i + 1, it["pn"], it["nama"] or it["nama_china"], it["qty"]]
                for i, it in enumerate(items)]
        _style_table(w, 4, ["No", "Part Number", "Nama Part", "Qty"],
                     data or [["", "(tidak ada)", "", ""]], fill, [6, 24, 60, 8])

    _diff_sheet("Hanya Rangka 1", f1, d["only1"], _HEAD_FILL)
    _diff_sheet("Hanya Rangka 2", f2, d["only2"],
                PatternFill("solid", fgColor="B35C00"))

    # ── Sheet 4: part sama (referensi) ──
    ws4 = wb.create_sheet("Part Sama")
    ws4.sheet_view.showGridLines = False
    _title(ws4, f"Part SAMA di kedua unit ({kat})",
           f"{len(d['same'])} part identik pada {f1} & {f2}", 4)
    same_data = [[i + 1, it["pn"], it["nama"] or it["nama_china"], it["qty"]]
                 for i, it in enumerate(d["same"])]
    _style_table(ws4, 4, ["No", "Part Number", "Nama Part", "Qty"],
                 same_data or [["", "(tidak ada)", "", ""]], _HEAD_FILL, [6, 24, 60, 8])

    buf = io.BytesIO()
    wb.save(buf)
    kat_sfx = "" if kat == "SEMUA part" else "_" + re.sub(r"[^A-Za-z0-9]+", "", kat)[:20]
    fname = f"Perbandingan_{f1}_vs_{f2}{kat_sfx}.xlsx"
    return buf.getvalue(), fname


# ── Export GENERIK dari asisten (tool AI `buat_excel`) ──────────────────────
# Asisten menyusun judul+kolom+baris dari HASIL TOOL percakapan; payload disimpan
# in-memory ber-TTL di sini, frontend mengunduh via GET /api/ai/excel/{id}.
# Uvicorn berjalan 1 worker (lihat Dockerfile) → dict module-level aman.
_STASH_TTL_SEC = 24 * 3600.0
_STASH_MAX = 200               # plafon entri (jaga memori)
_stash_lock = threading.Lock()
_stash: dict[str, dict] = {}   # id -> {at, judul, kolom, baris, filename}

# Hasil build yang BERAT (katalog bergambar bisa >2 MB, PNG exploded) di-cache ke
# DISK, bukan RAM: 200 entri × beberapa MB akan menggerus memori server (3,8 GB,
# backend sudah ~1 GB krn torch/DINOv2). Disk punya puluhan GB nganggur.
# File dibuang saat entri kedaluwarsa/ditendang, dan seluruh folder dibersihkan
# saat proses start (sisa dari proses sebelumnya yang mati mendadak).
_CACHE_DIR = Path(tempfile.gettempdir()) / "maspart_export_cache"


def _cache_init() -> None:
    try:
        shutil.rmtree(_CACHE_DIR, ignore_errors=True)
        _CACHE_DIR.mkdir(parents=True, exist_ok=True)
    except OSError:
        pass  # disk read-only/penuh → cache dinonaktifkan, build ulang tiap unduh


_cache_init()


def _cache_write(export_id: str, data: bytes) -> Path | None:
    """Tulis bytes hasil build ke disk. Return path, atau None bila gagal."""
    p = _CACHE_DIR / export_id
    try:
        _CACHE_DIR.mkdir(parents=True, exist_ok=True)
        tmp = p.with_suffix(".part")   # tulis atomik: cegah pembaca lihat file separuh
        tmp.write_bytes(data)
        tmp.replace(p)
        return p
    except OSError:
        return None


def _cache_read(p: Path) -> bytes | None:
    try:
        return p.read_bytes()
    except OSError:
        return None


def _cache_drop(entry: dict) -> None:
    for key in ("_path", "_src_path"):
        p = entry.get(key)
        if p:
            try:
                Path(p).unlink(missing_ok=True)
            except OSError:
                pass

# Header kolom yang isinya kode/PN → font mono di Excel.
_MONO_HEAD_RE = re.compile(r"\b(part\s*number|part\s*no|pn|nomor\s*part|kode)\b", re.IGNORECASE)

# ── Sel ANGKA vs sel TEKS ────────────────────────────────────────────────────
# ⛔ Aturan pemilik (2026-07-20): kolom Harga/Stok/Qty/Berat di file unduhan
# WAJIB berisi ANGKA asli, bukan teks "Rp 1.500.000" — kalau teks, SUM() dan
# rumus agregat lain mengembalikan 0 ("rumusnya error"). Tampilan "Rp …" hanya
# untuk JAWABAN CHAT & blok RINGKASAN, tidak pernah untuk nilai sel.
#
# Koersi dilakukan DI SINI (satu choke point saat menulis) supaya kolom milik
# user sendiri — yang di-stringkan ai_sheet._finish_parse saat mem-parse
# unggahan — ikut kembali jadi angka, tanpa membongkar logika pencocokan PN
# yang mengandalkan bentuk string.

# Format TAMPILAN sel angka — hanya kosmetik, NILAI sel tetap int/float murni
# sehingga SUM/rumus user tetap jalan. "#,##0" = pemisah ribuan mengikuti locale
# Excel user (titik di Indonesia). Ganti ke None utk angka polos "1500000".
_NUM_FMT: str | None = "#,##0"

# Header yang mengandung kata angka TAPI isinya naratif → jangan pernah dikoersi.
_TEXT_HEAD_RE = re.compile(
    r"\b(status|keterangan|catatan|remark|nama|name|deskripsi|description|"
    r"per\s*gudang|rincian|pemenuhan|dimensi|cross|pengganti|satuan|merek|"
    r"lokasi|gudang)\b", re.IGNORECASE)
# Header yang isinya memang angka. Sejajar _HEAD_HARGA/_HEAD_STOK/_HEAD_QTY
# di ai_sheet.py (deteksi peran kolom unggahan user).
_NUM_HEAD_RE = re.compile(
    r"\b(harga|price|rp|idr|cny|amount|nilai|subtotal|total|"
    r"stok|stock|ready|sisa|on\s*hand|"
    r"qty|quantity|jumlah|jml|pcs|order|"
    r"berat|weight|kg)\b", re.IGNORECASE)
# 'total'/'subtotal' SENGAJA tidak di sini: "Stok Total" bukan kolom uang, dan
# "Total Harga" sudah tertangkap lewat kata 'harga'. Salah menandai kolom sbg
# uang hanya mematikan sabuk pengaman anti-PN — jadi default-nya non-uang.
_UANG_HEAD_RE = re.compile(r"\b(harga|price|rp|idr|cny|amount|nilai)\b", re.IGNORECASE)

# Format angka PINTAR per-kolom (2026-07-22): tampilan mengikuti JENIS kolom,
# nilai sel tetap int/float asli (SUM/rumus user tetap jalan). Urutan cek:
# persen → berat → uang(CNY/USD/Rp) → default.
_PERSEN_HEAD_RE = re.compile(r"(%|persen|terpakai)", re.IGNORECASE)
_BERAT_HEAD_RE = re.compile(r"\b(berat|weight|kg)\b", re.IGNORECASE)
# Dicek HANYA setelah _UANG_HEAD_RE cocok (sudah pasti kolom uang) → substring
# cukup; \b gagal untuk nama ber-underscore spt 'total_cny'.
_CNY_HEAD_RE = re.compile(r"cny|rmb|yuan|¥", re.IGNORECASE)
_USD_HEAD_RE = re.compile(r"usd|\$", re.IGNORECASE)
_FMT_DEFAULT = "#,##0"


def num_format(header: str) -> str:
    """Format tampilan sel angka untuk sebuah kolom, dari NAMA header-nya.
    Nilai sel TIDAK diubah — hanya number_format Excel (kosmetik)."""
    # underscore → spasi agar batas kata (\b) tembus di nama snake_case
    # ('total_cny' → 'total cny' → \bcny\b cocok).
    s = re.sub(r"_+", " ", header or "")
    if _PERSEN_HEAD_RE.search(s):
        # Nilai kita literal (mis. 87.5) → tempel '%' sbg literal, JANGAN pakai
        # '0.0%' (Excel mengalikan 100 → salah).
        return '0.0"%"'
    if _BERAT_HEAD_RE.search(s):
        return '#,##0" kg"'
    if _UANG_HEAD_RE.search(s):
        if _CNY_HEAD_RE.search(s):
            return '#,##0.00" CNY"'   # CNY kerap berdesimal (7804.81)
        if _USD_HEAD_RE.search(s):
            return '"$"#,##0.00'
        return '"Rp"#,##0'            # rupiah = mata uang utama app
    return _FMT_DEFAULT

_KOSONG_ANGKA = {"", "-", "—", "–", "n/a", "na", "tidak ada"}
_MATA_UANG_RE = re.compile(r"^(?:rp\.?|idr|cny|usd|\$|¥)\s*", re.IGNORECASE)
_RIBUAN_RE = re.compile(r"^-?\d{1,3}(?:\.\d{3})+$")      # 1.500.000
_POLOS_RE = re.compile(r"^-?\d+$")                        # 1500000
_DESIMAL_RE = re.compile(r"^-?\d+,\d{1,2}$")              # 1,25
_RIBUAN_DESIMAL_RE = re.compile(r"^-?\d{1,3}(?:\.\d{3})+,\d{1,2}$")   # 1.500,25
# Angka polos sepanjang ini di kolom NON-uang hampir pasti identitas (PN),
# bukan kuantitas. Ambang 9 disamakan dengan deteksi PN numerik di ai_sheet.
_PN_DIGIT_MIN = 9


def kolom_angka(kolom: list[str]) -> dict[int, bool]:
    """{indeks_kolom_0based: apakah kolom UANG} untuk kolom yang boleh dikoersi.

    Deny menang mutlak: kolom identitas (PN/kode) dan kolom naratif
    (status/keterangan/'Stok per Gudang') tak pernah masuk, walau namanya
    mengandung kata angka.
    """
    out: dict[int, bool] = {}
    for i, h in enumerate(kolom):
        s = h or ""
        if _MONO_HEAD_RE.search(s) or _TEXT_HEAD_RE.search(s):
            continue
        if _NUM_HEAD_RE.search(s):
            out[i] = bool(_UANG_HEAD_RE.search(s))
        elif _PERSEN_HEAD_RE.search(s):
            out[i] = False        # persen = angka (non-uang) → koersi "87,5"→87.5
    return out


def ke_angka(v, uang: bool = False):
    """Nilai sel → int/float bila JELAS angka; selain itu dikembalikan apa adanya.

    Sengaja konservatif — ragu berarti biarkan teks. Yang TIDAK dikoersi:
    penanda kosong ('—', 'N/A'), leading zero ('0012' = identitas), persen
    (Excel mengubah maknanya), format ambigu ('1.50', '12 pcs'), dan angka
    panjang di kolom non-uang (PN seperti '1013133963').
    """
    if v is None or isinstance(v, (int, float, bool)):
        return v                      # idempoten; bool jangan disentuh
    if not isinstance(v, str):
        return v
    s = v.strip()
    if s.lower() in _KOSONG_ANGKA:
        return v
    inti = _MATA_UANG_RE.sub("", s).strip()
    if not inti:
        return v
    neg = inti.startswith("-")
    telanjang = inti[1:] if neg else inti
    if telanjang.startswith("0") and len(telanjang) > 1 and telanjang[1] not in ".,":
        return v                      # leading zero = semantik identitas
    try:
        if _POLOS_RE.match(inti):
            if not uang and len(telanjang) >= _PN_DIGIT_MIN:
                return v              # sabuk pengaman anti-PN
            return int(inti)
        if _RIBUAN_RE.match(inti):
            return int(inti.replace(".", ""))
        if _DESIMAL_RE.match(inti):
            return float(inti.replace(",", "."))
        if _RIBUAN_DESIMAL_RE.match(inti):
            return float(inti.replace(".", "").replace(",", "."))
    except ValueError:
        return v
    return v


def nama_file(judul: str, ext: str = "xlsx") -> str:
    """Judul → nama file unduhan (slug aman, ekstensi mengikuti file SUMBER)."""
    slug = re.sub(r"[^A-Za-z0-9]+", "_", judul).strip("_")[:60] or "Data_MASPART"
    return f"{slug}.{ext}"


def _stash_put(entry: dict, judul: str, ext: str = "xlsx") -> tuple[str, str]:
    now = time.monotonic()
    filename = nama_file(judul, ext)
    export_id = uuid.uuid4().hex
    with _stash_lock:
        for k in [k for k, v in _stash.items() if now - v["at"] > _STASH_TTL_SEC]:
            _cache_drop(_stash.pop(k, None) or {})
        while len(_stash) >= _STASH_MAX:   # buang paling tua
            _cache_drop(_stash.pop(min(_stash, key=lambda k: _stash[k]["at"]), None) or {})
        _stash[export_id] = {"at": now, "judul": judul, "filename": filename, **entry}
    return export_id, filename


def stash_export(judul: str, kolom: list[str], baris: list[list[str]]) -> tuple[str, str]:
    """Simpan payload export tabel polos → (export_id, filename)."""
    return _stash_put({"kolom": kolom, "baris": baris}, judul)


def stash_builder(judul: str, builder: dict, ext: str = "xlsx",
                  src: bytes | None = None) -> tuple[str, str]:
    """Simpan RESEP export yang dibangun SAAT DIUNDUH (mis. katalog bergambar —
    berat, jadi dibangun ketika kartu diklik lalu bytes-nya di-cache di entri).
    `ext` = ekstensi file hasil (xlsx/pdf) sesuai format yang dipilih user.

    `src` = workbook DASAR yang akan ditempeli gambar (file ASLI user yang sudah
    diisi kolom datanya). Ditulis ke cache DISK & umurnya mengikuti entri ini,
    jadi kartu unduh tetap sah walau lampiran chat-nya sudah kedaluwarsa."""
    export_id, filename = _stash_put({"builder": builder}, judul, ext=ext)
    if src:
        p = _cache_write(export_id + "_src", src)
        if p:
            with _stash_lock:
                e = _stash.get(export_id)
                if e is not None:
                    e["_src_path"] = str(p)
    return export_id, filename


def stash_raw(judul: str, data: bytes, filename: str) -> tuple[str, str]:
    """Simpan BYTES jadi (mis. PDF penawaran dari Accurate) → (export_id, filename).
    Bytes langsung ditulis ke cache disk; generic_excel menyajikannya apa adanya."""
    ext = filename.rsplit(".", 1)[-1] if "." in filename else "pdf"
    export_id, _fn = _stash_put({"raw": True}, judul, ext=ext)
    with _stash_lock:
        _stash[export_id]["filename"] = filename   # pertahankan nama asli
    p = _cache_write(export_id, data)
    if p:
        with _stash_lock:
            _stash[export_id]["_path"] = str(p)
    return export_id, filename


def generic_excel(export_id: str) -> tuple[bytes | None, str]:
    """Bangun workbook dari payload tersimpan. (bytes, filename) atau (None, pesan)."""
    with _stash_lock:
        d = _stash.get(export_id)
        if d and time.monotonic() - d["at"] > _STASH_TTL_SEC:
            _cache_drop(_stash.pop(export_id, None) or {})
            d = None
    if not d:
        return None, ("File sudah kedaluwarsa / tidak ditemukan. Minta asisten "
                      "buatkan Excel-nya lagi.")

    # Entri ber-'raw' = bytes jadi (PDF penawaran) sudah di cache disk.
    if d.get("raw"):
        p = d.get("_path")
        data = _cache_read(Path(p)) if p else None
        if data:
            return data, d["filename"]
        return None, "File sudah kedaluwarsa — minta asisten buatkan lagi."

    # Entri ber-'builder' = dibangun saat diunduh (berat); hasilnya di-cache ke
    # DISK supaya klik berikutnya instan tanpa menahan MB di RAM.
    if d.get("builder"):
        cached_path = d.get("_path")
        if cached_path:
            cached = _cache_read(Path(cached_path))
            if cached:
                return cached, d["filename"]
            with _stash_lock:      # file cache hilang → bangun ulang
                d.pop("_path", None)
        b = d["builder"]
        if b.get("kind") in ("katalog", "katalog_mesin"):
            src = "weichai" if b.get("kind") == "katalog_mesin" else "sinotruk"
            fmt = (b.get("fmt") or "excel").lower()
            ish = bool(b.get("isi_stok_harga"))  # hanya True bila admin minta (di-set saat tool dibuat)
            if fmt == "pdf":
                data, err = katalog_pdf(b.get("rangka", ""), b.get("kategori", ""), src, isi_stok_harga=ish)
            else:
                data, err = katalog_excel(b.get("rangka", ""), b.get("kategori", ""), src, isi_stok_harga=ish)
            if data is None:
                return None, err
            p = _cache_write(export_id, data)
            if p:
                with _stash_lock:
                    d["_path"] = str(p)
            return data, d["filename"]
        # Lampiran user yang DIISI DI TEMPAT lalu ditempeli gambar: dasarnya file
        # ASLI user (sudah berisi kolom data), disimpan di cache disk entri ini.
        if b.get("kind") == "sheet_gambar_isi":
            sp = d.get("_src_path")
            dasar = _cache_read(Path(sp)) if sp else None
            if not dasar:
                return None, ("File dasar sudah kedaluwarsa — minta asisten mengisi "
                              "Excel-nya sekali lagi.")
            data, err = gambar_di_tempat(dasar, b)
            if data is None:
                return None, err
            p = _cache_write(export_id, data)
            if p:
                with _stash_lock:
                    d["_path"] = str(p)
            return data, d["filename"]
        # SATU builder untuk semua Excel olahan lampiran (data + warna + rekap +
        # foto + gambar teknis). `kind` lama tetap dilayani: payload yang sudah
        # terlanjur ada di stash tak boleh mati.
        if b.get("kind") in ("sheet_gambar", "sheet_foto", "sheet_exploded", "sheet_status"):
            data, err = sheet_gambar_excel(b)
            if data is None:
                return None, err
            p = _cache_write(export_id, data)
            if p:
                with _stash_lock:
                    d["_path"] = str(p)
            return data, d["filename"]
        if b.get("kind") == "exploded":
            if b.get("source") == "weichai":
                data = exploded_png_weichai(b.get("svg", ""), b.get("rangka", ""), b.get("balon"))
            elif b.get("source") == "shantui":
                data = exploded_png_shantui(b.get("svg", ""), b.get("balon"))
            else:
                data = exploded_png(b.get("svg", ""), b.get("balon"))
            if data is None:
                return None, ("Gagal mengambil/menrender gambar exploded view EPC "
                              "(file gambar tak tersedia / resvg gagal).")
            p = _cache_write(export_id, data)
            if p:
                with _stash_lock:
                    d["_path"] = str(p)
            return data, d["filename"]
        return None, "Jenis export tidak dikenal."

    kolom: list[str] = d["kolom"]
    baris: list[list[str]] = d["baris"]
    mono_cols = {j for j, h in enumerate(kolom, start=1) if _MONO_HEAD_RE.search(h or "")}
    # Lebar kolom mengikuti isi (min 8, maks 60).
    widths = []
    for j, h in enumerate(kolom):
        # str() WAJIB: sel bisa berisi int/float (harga & stok kini numerik agar
        # rumus Excel user jalan) — len(int) melempar TypeError.
        w = max([len(h or "")] + [len(str(r[j])) for r in baris if j < len(r)] or [0])
        widths.append(max(8, min(60, w + 4)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.sheet_view.showGridLines = False
    start = _title(ws, _safe(d["judul"]),
                   f"{len(baris)} baris · MASPART Asisten AI", max(2, len(kolom)))
    for j, (h, w) in enumerate(zip(kolom, widths), start=1):
        c = ws.cell(row=start, column=j, value=_safe(h))
        c.fill = _HEAD_FILL
        c.font = _WHITE
        c.alignment = _CENTER
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(j)].width = w
    # Kolom Harga/Stok/Qty/Berat ditulis sebagai ANGKA (lihat ke_angka) supaya
    # rumus Excel user jalan. _safe() dipanggil SETELAH koersi: string yang tidak
    # jadi angka tetap dapat guard formula-injection, angka dilewatkan utuh.
    # Format tampilan angka PINTAR per-kolom (persen/berat/uang) dari nama header.
    num_cols = kolom_angka(kolom)
    col_fmts = [num_format(h) for h in kolom]
    r = start + 1
    for i, row in enumerate(baris):
        for j in range(1, len(kolom) + 1):
            val = row[j - 1] if j - 1 < len(row) else ""
            if j - 1 in num_cols:
                val = ke_angka(val, num_cols[j - 1])
            c = ws.cell(row=r, column=j, value=_safe(val))
            c.border = _BORDER
            angka = isinstance(c.value, (int, float)) and not isinstance(c.value, bool)
            if angka:
                c.alignment = _RIGHT
                c.number_format = col_fmts[j - 1]
            else:
                c.alignment = _CENTER if j == 1 or j in mono_cols else _LEFT
            c.font = _MONO if j in mono_cols else _INK
            if i % 2:
                c.fill = _ZEBRA
        r += 1
    ws.freeze_panes = ws.cell(row=start + 1, column=1)

    return _save_stable(wb), d["filename"]


# ── ISI DI TEMPAT: tulis hasil ke SALINAN FILE ASLI user ────────────────────
# ⛔ Aturan pemilik 2026-08-25: file Excel yang dikirim user JANGAN diubah
# formatnya — asisten hanya MENGISI data (atau membacanya). Semua jalur pengisi
# lampiran (`ai_sheet.fill_columns`, sheet_isi_part_number, sheet_cek_qty) lewat
# sini; workbook dibuka apa adanya lalu HANYA sel yang benar-benar diisi yang
# disentuh. Yang otomatis ikut selamat karena tak pernah dibangun ulang: baris
# kop/judul di atas header, baris kosong pemisah, rumus, warna/border/font,
# lebar kolom, merge, freeze pane, filter, sheet lain, gambar & logo, serta
# kolom di luar 40 kolom yang dibaca parser.
_SHEET_REKAP = "Ringkasan MASPART"


def _modified_asli(src: bytes) -> bytes:
    """Nilai `dcterms:modified` milik file ASLI (utk dipasang balik setelah save —
    openpyxl menimpanya dengan jam simpan). Fallback: konstanta byte-stabil."""
    try:
        with zipfile.ZipFile(io.BytesIO(src)) as z:
            m = _MODIFIED_RE.search(z.read(_CORE_XML))
        if m:
            inti = m.group(0)[len(m.group(1)):-len(m.group(2))]
            if inti:
                return inti
    except Exception:
        pass
    return _MODIFIED_TETAP


def _save_asli(wb, src: bytes) -> bytes:
    """Simpan workbook TURUNAN file user: properti dokumen (dibuat/diubah) tetap
    milik user, tapi bytes tetap deterministik (aman cache & test)."""
    buf = io.BytesIO()
    wb.save(buf)
    try:
        return _zip_stabil(buf.getvalue(), modified=_modified_asli(src))
    except Exception:  # pragma: no cover — file tetap sah walau tak ternormalisasi
        return buf.getvalue()


def _buka_asli(src: bytes, xlsm: bool = False):
    """Buka workbook user untuk DITULISI. data_only=False → RUMUS milik user tetap
    utuh (nilai cache-nya hilang, karena itu Excel disuruh hitung ulang saat dibuka)."""
    wb = load_workbook(io.BytesIO(src), keep_vba=xlsm)
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:  # pragma: no cover
        pass
    return wb


def _sel_ada(ws, row: int, col: int):
    """Sel yang MEMANG ADA di sheet, tanpa membuatnya. `ws.cell()` biasa menciptakan
    sel kosong — dipakai berulang atas file 5.000 baris, itu menggelembungkan file
    user dengan ratusan ribu sel hampa."""
    return getattr(ws, "_cells", {}).get((row, col))


def _kolom_terakhir(ws, hdr_row: int, rows: list[int]) -> int:
    """Kolom TERAKHIR yang benar-benar BERISI NILAI (bukan sekadar berformat) pada
    baris header + baris data. Kolom baru ditaruh setelahnya, jadi kolom user yang
    tak terbaca parser (tanpa header, atau kolom ke-41+) tak pernah tertimpa."""
    cells = getattr(ws, "_cells", None)
    if cells is None:                                  # pragma: no cover
        return ws.max_column or 0
    minat = {hdr_row, *rows}
    akhir = 0
    for (r, c), sel in cells.items():
        if c <= akhir or r not in minat:
            continue
        v = sel.value
        if v is not None and str(v).strip() != "":
            akhir = c
    return akhir


def _bisa_tulis(cell) -> bool:
    """Sel gabungan (merge) selain sel jangkarnya TIDAK bisa ditulis openpyxl."""
    from openpyxl.cell.cell import MergedCell
    return cell is not None and not isinstance(cell, MergedCell)


def isi_di_tempat(src: bytes, rencana: dict, status: list[str] | None = None,
                  ringkasan: list | None = None,
                  xlsm: bool = False) -> tuple[bytes | None, dict, dict, str]:
    """Terapkan isian ke SALINAN file ASLI user. Return (bytes, peta_kolom, lapor, error).

    `rencana` (dibuat `ai_sheet._rencana_isi`):
      sheet      nama sheet aktif saat file diparse
      hdr_row    baris header (1-based, nomor baris ASLI di sheet)
      row_map    [nomor baris asli] per baris `_body`
      ncol       jumlah kolom yang dibaca parser (kolom 1..ncol)
      headers    header hasil parse (untuk VERIFIKASI file masih file yang sama)
      kolom_baru [[idx_kolom_body, judul]] kolom yang ditambahkan di kanan
      isian      [[idx_baris_body, idx_kolom_body, nilai]] HANYA sel yang berubah
      fmt        {idx_kolom_body: number_format} untuk kolom BARU

    Aturan tulis (menjaga milik user):
      • kolom user + sel berisi RUMUS   → dilewati (rumus tak dirusak);
      • kolom user + hasil KOSONG       → dilewati (nilai lama user dibiarkan);
      • sel merge non-jangkar           → dilewati;
      • sisanya ditulis; number_format & gaya sel user TIDAK disentuh.
    """
    hdr_row = int(rencana.get("hdr_row") or 0)
    row_map = list(rencana.get("row_map") or [])
    ncol = int(rencana.get("ncol") or 0)
    if not hdr_row or not row_map or not ncol:
        return None, {}, {}, "peta baris/kolom file asli tidak tersedia"
    try:
        wb = _buka_asli(src, xlsm=xlsm)
    except Exception as e:
        return None, {}, {}, f"file asli tak bisa dibuka untuk ditulisi ({e})"

    try:
        nama = rencana.get("sheet") or ""
        ws = wb[nama] if nama in wb.sheetnames else wb.active

        # VERIFIKASI: baris header di file masih sama dengan yang diparse. Bila
        # tidak, peta baris/kolom tak bisa dipercaya → lebih baik gagal & jatuh ke
        # jalur cadangan daripada menulis ke sel yang salah.
        for j, h in enumerate(rencana.get("headers") or []):
            if j >= ncol or re.fullmatch(r"Kolom \d+", h or ""):
                continue
            sel = ws.cell(row=hdr_row, column=j + 1).value
            if str(sel if sel is not None else "").strip() != h:
                return None, {}, {}, "baris header file asli tak cocok dengan hasil baca"

        # Kolom BARU ditaruh setelah kolom terakhir yang benar-benar terpakai.
        mulai_baru = max(ncol, _kolom_terakhir(ws, hdr_row, row_map)) + 1
        peta_kolom = {j: (j + 1 if j < ncol else mulai_baru + (j - ncol))
                      for j in range(ncol + len(rencana.get("kolom_baru") or []))}

        # Header kolom baru — gaya menyontek sel header TERAKHIR milik user supaya
        # tabelnya tetap terlihat menyatu (sel user sendiri tak disentuh).
        contoh_hdr = _sel_ada(ws, hdr_row, ncol)
        for j, judul in (rencana.get("kolom_baru") or []):
            c = ws.cell(row=hdr_row, column=peta_kolom[j])
            if not _bisa_tulis(c):
                continue
            if _bisa_tulis(contoh_hdr):
                c._style = copy(contoh_hdr._style)
            c.value = _safe(judul)
            c.number_format = "General"
            lebar = max(10, min(60, len(str(judul or "")) + 4))
            ws.column_dimensions[get_column_letter(peta_kolom[j])].width = lebar

        fmt = {int(k): v for k, v in (rencana.get("fmt") or {}).items()}
        lapor = {"sel_diisi": 0, "sel_dilewati_rumus": 0,
                 "sel_dilewati_sudah_terisi": 0, "sel_dilewati_merge": 0}
        for i, j, val in (rencana.get("isian") or []):
            if i >= len(row_map) or j not in peta_kolom:
                continue
            r, c = row_map[i], peta_kolom[j]
            sel = ws.cell(row=r, column=c)
            if not _bisa_tulis(sel):
                lapor["sel_dilewati_merge"] += 1
                continue
            if j < ncol:                       # ── kolom MILIK user: hati-hati ──
                lama = sel.value
                if isinstance(lama, str) and lama.startswith("="):
                    lapor["sel_dilewati_rumus"] += 1
                    continue
                if (val is None or val == "") and lama is not None and str(lama).strip() != "":
                    lapor["sel_dilewati_sudah_terisi"] += 1
                    continue
            else:                              # ── kolom BARU: warisi gaya baris ──
                tetangga = _sel_ada(ws, r, ncol)
                if _bisa_tulis(tetangga):
                    sel._style = copy(tetangga._style)
            sel.value = _safe(val)
            if j >= ncol:
                angka = isinstance(sel.value, (int, float)) and not isinstance(sel.value, bool)
                sel.number_format = fmt.get(j, _FMT_DEFAULT) if angka else "General"
            lapor["sel_diisi"] += 1

        # WARNA status HANYA di kolom yang kita tambahkan — sel milik user tak
        # pernah dicat ulang. Statusnya tetap terbaca karena dwi-encode (ada kolom
        # teks "Status" di sebelahnya).
        kol_baru = [peta_kolom[j] for j, _h in (rencana.get("kolom_baru") or [])]
        for i, warna in enumerate(status or []):
            fill = _STATUS_FILL.get(warna or "")
            if not fill or i >= len(row_map):
                continue
            for c in kol_baru:
                sel = ws.cell(row=row_map[i], column=c)
                if _bisa_tulis(sel):
                    sel.fill = fill

        # REKAP ditaruh di SHEET BARU, bukan disisipkan di bawah data user (di sana
        # kerap sudah ada baris TOTAL / catatan miliknya sendiri).
        if ringkasan:
            _tulis_rekap(wb, ringkasan)

        lapor["sheet"] = ws.title
        lapor["kolom_ditambah"] = [h for _j, h in (rencana.get("kolom_baru") or [])]
        return _save_asli(wb, src), peta_kolom, lapor, ""
    except Exception as e:  # pragma: no cover — jaring terakhir, ada jalur cadangan
        return None, {}, {}, f"gagal menulis ke file asli ({e})"
    finally:
        try:
            wb.close()
        except Exception:  # pragma: no cover
            pass


def _tulis_rekap(wb, ringkasan: list) -> None:
    """Blok RINGKASAN → sheet TERSENDIRI (sheet user tak ditambahi baris)."""
    nama, n = _SHEET_REKAP, 2
    while nama in wb.sheetnames:
        nama, n = f"{_SHEET_REKAP} {n}", n + 1
    ws = wb.create_sheet(nama)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 30
    hc = ws.cell(row=1, column=1, value="RINGKASAN")
    hc.font = Font(bold=True, color=_BRAND_DK, size=12)
    r = 2
    for item in ringkasan:
        label, nilai = item[0], item[1]
        warna = item[2] if len(item) > 2 else ""
        lc = ws.cell(row=r, column=1, value=_safe(str(label)))
        lc.font = _BOLD
        lc.fill = _STATUS_FILL.get(warna, _SUB1_FILL)
        lc.border = _BORDER
        lc.alignment = _LEFT
        vc = ws.cell(row=r, column=2, value=_safe(nilai))
        vc.font = _INK
        vc.border = _BORDER
        vc.alignment = _LEFT
        r += 1


def sheet_status_excel(b: dict) -> tuple[bytes | None, str]:
    """Excel olahan lampiran BERWARNA STATUS + blok RINGKASAN, tanpa gambar —
    pembungkus `sheet_gambar_excel`. Dipertahankan untuk payload lama di stash
    (kind 'sheet_status') & pemanggil yang memang tak butuh gambar."""
    if not (b.get("kolom") or []):
        return None, "Payload sheet_status kosong."
    return sheet_gambar_excel(b)


# ── Excel UNGGAHAN USER + FOTO part (tool `sheet_isi_foto`) ─────────────────
# Foto SIMS berukuran besar (ada yang 6000 px / >10 MB) → WAJIB diciutkan sebelum
# ditempel, kalau tidak workbook bisa ratusan MB & server (3,8 GB RAM) tumbang.
_FOTO_H_PX = 105               # tinggi foto di sel
_FOTO_COL_W = 17               # lebar kolom foto
_FOTO_MAX_TOTAL = 700          # plafon gambar per file (300 PN × 2 + sisa)
_FOTO_TIMEOUT = 25


def _foto_thumb(url: str) -> bytes | None:
    """Unduh 1 foto SIMS → JPEG kecil (tinggi _FOTO_H_PX). None bila gagal."""
    try:
        import requests
        from PIL import Image as PILImage
    except Exception:
        return None
    try:
        r = requests.get(url, timeout=_FOTO_TIMEOUT)
        if r.status_code != 200 or not r.content:
            return None
        pil = PILImage.open(io.BytesIO(r.content))
        if pil.mode not in ("RGB", "L"):
            pil = pil.convert("RGB")
        w, h = pil.size
        if h > _FOTO_H_PX:
            w, h = max(1, int(w * _FOTO_H_PX / h)), _FOTO_H_PX
            pil = pil.resize((w, h), PILImage.LANCZOS)
        buf = io.BytesIO()
        pil.save(buf, format="JPEG", quality=82, optimize=True)
        return buf.getvalue()
    except Exception:
        return None


def sheet_foto_excel(b: dict) -> tuple[bytes | None, str]:
    """Excel unggahan user + FOTO part saja — pembungkus `sheet_gambar_excel`.
    Dipertahankan untuk payload lama yang masih ada di stash (kind 'sheet_foto')."""
    return sheet_gambar_excel(b)


# ── Excel UNGGAHAN USER + GAMBAR TEKNIS exploded view (tool `sheet_isi_exploded`) ──
# Ukuran tampil & lebar kolom SENGAJA sama dengan kolom "Exploded View" Batch
# Download (services/catalog.py): balon harus terbaca dan file harus tetap bisa
# dicetak. PNG-nya TIDAK di-resample — hanya ukuran TAMPIL yang diatur, jadi zoom
# di Excel tetap tajam.
_EXPL_IMG_W = 620          # px lebar tampil maks
_EXPL_IMG_H = 500          # px tinggi tampil maks
_EXPL_COL_W = 88           # satuan lebar openpyxl (±7 px/satuan, Calibri 11)
_EXPL_INFO_COL_W = 46      # kolom teks "Info Gambar Teknis"
_EXPL_ROW_PT_MAX = 409     # batas KERAS tinggi baris Excel = 409,5 pt (±546 px)
# Anggaran waktu SATU file (lebih pendek dari batch download 15 menit): unduhan
# kartu chat menunggu di satu request HTTP. PN yang tak terjangkau ditandai jujur.
_EXPL_ANGGARAN = 600


def _info_exploded_sel(d: dict, rangka: str) -> str:
    """Isi sel "Info Gambar Teknis". Sumber teks = catalog._teks_info_exploded
    (satu sumber dengan Batch Download) + baris asal gambar untuk jalur per-VIN.
    ⛔ Tak ditemukan/gagal → tulis ALASANNYA apa adanya, jangan karang figure."""
    from . import catalog          # impor lokal: hindari siklus impor saat modul dimuat
    teks = catalog._teks_info_exploded(d)
    if rangka and d and d.get("found"):
        frame = d.get("frame_number") or rangka
        return f"Sumber: figure unit {frame} (per-VIN)\n{teks}"
    return teks


def sheet_exploded_excel(b: dict) -> tuple[bytes | None, str]:
    """Excel unggahan user + GAMBAR TEKNIS saja — pembungkus `sheet_gambar_excel`.
    Dipertahankan untuk payload lama di stash (kind 'sheet_exploded')."""
    return sheet_gambar_excel(b)


def _ambil_gambar(foto: list, n_kol_foto: int, pns: list, rangka: str,
                  ada_expl: bool) -> tuple[dict, dict, str]:
    """Ambil bahan GAMBAR sekali untuk kedua jalur (file baru & isi-di-tempat):
    (a) foto SIMS diunduh paralel lalu diciutkan, (b) exploded EPC lewat gerbang
    satu-batch (server 1 vCPU & EPC mudah menolak). Return (thumb, peta, error)."""
    from . import exploded_view       # impor lokal: hindari siklus saat modul dimuat

    thumb: dict[tuple[int, int], bytes] = {}
    if n_kol_foto:
        tugas: list[tuple[int, int, str]] = []   # (baris_i, urutan_foto, url)
        for i, urls in enumerate(foto or []):
            for k, u in enumerate((urls or [])[:n_kol_foto]):
                if u and len(tugas) < _FOTO_MAX_TOTAL:
                    tugas.append((i, k, u))
        if tugas:
            with ThreadPoolExecutor(max_workers=8) as ex:
                for (i, k, _u), png in zip(tugas, ex.map(lambda t: _foto_thumb(t[2]), tugas)):
                    if png:
                        thumb[(i, k)] = png

    peta: dict[str, dict] = {}
    if ada_expl:
        unik = [p for p in dict.fromkeys(pns or []) if p]
        try:
            peta = exploded_view.bangun_dengan_gerbang(
                exploded_view.png_batch, unik, anggaran=_EXPL_ANGGARAN, rangka=rangka)
        except exploded_view.SedangSibuk:
            return {}, {}, ("Server sedang menyusun satu batch gambar exploded lain (hanya "
                            "boleh satu sekaligus — server 1 vCPU dan server EPC mudah "
                            "menolak). Coba klik unduh lagi beberapa menit lagi.")
        except Exception:
            return {}, {}, "Gagal mengambil gambar exploded dari EPC. Coba lagi sebentar."
    return thumb, peta, ""


def gambar_di_tempat(src: bytes, b: dict) -> tuple[bytes | None, str]:
    """Tempel FOTO & GAMBAR TEKNIS ke SALINAN FILE ASLI user (yang kolom datanya
    sudah diisi `isi_di_tempat`). Format file user tetap: yang berubah hanya sel
    gambar di kolom BARU, lebar kolom gambar itu, dan tinggi baris yang memang
    kebagian gambar (kalau tidak, gambarnya terpotong).

    Dipanggil SAAT KARTU DIUNDUH — satu PN exploded dingin bisa puluhan detik."""
    from openpyxl.drawing.image import Image as XLImage

    from . import exploded_view

    row_map = list(b.get("row_map") or [])
    kol = {int(k): int(v) for k, v in (b.get("kol") or {}).items()}
    kol_foto = [kol[j] for j in (b.get("kol_foto") or []) if j in kol]
    pns: list[str] = b.get("pns") or []
    rangka = (b.get("rangka") or "").strip()
    j_info, j_gambar = b.get("kol_info"), b.get("kol_gambar")
    kol_info = kol.get(j_info) if j_info is not None else None
    kol_gambar = kol.get(j_gambar) if j_gambar is not None else None
    ada_expl = kol_gambar is not None

    thumb, peta, err = _ambil_gambar(b.get("foto") or [], len(kol_foto), pns,
                                     rangka, ada_expl)
    if err:
        return None, err
    try:
        wb = _buka_asli(src, xlsm=bool(b.get("xlsm")))
    except Exception as e:  # pragma: no cover
        return None, f"File hasil tak bisa dibuka untuk ditempeli gambar ({e})."
    try:
        nama = b.get("sheet") or ""
        ws = wb[nama] if nama in wb.sheetnames else wb.active
        for c in kol_foto:
            ws.column_dimensions[get_column_letter(c)].width = _FOTO_COL_W
        if kol_gambar:
            ws.column_dimensions[get_column_letter(kol_gambar)].width = _EXPL_COL_W
        if kol_info:
            ws.column_dimensions[get_column_letter(kol_info)].width = _EXPL_INFO_COL_W

        for i, r in enumerate(row_map):
            pn = pns[i] if i < len(pns) else ""
            d = peta.get(exploded_view.kunci(pn, rangka)) if (ada_expl and pn) else None
            if kol_info:
                sel = ws.cell(row=r, column=kol_info)
                if _bisa_tulis(sel):
                    sel.value = _safe(_info_exploded_sel(d, rangka))
                    sel.alignment = _LEFT
            tinggi = 0.0
            for k, c in enumerate(kol_foto):
                png = thumb.get((i, k))
                if not png:
                    continue
                img = XLImage(io.BytesIO(png))
                ratio = img.width / img.height if img.height else 1
                img.height = _FOTO_H_PX
                img.width = int(_FOTO_H_PX * ratio)
                ws.add_image(img, f"{get_column_letter(c)}{r}")
                tinggi = max(tinggi, _FOTO_H_PX * 0.78)
            png = (d or {}).get("png")
            if png and kol_gambar:
                img = XLImage(io.BytesIO(png))
                w, h = int(img.width or 1), int(img.height or 1)
                skala = min(_EXPL_IMG_W / w, _EXPL_IMG_H / h, 1.0)
                img.width, img.height = max(1, int(w * skala)), max(1, int(h * skala))
                ws.add_image(img, f"{get_column_letter(kol_gambar)}{r}")
                tinggi = max(tinggi, min(_EXPL_ROW_PT_MAX, img.height * 0.78 + 8))
            if tinggi:
                ws.row_dimensions[r].height = tinggi
        return _save_asli(wb, src), ""
    except Exception as e:  # pragma: no cover
        return None, f"Gagal menempelkan gambar ke file ({e})."
    finally:
        try:
            wb.close()
        except Exception:  # pragma: no cover
            pass


def sheet_gambar_excel(b: dict) -> tuple[bytes | None, str]:
    """SATU builder untuk SEMUA Excel olahan lampiran user: kolom data (stok/harga/
    dst, lengkap dengan WARNA status & blok RINGKASAN) + FOTO fisik part (SIMS) +
    GAMBAR TEKNIS exploded view (EPC) — semuanya boleh berada di file yang SAMA.

    Dulu ada tiga builder terpisah, dan itulah sebab keluhan pemilik 2026-08-06:
    'isikan stok Jakarta & Pekanbaru, foto, dan exploded' menghasilkan beberapa
    kartu unduh karena tiap jalur menstash filenya sendiri. Satu builder = satu file.

    Dipanggil saat kartu unduh diklik (lihat generic_excel), BUKAN saat tool
    dijalankan: satu PN exploded dingin bisa makan puluhan detik & foto SIMS bisa
    >10 MB — menahannya di giliran chat akan membuat asisten tampak menggantung.

    Payload (bagian yang tak dipakai boleh absen):
      {kind:"sheet_gambar", judul, sub?, kolom, baris,
       status:[per-baris ""|"hijau"|"merah"|"kuning"],             ← WARNA baris
       ringkasan:[(label, nilai, warna)],                          ← blok REKAP
       foto:[per-baris [url,…]], kol_foto:[idx,…],                 ← bagian FOTO
       pns:[per-baris], rangka:""|VIN, kol_info, kol_gambar}       ← bagian EXPLODED
    Return (bytes, pesan_error)."""
    from openpyxl.drawing.image import Image as XLImage

    from . import exploded_view      # impor lokal: hindari siklus saat modul dimuat

    kolom: list[str] = b.get("kolom") or []
    baris: list[list] = b.get("baris") or []
    if not kolom:
        return None, "Data sheet tidak ditemukan — minta asisten membuat ulang."
    status: list[str] = b.get("status") or []
    ringkasan: list = b.get("ringkasan") or []
    foto: list[list[str]] = b.get("foto") or []
    kol_foto: list[int] = b.get("kol_foto") or []
    pns: list[str] = b.get("pns") or []
    rangka = (b.get("rangka") or "").strip()
    kol_info = b.get("kol_info")
    kol_gambar = b.get("kol_gambar")
    ada_foto = bool(kol_foto)
    ada_expl = kol_gambar is not None

    thumb, peta, err = _ambil_gambar(foto, len(kol_foto), pns, rangka, ada_expl)
    if err:
        return None, err

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.sheet_view.showGridLines = False
    bagian = []
    if ada_foto:
        bagian.append("foto resmi SIMS (dicocokkan per Part Number)")
    if ada_expl:
        bagian.append(f"gambar teknis EPC per-VIN (unit {rangka})" if rangka
                      else "gambar teknis EPC LINTAS MODEL (bukan milik unit tertentu)")
    judul_default = ("Data + Foto" if ada_foto and not ada_expl else
                     "Data + Gambar Teknis" if ada_expl and not ada_foto else
                     "Data + Foto & Gambar Teknis" if ada_foto and ada_expl else
                     "Data MASPART")
    sub = b.get("sub") or (f"{len(baris)} baris"
                           + ("" if not bagian else " · " + " · ".join(bagian))
                           + " · MASPART Asisten AI")
    start = _title(ws, _safe(b.get("judul") or judul_default), sub, max(2, len(kolom)))

    mono_cols = {j for j, h in enumerate(kolom, start=1) if _MONO_HEAD_RE.search(h or "")}
    foto_cols = {j + 1 for j in kol_foto}      # 1-based
    for j, h in enumerate(kolom, start=1):
        c = ws.cell(row=start, column=j, value=_safe(h))
        c.fill = _HEAD_FILL
        c.font = _WHITE
        c.alignment = _CENTER
        c.border = _BORDER
        if j in foto_cols:
            ws.column_dimensions[get_column_letter(j)].width = _FOTO_COL_W
        elif ada_expl and j - 1 == kol_gambar:
            ws.column_dimensions[get_column_letter(j)].width = _EXPL_COL_W
        elif ada_expl and j - 1 == kol_info:
            ws.column_dimensions[get_column_letter(j)].width = _EXPL_INFO_COL_W
        else:
            w = max([len(str(h or ""))]
                    + [len(str(r[j - 1])) for r in baris if j - 1 < len(r)] or [0])
            ws.column_dimensions[get_column_letter(j)].width = max(8, min(60, w + 4))

    # Kolom Harga/Stok/Qty milik FILE USER tetap ditulis sebagai ANGKA (rumus
    # user harus tetap jalan) — sama seperti generic_excel; kolom foto/info/gambar
    # tak pernah kena koersi (nama headernya bukan kolom angka).
    num_cols = kolom_angka(kolom)
    col_fmts = [num_format(h) for h in kolom]
    r = start + 1
    for i, row in enumerate(baris):
        pn = pns[i] if i < len(pns) else ""
        d = peta.get(exploded_view.kunci(pn, rangka)) if (ada_expl and pn) else None
        # WARNA status menang atas zebra (warna MENYERTAI kolom teks 'Status' —
        # dwi-encode, warna bukan satu-satunya sinyal).
        fill_status = _STATUS_FILL.get(status[i] if i < len(status) else "")
        row = list(row)
        if ada_expl and kol_info is not None and kol_info < len(row):
            # Baris tanpa PN → '—' (dari _teks_info_exploded(None)), bukan sel kosong
            # yang tampak seperti "sedang diproses".
            row[kol_info] = _info_exploded_sel(d, rangka)
        for j in range(1, len(kolom) + 1):
            val = row[j - 1] if j - 1 < len(row) else ""
            if j - 1 in num_cols:
                val = ke_angka(val, num_cols[j - 1])
            c = ws.cell(row=r, column=j, value=_safe(val))
            c.border = _BORDER
            angka = isinstance(c.value, (int, float)) and not isinstance(c.value, bool)
            if angka:
                c.alignment = _RIGHT
                c.number_format = col_fmts[j - 1]
            else:
                c.alignment = (_CENTER if j == 1 or j in mono_cols or j in foto_cols
                               or (ada_expl and j - 1 == kol_gambar) else _LEFT)
            c.font = _MONO if j in mono_cols else _INK
            if fill_status:
                c.fill = fill_status
            elif i % 2:
                c.fill = _ZEBRA
        # Tinggi baris = yang TERTINGGI di antara foto & gambar teknis (kalau
        # keduanya ada, memakai tinggi foto akan memotong figure exploded).
        tinggi = 0.0
        for k, col0 in enumerate(kol_foto):
            png = thumb.get((i, k))
            if not png:
                continue
            img = XLImage(io.BytesIO(png))
            ratio = img.width / img.height if img.height else 1
            img.height = _FOTO_H_PX
            img.width = int(_FOTO_H_PX * ratio)
            ws.add_image(img, f"{get_column_letter(col0 + 1)}{r}")
            tinggi = max(tinggi, _FOTO_H_PX * 0.78)
        png = (d or {}).get("png")
        if png:
            img = XLImage(io.BytesIO(png))
            w, h = int(img.width or 1), int(img.height or 1)
            skala = min(_EXPL_IMG_W / w, _EXPL_IMG_H / h, 1.0)
            img.width, img.height = max(1, int(w * skala)), max(1, int(h * skala))
            ws.add_image(img, f"{get_column_letter(kol_gambar + 1)}{r}")
            tinggi = max(tinggi, min(_EXPL_ROW_PT_MAX, img.height * 0.78 + 8))
        if tinggi:
            ws.row_dimensions[r].height = tinggi
        r += 1
    ws.freeze_panes = ws.cell(row=start + 1, column=1)

    # Blok RINGKASAN (rekap) di bawah tabel.
    if ringkasan:
        r += 1
        hc = ws.cell(row=r, column=1, value="RINGKASAN")
        hc.font = Font(bold=True, color=_BRAND_DK, size=12)
        r += 1
        for item in ringkasan:
            label, nilai = item[0], item[1]
            warna = item[2] if len(item) > 2 else ""
            lc = ws.cell(row=r, column=1, value=_safe(str(label)))
            lc.font = _BOLD
            lc.fill = _STATUS_FILL.get(warna, _SUB1_FILL)
            lc.border = _BORDER
            lc.alignment = _LEFT
            vc = ws.cell(row=r, column=2, value=_safe(nilai))
            vc.font = _INK
            vc.border = _BORDER
            vc.alignment = _LEFT
            r += 1

    return _save_stable(wb), ""


# ── KATALOG BERGAMBAR per kategori (exploded view EPC) ──────────────────────
# Plafon seksi figure — katalog LENGKAP unit besar bisa ~500 (PB087964: 477).
_KATALOG_MAX_FIGURES = 800
_KATALOG_IMG_WIDTH = 1400      # px render SVG→PNG (tajam saat di-zoom)
_KATALOG_IMG_VIEW = 660        # px lebar tampil di sheet


def _svg_to_png(svg: bytes, width: int = _KATALOG_IMG_WIDTH) -> bytes | None:
    """Render SVG exploded view EPC (Creo Illustrate) → PNG via resvg.
    resvg menolak width/height ber-satuan mm di tag root → buang atributnya
    (viewBox tetap menjaga proporsi). None bila gagal/resvg tak terpasang."""
    if not svg:
        return None
    try:
        import resvg_py
    except Exception:
        return None
    try:
        i = svg.find(b"<svg")
        if i < 0:
            return None
        txt = svg[i:].decode("utf-8", "replace")
        head, rest = txt.split(">", 1)
        head = re.sub(r'\s(width|height)="[^"]*"', "", head)
        return bytes(resvg_py.svg_to_bytes(svg_string=head + ">" + rest, width=width))
    except Exception:
        return None


def _highlight_ball(svg: bytes, ball) -> bytes:
    """Sisipkan LINGKARAN KUNING di belakang teks NOMOR BALON `ball` di SVG EPC —
    agar part yang diminta MENONJOL di gambar. Balon = <text x y font-size>N</text>
    (Arial); disisipkan <circle> tepat SEBELUM teks itu (paint order → di belakang).
    Aman: bila nomor tak ketemu / ball kosong → SVG dikembalikan apa adanya."""
    if not svg or ball in (None, ""):
        return svg
    b = re.sub(r"\s+", "", str(ball))
    if not b:
        return svg
    try:
        txt = svg.decode("utf-8", "replace")
    except Exception:
        return svg

    def _repl(m: "re.Match") -> str:
        tag = m.group(0)
        inner = re.sub(r"\s+", "", re.sub(r"<[^>]+>", "", tag))
        if inner != b:
            return tag
        mx, my = re.search(r'\bx="([-\d.]+)"', tag), re.search(r'\by="([-\d.]+)"', tag)
        if not mx or not my:
            return tag
        x, y = float(mx.group(1)), float(my.group(1))
        mfs = re.search(r'font-size="([-\d.]+)"', tag)
        fs = float(mfs.group(1)) if mfs else 4.2
        n = len(b)
        cx = x + fs * 0.30 * n           # tengah horizontal angka (x = tepi kiri)
        cy = y - fs * 0.33               # y = baseline → naik ke tengah tinggi angka
        r = fs * (0.95 + 0.20 * (n - 1))  # sedikit lebih besar dari angka
        circle = (f'<circle cx="{cx:.3f}" cy="{cy:.3f}" r="{r:.3f}" fill="#FFD400" '
                  f'fill-opacity="0.9" stroke="#E39A00" stroke-width="{fs*0.09:.3f}"/>')
        return circle + tag              # lingkaran DI BELAKANG teks nomor

    try:
        out = re.sub(r"<text\b[^>]*>.*?</text>", _repl, txt, flags=re.S)
        return out.encode("utf-8")
    except Exception:
        return svg


def exploded_png(svg_name: str, ball=None) -> bytes | None:
    """Unduh SATU file SVG exploded-view EPC (nama dari field d2s) → render PNG,
    dengan NOMOR BALON `ball` di-HIGHLIGHT kuning bila diberikan. Dipakai fitur
    'tampilkan gambar exploded view part ini' (gambar inline di chat).
    None bila nama kosong / file tak ada / resvg gagal."""
    if not svg_name:
        return None
    try:
        svg = epc_bom.fetch_file(svg_name)
    except Exception:
        return None
    if not svg:
        return None
    return _svg_to_png(_highlight_ball(svg, ball))


def exploded_png_weichai(svg_file_id: str, rangka: str, ball=None) -> bytes | None:
    """Versi MESIN Weichai: unduh SVG exploded-view via dePreview (svgFileId +
    token yang di-mint ulang dari `rangka` saat unduh — token pendek), highlight
    nomor balon (orderNo), render PNG. None bila gagal."""
    if not svg_file_id:
        return None
    try:
        from . import epc_weichai as _wc
        tok = _wc._ensure_token(rangka)
        if not tok:
            return None
        svg = _wc.fetch_svg(svg_file_id, tok)
    except Exception:
        return None
    if not svg:
        return None
    return _svg_to_png(_highlight_ball(svg, ball))


def exploded_png_shantui(svg_name: str, ball=None) -> bytes | None:
    """Versi ALAT BERAT Shantui: unduh SVG exploded-view (nama file d2s '.EN.svg')
    via epc_shantui.fetch_file (butuh token Shantui), highlight nomor balon, render
    PNG. None bila token kedaluwarsa / file tak ada / resvg gagal."""
    if not svg_name:
        return None
    try:
        from . import epc_shantui as _sh
        svg = _sh.fetch_file(svg_name)
    except Exception:
        return None
    if not svg:
        return None
    return _svg_to_png(_highlight_ball(svg, ball))


def _katalog_source(rangka: str, kategori: str, source: str):
    """Ambil hasil walk + fungsi pengambil-SVG sesuai sumber katalog.
    'sinotruk' → epc_bom.catalog_walk + fetch_file(nama); 'weichai' →
    epc_weichai.catalog_walk + fetch_svg(svgFileId, token). Return (d, fetch)."""
    if source == "weichai":
        from . import epc_weichai as _wc
        d = _wc.catalog_walk(rangka, kategori)
        tok = d.get("_token")
        return d, (lambda ref: _wc.fetch_svg(ref, tok))
    from . import epc_bom as _epc
    return _epc.catalog_walk(rangka, kategori), _epc.fetch_file


def katalog_excel(rangka: str, kategori: str, source: str = "sinotruk",
                  isi_stok_harga: bool = False) -> tuple[bytes | None, str]:
    """KATALOG PART BERGAMBAR satu kategori per-VIN: satu sheet per FIGURE
    (gambar exploded view EPC + tabel part ber-nomor balon) + sheet Ringkasan.
    Kolom Stok & Harga SELALU ADA tapi ISINYA KOSONG secara default; hanya diisi
    bila isi_stok_harga=True (admin minta). Return (bytes, filename) atau
    (None, pesan_error). source='weichai' → katalog MESIN Weichai."""
    d, _fetch = _katalog_source(rangka, kategori, source)

    if not d.get("found"):
        return None, (d.get("message") or "Gagal mengambil katalog dari EPC. Coba lagi.")
    figures = (d.get("figures") or [])[:_KATALOG_MAX_FIGURES]
    if not figures:
        return None, f"Tidak ada figure untuk kategori '{kategori}' di unit ini."
    frame = d.get("frame_number") or ""

    # 1) Unduh + render SEMUA gambar paralel (referensi gambar → PNG bytes).
    svg_names = list(dict.fromkeys(f["svg"] for f in figures if f.get("svg")))

    def _render(name: str) -> tuple[str, bytes | None]:
        return name, _svg_to_png(_fetch(name))

    pngs: dict[str, bytes] = {}
    if svg_names:
        with ThreadPoolExecutor(max_workers=8) as ex:
            for name, png in ex.map(_render, svg_names):
                if png:
                    pngs[name] = png

    # 2) Stok/harga lokal utk SEMUA PN (sekali query).
    all_pns = list({it["pn"] for f in figures for it in f["items"]})
    local: dict[str, dict] = {}
    for r in part_index.search_exact_pns(all_pns):
        pn = (r.get("part_number") or "").upper()
        if pn and pn not in local:
            local[pn] = r

    # ── SATU SHEET, alur vertikal per figure: bar seksi → GAMBAR → TABEL.
    #    Plus DAFTAR ISI ber-hyperlink di atas & link "↑ Daftar Isi" di tiap seksi,
    #    supaya 60+ seksi mudah dinavigasi dan urutan bacanya tak membingungkan. ──
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.worksheet.hyperlink import Hyperlink

    wb = Workbook()
    ws = wb.active
    ws.title = "Katalog"
    ws.sheet_view.showGridLines = False
    n_part = sum(len(f["items"]) for f in figures)

    headers = ["No. Balon", "Part Number", "Nama Part", "Qty", "Stok", "Harga", "Pengganti"]
    widths = [10, 20, 52, 6, 8, 14, 20]
    ncol = len(headers)
    for j, wd in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = wd

    kat_title = ("Lengkap (Semua Kategori)" if d.get("lengkap") else kategori.title())
    _title(ws, f"Katalog {kat_title} — Unit {frame}",
           f"{len(figures)} figure · {n_part} part · gambar exploded view resmi EPC "
           "(Parts Atlas per-VIN) · MASPART Asisten AI", ncol)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncol)
    note = ws.cell(row=3, column=1,
                   value="ℹ️ Cara baca: tiap figure = GAMBAR lalu TABEL part-nya di bawah. "
                         "Angka pada gambar = kolom 'No. Balon' — baris itulah Part Number-nya.")
    note.font = Font(color=_BRAND_DK, size=10, bold=True)
    note.fill = _SUB1_FILL
    note.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.freeze_panes = "A4"   # judul + cara baca tetap terlihat saat scroll

    _ROW_PX = 19.0    # tinggi baris default Excel ±19px — utk melewati tinggi gambar
    _LINK_FONT = Font(color="0563C1", underline="single", size=10)

    def _num(v):
        """Angka murni ditulis sbg NUMBER (hindari segitiga hijau 'number as text')."""
        if isinstance(v, (int, float)):
            return v
        s = str(v or "").strip()
        return int(s) if s.isdigit() else (s or None)

    # ── DAFTAR ISI (target hyperlink diisi setelah posisi seksi diketahui) ──
    TOC_BAR = 5
    ws.merge_cells(start_row=TOC_BAR, start_column=1, end_row=TOC_BAR, end_column=ncol)
    tb = ws.cell(row=TOC_BAR, column=1,
                 value="DAFTAR ISI — klik nama figure untuk melompat ke bagiannya")
    tb.fill = _HEAD_FILL
    tb.font = Font(color="FFFFFF", bold=True, size=12)
    tb.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[TOC_BAR].height = 22
    toc_first = TOC_BAR + 1
    for i, f in enumerate(figures):
        r = toc_first + i
        cno = ws.cell(row=r, column=1, value=i + 1)
        cno.alignment = _CENTER
        cno.border = _BORDER
        cno.font = _INK
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        cn = ws.cell(row=r, column=2, value=_safe(f["nama"] + ("" if f.get("svg") else "  (tanpa gambar di EPC)")))
        cn.font = _LINK_FONT
        cn.alignment = _LEFT
        cn.border = _BORDER
        cq = ws.cell(row=r, column=4, value=len(f["items"]))
        cq.alignment = _CENTER
        cq.border = _BORDER
        cq.font = _INK
        # Kolom kelompok — penting utk katalog lengkap (kabin/sasis/kelistrikan/…).
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
        ck = ws.cell(row=r, column=5, value=_safe(f.get("kategori") or ""))
        ck.font = Font(color="535B56", size=9)
        ck.alignment = _LEFT
        ck.border = _BORDER
        if i % 2:
            for j in range(1, 8):
                ws.cell(row=r, column=j).fill = _ZEBRA
    row = toc_first + len(figures) + 2

    # ── Seksi per figure: bar → gambar → tabel ──
    sec_rows: list[int] = []
    for i, f in enumerate(figures):
        sec_rows.append(row)
        # Bar seksi: judul (A..F) + link kembali ke daftar isi (G).
        kat_lbl = (f.get("kategori") or "").strip()
        kat_sfx = f"  ·  {kat_lbl}" if kat_lbl and kat_lbl not in f["nama"] else ""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol - 1)
        c = ws.cell(row=row, column=1,
                    value=f"{i + 1:02d}. {f['nama']}  ·  Figure {f['kode']} · {len(f['items'])} part"
                          + kat_sfx + ("" if f.get("svg") else "  ·  (tanpa gambar di EPC)"))
        c.fill = _HEAD_FILL
        c.font = Font(color="FFFFFF", bold=True, size=12)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        back = ws.cell(row=row, column=ncol, value="↑ Daftar Isi")
        back.fill = _HEAD_FILL
        back.font = Font(color="FFFFFF", underline="single", size=10)
        back.alignment = _CENTER
        back.hyperlink = Hyperlink(ref=f"{get_column_letter(ncol)}{row}",
                                   location=f"Katalog!A{TOC_BAR}")
        ws.row_dimensions[row].height = 22
        row += 1

        # GAMBAR tepat di bawah bar seksi (kolom B) — alur baca vertikal.
        png = pngs.get(f.get("svg") or "")
        if png:
            img = XLImage(io.BytesIO(png))
            ratio = img.height / img.width if img.width else 1
            img.width = _KATALOG_IMG_VIEW
            img.height = int(_KATALOG_IMG_VIEW * ratio)
            ws.add_image(img, f"B{row}")
            row += int(img.height / _ROW_PX) + 2

        # TABEL part figure ini.
        for j, h in enumerate(headers, start=1):
            hc = ws.cell(row=row, column=j, value=h)
            hc.fill = _SUB1_FILL
            hc.font = Font(bold=True, color=_BRAND_DK, size=10)
            hc.alignment = _CENTER
            hc.border = _BORDER
        row += 1
        items = sorted(f["items"], key=lambda x: (x.get("balon") is None, x.get("balon") or 0))
        for k, it in enumerate(items):
            lr = local.get(it["pn"], {})
            nama = " ".join((lr.get("part_name") or it["nama"] or it["nama_cn"]).split())
            # Stok & Harga: KOSONG kecuali admin minta (isi_stok_harga). Kolomnya
            # tetap ada agar layout konsisten; hanya nilainya yang ditahan.
            vals = [_num(it.get("balon")), it["pn"], nama, _num(it.get("qty")),
                    (_num(lr.get("stok")) if (isi_stok_harga and lr) else None),
                    ((lr.get("harga") or None) if (isi_stok_harga and lr) else None),
                    ", ".join(it.get("pengganti") or []) or None]
            for j, v in enumerate(vals, start=1):
                dc = ws.cell(row=row, column=j, value=_safe(v))
                dc.border = _BORDER
                dc.font = _MONO if j == 2 else _INK
                dc.alignment = _CENTER if j in (1, 4, 5) else _LEFT
                if k % 2:
                    dc.fill = _ZEBRA
            row += 1
        row += 2   # jeda antar seksi

    # Isi hyperlink DAFTAR ISI → baris bar tiap seksi.
    for i, target in enumerate(sec_rows):
        r = toc_first + i
        ws.cell(row=r, column=2).hyperlink = Hyperlink(
            ref=f"B{r}", location=f"Katalog!A{target}")

    buf = io.BytesIO()
    wb.save(buf)
    kat_slug = re.sub(r"[^A-Za-z0-9]+", "_", kategori).strip("_")[:24] or "Kategori"
    return buf.getvalue(), f"Katalog_{kat_slug}_{frame}.xlsx"


# ── KATALOG BERGAMBAR versi PDF (cetak) ─────────────────────────────────────
def _pdf_esc(v) -> str:
    s = "" if v is None else str(v)
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def katalog_pdf(rangka: str, kategori: str, source: str = "sinotruk",
                isi_stok_harga: bool = False) -> tuple[bytes | None, str]:
    """Versi PDF (siap cetak/kirim) dari katalog_excel: satu bagian per FIGURE —
    judul, gambar exploded view EPC, lalu tabel part ber-nomor balon. Kolom Stok
    & Harga SELALU ADA tapi ISINYA KOSONG kecuali isi_stok_harga=True (admin).
    Return (bytes, filename) atau (None, pesan_error).
    source='weichai' → katalog MESIN Weichai."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        from reportlab.platypus import (Image as RLImage, Paragraph, SimpleDocTemplate,
                                        Spacer, Table, TableStyle)
    except Exception:
        return None, "Modul PDF (reportlab) belum terpasang di server."

    d, _fetch = _katalog_source(rangka, kategori, source)

    if not d.get("found"):
        return None, (d.get("message") or "Gagal mengambil katalog dari EPC. Coba lagi.")
    figures = (d.get("figures") or [])[:_KATALOG_MAX_FIGURES]
    if not figures:
        return None, f"Tidak ada figure untuk kategori '{kategori}' di unit ini."
    frame = d.get("frame_number") or ""

    svg_names = list(dict.fromkeys(f["svg"] for f in figures if f.get("svg")))

    def _render(name: str) -> tuple[str, bytes | None]:
        return name, _svg_to_png(_fetch(name))

    pngs: dict[str, bytes] = {}
    if svg_names:
        with ThreadPoolExecutor(max_workers=8) as ex:
            for name, png in ex.map(_render, svg_names):
                if png:
                    pngs[name] = png

    all_pns = list({it["pn"] for f in figures for it in f["items"]})
    local: dict[str, dict] = {}
    for r in part_index.search_exact_pns(all_pns):
        pn = (r.get("part_number") or "").upper()
        if pn and pn not in local:
            local[pn] = r

    buf = io.BytesIO()
    page = landscape(A4)
    doc = SimpleDocTemplate(buf, pagesize=page, leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm,
                            title=f"Katalog {kategori} {frame}")
    ss = getSampleStyleSheet()
    st_title = ParagraphStyle("cc_t", parent=ss["Title"], fontSize=17,
                              textColor=colors.HexColor("#026A0E"))
    st_sub = ParagraphStyle("cc_sub", parent=ss["Normal"], fontSize=8.5,
                            textColor=colors.HexColor("#535B56"))
    st_sec = ParagraphStyle("cc_sec", parent=ss["Heading2"], fontSize=11.5,
                            spaceBefore=10, spaceAfter=3, textColor=colors.HexColor("#0F1411"))
    st_cell = ParagraphStyle("cc_c", parent=ss["Normal"], fontSize=7.3, leading=9)
    st_head = ParagraphStyle("cc_h", parent=ss["Normal"], fontSize=7.3, leading=9,
                             textColor=colors.HexColor("#026A0E"))

    avail_w = page[0] - 24 * mm
    story: list = [
        Paragraph(f"Katalog Part — {_pdf_esc(kategori.title())}", st_title),
        Paragraph(f"Unit / VIN: {_pdf_esc(frame)} · {len(figures)} figure · Sumber: "
                  + ("EPC Weichai resmi (mesin, per-VIN)" if source == "weichai"
                     else "EPC Parts Atlas resmi (Sinotruk)") + " — MASPART", st_sub),
        Spacer(1, 6),
    ]

    # Lebar kolom tabel: No, PN, Nama (fleksibel), Qty, Stok, Harga, Pengganti.
    fixed = [11 * mm, 36 * mm, 13 * mm, 15 * mm, 24 * mm, 30 * mm]
    col_w = [fixed[0], fixed[1], avail_w - sum(fixed), fixed[2], fixed[3], fixed[4], fixed[5]]
    head = ["No.", "Part Number", "Nama", "Qty", "Stok", "Harga", "Pengganti"]

    for i, f in enumerate(figures):
        story.append(Paragraph(
            f"{i + 1:02d}. {_pdf_esc(f['nama'])} · Figure {_pdf_esc(f.get('kode') or '')} · "
            f"{len(f['items'])} part", st_sec))
        png = pngs.get(f.get("svg") or "")
        if png:
            try:
                iw, ih = ImageReader(io.BytesIO(png)).getSize()
                w = min(avail_w, 165 * mm)
                h = (w * ih / iw) if iw else 70 * mm
                if h > 92 * mm:
                    h = 92 * mm
                    w = h * iw / ih if ih else w
                story.append(RLImage(io.BytesIO(png), width=w, height=h))
                story.append(Spacer(1, 3))
            except Exception:
                pass
        rows = [[Paragraph(f"<b>{c}</b>", st_head) for c in head]]
        items = sorted(f["items"], key=lambda x: (x.get("balon") is None, x.get("balon") or 0))
        for it in items:
            lr = local.get(it["pn"], {})
            nama = " ".join((lr.get("part_name") or it.get("nama") or it.get("nama_cn") or "").split())
            rows.append([
                Paragraph(_pdf_esc(it.get("balon")), st_cell),
                Paragraph(_pdf_esc(it["pn"]), st_cell),
                Paragraph(_pdf_esc(nama), st_cell),
                Paragraph(_pdf_esc(it.get("qty")), st_cell),
                Paragraph(_pdf_esc(lr.get("stok") if (isi_stok_harga and lr) else ""), st_cell),
                Paragraph(_pdf_esc(lr.get("harga") if (isi_stok_harga and lr) else ""), st_cell),
                Paragraph(_pdf_esc(", ".join(it.get("pengganti") or [])), st_cell),
            ])
        tbl = Table(rows, colWidths=col_w, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAF6EC")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E1E4E1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9F7")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 9))

    try:
        doc.build(story)
    except Exception as e:
        return None, f"Gagal membangun PDF katalog: {e}"
    kat_slug = re.sub(r"[^A-Za-z0-9]+", "_", kategori).strip("_")[:24] or "Kategori"
    return buf.getvalue(), f"Katalog_{kat_slug}_{frame}.pdf"
