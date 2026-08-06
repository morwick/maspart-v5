"""
Service Accurate Online — stok live dari ERP Accurate (iris.accurate.id).

Auth Accurate = sesi SSO: cookie **JSESSIONID** + parameter **_dsi** (keduanya
didapat dari login SAML di account.accurate.id). Sesi disimpan di
``data/accurate_session.json`` dan **dibaca segar tiap panggil** (pola sama dengan
``epc_token.txt``) → sesi baru cukup ditimpa ke file tanpa restart.

Endpoint data (dibedah dari lalu-lintas web resmi, terverifikasi live):

    POST {BASE}/accurate/inventory/search-item.do
    body: _dsi=<dsi>&keywords=<q>&resetFilter=false
          &sp.pageSize=<N>&sp.start=<OFF>&sp.limit=<N>
    → {"s": true, "d": [ {item...} ], "sp": {rowCount, pageCount, ...}}

Sesi mati → server MEMANTULKAN ke SSO dan membalas **HTML** (form SAML), bukan
JSON. Terdeteksi via ``_looks_expired`` → naikkan ``AccurateSessionExpired``.

Field penting per item (dari response nyata):
  no             kode barang, mis. "003330.WG9725220536+305" (ada prefix "NNNNNN.")
  name           nama/uraian barang
  availableToSell stok yang dapat dijual (dipakai sebagai "stok")
  quantity        kuantitas gudang
  unit1.name      satuan (Pc/Liter/Kg…)
  itemTypeName    "Persediaan" / "Jasa" / dll
  id              id internal Accurate
"""
from __future__ import annotations

import base64
import datetime as _dt
import html as _html
import json
import logging
import re
import threading
import time
from pathlib import Path
from typing import Any, Iterable

import requests

from ..core.config import get_settings

logger = logging.getLogger("maspart.accurate")

# ── Konfigurasi ────────────────────────────────────────────────────────────
def _base() -> str:
    """Base URL app perusahaan (host/zona dari config, mis. iris.accurate.id)."""
    return f"https://{get_settings().accurate_host}"

_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36"
)

# Ukuran halaman saat menarik seluruh katalog (server terbukti melayani 200/҂call).
_PAGE_SIZE = 200
_HTTP_TIMEOUT = 30

# (Deprecated) TTL lama 5 jam. Sejak 2026-07-10 indeks TIDAK lagi berbasis TTL:
# ditarik pada JAM WIB TETAP 3×/hari (_REFRESH_HOURS_WIB) & dimuat dari disk saat
# restart. Konstanta dipertahankan hanya utk referensi doc lama.
_INDEX_TTL = 5 * 60 * 60


def _session_file() -> Path:
    return get_settings().data_path / "accurate_session.json"


# ── Sesi (JSESSIONID + _dsi) ───────────────────────────────────────────────
class AccurateError(RuntimeError):
    """Kesalahan umum saat berkomunikasi dengan Accurate."""


class AccurateSessionMissing(AccurateError):
    """File sesi belum ada / tak lengkap (perlu diisi JSESSIONID + _dsi)."""


class AccurateSessionExpired(AccurateError):
    """Sesi Accurate kadaluarsa — server memantulkan ke login SSO."""


def load_session() -> dict[str, str]:
    """Baca ``data/accurate_session.json`` → {"jsessionid", "dsi", ...}.

    Dibaca SEGAR tiap panggil (tak di-cache) supaya penggantian sesi via file
    langsung berlaku tanpa restart. Naikkan ``AccurateSessionMissing`` bila file
    tak ada atau field wajib kosong.
    """
    fp = _session_file()
    try:
        raw = json.loads(fp.read_text(encoding="utf-8"))
    except FileNotFoundError:
        raise AccurateSessionMissing(
            f"File sesi Accurate belum ada: {fp}. Isi JSESSIONID + _dsi "
            "(lihat save_session / tools/accurate_set_session)."
        )
    except (OSError, json.JSONDecodeError) as e:
        raise AccurateSessionMissing(f"File sesi Accurate tak terbaca: {e}") from e

    jsid = str(raw.get("jsessionid") or raw.get("JSESSIONID") or "").strip()
    dsi = str(raw.get("dsi") or raw.get("_dsi") or "").strip()
    if not jsid or not dsi:
        raise AccurateSessionMissing(
            "Sesi Accurate tak lengkap: butuh 'jsessionid' dan 'dsi' di "
            f"{fp}."
        )
    return {"jsessionid": jsid, "dsi": dsi, **{k: v for k, v in raw.items()
                                              if k not in ("jsessionid", "dsi")}}


def save_session(jsessionid: str, dsi: str) -> Path:
    """Tulis pasangan sesi ke file (dipakai auto-login tahap-2 & pengisian manual)."""
    jsid = (jsessionid or "").strip()
    d = (dsi or "").strip()
    if not jsid or not d:
        raise ValueError("jsessionid dan dsi wajib diisi.")
    fp = _session_file()
    fp.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "jsessionid": jsid,
        "dsi": d,
        "updated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    fp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return fp


def session_status() -> dict[str, Any]:
    """Ringkasan status sesi tanpa membocorkan nilai penuh (untuk endpoint admin)."""
    try:
        s = load_session()
    except AccurateSessionMissing as e:
        return {"configured": False, "detail": str(e)}
    return {
        "configured": True,
        "jsessionid_tail": s["jsessionid"][-8:],
        "dsi_len": len(s["dsi"]),
        "updated_at": s.get("updated_at"),
    }


# ── Auto-login SSO (account.accurate.id → SAML → iris; per-VIN dsi) ─────────
# Alur (dibedah & terverifikasi live):
#   GET  account.accurate.id/            → seed cookie
#   POST /pre-login.do (account,password="up"+b64{v,p,d})       → {d.permit}
#   POST /auth.do (j_username,j_password="ua"+b64{v,p,c=permit,t:null,d})  → 302 /manage
#   GET  iris/accurate/open.do?uid=<uid>&product=aol            → form SAML auto-submit
#        → POST account/idp/sso (SAMLRequest) → POST iris/accurate/saml/SSO (SAMLResponse)
#   POST iris/accurate/open-database.do (uid,product=aol)       → {dsi}
#   JSESSIONID iris diambil dari cookie jar.
_ACCOUNT_BASE = "https://account.accurate.id"
_login_lock = threading.Lock()


def credentials_configured() -> bool:
    return get_settings().accurate_login_configured


def _b64(x: str) -> str:
    return base64.b64encode(x.encode("utf-8")).decode("ascii")


def _parse_form_inputs(fragment: str) -> dict[str, str]:
    out: dict[str, str] = {}
    for tag in re.findall(r"<input\b[^>]*>", fragment, re.I):
        n = re.search(r'name=["\']([^"\']+)["\']', tag)
        v = re.search(r'value=["\']([^"\']*)["\']', tag)
        if n:
            out[_html.unescape(n.group(1))] = _html.unescape(v.group(1) if v else "")
    return out


def _get_form(body: str) -> tuple[str | None, str | None]:
    m = re.search(r"<form\b([^>]*)>(.*?)</form>", body, re.S | re.I)
    if not m:
        return None, None
    am = re.search(r'action=["\']([^"\']+)["\']', m.group(1))
    return (_html.unescape(am.group(1)) if am else None), m.group(2)


def _follow_saml(sess: requests.Session, resp: requests.Response, max_hops: int = 6) -> requests.Response:
    """Ikuti rantai form SAML auto-submit sampai tak ada form lagi."""
    for _ in range(max_hops):
        action, inner = _get_form(resp.text)
        if not action or inner is None:
            break
        resp = sess.post(action, data=_parse_form_inputs(inner), timeout=_HTTP_TIMEOUT)
    return resp


def login() -> dict[str, str]:
    """Login otomatis penuh (SSO) → tulis & kembalikan {jsessionid, dsi}.

    Naikkan ``AccurateSessionMissing`` bila kredensial tak lengkap, atau
    ``AccurateError`` bila alur gagal (mis. akun minta 2FA/OTP).
    """
    cfg = get_settings()
    if not cfg.accurate_login_configured:
        raise AccurateSessionMissing(
            "Kredensial auto-login Accurate belum lengkap (ACCURATE_USERNAME/"
            "PASSWORD/DEVICE_ID/UID)."
        )
    user, pw = cfg.accurate_username, cfg.accurate_password
    device, uid, host = cfg.accurate_device_id, cfg.accurate_uid, cfg.accurate_host
    s = requests.Session()
    s.headers.update({"User-Agent": _UA, "Accept-Language": "id,en-US;q=0.9,en;q=0.8"})
    try:
        s.get(f"{_ACCOUNT_BASE}/", timeout=_HTTP_TIMEOUT)
        # 1) pre-login → challenge 'permit'
        pw_pre = "up" + _b64(json.dumps({"v": 1, "p": pw, "d": device}, separators=(",", ":")))
        r = s.post(f"{_ACCOUNT_BASE}/pre-login.do",
                   data={"account": user, "password": pw_pre},
                   headers={"X-Requested-With": "XMLHttpRequest", "Referer": f"{_ACCOUNT_BASE}/"},
                   timeout=_HTTP_TIMEOUT)
        try:
            d = (r.json() or {}).get("d") or {}
        except ValueError:
            raise AccurateError("pre-login.do gagal (respons bukan JSON).")
        permit = d.get("permit")
        if d.get("requireTotp") or d.get("isEmail2FA"):
            raise AccurateError("Akun Accurate meminta 2FA/OTP — auto-login tak didukung.")
        if not permit:
            raise AccurateError(f"pre-login.do tak mengembalikan challenge: {str(d)[:150]}")
        # 2) auth.do (Spring Security) → sesi account terautentikasi
        jpw = "ua" + _b64(json.dumps({"v": 1, "p": pw, "c": permit, "t": None, "d": device},
                                     separators=(",", ":")))
        s.post(f"{_ACCOUNT_BASE}/auth.do",
               data={"j_username": user, "j_password": jpw},
               headers={"Referer": f"{_ACCOUNT_BASE}/"}, timeout=_HTTP_TIMEOUT)
        # 3) open.do → SAML → sesi iris terautentikasi (JSESSIONID)
        resp = s.get(f"https://{host}/accurate/open.do?uid={uid}&product=aol", timeout=_HTTP_TIMEOUT)
        resp = _follow_saml(s, resp)
        # 4) open-database.do → dsi segar
        od = s.post(f"https://{host}/accurate/open-database.do",
                    data={"uid": uid, "product": "aol"},
                    headers={"X-Requested-With": "XMLHttpRequest", "Referer": resp.url},
                    timeout=_HTTP_TIMEOUT)
        try:
            oj = od.json()
        except ValueError:
            raise AccurateError("open-database.do bukan JSON — login/SAML kemungkinan gagal.")
        dsi = oj.get("dsi")
        if not dsi:
            raise AccurateError(f"open-database.do tak mengembalikan dsi: {str(oj)[:150]}")
    except requests.RequestException as e:
        raise AccurateError(f"Auto-login Accurate gagal (jaringan): {e}") from e

    jsid = next((c.value for c in s.cookies
                 if c.name == "JSESSIONID" and host in (c.domain or "")), None)
    if not jsid:
        jsid = next((c.value for c in s.cookies if c.name == "JSESSIONID"), None)
    if not jsid:
        raise AccurateError("JSESSIONID iris tak ditemukan setelah login.")
    save_session(jsid, dsi)
    return {"jsessionid": jsid, "dsi": dsi}


# Cooldown login: setelah login GAGAL (mis. Accurate throttle → errorTimeout),
# JANGAN coba login lagi selama _LOGIN_COOLDOWN_SEC — pakai fallback lokal dulu.
# Mencegah hantaman login berulang yang bisa memperpanjang throttle / terlihat abnormal.
_LOGIN_COOLDOWN_SEC = 300        # backoff login latar (refresh stok). Aksi user
                                 # (buat penawaran) memakai login_now() yg MENGABAIKAN
                                 # cooldown ini → penawaran tak terblokir backoff latar.
_login_fail_until = 0.0

# Setelah membuat penawaran, admin ingin membuka Accurate MANUAL (akun 1-sesi).
# Selama jendela ini MASPART MENAHAN diri auto-login latar (stok/harga pakai cache
# indeks — tak butuh sesi) agar sesi tetap kosong utk admin. login_now() (penawaran
# berikutnya) TETAP bisa (mengabaikan penahanan ini).
_POST_QUOTATION_HOLD_SEC = 10 * 60      # 10 menit
_login_suppress_until = 0.0             # monotonic


def suppress_autologin(seconds: float = _POST_QUOTATION_HOLD_SEC) -> None:
    """Tahan auto-login LATAR selama `seconds` (dipanggil setelah buat penawaran)."""
    global _login_suppress_until
    _login_suppress_until = time.monotonic() + seconds


def _login_guarded() -> dict[str, str]:
    """login() LATAR dgn gerbang cooldown + penahanan pasca-penawaran (di dalam
    _login_lock). Aksi user memakai login_now() yang melewati gerbang ini."""
    global _login_fail_until
    if time.monotonic() < _login_suppress_until:
        raise AccurateSessionExpired(
            "Login MASPART ditahan sementara agar admin bisa memakai Accurate manual "
            "setelah membuat penawaran — pakai fallback cache."
        )
    if time.time() < _login_fail_until:
        raise AccurateSessionExpired(
            "Login Accurate sedang cooldown setelah gagal berturut — pakai fallback."
        )
    try:
        return login()
    except AccurateError:
        _login_fail_until = time.time() + _LOGIN_COOLDOWN_SEC
        raise


# ── Logout & idle-logout ────────────────────────────────────────────────────
# Aturan keras pemilik: akun Accurate hanya boleh 1 SESI/perangkat. Selama MASPART
# memegang sesi, orang lain TAK BISA login untuk membuat penawaran. Maka MASPART
# harus melepas sesi secepatnya: logout SETELAH membuat penawaran, dan otomatis
# logout saat IDLE. Aman: _ensure_session auto-login lagi saat file sesi hilang,
# jadi stok/harga tetap jalan (login on-demand). Cache data indeks TAK terpengaruh.
# Ambang idle 2 mnt = kompromi: cukup singkat agar orang lain cepat bisa masuk,
# tapi tak thrashing login (tiap login SSO menendang sesi orang lain).
_IDLE_LOGOUT_SEC = 120              # 2 menit tanpa aktivitas → logout
_last_activity = time.monotonic()
_idle_started = False
_idle_lock = threading.Lock()


def _mark_activity() -> None:
    global _last_activity
    _last_activity = time.monotonic()


def logout() -> bool:
    """Tutup sesi Accurate (best-effort): close-database.do lalu hapus file sesi
    → panggilan berikutnya auto-login segar. TAK PERNAH melempar."""
    try:
        s = load_session()
    except AccurateSessionMissing:
        return True                 # sudah tak ada sesi
    try:
        requests.post(f"{_base()}/accurate/close-database.do",
                      data={"_dsi": s["dsi"]},
                      headers={"User-Agent": _UA, "X-Requested-With": "XMLHttpRequest",
                               "Origin": _base(), "Referer": f"{_base()}/accurate/?_dsi={s['dsi']}",
                               "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                               "Cookie": f"JSESSIONID={s['jsessionid']}"},
                      timeout=_HTTP_TIMEOUT, allow_redirects=False)
    except Exception:
        pass                        # server mungkin sudah menutup; tetap hapus file
    try:
        _session_file().unlink(missing_ok=True)
    except OSError:
        pass
    logger.info("Accurate: sesi ditutup (logout).")
    return True


def _idle_logout_loop() -> None:
    while True:
        time.sleep(60)
        try:
            idle = time.monotonic() - _last_activity
            if idle > _IDLE_LOGOUT_SEC and _session_file().exists():
                logout()
        except Exception:           # daemon tak boleh mati
            pass


def start_idle_logout() -> bool:
    """Mulai daemon auto-logout saat idle. Idempoten (aman dipanggil berulang)."""
    global _idle_started
    with _idle_lock:
        if _idle_started:
            return False
        _idle_started = True
    threading.Thread(target=_idle_logout_loop, daemon=True,
                     name="accurate-idle-logout").start()
    return True


def _ensure_session() -> dict[str, str]:
    """Sesi siap-pakai: dari file, atau auto-login bila file kosong & kredensial ada."""
    try:
        return load_session()
    except AccurateSessionMissing:
        if not credentials_configured():
            raise
    with _login_lock:
        try:  # thread lain mungkin sudah login
            return load_session()
        except AccurateSessionMissing:
            return _login_guarded()


def _refresh_session() -> dict[str, str]:
    """Paksa login ulang (dipanggil saat sesi kadaluarsa). Thread-safe + cooldown."""
    with _login_lock:
        return _login_guarded()


def login_now() -> dict[str, str]:
    """Login SEGERA, MENGABAIKAN cooldown backoff — untuk aksi yang DIPICU USER
    (mis. buat penawaran) yang tak boleh terblokir backoff refresh-stok latar.
    Tetap thread-safe & tetap MENGARM cooldown bila gagal (agar latar tak hajar)."""
    global _login_fail_until
    with _login_lock:
        try:
            return login()
        except AccurateError:
            _login_fail_until = time.time() + _LOGIN_COOLDOWN_SEC
            raise


def ensure_session_force() -> dict[str, str]:
    """Sesi siap-pakai untuk aksi user-triggered: pakai file bila ada, kalau tidak
    login SEGERA (abaikan cooldown). Raise bila kredensial tak ada / login gagal."""
    try:
        return load_session()
    except AccurateSessionMissing:
        if not credentials_configured():
            raise
        return login_now()


# ── HTTP ───────────────────────────────────────────────────────────────────
def _looks_expired(resp: requests.Response) -> bool:
    """True bila respons adalah pantulan SSO (HTML/SAML), bukan JSON data."""
    ctype = (resp.headers.get("Content-Type") or "").lower()
    if "application/json" in ctype:
        return False
    head = resp.text[:600].lstrip()
    if head.startswith("{"):
        return False
    # HTML SAML auto-submit atau apa pun non-JSON = sesi tak dikenali.
    return True


def _call(session: dict[str, str], path: str, data: dict[str, Any]) -> dict[str, Any]:
    """POST ke satu endpoint `.do` Accurate (auth via _dsi body + JSESSIONID cookie).

    Naikkan ``AccurateSessionExpired`` bila server memantulkan ke SSO (HTML/redirect),
    ``AccurateError`` bila respons bukan JSON atau ``s=false``.
    """
    dsi = session["dsi"]
    headers = {
        "User-Agent": _UA,
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Origin": _base(),
        "Referer": f"{_base()}/accurate/?_dsi={dsi}",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "Cookie": f"JSESSIONID={session['jsessionid']}",
    }
    _mark_activity()   # jaga sesi hidup selama masih dipakai (idle-logout menunda)
    try:
        resp = requests.post(f"{_base()}/accurate/{path}", data={"_dsi": dsi, **data},
                             headers=headers, timeout=_HTTP_TIMEOUT, allow_redirects=False)
    except requests.RequestException as e:
        raise AccurateError(f"Gagal menghubungi Accurate: {e}") from e
    if resp.status_code in (301, 302, 303, 307, 308):
        raise AccurateSessionExpired("Sesi Accurate kadaluarsa (redirect ke SSO).")
    if _looks_expired(resp):
        raise AccurateSessionExpired("Sesi Accurate kadaluarsa — server memantulkan ke login SSO.")
    try:
        j = resp.json()
    except ValueError as e:
        raise AccurateError(f"Respons Accurate bukan JSON valid: {e}") from e
    if not j.get("s", False):
        raise AccurateError(f"Accurate menolak permintaan: {str(j)[:200]}")
    return j


def _search_items_raw(session: dict[str, str], *, keywords: str = "", start: int = 0,
                      limit: int = _PAGE_SIZE) -> dict[str, Any]:
    """Satu panggilan search-item.do → dict JSON ter-parse."""
    return _call(session, "inventory/search-item.do", {
        "keywords": keywords or "", "resetFilter": "false",
        "sp.pageSize": str(limit), "sp.start": str(start), "sp.limit": str(limit),
    })


def _warehouse_raw(session: dict[str, str], item_id: Any) -> dict[str, Any]:
    """view-itemstock-bywarehouse.do → stok per gudang TERKINI untuk 1 item id.
    Endpoint resmi UI Accurate (tab Stok barang). Field: detailWarehouseData[]."""
    return _call(session, "inventory/view-itemstock-bywarehouse.do", {"id": str(item_id)})


def _search_retry(*, keywords: str = "", start: int = 0, limit: int = _PAGE_SIZE) -> dict[str, Any]:
    """Satu panggilan search dengan auto-login: bila sesi kadaluarsa & kredensial
    tersedia → login ulang sekali lalu ulangi. Bila tak ada kredensial, error sesi
    diteruskan (mode sesi-manual)."""
    sess = _ensure_session()
    try:
        return _search_items_raw(sess, keywords=keywords, start=start, limit=limit)
    except AccurateSessionExpired:
        if not credentials_configured():
            raise
        sess = _refresh_session()
        return _search_items_raw(sess, keywords=keywords, start=start, limit=limit)


def search_by_keyword(keyword: str, *, limit: int = 50) -> list[dict[str, Any]]:
    """Cari barang via filter ``keywords`` Accurate (server-side, cepat).

    ``keyword`` dicocokkan server ke kode/nama barang. Dipakai untuk lookup 1 PN
    on-demand tanpa menarik seluruh katalog. Auto-login saat sesi habis.
    """
    kw = (keyword or "").strip()
    if not kw:
        return []
    data = _search_retry(keywords=kw, start=0, limit=limit)
    return [normalize_item(x) for x in (data.get("d") or [])]


def stock_for_live(part_number: str) -> dict[str, Any] | None:
    """Stok Accurate untuk 1 PN via pencarian keyword server-side (cepat, no cache).

    Mengembalikan barang yang PN-ternormalisasinya PERSIS sama; bila tak ada yang
    persis, None (hindari salah-cocok ke barang lain yang kebetulan ikut terfilter).
    Bila ada beberapa yang persis, pilih stok tertinggi.
    """
    want = _norm_pn(part_number)
    if not want:
        return None
    hits = search_by_keyword(part_number, limit=50)
    exact = [h for h in hits if _norm_pn(h["pn"]) == want]
    if not exact:
        return None
    return max(exact, key=lambda h: h["available_to_sell"])


def _warehouse_retry(item_id: Any) -> dict[str, Any]:
    """Panggil view-itemstock-bywarehouse.do dgn auto-login saat sesi habis."""
    sess = _ensure_session()
    try:
        return _warehouse_raw(sess, item_id)
    except AccurateSessionExpired:
        if not credentials_configured():
            raise
        return _warehouse_raw(_refresh_session(), item_id)


def stock_full(part_number: str) -> dict[str, Any] | None:
    """Stok Accurate 1 PN: agregat + harga + rincian per-gudang, SEMUA dari
    **INDEKS BERSAMA** (tarikan terjadwal 3×/hari jam WIB tetap; dimuat dari disk
    saat restart).

    ⛔ ATURAN KERAS PEMILIK: BACA HANYA DARI CACHE — TIDAK PERNAH login/menembak
    Accurate live per-PN. Semua login untuk baca terjadi SEKALI per 5 jam
    (refresh + enrichment). Bila rincian per-gudang belum ada di indeks (mis.
    ~8 mnt setelah boot sebelum enrichment tuntas), `per_gudang` KOSONG dulu —
    terisi otomatis di siklus enrichment berikutnya. Login live HANYA untuk
    membuat penawaran.

    Non-blocking: indeks belum siap / PN tak ada → None → pemanggil fallback Excel.
    """
    # PN ber-suffix varian ('WG9525160004/2') dicocokkan ke PN dasar Accurate.
    want = index_key(part_number)
    if not want:
        return None
    entry = (_index_cache.get("by_pn") or {}).get(want)
    if not entry:
        return None
    base = dict(entry)
    # Rincian per-gudang HANYA dari indeks (enrichment 5-jam). Belum ter-enrich →
    # kosong (TANPA panggilan live — dulu di sini ada _warehouse_retry yang login).
    gmap = (_index_cache.get("by_gudang") or {}).get(want)
    base["per_gudang"] = (
        [{"gudang": g, "qty": q} for g, q in sorted(gmap.items(), key=lambda kv: kv[1], reverse=True)]
        if gmap else []
    )
    # KELUARGA VARIAN PEMASOK (2026-08-05). Bentuk lama dipertahankan APA ADANYA;
    # dua field ini HANYA ditambahkan bila part ini memang punya kartu saudara —
    # jadi part biasa tak berubah sedikit pun. Tanpa ini, halaman part hanya
    # memperlihatkan stok satu kartu (mis. 2 pc dari kartu '/SH' yang mati)
    # padahal saudaranya menumpuk 1.069 pc.
    fam = [k for k in ((_index_cache.get("by_base") or {}).get(
        _pn_base_key(entry.get("pn") or want)) or []) if k in (_index_cache.get("by_pn") or {})]
    if len(fam) > 1:
        by = _index_cache["by_pn"]
        base["varian_lain"] = [
            {"pn": by[k].get("pn"), "no": by[k].get("no"),
             "available_to_sell": _num(by[k].get("available_to_sell")),
             "price": _num(by[k].get("price"))}
            for k in fam if k != want
        ]
        base["stok_semua_varian"] = sum(_num(by[k].get("available_to_sell")) for k in fam)
    return base


def stock_family(part_number: str) -> dict[str, Any]:
    """SEMUA kartu barang Accurate untuk satu part fisik (keluarga varian pemasok).

    LATAR LIVE 2026-08-05: PN 1000442956 punya TIGA kartu — base 482 pc @300rb,
    '/SN' 1.069 pc @285rb, '/SH' 2 pc @455rb. Selama ini lookup hanya bisa
    memilih SATU (dan sering kartu yang salah), sehingga stok terlihat 2 pc dan
    harga terlihat termahal. Fungsi ini menyajikan semuanya apa adanya —
    keputusan kartu mana yang dipakai ada di manusia/pemanggil, bukan ditebak.

    Baca INDEKS saja (⛔ aturan keras pemilik: tak pernah login/live per-PN).
    Return: {found, base, varian[…+per_gudang], total_available, harga_min, harga_max}.
    """
    kosong: dict[str, Any] = {"found": False, "base": "", "varian": [],
                              "total_available": 0.0, "harga_min": None, "harga_max": None}
    by = _index_cache.get("by_pn") or {}
    by_base = _index_cache.get("by_base") or {}
    if not by:
        return kosong                     # indeks belum siap → pemanggil fallback
    kb = _pn_base_key(part_number)
    keys = [k for k in (by_base.get(kb) or []) if k in by]
    if not keys:
        # Query berupa KUNCI PERSIS yang sudah ternormalisasi ('1000442956SN' —
        # tanpa '/'), jadi _pn_base_key tak bisa memotongnya. Ambil keluarga dari
        # entri itu sendiri (field pn-nya masih memuat '/').
        k = _norm_pn(part_number)
        ent = by.get(k)
        if ent:
            kb = _pn_base_key(ent.get("pn") or k)
            keys = [x for x in (by_base.get(kb) or []) if x in by] or [k]
    if not keys:
        return kosong
    by_g = _index_cache.get("by_gudang") or {}
    varian: list[dict[str, Any]] = []
    for k in keys:                        # by_base sudah terurut stok desc
        v = dict(by[k])
        gmap = by_g.get(k) or {}
        v["per_gudang"] = [{"gudang": g, "qty": q}
                           for g, q in sorted(gmap.items(), key=lambda kv: kv[1], reverse=True)]
        varian.append(v)
    harga = [int(_num(v.get("price"))) for v in varian if _num(v.get("price")) > 0]
    return {
        "found": True,
        "base": kb,
        "varian": varian,
        "total_available": sum(_num(v.get("available_to_sell")) for v in varian),
        "harga_min": min(harga) if harga else None,
        "harga_max": max(harga) if harga else None,
    }


def family_summary(part_number: str) -> dict[str, Any] | None:
    """RINGKASAN MURAH keluarga varian pemasok — ``None`` bila part ini sendirian.

    Kembaran ringan `stock_family`: baca `by_base` + view `snap` saja, TANPA
    menyusun rincian per-gudang. Alasannya ada di jalur pemakainya — TIAP BARIS
    hasil pencarian (ratusan per request; ditambah tiap PN di cek massal) perlu
    tahu "punya kartu saudara atau tidak". Memanggil stock_family di sana berarti
    membangun daftar per-gudang ratusan kali untuk data yang langsung dibuang.

    ``None`` (bukan dict kosong) = keluarga ≤1 anggota → pemanggil memakai jalur
    lamanya APA ADANYA, jadi part biasa tak berubah perilakunya sedikit pun.
    Return: {n, stok_total, harga_min, harga_max, kunci[…]}.
    """
    by_base = _index_cache.get("by_base") or {}
    if not by_base:
        return None                       # indeks belum siap → jangan mengarang
    by = _index_cache.get("by_pn") or {}
    keys = [k for k in (by_base.get(_pn_base_key(part_number)) or []) if k in by]
    if not keys:
        # Query berupa KUNCI ternormalisasi ('1000442956SN', tanpa '/') → _pn_base_key
        # tak bisa memotongnya; ambil basis dari entri itu sendiri (sama spt stock_family).
        ent = by.get(_norm_pn(part_number))
        if ent:
            keys = [k for k in (by_base.get(_pn_base_key(ent.get("pn") or "")) or [])
                    if k in by]
    if len(keys) < 2:
        return None
    snap = _index_cache.get("snap") or {}
    total = 0.0
    harga: list[int] = []
    for k in keys:
        s = snap.get(k)
        if s is None:
            # snap = VIEW turunan by_pn (dibangun di refresh). Bila timpang — mis.
            # entri hasil rekonsiliasi yang masuk sebelum snap disentuh — jatuh ke
            # by_pn, jangan melaporkan stok 0 palsu untuk kartu yang jelas ada.
            e = by.get(k) or {}
            q, p = _num(e.get("available_to_sell")), _num(e.get("price"))
        else:
            q, p = _num(s.get("stok")), _num(s.get("harga"))
        total += q
        if p > 0:
            harga.append(int(p))
    return {"n": len(keys), "stok_total": total,
            "harga_min": min(harga) if harga else None,
            "harga_max": max(harga) if harga else None,
            "kunci": list(keys)}


# Berapa kali katalog disapu ulang bila sapuan sebelumnya masih kurang dari
# rowCount. 3 = cukup untuk menambal geseran offset tanpa memperpanjang window
# login (tiap sapuan ~27 panggilan, hitungan detik).
_MAX_SWEEPS = 3


def _item_key(it: dict[str, Any]) -> Any:
    """Kunci dedup lintas halaman: id internal Accurate; fallback kode barang
    (`no`) bila id kosong — supaya barang yang terbaca 2× (efek geser offset)
    tidak dihitung dua kali dan tidak menutupi barang lain."""
    i = it.get("id")
    return i if i is not None else ("no", str(it.get("no") or ""))


def fetch_all_items() -> list[dict[str, Any]]:
    """Tarik SELURUH barang (sapuan berulang + dedup). Untuk sync massal.
    Auto-login saat sesi habis (per halaman).

    ⚠️ BUKTI PRODUKSI 2026-08-05 — INDEKS BOCOR: server melaporkan
    ``rowCount=5229`` tetapi sapuan offset polos hanya memungut **5200** (persis
    26×200). 29 barang hilang dari indeks TIAP siklus, dan yang hilang justru
    yang PALING LAKU (PN 1000442956 'Fuel Filter' 482 pc, dan kartu varian
    pemasok '/SN' 1.069 pc), sementara kartu mati ('/SH', 2 pc) tertangkap.
    Akibatnya halaman part & asisten menjawab "stok —" untuk part yang paling
    laku. Dua sebabnya:

      1. ``search-item.do`` di-page dengan OFFSET tanpa urutan stabil. Barang
         yang BERTRANSAKSI selama sapuan berpindah posisi: yang bergeser MAJU
         melewati offset yang sudah dilewati → tak pernah terbaca; yang mundur
         terbaca dua kali (karena itu jumlahnya pas kelipatan halaman).
      2. ``if not batch: break`` — SATU halaman yang kebetulan balas kosong
         (glitch sesaat) MEMOTONG SELURUH EKOR katalog.

    Penawarnya di sini: halaman kosong di TENGAH tidak menghentikan sapuan
    (retry sekali lalu lanjut ke offset berikutnya), hasil di-DEDUP by id, dan
    bila jumlah unik masih kurang dari rowCount → sapuan penuh diulang (maks
    ``_MAX_SWEEPS``) sampai genap atau tak ada tambahan lagi. rowCount dibaca
    ULANG tiap sapuan karena katalog memang bisa bertambah/berkurang.
    """
    found: dict[Any, dict[str, Any]] = {}
    row_count = 0
    for sweep in range(1, _MAX_SWEEPS + 1):
        before = len(found)
        first = _search_retry(start=0, limit=_PAGE_SIZE)
        row_count = int((first.get("sp") or {}).get("rowCount") or 0)
        for it in (first.get("d") or []):
            found[_item_key(it)] = it
        # Guard loop: server aneh (rowCount melonjak / offset tak pernah habis)
        # tak boleh membuat sapuan berputar tanpa batas.
        max_pages = row_count // _PAGE_SIZE + 3
        start, pages = _PAGE_SIZE, 1
        while start < row_count and pages < max_pages:
            batch = _search_retry(start=start, limit=_PAGE_SIZE).get("d") or []
            if not batch:
                # Kosong DI TENGAH (start < rowCount) = bukan akhir data. Coba
                # sekali lagi; masih kosong → LEWATI offset ini dan teruskan.
                # ⛔ JANGAN break: itu yang dulu memotong ekor katalog.
                batch = _search_retry(start=start, limit=_PAGE_SIZE).get("d") or []
                if not batch:
                    logger.warning(
                        "[accurate] halaman kosong di offset %d (rowCount %d) — "
                        "dilewati, sapuan diteruskan (ekor katalog tak dipotong)",
                        start, row_count)
            for it in batch:
                found[_item_key(it)] = it
            start += _PAGE_SIZE
            pages += 1
        if not row_count or len(found) >= row_count:
            break
        if len(found) == before:
            break                      # sapuan ulang tak menambah apa pun → sudahi
        logger.info("[accurate] sapuan %d baru %d dari rowCount %d — sapu ulang",
                    sweep, len(found), row_count)
    if row_count and len(found) < row_count:
        logger.warning(
            "[accurate] sapuan katalog KURANG: %d unik dari rowCount %d — sebagian "
            "barang tak masuk indeks siklus ini (bukti 2026-08-05: 5200 vs 5229, "
            "yang hilang justru barang paling laku)", len(found), row_count)
    return list(found.values())


# ── Parsing PN ─────────────────────────────────────────────────────────────
# Kode barang Accurate: "NNNNNN.<PN><+suffix>", mis. "003330.WG9725220536+305".
# Prefix "NNNNNN." = nomor urut internal Accurate (bukan bagian PN). Untuk barang
# non-part (oli/grease) sisa setelah titik berupa nama produk, bukan PN — pemanggil
# yang mencocokkan ke katalog memutuskan validitasnya.
_CODE_PREFIX_RE = re.compile(r"^\d{4,}\.")
_PN_SUFFIX_RE = re.compile(r"\+\d+$")


def parse_pn(no: str) -> str:
    """Ambil kandidat Part Number dari field ``no`` Accurate.

    "003330.WG9725220536+305" → "WG9725220536"
    "000674.SNPC TULUX ... 200L" → "SNPC TULUX ... 200L" (bukan PN — biar pemanggil menilai)
    """
    s = (no or "").strip()
    s = _CODE_PREFIX_RE.sub("", s, count=1)  # buang "003330."
    s = _PN_SUFFIX_RE.sub("", s.strip())     # buang "+305"
    return s.strip()


def _norm_pn(pn: str) -> str:
    """Normalisasi PN untuk pencocokan (uppercase, buang spasi & pemisah umum)."""
    return re.sub(r"[\s\-_/]", "", (pn or "").upper())


def norm_pn(pn: str) -> str:
    """Publik: normalisasi PN (dipakai router untuk cocokkan ke snapshot)."""
    return _norm_pn(pn)


def _pn_base_key(pn: str) -> str:
    """Kunci KELUARGA VARIAN dari sebuah PN: potong di '/' atau '+' PERTAMA lalu
    normalisasi. '1000442956/SN' → '1000442956'; 'WG9525160004' → dirinya.

    Dipakai atas field ``pn`` (hasil parse_pn), BUKAN ``no`` — ``no`` masih
    berprefix nomor urut Accurate ('000951.…') yang bukan bagian PN."""
    base = re.split(r"[/+]", (pn or "").strip())[0]
    return _norm_pn(base)


def _build_by_base(by_pn: dict[str, dict[str, Any]]) -> dict[str, list[str]]:
    """Peta KELUARGA: kunci PN dasar → daftar kunci by_pn anggotanya (stok desc).

    FAKTA LIVE 2026-08-05: satu part fisik bisa dipecah jadi BEBERAPA kartu
    barang Accurate per PEMASOK, dibedakan suffix huruf. Contoh PN 1000442956
    (Fuel Filter): '000969.1000442956' 482 pc @Rp 300rb, '000951.1000442956/SN'
    1.069 pc @Rp 285rb, '000993.1000442956/SH' 2 pc @Rp 455rb — part SAMA
    (kata pemilik), pemasok & harga beda.

    Semua entri masuk peta, termasuk keluarga beranggota satu, supaya pemanggil
    tak perlu membedakan 'punya keluarga' vs 'tidak'. Peta ini TURUNAN dari
    by_pn (jangan dipersist — selalu dibangun ulang)."""
    out: dict[str, list[str]] = {}
    for key, entry in (by_pn or {}).items():
        kb = _pn_base_key((entry or {}).get("pn") or key)
        if not kb:
            continue
        out.setdefault(kb, []).append(key)
    for kb, keys in out.items():
        # Stok terbanyak dulu = kartu pemasok yang paling mungkin dipakai;
        # kunci sbg pemecah seri agar urutan DETERMINISTIK antar-restart.
        keys.sort(key=lambda k: (-_num(((by_pn.get(k) or {}).get("available_to_sell"))), k))
    return out


def index_key(pn: str) -> str:
    """Kunci indeks Accurate untuk sebuah PN — PEMAAF terhadap SUFFIX VARIAN.

    Katalog & EPC memakai PN ber-suffix varian pemasok/halaman ('WG9525160004/2',
    'YG9525230005/1'), sementara Accurate menyimpan PN DASAR ('WG9525160004').
    Normalisasi biasa membuang '/' → 'WG95251600042' ≠ 'WG9525160004', jadi stok &
    harga part yang ADA dilaporkan '—' (kasus nyata: kampas kopling, 11 pc).
    Urutan: kunci apa adanya → kunci PN dasar (potong di '/' atau '+') → satu-satunya
    kartu varian dari keluarga PN dasar itu. '' bila semuanya tak ada di indeks."""
    by = _index_cache.get("by_pn") or {}
    k = _norm_pn(pn)
    if not by:
        return k          # indeks belum dimuat → tak bisa verifikasi; pakai kunci apa adanya
    if k and k in by:
        return k
    base = re.split(r"[/+]", (pn or "").strip().upper())[0]
    kb = _norm_pn(base)
    if kb and kb in by:
        return kb
    # Langkah-3 (2026-08-05) — arah SEBALIKNYA. Sampai hari ini pemaafan hanya
    # SATU ARAH (query ber-suffix → PN dasar), jadi query PN dasar '1000442956'
    # MELESET dari kunci '1000442956SN'/'1000442956SH' yang ada di indeks. Efek
    # nyata di halaman part: total stok '—' padahal tabel per-gudang terisi
    # (by_gudang dari report memuat ketiga kartu, by_pn cuma satu).
    # Bila keluarga PN dasar itu hanya punya SATU kartu → itu jawabannya
    # (aturan pemilik: "PN suffix varian = part SAMA").
    # ⛔ ≥2 kartu (kasus nyata 1000442956 punya 3) = AMBIGU harga/stok → JANGAN
    # menebak salah satu; kembalikan '' dan biarkan stock_family menyajikan
    # semuanya apa adanya.
    cands = [c for c in ((_index_cache.get("by_base") or {}).get(kb) or []) if c in by]
    return cands[0] if len(cands) == 1 else ""


def _num(v: Any) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def normalize_item(it: dict[str, Any]) -> dict[str, Any]:
    """Bentuk ringkas 1 barang Accurate untuk konsumsi internal."""
    no = str(it.get("no") or "")
    unit = ((it.get("unit1") or {}) or {}).get("name") or it.get("unit1Name") or ""
    return {
        "no": no,
        "pn": parse_pn(no),
        "name": (it.get("name") or "").strip(),
        "available_to_sell": _num(it.get("availableToSell")),
        "quantity": _num(it.get("quantity")),
        "unit": str(unit).strip(),
        "item_type": (it.get("itemTypeName") or "").strip(),
        # Harga JUAL satuan-1 dari Accurate; fallback ke branchPrice bila unitPrice 0.
        "price": _num(it.get("unitPrice")) or _num(it.get("branchPrice")),
        "accurate_id": it.get("id"),
    }


# ── Indeks stok (cache TTL) ────────────────────────────────────────────────
_index_lock = threading.Lock()
# by_gudang: {norm_pn: {warehouseName: qty}} — rincian per-gudang, diisi enrichment
# latar (enrich_warehouses) sekali per siklus 5-jam & DIBAGI ke semua fitur (stock_full,
# stok_gudang) tanpa panggilan live per-PN. gudang_ts = kapan enrichment terakhir tuntas.
# by_base: {pn_dasar: [kunci by_pn…]} — peta KELUARGA VARIAN PEMASOK (satu part
# dipecah jadi beberapa kartu barang bersuffix '/SN', '/SH'…). TURUNAN dari by_pn,
# dibangun ulang tiap refresh/_load_index/rekonsiliasi — TIDAK dipersist ke disk.
_index_cache: dict[str, Any] = {"ts": 0.0, "items": [], "by_pn": {}, "by_gudang": {},
                                "by_base": {}, "gudang_ts": 0.0}


def _build_by_pn(items: Iterable[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """Peta PN-ternormalisasi → barang stok tertinggi (bila duplikat PN)."""
    by_pn: dict[str, dict[str, Any]] = {}
    for raw in items:
        n = normalize_item(raw)
        key = _norm_pn(n["pn"])
        if not key:
            continue
        prev = by_pn.get(key)
        if prev is None or n["available_to_sell"] > prev["available_to_sell"]:
            by_pn[key] = n
    return by_pn


def refresh(force: bool = False) -> dict[str, Any]:
    """Bangun/segarkan indeks stok penuh dari Accurate.

    ⛔ ATURAN KERAS PEMILIK: indeks HANYA diperbarui pada jadwal tetap (3×/hari)
    ATAU permintaan admin eksplisit (force=True). ``force=False`` SELALU
    mengembalikan cache apa adanya — TAK PERNAH menarik/login on-demand, sekalipun
    cache tua. Jadi buka part/cari/harga tak pernah menyentuh Accurate."""
    if not force:
        return _index_cache
    with _index_lock:
        items = fetch_all_items()
        if not items:
            # Glitch sesi/izin bisa balas 200-KOSONG. Menimpa indeks dgn [] akan
            # mengosongkan etalase + harga checkout sampai window terjadwal berikut
            # (≤12 jam) — dan _save_index mem-persist kosong itu. Pertahankan indeks
            # lama; jangan timpa, jangan simpan. Cache tua > cache kosong.
            logger.warning(
                "[accurate] refresh dapat 0 barang — indeks lama (%d) DIPERTAHANKAN",
                len(_index_cache.get("items") or []),
            )
            return _index_cache
        _index_cache["items"] = [normalize_item(x) for x in items]
        _index_cache["by_pn"] = _build_by_pn(items)
        _index_cache["by_base"] = _build_by_base(_index_cache["by_pn"])
        # View ringkas utk overlay hasil pencarian (dipakai snapshot()) — dibangun
        # sekali di sini agar tiap request pencarian tinggal baca dict jadi.
        _index_cache["snap"] = {
            k: {"stok": v["available_to_sell"], "harga": v["price"], "unit": v["unit"]}
            for k, v in _index_cache["by_pn"].items()
        }
        _index_cache["ts"] = time.time()
        return _index_cache


# ── Enrichment stok PER-GUDANG ke indeks ────────────────────────────────────
# Rincian per-gudang TIDAK ikut tarikan massal search-item.do (itu agregat "[Semua
# Cabang]"); hanya endpoint per-item (view-itemstock-bywarehouse.do) yang punya. Karena
# Accurate men-serialkan panggilan per-sesi (~0,12 dtk/barang; ~8 mnt utk ~3.700 barang
# berstok), enrichment dijalankan SERIAL di THREAD LATAR (bukan di refresh() yang bisa
# dipicu on-demand) sekali per siklus 5-jam, lalu HASILNYA DIBAGI ke semua fitur.
_GUDANG_ENRICH_PUBLISH_EVERY = 100   # terbitkan hasil parsial tiap N barang
_gudang_enrich_running = False
_gudang_enrich_lock = threading.Lock()


def _warehouse_map(item_id: Any) -> dict[str, float]:
    """{warehouseName: qty>0} untuk 1 item id (live, 1 panggilan). {} bila gagal."""
    try:
        wd = _warehouse_retry(item_id).get("d") or {}
    except AccurateError:
        return {}
    out: dict[str, float] = {}
    for w in wd.get("detailWarehouseData") or []:
        qty = _num(w.get("balance"))
        if qty > 0:
            name = w.get("warehouseName") or w.get("name") or ""
            if name:
                out[name] = qty
    return out


def gudang_breakdown(part_number: str) -> dict[str, float]:
    """Rincian stok {warehouseName: qty} 1 PN dari INDEKS (hasil enrichment 5-jam).
    {} bila belum ter-enrich / PN tak ada. Non-blocking, tanpa panggilan live.
    PN ber-suffix varian ('…/2') dicocokkan ke PN dasar lewat index_key."""
    key = index_key(part_number) or _norm_pn(part_number)
    if not key:
        return {}
    return dict((_index_cache.get("by_gudang") or {}).get(key, {}))


def gudang_enriched_count() -> int:
    """Jumlah PN yang sudah punya rincian per-gudang di indeks (0 = belum siap)."""
    return len(_index_cache.get("by_gudang") or {})


# Batas tambalan per siklus. Rekonsiliasi = 1 panggilan search per kunci; 80 sudah
# jauh di atas kebocoran nyata (29 barang, 2026-08-05) dan menjaga window login
# tetap pendek. Lebih dari itu = ada yang salah di sapuan → dicatat, bukan diborong.
_RECONCILE_MAX = 80

# Kunci indeks = PN TANPA tanda pisah ('1000442956/SN' → '1000442956SN'), padahal
# kode barang di Accurate MASIH memakai '/'. Pola ini memisahkan ekor huruf agar
# kita punya kata kunci cadangan yang pasti cocok di sisi server.
_EKOR_HURUF_RE = re.compile(r"^(.{5,}?)([A-Z]{1,3})$")


def _keyword_kandidat(key: str) -> list[str]:
    """Kata kunci pencarian untuk satu KUNCI indeks, urut dari paling tepat.

    Mencari '1000442956SN' apa adanya bisa 0 hasil karena kode aslinya
    '000951.1000442956/SN'; kata kunci cadangan '1000442956' pasti mengembalikan
    SELURUH kartu keluarga. Hasilnya tetap disaring kecocokan PERSIS, jadi kata
    kunci yang lebih longgar tak berisiko salah-cocok."""
    out = [key]
    m = _EKOR_HURUF_RE.match(key)
    if m and m.group(1) not in out:
        out.append(m.group(1))
    return out


def _reconcile_report_keys(by_g: dict[str, Any]) -> int:
    """Tambal barang yang LOLOS dari sapuan katalog, pakai Report per-gudang
    sebagai daftar pembanding.

    Report 'Kuantitas Barang per Gudang' adalah satu query server-side (bukan
    paging offset), jadi ia TIDAK ikut bocor seperti fetch_all_items. Kunci yang
    ada di report tapi TIDAK ada di by_pn = bukti hitam-putih barang itu hilang
    dari indeks siklus ini (bukti 2026-08-05: PN 1000442956 & varian '/SN' —
    justru yang paling laku). Untuk tiap kunci begitu kita tarik SATU barang via
    ``search_by_keyword`` (server-side, bukan paging) lalu sisipkan ke indeks.

    ⛔ ATURAN KERAS PEMILIK: HANYA dipanggil dari jalur enrichment/refresh
    TERJADWAL — TIDAK PERNAH dari jalur request (request selalu baca cache saja).
    Best-effort: kegagalan jaringan dilewati, tak pernah menjatuhkan enrichment.
    Return jumlah barang yang berhasil ditambal."""
    by_pn = _index_cache.get("by_pn") or {}
    if not by_g or not by_pn:
        # by_pn kosong = indeks memang belum dibangun; menambal satu-satu di sini
        # bukan obatnya (dan bisa ribuan panggilan). Diamkan.
        return 0
    kurang = sorted(k for k in by_g if k not in by_pn)
    if not kurang:
        return 0
    if len(kurang) > _RECONCILE_MAX:
        logger.warning("[accurate] %d kunci report tak ada di indeks — hanya %d "
                       "pertama yang ditambal siklus ini", len(kurang), _RECONCILE_MAX)
        kurang = kurang[:_RECONCILE_MAX]
    items = _index_cache.setdefault("items", [])
    snap = _index_cache.setdefault("snap", {})
    tertambal = 0
    gagal_beruntun = 0
    for key in kurang:
        cocok: list[dict[str, Any]] = []
        try:
            for kw in _keyword_kandidat(key):
                cocok = [h for h in search_by_keyword(kw, limit=20)
                         if _norm_pn(h.get("pn") or "") == key]
                if cocok:
                    break
            gagal_beruntun = 0
        except AccurateError as e:
            gagal_beruntun += 1
            logger.warning("[accurate] rekonsiliasi '%s' gagal: %s", key, e)
            if gagal_beruntun >= 3:
                logger.warning("[accurate] rekonsiliasi dihentikan (3 kegagalan beruntun)")
                break
            continue
        if not cocok:
            continue                     # kunci report tak terpetakan ke barang → lewati
        n = max(cocok, key=lambda h: _num(h.get("available_to_sell")))
        by_pn[key] = n
        items.append(n)
        snap[key] = {"stok": n["available_to_sell"], "harga": n["price"], "unit": n["unit"]}
        tertambal += 1
    if tertambal:
        _index_cache["by_base"] = _build_by_base(by_pn)   # keluarga ikut berubah
        logger.info("[accurate] rekonsiliasi report→indeks menambal %d barang yang "
                    "lolos dari sapuan katalog", tertambal)
    return tertambal


def enrich_warehouses() -> int:
    """Isi indeks per-gudang: tarik rincian gudang utk SEMUA barang berstok>0 (serial,
    santun ke Accurate) → simpan ke _index_cache['by_gudang'] = {norm_pn:{gudang:qty}}.
    Terbitkan parsial berkala agar query dapat data lebih awal. Return jumlah PN
    ter-enrich. Dipanggil dari thread latar terjadwal; skip bila sedang berjalan."""
    global _gudang_enrich_running
    with _gudang_enrich_lock:
        if _gudang_enrich_running:
            return gudang_enriched_count()
        _gudang_enrich_running = True
    try:
        if not available():
            return gudang_enriched_count()
        idx = refresh(force=False)
        in_stock = [(k, v) for k, v in (idx.get("by_pn") or {}).items()
                    if _num(v.get("available_to_sell")) > 0 and v.get("accurate_id") is not None]
        built: dict[str, dict[str, float]] = {}
        for i, (key, v) in enumerate(in_stock, 1):
            built[key] = _warehouse_map(v.get("accurate_id"))
            if i % _GUDANG_ENRICH_PUBLISH_EVERY == 0:
                _index_cache["by_gudang"] = dict(built)   # publish parsial
        if not built:
            # Nol PN ter-enrich (mis. by_pn kosong / semua available_to_sell 0)
            # — jangan timpa indeks per-gudang lama yang bagus dgn {}.
            prev = gudang_enriched_count()
            if prev:
                logger.warning("[accurate] enrichment per-PN 0 hasil — indeks lama "
                               "(%d) DIPERTAHANKAN", prev)
                return prev
        _index_cache["by_gudang"] = built
        _index_cache["gudang_ts"] = time.time()
        try:                            # tambal barang yang lolos sapuan katalog
            _reconcile_report_keys(built)
        except Exception as e:  # pragma: no cover - jaring pengaman
            logger.warning("[accurate] rekonsiliasi report→indeks gagal: %s", e)
        return len(built)
    finally:
        with _gudang_enrich_lock:
            _gudang_enrich_running = False


# ── Enrichment per-gudang CEPAT via Report (Kuantitas Barang per Gudang) ──────
# Gantikan enrich_warehouses() yang 3.700 panggilan per-PN (~8–15 mnt) dengan
# SATU laporan server-side yang di-export XLS (crosstab item×gudang) — ~4
# panggilan, hitungan detik. Semua di host iris-report. Lihat memory
# accurate-stok-report-cepat.
_STOCK_REPORT_ID = 503
_STOCK_REPORT_PLAN = "QuantityItemByWarehouseReport"


def _stock_report_xls(session: dict[str, str]) -> bytes:
    """Jalankan Report Kuantitas Barang per Gudang → XLSX. Best-effort raise."""
    rep, dsi, jsid = _report_base(), session["dsi"], session["jsessionid"]
    hdr = _sq_headers(dsi, jsid)
    usi = _harvest_usi(session)

    # 1) input default laporan (berisi 31 gudang + role + [Semua Cabang]).
    #    init-report-input.do ada di host UTAMA (iris), bukan iris-report.
    r1 = requests.post(f"{_base()}/accurate/report/init-report-input.do",
                       data={"id": _STOCK_REPORT_ID, "planId": _STOCK_REPORT_PLAN, "_dsi": dsi},
                       headers=hdr, timeout=_HTTP_TIMEOUT, allow_redirects=False)
    if _looks_expired(r1):
        raise AccurateSessionExpired("Sesi kadaluarsa saat init report stok.")
    ri_str = ((r1.json().get("d") or {}) or {}).get("reportInput")
    if not ri_str:
        raise AccurateError("init-report-input tak mengembalikan reportInput.")
    ri = json.loads(ri_str)
    ri.setdefault("param", {})["asOfDate"] = time.strftime("%d/%m/%Y")

    # 2) jalankan (background) → bgPid
    r2 = requests.post(f"{rep}/accurate/report/bg-execute-report.do",
                       data={"id": _STOCK_REPORT_ID, "planId": _STOCK_REPORT_PLAN,
                             "reportInput": json.dumps(ri, ensure_ascii=False),
                             "pageIndex": 0, "_usi": usi, "_dsi": dsi},
                       headers=hdr, timeout=_HTTP_TIMEOUT, allow_redirects=False)
    bgpid = r2.json().get("b")
    if not bgpid:
        raise AccurateError(f"bg-execute-report gagal: {r2.text[:150]}")

    # 3) poll sampai FINISHED → cacheId
    cache_id = None
    for _ in range(120):                      # maks ~2 mnt
        rp = requests.post(f"{rep}/accurate/company/bg-proc-response.do",
                           data={"bgPid": bgpid, "keepCache": "true", "_usi": usi, "_dsi": dsi},
                           headers=hdr, timeout=_HTTP_TIMEOUT, allow_redirects=False)
        d = rp.json().get("d") or {}
        if d.get("status") == "FINISHED":
            cache_id = ((d.get("response") or {}) or {}).get("cacheId")
            break
        time.sleep(1)
    if not cache_id:
        raise AccurateError("Report stok tak selesai (timeout poll).")

    # 4) export XLSX (crosstab item×gudang)
    r4 = requests.post(f"{rep}/accurate/report/export-report.do",
                       data={"cacheId": cache_id, "exportType": "xls", "_usi": usi, "_dsi": dsi},
                       headers=hdr, timeout=_HTTP_TIMEOUT + 45, allow_redirects=False)
    if r4.content[:2] != b"PK":               # xlsx = arsip ZIP (magic PK)
        raise AccurateError("Export report bukan XLSX.")
    return r4.content


def _parse_stock_report(data: bytes) -> dict[str, dict[str, float]]:
    """Parse XLSX crosstab → {norm_pn: {gudang: qty>0}}. Kolom TOTAL diabaikan.
    Kode barang ('NNNNNN.PN') → PN via parse_pn (SAMA dgn by_pn → kunci cocok)."""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True)   # read_only gagal utk file ini
    ws = wb.active
    mr, mc = ws.max_row, ws.max_column
    hi = next((i for i in range(1, 20)
               if str(ws.cell(i, 1).value or "").strip().lower() == "kode barang"), None)
    if not hi:
        raise AccurateError("Header 'Kode Barang' tak ditemukan di report XLS.")
    # kolom gudang = kolom 6+ ber-header, KECUALI kolom Total.
    wh_cols = []
    for j in range(6, mc + 1):
        name = str(ws.cell(hi, j).value or "").strip()
        if name and "total" not in name.lower():
            wh_cols.append((j, name))
    out: dict[str, dict[str, float]] = {}
    for i in range(hi + 1, mr + 1):
        code = ws.cell(i, 1).value
        if not code:
            continue
        pn = _norm_pn(parse_pn(str(code)))
        if not pn:
            continue
        g = {}
        for j, name in wh_cols:
            q = _num(ws.cell(i, j).value)
            if q > 0:
                g[name] = q
        if g:
            out[pn] = g
    return out


def enrich_warehouses_via_report() -> int:
    """Isi indeks per-gudang via Report (CEPAT). Fallback ke enrich_warehouses
    (per-PN lambat) bila report gagal. Return jumlah PN ter-enrich."""
    def _do(sess):
        return _parse_stock_report(_stock_report_xls(sess))

    sess = _ensure_session()
    try:
        by_g = _do(sess)
    except AccurateSessionExpired:
        if not credentials_configured():
            raise
        by_g = _do(_refresh_session())
    except AccurateError as e:
        logger.warning("[accurate] enrichment via report gagal (%s) — fallback per-PN", e)
        return enrich_warehouses()
    if not by_g:
        # Report HTTP sukses tapi parsing 0 PN (format report berubah / baris
        # bergeser / semua qty 0). Menimpa by_gudang dgn {} membuat etalase &
        # checkout 'habis' semua (gudang_breakdown kosong) sampai window
        # berikutnya, DAN _save_index mem-persist kosong itu. Guard sama dgn
        # refresh: cache tua > cache kosong.
        prev = gudang_enriched_count()
        if prev:
            logger.warning("[accurate] report enrichment 0 PN — indeks per-gudang "
                           "lama (%d) DIPERTAHANKAN, tak ditimpa kosong", prev)
            return prev
        logger.warning("[accurate] report enrichment 0 PN & belum ada indeks — "
                       "fallback per-PN")
        return enrich_warehouses()
    _index_cache["by_gudang"] = by_g
    _index_cache["gudang_ts"] = time.time()
    # Report tak ikut bocor spt paging offset → pakai sbg pembanding untuk
    # menambal barang yang lolos dari sapuan katalog. Best-effort: gagal di sini
    # tak boleh membatalkan enrichment yang sudah berhasil.
    try:
        _reconcile_report_keys(by_g)
    except Exception as e:  # pragma: no cover - jaring pengaman
        logger.warning("[accurate] rekonsiliasi report→indeks gagal: %s", e)
    logger.info("[accurate] enrichment per-gudang via REPORT OK (%d PN, ~detik)", len(by_g))
    return len(by_g)


def stock_for(part_number: str, force: bool = False) -> dict[str, Any] | None:
    """Stok Accurate untuk 1 PN (atau None bila tak ada di Accurate)."""
    key = _norm_pn(part_number)
    if not key:
        return None
    idx = refresh(force=force)
    return idx["by_pn"].get(key)


def all_items(force: bool = False) -> list[dict[str, Any]]:
    """Seluruh barang Accurate (ternormalisasi) dari indeks ber-cache TTL.

    Dipakai menu Stok untuk menampilkan katalog stok penuh. Auto-login saat sesi
    habis (via ``refresh`` → ``fetch_all_items``)."""
    return list(refresh(force=force)["items"])


def available() -> bool:
    """True bila Accurate bisa dipakai: ada kredensial auto-login ATAU file sesi
    manual (tanpa menembak jaringan)."""
    if credentials_configured():
        return True
    try:
        load_session()
        return True
    except AccurateSessionMissing:
        return False


# ── Refresh indeks TERJADWAL (jam WIB tetap) + PERSISTENSI DISK ──────────────
# ⛔ ATURAN KERAS PEMILIK: indeks ditarik dari Accurate HANYA 3× sehari pada jam
# WIB tetap (07:00, 12:00, 19:00). DEPLOY/RESTART TIDAK menarik ulang — indeks
# dimuat dari DISK (tanpa login). Jadi sesi Accurate hanya dipegang di 3 window
# itu (± enrichment ~8 mnt), lalu LANGSUNG dilepas (logout). Bootstrap SEKALI
# hanya bila disk kosong sama sekali (agar stok/harga tak kosong).
_REFRESH_HOURS_WIB = (7, 12, 19)
_WIB = _dt.timezone(_dt.timedelta(hours=7))
_sched_lock = threading.Lock()
_sched_started = False


def _index_file() -> Path:
    return get_settings().data_path / "accurate_index.json"


def _save_index() -> None:
    """Simpan indeks ke disk (JSON, atomik) → restart berikutnya muat tanpa login."""
    try:
        c = _index_cache
        payload = {k: c.get(k) for k in ("items", "by_pn", "by_gudang", "snap", "ts", "gudang_ts")}
        fp = _index_file()
        fp.parent.mkdir(parents=True, exist_ok=True)
        tmp = fp.with_suffix(".tmp")
        tmp.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        tmp.replace(fp)
    except Exception as e:  # pragma: no cover
        logger.warning("[accurate] simpan indeks ke disk gagal: %s", e)


def _load_index() -> bool:
    """Muat indeks dari disk (TANPA login). True bila ada & berisi."""
    try:
        raw = json.loads(_index_file().read_text(encoding="utf-8"))
    except (FileNotFoundError, OSError, ValueError):
        return False
    if not raw.get("items"):
        return False
    with _index_lock:
        for k in ("items", "by_pn", "by_gudang", "snap", "ts", "gudang_ts"):
            if k in raw:
                _index_cache[k] = raw[k]
        # by_base TURUNAN — dibangun ulang di sini, sengaja TIDAK ikut dipersist
        # (kalau ikut, file lama tanpa kunci itu akan menyisakan peta basi).
        _index_cache["by_base"] = _build_by_base(_index_cache.get("by_pn") or {})
    logger.info("[accurate] indeks dimuat dari DISK (%d barang) — tanpa login Accurate",
                len(raw["items"]))
    return True


def _seconds_until_next_refresh() -> float:
    """Detik sampai jam refresh WIB terjadwal berikutnya (07/12/19 WIB)."""
    now = _dt.datetime.now(_WIB)
    cands = []
    for d in (0, 1):
        day = (now + _dt.timedelta(days=d))
        for h in _REFRESH_HOURS_WIB:
            cands.append(day.replace(hour=h, minute=0, second=0, microsecond=0))
    nxt = min(t for t in cands if t > now)
    return max(1.0, (nxt - now).total_seconds())


def _do_scheduled_refresh() -> None:
    """Satu refresh terjadwal: tarik agregat + enrichment per-gudang → SIMPAN ke
    disk → LEPAS sesi (logout). Best-effort; kegagalan tak menjatuhkan loop."""
    if not available():
        return
    try:
        refresh(force=True)
        logger.info("[accurate] refresh terjadwal OK (%d barang); enrichment per-gudang…",
                    len(_index_cache["items"]))
        try:
            n = enrich_warehouses_via_report()   # CEPAT: 1 laporan XLS, ~detik
            logger.info("[accurate] enrichment per-gudang OK (%d PN)", n)
        except Exception as e:
            logger.warning("[accurate] enrichment per-gudang gagal (%s) — agregat tetap dipakai", e)
        _save_index()
    except Exception as e:
        logger.warning("[accurate] refresh terjadwal gagal: %s", e)
    finally:
        # Lepas sesi SEGERA setelah window — jangan pegang di luar 3 jam terjadwal
        # (akun 1-sesi; admin harus bisa buka Accurate manual).
        try:
            logout()
        except Exception:  # pragma: no cover
            pass


def _scheduled_refresh_loop() -> None:
    while True:
        time.sleep(_seconds_until_next_refresh())
        _do_scheduled_refresh()


def start_scheduled_refresh() -> bool:
    """Mulai penjadwalan refresh indeks. Idempoten. Muat indeks dari DISK dulu
    (tanpa login); bootstrap SEKALI hanya bila disk kosong. Loop menunggu jam
    WIB terjadwal berikutnya — TIDAK menarik saat start/deploy."""
    global _sched_started
    with _sched_lock:
        if _sched_started:
            return False
        _sched_started = True
    loaded = _load_index()
    if not loaded and available():
        # Disk kosong (mis. pertama kali) → bootstrap SEKALI agar stok tak kosong.
        threading.Thread(target=_do_scheduled_refresh, daemon=True,
                         name="accurate-index-bootstrap").start()
    threading.Thread(target=_scheduled_refresh_loop, daemon=True,
                     name="accurate-index-refresh").start()
    return True


# ── Snapshot utk overlay hasil PENCARIAN — VIEW dari indeks bersama ─────────
# Sumber SATU: indeks terjadwal 3×/hari (start_scheduled_refresh), dibagi ke
# semua fitur. View "snap" dibangun saat refresh() — di sini tinggal baca,
# non-blocking, request pencarian tak pernah menunggu tarikan / login.


def warehouse_names() -> list[str]:
    """Semua nama gudang yang muncul di rincian per-gudang indeks (mis. '01.Jakarta').
    Nama-nama ini identik dengan label di gudang_config/no_ship — sumber tunggal
    daftar gudang untuk scoping & fallback terdekat."""
    names: set[str] = set()
    for m in (_index_cache.get("by_gudang") or {}).values():
        names.update(m)
    return sorted(names)


def snapshot() -> dict[str, dict[str, Any]]:
    """{norm_pn: {stok,harga,unit}} dari indeks 5-jam bersama (tanpa tarikan
    terpisah). Bisa kosong (cold start / Accurate down) — pemanggil fallback Excel."""
    return _index_cache.get("snap") or {}


def index_stamp() -> tuple[float, float]:
    """(ts agregat, ts enrichment per-gudang) dari indeks aktif. Berubah TIAP
    refresh yang benar-benar menimpa data — dipakai fingerprint etalase agar
    perubahan HARGA (jumlah item sama) tetap memicu rebuild katalog."""
    return (
        float(_index_cache.get("ts") or 0.0),
        float(_index_cache.get("gudang_ts") or 0.0),
    )


# Batas umur indeks agregat untuk TRANSAKSI UANG (checkout). Refresh terjadwal
# 3×/hari; 24 jam = beberapa window gagal berturut-turut → harga/stok tak bisa
# dipertanggungjawabkan untuk menagih pembeli. TIDAK dipakai memblokir
# pencarian/etalase (yang boleh best-effort).
_INDEX_MAX_AGE_CHECKOUT = 24 * 3600.0


def index_too_old_for_checkout() -> bool:
    """True bila indeks agregat Accurate terlalu tua (atau belum ada) untuk jadi
    dasar checkout — harga bisa basi & stok bisa oversell vs ERP. `ts` dipersist
    ke disk, jadi restart dgn indeks masih segar TIDAK ikut terblokir."""
    ts = float(_index_cache.get("ts") or 0.0)
    if ts <= 0:
        return True
    return (time.time() - ts) > _INDEX_MAX_AGE_CHECKOUT


def items_matching(terms: Iterable[str], *, limit: int = 400) -> list[dict[str, Any]]:
    """Barang di indeks yang NAMA/PN-nya memuat SALAH SATU `terms` (word-boundary,
    case-insensitive). Beda dari search_index (yg skor+batas kecil utk aftermarket):
    ini mengumpulkan SEMUA yang cocok (utk daftar kategori spt 'kopling'/'rem') —
    in-memory, cepat (satu pindai). Return normalize_item (pn, name, price…)."""
    items = _index_cache.get("items") or []
    if not items:
        return []
    pats = [re.compile(r"(?<!\w)" + re.escape(t.strip().lower()) + r"(?!\w)")
            for t in terms if t and len(t.strip()) >= 3]
    if not pats:
        return []
    out: list[dict[str, Any]] = []
    for it in items:
        hay = f"{it.get('name') or ''} {it.get('pn') or ''}".lower()
        if any(p.search(hay) for p in pats):
            out.append(it)
            if len(out) >= limit:
                break
    return out


def search_index(terms: Iterable[str], *, limit: int = 8) -> list[dict[str, Any]]:
    """Cari INDEKS stok bersama per NAMA/PN — substring, case-insensitive, murah
    (in-memory, NON-BLOCKING: baca cache seperti snapshot(), TIDAK memicu tarikan).

    Indeks stok memuat barang yang TIDAK ada di katalog Sinotruk (aftermarket/
    lokal, sering bernama Indonesia: 'Kaca Spion LH', 'Alternator Regulator',
    'Cucuk Per Depan Faw') — satu-satunya jalan menemukannya per kata kunci.
    `terms` = query asli + ekspansi sinonim. Skor: frasa utuh di nama/PN lebih
    tinggi dari kecocokan kata tunggal; hasil diurut skor, lalu stok, lalu nama.
    [] bila indeks kosong (cold start) — pemanggil tinggal melewatkan."""
    items = _index_cache.get("items") or []
    if not items:
        return []
    # Frasa pendek/generik ('per', 'oli') terlalu banjir → minimal 4 huruf, dan
    # cocok per BATAS KATA (bukan substring di tengah kata: 'per' ≠ 'Super').
    frasa = [t.strip().lower() for t in terms if t and len(t.strip()) >= 4]
    if not frasa:
        return []
    kata = {w for f in frasa for w in re.split(r"\s+", f) if len(w) >= 3}

    def _cocok(needle: str, hay: str) -> bool:
        return re.search(r"(?<!\w)" + re.escape(needle) + r"(?!\w)", hay) is not None

    scored: list[tuple[float, dict[str, Any]]] = []
    for it in items:
        hay = f"{it.get('name') or ''} {it.get('pn') or ''} {it.get('no') or ''}".lower()
        score = 0.0
        for f in frasa:
            hit = (f in hay) if " " in f else _cocok(f, hay)
            if hit:
                score += 10.0 + len(f) / 10.0   # frasa lebih panjang = lebih spesifik
        if not score:
            score = sum(2.0 for w in kata if _cocok(w, hay))
            if score < 4.0:                     # < 2 kata cocok = terlalu lemah (noise)
                score = 0.0
        if score:
            scored.append((score, it))
    scored.sort(key=lambda s: (-s[0], -(s[1].get("available_to_sell") or 0.0),
                               s[1].get("name") or ""))
    return [it for _s, it in scored[:limit]]


# ═══════════════════════════════════════════════════════════════════════════
#  PENAWARAN PENJUALAN (Sales Quotation / SQ) — buat dokumen + tarik PDF resmi
#
#  Dibedah dari HAR UI Accurate (2026-07-10). Kunci yang tak terlihat dari luar:
#    • SEMUA field header/detail berprefix `param.` (form-urlencoded, bukan JSON)
#    • WAJIB token `_usi` (user-session-id) DI SAMPING `_dsi`; `_usi` dipanen dari
#      respons init-sales-quotation.do (tertanam sbg `_usi=…` di dalam teks)
#    • Total & pajak DIHITUNG Accurate via calculate-header (config-agnostic) —
#      kita TIDAK menebak rumus PPN
#  PDF resmi ditarik 2 langkah dari host `*-report.accurate.id`.
#  Lihat memory: accurate-penawaran-api.
# ═══════════════════════════════════════════════════════════════════════════
_USI_RE = re.compile(r"_usi=([A-Za-z0-9+/=]{40,})")
_qdefaults_cache: dict[str, Any] = {"at": 0.0, "val": None}
_QDEFAULTS_TTL = 6 * 3600


def _report_base() -> str:
    """Host mesin cetak: iris.accurate.id → iris-report.accurate.id."""
    return f"https://{get_settings().accurate_host}".replace(".accurate.id", "-report.accurate.id")


def _harvest_usi(session: dict[str, str],
                 path: str = "customer/init-sales-quotation.do") -> str:
    """Ambil token `_usi` dari respons init-<modul>.do (satu-satunya sumber).
    `path` bisa modul lain (mis. vendor/init-purchase-requisition.do) — tokennya
    milik SESI, tapi hanya keluar lewat halaman init sebuah modul."""
    dsi = session["dsi"]
    headers = {
        "User-Agent": _UA, "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest", "Origin": _base(),
        "Referer": f"{_base()}/accurate/?_dsi={dsi}",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "Cookie": f"JSESSIONID={session['jsessionid']}",
    }
    r = requests.post(f"{_base()}/accurate/{path}",
                      data={"_dsi": dsi}, headers=headers, timeout=_HTTP_TIMEOUT, allow_redirects=False)
    if _looks_expired(r):
        raise AccurateSessionExpired("Sesi Accurate kadaluarsa saat init dokumen.")
    m = _USI_RE.search(r.text)
    if not m:
        raise AccurateError("Gagal memperoleh token _usi dari init penawaran.")
    return m.group(1)


def _company_defaults(session: dict[str, str]) -> dict[str, Any]:
    """Default header perusahaan (branch/currency/tax1/warehouse) dari dokumen
    terbaru — dibaca sekali & di-cache. Menghindari hardcode id yang bisa beda
    antar-perusahaan. countAutoNumber dari view dokumen."""
    now = time.time()
    if _qdefaults_cache["val"] and (now - _qdefaults_cache["at"]) < _QDEFAULTS_TTL:
        return _qdefaults_cache["val"]
    rows = _call(session, "customer/search-sales-quotation.do", {"start": 0, "limit": 1}).get("d") or []
    if not rows:
        raise AccurateError("Tak ada dokumen penawaran acuan untuk membaca default perusahaan.")
    v = _call(session, "customer/view-sales-quotation.do", {"id": rows[0]["id"]}).get("d") or {}
    det0 = (v.get("detailItem") or [{}])[0]
    val = {
        "branchId": v.get("branchId") or 50,
        "currencyId": v.get("currencyId") or 50,
        "tax1Id": v.get("tax1Id") or 50,
        "warehouseId": det0.get("warehouseId") or v.get("branchId") or 50,
        "countAutoNumber": v.get("countAutoNumber") or 3,
    }
    _qdefaults_cache.update(at=now, val=val)
    return val


def _sq_headers(dsi: str, jsessionid: str) -> dict[str, str]:
    return {
        "User-Agent": _UA, "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest", "Origin": _base(),
        "Referer": f"{_base()}/accurate/?_dsi={dsi}",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "Cookie": f"JSESSIONID={jsessionid}",
    }


def search_customers(keyword: str, *, limit: int = 15) -> list[dict[str, Any]]:
    """Cari pelanggan (untuk field 'Dipesan oleh'). Auto-login saat sesi habis."""
    kw = (keyword or "").strip()
    if not kw:
        return []

    def _do(sess):
        rows = _call(sess, "customer/search-customer.do",
                     {"start": 0, "limit": limit, "keywords": kw}).get("d") or []
        out = []
        for r in rows:
            # Alamat ikut dikirim supaya admin bisa membedakan pelanggan bernama
            # mirip — persis daftar 'Dipesan oleh' di Accurate sendiri.
            alamat = (r.get("fullShipAddress") or "").strip()
            if not alamat:
                bill = r.get("billAddress") or {}
                alamat = str(bill.get("address") or "").strip() if isinstance(bill, dict) else ""
            out.append({"id": r.get("id"), "no": r.get("customerNo"),
                        "name": (r.get("name") or "").strip(), "address": alamat})
        return out

    sess = _ensure_session()
    try:
        return _do(sess)
    except AccurateSessionExpired:
        if not credentials_configured():
            raise
        return _do(_refresh_session())


def item_for_quotation(part_number: str) -> dict[str, Any] | None:
    """Cari 1 barang untuk baris penawaran: {id, no, pn, name, unit_id, unit, price}.
    PN harus PERSIS cocok (hindari salah barang). None bila tak ada."""
    want = _norm_pn(part_number)
    if not want:
        return None
    hits = search_by_keyword(part_number, limit=50)  # normalize_item; punya accurate_id
    exact = [h for h in hits if _norm_pn(h["pn"]) == want]
    if not exact:
        return None
    best = max(exact, key=lambda h: h["available_to_sell"])
    # unit1.id tak disimpan normalize_item → ambil dari hasil mentah search-item sekali.
    unit_id = 100
    sess = _ensure_session()
    try:
        rd = _search_items_raw(sess, keywords=part_number, start=0, limit=50).get("d") or []
    except AccurateSessionExpired:
        rd = _search_items_raw(_refresh_session(), keywords=part_number, start=0, limit=50).get("d") or []
    for it in rd:
        if it.get("id") == best["accurate_id"]:
            unit_id = ((it.get("unit1") or {}) or {}).get("id") or 100
            break
    return {
        "id": best["accurate_id"], "no": best["no"], "pn": best["pn"],
        "name": best["name"], "unit_id": unit_id, "unit": best["unit"],
        "price": best["price"], "available": best["available_to_sell"],
    }


def _calc_totals(session, usi, defaults, customer_id, lines, taxable, inclusive):
    """Minta Accurate menghitung subTotal/pajak/total (config-agnostic)."""
    body = {
        "_dsi": session["dsi"], "_usi": usi,
        "param.customerId": customer_id, "param.currencyId": defaults["currencyId"],
        "param.rate": 1, "param.branchId": defaults["branchId"],
        "param.tax1Id": defaults["tax1Id"],
        "param.taxable": "true" if taxable else "false",
        "param.inclusiveTax": "true" if inclusive else "false",
        "param.forceCalculatePercentTaxable": "true",
    }
    for i, ln in enumerate(lines):
        p = f"param.detailItem[{i}]."
        body[p + "seq"] = i + 1
        body[p + "itemId"] = ln["item_id"]
        body[p + "quantity"] = ln["qty"]
        body[p + "unitPrice"] = ln["unit_price"]
        body[p + "itemUnitId"] = ln.get("unit_id", 100)
        body[p + "useTax1"] = "true" if taxable else "false"
    r = requests.post(f"{_base()}/accurate/customer/calculate-header-sales-quotation.do",
                      data=body, headers=_sq_headers(session["dsi"], session["jsessionid"]),
                      timeout=_HTTP_TIMEOUT, allow_redirects=False)
    if _looks_expired(r):
        raise AccurateSessionExpired("Sesi kadaluarsa saat hitung total.")
    j = r.json()
    if not j.get("s"):
        raise AccurateError(f"Hitung total ditolak: {str(j.get('d'))[:150]}")
    return j.get("d") or {}


def _save_quotation(session, usi, defaults, *, number, customer_id, lines,
                    transdate, taxable, inclusive, description, totals,
                    ignore_warning=False):
    """POST save-sales-quotation.do. Return dict hasil ('r').
    Accurate bisa membalas PERINGATAN (d.w_) — mis. anjuran setel DPP 11/12 —
    yang di UI ditembus dgn tombol 'lanjutkan'. Kita kirim ulang sekali dgn
    ignoreWarning=true (setara klik 'lanjutkan'), bukan menganggapnya gagal."""
    body = {
        "_dsi": session["dsi"], "_usi": usi,
        "param.uniqueDataNumber": int(time.time() * 1000),
        "param.needDetailResult": "false",
        "param.ignoreWarning": "true" if ignore_warning else "false",
        "param.number": number,                       # MANUAL — bukan penomoran otomatis
        "param.countAutoNumber": defaults["countAutoNumber"],
        "param.customerId": customer_id,
        "param.currencyId": defaults["currencyId"], "param.rate": 1,
        "param.transDate": transdate,
        "param.branchId": defaults["branchId"],
        "param.tax1Id": defaults["tax1Id"],
        "param.taxable": "true" if taxable else "false",
        "param.inclusiveTax": "true" if inclusive else "false",
        "param.forceCalculatePercentTaxable": "true",
        "param.saveAsStatusType": "UNAPPROVED",
        "param.transactionCurrencyId": defaults["currencyId"],
        "param.attachments": "[]",
        "param.description": description or "",
        "param.cashDiscount": 0, "param.lastCashDiscount": 0, "param.totalExpense": 0,
        # total & pajak DIHITUNG Accurate (bukan tebakan kita)
        "param.subTotal": totals.get("subTotal", 0),
        "param.totalAmount": totals.get("subTotal", 0) if inclusive else totals.get("totalAmount", 0),
        "param.tax1Amount": totals.get("tax1Amount", 0),
        "param.tax2Amount": 0, "param.tax3Amount": 0, "param.tax4Amount": 0,
        "param.tax1Rate": totals.get("tax1Rate", 0),
        "param.percentTaxable": totals.get("percentTaxable", 100),
    }
    for i, ln in enumerate(lines):
        p = f"param.detailItem[{i}]."
        qty, price = ln["qty"], ln["unit_price"]
        body[p + "_status"] = "insert"
        body[p + "seq"] = i + 1
        body[p + "itemId"] = ln["item_id"]
        body[p + "detailName"] = ln.get("name", "")
        body[p + "quantity"] = qty
        body[p + "itemUnitId"] = ln.get("unit_id", 100)
        body[p + "unitRatio"] = 1
        body[p + "unitPrice"] = price
        body[p + "totalPrice"] = round(qty * price, 2)
        body[p + "warehouseId"] = defaults["warehouseId"]
        body[p + "useTax1"] = "true" if taxable else "false"
        body[p + "useTax2"] = "false"
        body[p + "useTax3"] = "false"
        body[p + "useTax4"] = "false"
    r = requests.post(f"{_base()}/accurate/customer/save-sales-quotation.do",
                      data=body, headers=_sq_headers(session["dsi"], session["jsessionid"]),
                      timeout=_HTTP_TIMEOUT + 15, allow_redirects=False)
    if _looks_expired(r):
        raise AccurateSessionExpired("Sesi kadaluarsa saat simpan penawaran.")
    j = r.json()
    if not j.get("s"):
        d = j.get("d") or {}
        # PERINGATAN saja (bukan error validasi) & belum coba abaikan → ulangi
        # sekali dgn ignoreWarning=true (setara tombol 'lanjutkan' di UI).
        if not ignore_warning and isinstance(d, dict) and set(d.keys()) <= {"w_"}:
            return _save_quotation(session, usi, defaults, number=number, customer_id=customer_id,
                                   lines=lines, transdate=transdate, taxable=taxable,
                                   inclusive=inclusive, description=description, totals=totals,
                                   ignore_warning=True)
        raise AccurateError(f"Accurate menolak simpan penawaran: {str(d)[:200]}")
    return j.get("r") or {}


_MASPART_NUM_RE = re.compile(r"MASPART-0*(\d+)\s*$", re.I)


def next_quotation_number() -> str:
    """Nomor penawaran MASPART BERIKUTNYA: 'MASPART-NN' (NN = nomor MASPART
    tertinggi yang sudah ada + 1, mulai 01). Dihitung dari dokumen Accurate →
    tahan restart & tak pernah bentrok.

    ⛔⛔ PENOMORAN OTOMATIS ACCURATE TIDAK PERNAH DIPAKAI. MASPART selalu memasok
    nomornya sendiri (dikirim sbg param.number manual di _save_quotation). Aturan
    keras dari pemilik: toggle 'Penomoran Otomatis' harus SELALU MATI."""
    def _do(sess):
        rows = _call(sess, "customer/search-sales-quotation.do",
                     {"start": 0, "limit": 200, "keywords": "MASPART"}).get("d") or []
        mx = 0
        for r in rows:
            m = _MASPART_NUM_RE.match(str(r.get("number") or "").strip())
            if m:
                mx = max(mx, int(m.group(1)))
        return f"MASPART-{mx + 1:02d}"

    sess = _ensure_session()
    try:
        return _do(sess)
    except AccurateSessionExpired:
        if not credentials_configured():
            raise
        return _do(_refresh_session())


def create_sales_quotation(*, number: str, customer_id: int, lines: list[dict],
                           transdate: str, taxable: bool = True, inclusive: bool = True,
                           description: str = "") -> dict[str, Any]:
    """Buat Penawaran Penjualan. `lines` = [{item_id, name, qty, unit_price, unit_id}].
    `number` WAJIB (manual). Return {id, number, total}. Auto-login saat sesi habis."""
    if not number.strip():
        raise AccurateError("Nomor penawaran wajib diisi (manual).")
    if not lines:
        raise AccurateError("Minimal satu baris barang.")

    def _flow(sess):
        usi = _harvest_usi(sess)
        defs = _company_defaults(sess)
        totals = _calc_totals(sess, usi, defs, customer_id, lines, taxable, inclusive)
        return _save_quotation(sess, usi, defs, number=number, customer_id=customer_id,
                               lines=lines, transdate=transdate, taxable=taxable,
                               inclusive=inclusive, description=description, totals=totals)

    sess = _ensure_session()
    try:
        r = _flow(sess)
    except AccurateSessionExpired:
        if not credentials_configured():
            raise
        r = _flow(_refresh_session())
    return {"id": r.get("id"), "number": r.get("number") or number,
            "total": _num(r.get("totalAmount"))}


def sales_quotation_pdf(quotation_id: int, *, layout_id: int = 50) -> bytes:
    """Tarik PDF RESMI Accurate untuk 1 penawaran (2 langkah, host *-report).
    layout_id: 50=Default, 551=after disc, 700=Service-JNT, 450=Proforma."""

    def _flow(sess):
        usi = _harvest_usi(sess)
        rep, dsi = _report_base(), sess["dsi"]
        hdr = _sq_headers(dsi, sess["jsessionid"])
        r1 = requests.post(f"{rep}/accurate/company/view-print-layout-execute.do",
                           data={"dataId": quotation_id, "printLayoutId": layout_id,
                                 "transactionType": "SQ", "_usi": usi, "_dsi": dsi},
                           headers=hdr, timeout=_HTTP_TIMEOUT, allow_redirects=False)
        if _looks_expired(r1):
            raise AccurateSessionExpired("Sesi kadaluarsa saat siapkan cetak.")
        cache_id = ((r1.json().get("d") or {}) or {}).get("cacheId")
        if not cache_id:
            raise AccurateError(f"Gagal menyiapkan cetak: {r1.text[:150]}")
        r2 = requests.get(f"{rep}/accurate/report/export-report.do",
                          params={"_dsi": dsi, "_usi": usi, "cacheId": cache_id, "exportType": "pdf"},
                          headers=hdr, timeout=_HTTP_TIMEOUT + 30, allow_redirects=False)
        if r2.content[:4] != b"%PDF":
            raise AccurateError("Respons cetak bukan PDF (sesi/izin?).")
        return r2.content

    sess = _ensure_session()
    try:
        return _flow(sess)
    except AccurateSessionExpired:
        if not credentials_configured():
            raise
        return _flow(_refresh_session())


# ═══════════════════════════════════════════════════════════════════════════
#  PERMINTAAN BARANG (Purchase Requisition / PR) — modul `vendor/`
#
#  Dibedah dari HAR UI Accurate (iris.accurate.id11.har, 2026-08-06). Sama pola
#  dengan Penawaran (param.* form-urlencoded + _dsi & _usi), dengan tiga beda:
#    • SIMPAN LEWAT LATAR: bg-save-purchase-requisition.do balas {b: bgPid},
#      hasilnya baru muncul saat polling company/bg-proc-response.do FINISHED.
#    • Header khas: requisitionType=PURCHASE, canProcess/canTransfer, TANPA
#      pelanggan/gudang (ini dokumen permintaan, bukan transaksi barang).
#    • TIGA kolom WAJIB di tiap baris (custom field perusahaan ini — labelnya
#      dibaca dari init-purchase-requisition.do):
#         charField1 = "Sektor"  charField2 = "No Unit"  charField4 = "Kts Jkt"
#      Kosongkan salah satunya → Accurate menolak simpan.
#  Aturan pemilik 2026-08-06: No Unit selalu "STOK", Kts Jkt = stok Jakarta saat
#  ini (otomatis dari indeks), nomor dokumen dipasok MASPART (PERMINTAAN-NN).
# ═══════════════════════════════════════════════════════════════════════════
_PR_NUM_RE = re.compile(r"PERMINTAAN-0*(\d+)\s*$", re.I)
_prdefaults_cache: dict[str, Any] = {"at": 0.0, "val": None}
_PR_SEKTOR_DEFAULT = "MASPART"
_PR_NO_UNIT_DEFAULT = "STOK"


def _pr_defaults(session: dict[str, str]) -> dict[str, Any]:
    """Default header (branch/currency/tax/countAutoNumber) dari PR TERAKHIR —
    dibaca sekali & di-cache. Pola sama dgn _company_defaults: jangan hardcode id
    yang bisa berbeda antar perusahaan."""
    now = time.time()
    if _prdefaults_cache["val"] and (now - _prdefaults_cache["at"]) < _QDEFAULTS_TTL:
        return _prdefaults_cache["val"]
    rows = _call(session, "vendor/search-purchase-requisition.do",
                 {"start": 0, "limit": 1, "sp.pageSize": 1, "sp.start": 0,
                  "sp.limit": 1}).get("d") or []
    if not rows:
        raise AccurateError("Tak ada dokumen Permintaan Barang acuan untuk membaca "
                            "default perusahaan.")
    v = _call(session, "vendor/view-purchase-requisition.do",
              {"id": rows[0]["id"]}).get("d") or {}
    val = {
        "branchId": v.get("branchId") or 50,
        "currencyId": v.get("currencyId") or 50,
        "tax1Id": v.get("tax1Id") or 50,
        "countAutoNumber": v.get("countAutoNumber") or 4,
    }
    _prdefaults_cache.update(at=now, val=val)
    return val


def next_pr_number() -> str:
    """Nomor Permintaan Barang BERIKUTNYA: 'PERMINTAAN-NN'. Dihitung dari dokumen
    Accurate sendiri (tahan restart, tak bentrok) — sama seperti penawaran,
    penomoran otomatis Accurate TIDAK dipakai.

    ⚠️ `keywords` TIDAK menyaring berdasarkan NOMOR di endpoint ini (terbukti live
    2026-08-06: kirim 'PERMINTAAN' tetap membalas dokumen ber-nomor P.R.U./PRD.),
    jadi penyaringan dilakukan DI SINI atas 200 dokumen terbaru bertipe PURCHASE
    — jangan mengandalkan filter server."""
    def _do(sess):
        rows = _call(sess, "vendor/search-purchase-requisition.do",
                     {"start": 0, "limit": 200, "keywords": "",
                      "purchaseRequisitionTypeFilter": json.dumps(["PURCHASE"]),
                      "sp.pageSize": 200, "sp.start": 0, "sp.limit": 200}).get("d") or []
        mx = 0
        for r in rows:
            m = _PR_NUM_RE.match(str(r.get("number") or "").strip())
            if m:
                mx = max(mx, int(m.group(1)))
        return f"PERMINTAAN-{mx + 1:02d}"

    sess = _ensure_session()
    try:
        return _do(sess)
    except AccurateSessionExpired:
        if not credentials_configured():
            raise
        return _do(_refresh_session())


def _pr_detail_body(lines: list[dict], transdate: str) -> dict[str, Any]:
    """Bagian detailItem[i].* untuk calculate & save (satu sumber, jadi angka yang
    dihitung Accurate persis atas baris yang akan disimpan)."""
    body: dict[str, Any] = {}
    for i, ln in enumerate(lines):
        p = f"param.detailItem[{i}]."
        body[p + "_status"] = "insert"
        body[p + "seq"] = i + 1
        body[p + "itemId"] = ln["item_id"]
        body[p + "detailName"] = ln.get("name", "")
        body[p + "quantity"] = ln["qty"]
        body[p + "itemUnitId"] = ln.get("unit_id", 100)
        body[p + "unitRatio"] = 1
        # Permintaan Barang tak memutuskan harga — biar bagian pembelian yang
        # mengisinya. Semua nol, persis seperti dokumen contoh pemilik.
        body[p + "unitPrice"] = 0
        body[p + "totalPrice"] = 0
        body[p + "useTax1"] = "true"
        body[p + "useTax2"] = "false"
        body[p + "useTax3"] = "false"
        body[p + "taxableAmount3"] = 0
        body[p + "requiredDate"] = ln.get("required_date") or transdate
        body[p + "orderedQuantity"] = 0
        body[p + "receivedQuantity"] = 0
        body[p + "transferQuantity"] = 0
        body[p + "manualClosed"] = "false"
        body[p + "manualClosedVisible"] = "false"
        # ⛔ WAJIB — Accurate menolak baris bila salah satu kosong.
        body[p + "charField1"] = ln.get("sektor") or _PR_SEKTOR_DEFAULT
        body[p + "charField2"] = ln.get("no_unit") or _PR_NO_UNIT_DEFAULT
        body[p + "charField4"] = ln.get("kts_jkt", 0)
    return body


def _calc_totals_pr(session, usi, defaults, lines, transdate) -> dict[str, Any]:
    """Minta Accurate menghitung total/pajak PR (kita tak menebak rumus PPN)."""
    body = {
        "_dsi": session["dsi"], "_usi": usi,
        "param.uniqueDataNumber": int(time.time() * 1000),
        "param.needDetailResult": "false",
        "param.transDate": transdate,
        "param.requisitionType": "PURCHASE",
        "param.currencyId": defaults["currencyId"], "param.rate": 1,
        "param.branchId": defaults["branchId"], "param.tax1Id": defaults["tax1Id"],
        "param.taxable": "true", "param.forceCalculatePercentTaxable": "true",
        "param.canProcess": "true", "param.canTransfer": "true",
        **_pr_detail_body(lines, transdate),
    }
    r = requests.post(f"{_base()}/accurate/vendor/calculate-header-purchase-requisition.do",
                      data=body, headers=_sq_headers(session["dsi"], session["jsessionid"]),
                      timeout=_HTTP_TIMEOUT, allow_redirects=False)
    if _looks_expired(r):
        raise AccurateSessionExpired("Sesi kadaluarsa saat hitung total permintaan.")
    j = r.json()
    if not j.get("s"):
        raise AccurateError(f"Hitung total permintaan ditolak: {str(j.get('d'))[:150]}")
    return j.get("d") or {}


def _bg_tunggu(session: dict[str, str], usi: str, bgpid: str,
               batas_detik: int = 90) -> dict[str, Any]:
    """Tunggu proses LATAR Accurate selesai → isi `d.response`. Simpan PR memakai
    jalur ini (bg-save), jadi 's:true' pertama BELUM berarti tersimpan."""
    hdr = _sq_headers(session["dsi"], session["jsessionid"])
    for _ in range(batas_detik):
        rp = requests.post(f"{_base()}/accurate/company/bg-proc-response.do",
                           data={"bgPid": bgpid, "keepCache": "true",
                                 "_usi": usi, "_dsi": session["dsi"]},
                           headers=hdr, timeout=_HTTP_TIMEOUT, allow_redirects=False)
        if _looks_expired(rp):
            raise AccurateSessionExpired("Sesi kadaluarsa saat menunggu proses latar.")
        d = rp.json().get("d") or {}
        if d.get("status") == "FINISHED":
            return d.get("response") or {}
        if d.get("status") in ("FAILED", "ERROR"):
            raise AccurateError(f"Proses latar Accurate gagal: {str(d)[:200]}")
        time.sleep(1)
    raise AccurateError("Proses simpan di Accurate tak selesai (timeout).")


def _save_pr(session, usi, defaults, *, number, lines, transdate, description,
             totals) -> dict[str, Any]:
    """POST bg-save-purchase-requisition.do lalu tunggu hasil latar → dict 'r'."""
    body = {
        "_dsi": session["dsi"], "_usi": usi,
        "param.uniqueDataNumber": int(time.time() * 1000),
        "param.needDetailResult": "false",
        "param.attachmentCount": 0, "param.commentCount": 0,
        "param.number": number,                       # MANUAL — bukan penomoran otomatis
        "param.countAutoNumber": defaults["countAutoNumber"],
        "param.transDate": transdate,
        "param.description": description or "",
        "param.requisitionType": "PURCHASE",
        "param.currencyId": defaults["currencyId"], "param.rate": 1,
        "param.branchId": defaults["branchId"], "param.tax1Id": defaults["tax1Id"],
        "param.taxable": "true", "param.forceCalculatePercentTaxable": "true",
        "param.saveAsStatusType": "UNAPPROVED",
        "param.canProcess": "true", "param.canTransfer": "true",
        "param.attachments": "[]",
        "param.manualClosed": "false", "param.manualClosedVisible": "false",
        # total & pajak DIHITUNG Accurate (bukan tebakan kita)
        "param.subTotal": totals.get("subTotal", 0),
        "param.totalAmount": totals.get("totalAmount", 0),
        "param.tax1Amount": totals.get("tax1Amount", 0),
        "param.tax1Rate": totals.get("tax1Rate", 0),
        "param.percentTaxable": totals.get("percentTaxable", 100),
        **_pr_detail_body(lines, transdate),
    }
    r = requests.post(f"{_base()}/accurate/vendor/bg-save-purchase-requisition.do",
                      data=body, headers=_sq_headers(session["dsi"], session["jsessionid"]),
                      timeout=_HTTP_TIMEOUT + 15, allow_redirects=False)
    if _looks_expired(r):
        raise AccurateSessionExpired("Sesi kadaluarsa saat simpan permintaan barang.")
    j = r.json()
    if not j.get("s") or not j.get("b"):
        raise AccurateError(f"Accurate menolak simpan permintaan: {str(j.get('d'))[:200]}")
    resp = _bg_tunggu(session, usi, j["b"])
    if not resp.get("s"):
        raise AccurateError(f"Simpan permintaan gagal: {str(resp.get('d'))[:200]}")
    return resp.get("r") or {}


def create_purchase_requisition(*, number: str, lines: list[dict], transdate: str,
                                description: str = "") -> dict[str, Any]:
    """Buat PERMINTAAN BARANG (Purchase Requisition).

    `lines` = [{item_id, name, qty, unit_id, sektor, no_unit, kts_jkt}] —
    tiga yang terakhir mengisi kolom WAJIB Sektor/No Unit/Kts Jkt.
    `number` WAJIB (manual, lihat next_pr_number). Return {id, number, jumlah_baris}."""
    if not (number or "").strip():
        raise AccurateError("Nomor permintaan barang wajib diisi (manual).")
    if not lines:
        raise AccurateError("Minimal satu baris barang.")

    def _flow(sess):
        usi = _harvest_usi(sess, "vendor/init-purchase-requisition.do")
        defs = _pr_defaults(sess)
        totals = _calc_totals_pr(sess, usi, defs, lines, transdate)
        return _save_pr(sess, usi, defs, number=number, lines=lines, transdate=transdate,
                        description=description, totals=totals)

    sess = _ensure_session()
    try:
        r = _flow(sess)
    except AccurateSessionExpired:
        if not credentials_configured():
            raise
        r = _flow(_refresh_session())
    return {"id": r.get("id"), "number": r.get("number") or number,
            "jumlah_baris": len(lines)}


# ⛔⛔ SENGAJA TIDAK ADA fungsi hapus/ubah penawaran (delete/update/void). Aturan
# keras pemilik: aplikasi HANYA boleh MEMBUAT penawaran (kuantitas part) — tak
# boleh mengubah atau menghapus apa pun di Accurate. Jangan menambah endpoint
# tulis Accurate lain tanpa persetujuan pemilik.


# ── CLI selftest (tanpa server) ────────────────────────────────────────────
if __name__ == "__main__":  # pragma: no cover
    import sys

    if len(sys.argv) > 1 and sys.argv[1] == "--all":
        items = fetch_all_items()
        with_stock = [i for i in items if _num(i.get("availableToSell")) > 0]
        print(f"Total barang: {len(items)} | berstok>0: {len(with_stock)}")
    else:
        pn = sys.argv[1] if len(sys.argv) > 1 else "WG9725220536"
        hit = stock_for_live(pn)
        print(f"{pn} -> {hit}")
