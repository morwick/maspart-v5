"""
Service: Catalog BOM — baca data/catalog_bom.json (dibangun oleh
backend/tools/build_catalog_bom.py dari SEMUA sheet kategori tiap unit).

Mendukung dua sumbu perbandingan:
  1. ANTAR-UNIT per kategori   — compare_units(unit1, unit2, kategori)
  2. ANTAR-PN ASSY             — compare_assy(pn1, pn2)   (gearbox/clutch/axle/…)

File ~7-8MB → di-cache per-mtime (parse sekali, dipakai ulang sampai file berubah),
tetap 'segar' bila data di-scp ulang (mtime berubah → reload otomatis).
"""
from __future__ import annotations

import glob
import json
import os
import re
from pathlib import Path

from ..core.config import get_settings

# ── Cache per-mtime ──────────────────────────────────────────────────────────
_CACHE: dict = {"mtime": None, "data": {}}


def _path():
    return get_settings().data_path / "catalog_bom.json"


def _load() -> dict:
    try:
        p = _path()
        mt = p.stat().st_mtime if p.exists() else None
    except Exception:
        mt = None
    if mt != _CACHE["mtime"]:
        data = {}
        try:
            if mt is not None:
                data = json.loads(_path().read_text(encoding="utf-8")) or {}
        except Exception:
            data = {}
        _CACHE.update(mtime=mt, data=data)
    return _CACHE["data"]


def available() -> bool:
    return bool(_load().get("units"))


def _norm(s: str) -> str:
    return re.sub(r"[\s_\-/]", "", (s or "")).upper()


# ── Kategori: kode 01..12 + sinonim istilah lapangan Indonesia ───────────────
# Urutan penting (frasa spesifik dulu). Tiap entri: (kode, [trigger...]).
_KAT_SINONIM = [
    ("06", ["poros penumpu", "driven axle", "gardan depan", "as depan", "axle depan", "front axle", "从动桥"]),
    ("07", ["poros penggerak", "drive axle", "gardan belakang", "gardan penggerak", "as belakang",
            "differential", "diferensial", "gardan", "drive", "驱动桥"]),
    ("01", ["kabin", "cabin", "cab", "ruang kemudi", "kepala", "驾驶室"]),
    ("02", ["powertrain", "engine", "mesin", "动力总成"]),
    ("03", ["aksesori powertrain", "aksesoris mesin", "kelengkapan mesin", "动力总成附件"]),
    ("04", ["kopling", "clutch", "matahari", "离合器"]),
    ("05", ["transmisi", "gearbox", "gear box", "persneling", "perseneling", "girboks", "变速箱", "变速器"]),
    ("08", ["kelistrikan", "kelistikan", "listrik", "electrical", "electric", "kabel", "电器"]),
    ("09", ["rem", "brake", "pengereman", "制动"]),
    ("10", ["sasis", "chassis", "rangka", "底盘"]),
    ("11", ["lainnya", "lain-lain", "others", "其他"]),
    ("12", ["karoseri", "bak", "upper body", "上装", "truck loading"]),
]


def categories() -> dict:
    return _load().get("kategori", {})


def kategori_nama(code: str) -> str:
    return categories().get(code, code)


def resolve_kategori(query: str) -> str | None:
    """Query (kode / nama / istilah lapangan) -> kode kategori '01'..'12'."""
    q = (query or "").strip().lower()
    if not q:
        return None
    cats = categories()
    m = re.match(r"\s*(\d{2})\b", q)
    if m and m.group(1) in cats:
        return m.group(1)
    for code, triggers in _KAT_SINONIM:
        if code not in cats:
            continue
        if any(t in q for t in triggers):
            return code
    # fallback: cocokkan ke nama tampil kategori
    for code, nama in cats.items():
        if q in nama.lower():
            return code
    return None


# ── Unit ─────────────────────────────────────────────────────────────────────
def resolve_unit(query: str) -> list[str]:
    """Query -> daftar nama unit yang cocok (persis dulu, lalu substring)."""
    units = _load().get("units", {})
    if not units or not (query or "").strip():
        return []
    qn = _norm(query)
    exact = [u for u in units if _norm(u) == qn]
    if exact:
        return exact
    return sorted(u for u in units if qn in _norm(u))


def list_units() -> list[str]:
    return sorted(_load().get("units", {}).keys())


def _cat_parts(unit: str, code: str) -> list[dict]:
    return (_load().get("units", {}).get(unit, {}).get("kategori", {})
            .get(code, {}).get("parts", []))


def _pnmap(parts: list[dict]) -> dict[str, str]:
    out: dict[str, str] = {}
    for p in parts:
        pn = _norm(p.get("pn", ""))
        if pn:
            out.setdefault(pn, p.get("nama") or "")
    return out


def _verdict(jaccard: float, only1: int, only2: int) -> tuple[str, str]:
    pct = round(jaccard * 100, 1)
    if only1 == 0 and only2 == 0:
        return "identik", "Isi part 100% sama persis."
    if jaccard >= 0.95:
        return "praktis_identik", f"Praktis identik ({pct}% sama) — beda hanya segelintir part."
    if jaccard >= 0.75:
        return "sangat_mirip", f"Sangat mirip ({pct}% sama) — inti sama, beda minor (kemungkinan varian/versi)."
    if jaccard >= 0.45:
        return "mirip_satu_keluarga", f"Mirip / satu keluarga ({pct}% sama) tapi cukup banyak part berbeda."
    return "berbeda", f"Berbeda signifikan (hanya {pct}% sama)."


def _diff(s1: dict[str, str], s2: dict[str, str], cap: int) -> dict:
    set1, set2 = set(s1), set(s2)
    inter, only1, only2 = set1 & set2, set1 - set2, set2 - set1
    union = set1 | set2
    if not union:
        # Kedua sisi tak punya PN yg bisa dibandingkan (data kosong/garbled). JANGAN
        # klaim 'identik 100%' — itu kesamaan palsu.
        return {
            "jumlah_1": 0, "jumlah_2": 0, "jumlah_part_sama": 0,
            "jumlah_hanya_di_1": 0, "jumlah_hanya_di_2": 0, "persen_kesamaan": 0.0,
            "verdict": "tak_dapat_dibandingkan",
            "ringkasan": "Tidak ada PN yang bisa dibandingkan (data part kosong).",
            "hanya_di_1": [], "hanya_di_2": [],
            "hanya_di_1_terpotong": 0, "hanya_di_2_terpotong": 0,
        }
    jac = (len(inter) / len(union))
    kode, ringkas = _verdict(jac, len(only1), len(only2))

    def fmt(pns, src):
        return [{"pn": p, "nama": src.get(p, "")} for p in sorted(pns)[:cap]]

    return {
        "jumlah_1": len(set1), "jumlah_2": len(set2),
        "jumlah_part_sama": len(inter),
        "jumlah_hanya_di_1": len(only1), "jumlah_hanya_di_2": len(only2),
        "persen_kesamaan": round(jac * 100, 1),
        "verdict": kode, "ringkasan": ringkas,
        "hanya_di_1": fmt(only1, s1), "hanya_di_2": fmt(only2, s2),
        "hanya_di_1_terpotong": max(0, len(only1) - cap),
        "hanya_di_2_terpotong": max(0, len(only2) - cap),
    }


# ── 1) Banding ANTAR-UNIT per kategori ───────────────────────────────────────
def compare_units(q_unit1: str, q_unit2: str, q_kat: str, cap: int = 50) -> dict:
    if not available():
        return {"error": "Data katalog BOM belum tersedia di server."}
    code = resolve_kategori(q_kat)
    if not code:
        return {"error": f"Kategori '{q_kat}' tak dikenal. Pilihan: "
                         + ", ".join(f"{k}={v}" for k, v in categories().items())}
    u1, u2 = resolve_unit(q_unit1), resolve_unit(q_unit2)
    if not u1:
        return {"error": f"Unit '{q_unit1}' tidak ditemukan."}
    if not u2:
        return {"error": f"Unit '{q_unit2}' tidak ditemukan."}
    if len(u1) > 1:
        return {"ambigu": 1, "kandidat": u1, "catatan": f"'{q_unit1}' cocok ke banyak unit — sebutkan persis."}
    if len(u2) > 1:
        return {"ambigu": 2, "kandidat": u2, "catatan": f"'{q_unit2}' cocok ke banyak unit — sebutkan persis."}
    u1, u2 = u1[0], u2[0]
    p1, p2 = _cat_parts(u1, code), _cat_parts(u2, code)
    if not p1 and not p2:
        return {"error": f"Kategori '{kategori_nama(code)}' tak ada di kedua unit."}
    res = _diff(_pnmap(p1), _pnmap(p2), cap)
    return {"mode": "antar_unit", "kategori": code, "kategori_nama": kategori_nama(code),
            "unit_1": u1, "unit_2": u2, **res}


def category_parts(q_unit: str, q_kat: str, cap: int = 80) -> dict:
    if not available():
        return {"error": "Data katalog BOM belum tersedia di server."}
    code = resolve_kategori(q_kat)
    if not code:
        return {"error": f"Kategori '{q_kat}' tak dikenal."}
    u = resolve_unit(q_unit)
    if not u:
        return {"error": f"Unit '{q_unit}' tidak ditemukan."}
    if len(u) > 1:
        return {"ambigu": True, "kandidat": u, "catatan": f"'{q_unit}' cocok ke banyak unit."}
    u = u[0]
    cat = _load().get("units", {}).get(u, {}).get("kategori", {}).get(code, {})
    parts = cat.get("parts", [])
    return {"unit": u, "kategori": code, "kategori_nama": kategori_nama(code),
            "assy_pn": cat.get("assy_pn"), "jumlah_part": len(parts),
            "parts": parts[:cap], "terpotong": max(0, len(parts) - cap)}


# ── 2) Banding ANTAR-PN ASSY (kategori assembly) ─────────────────────────────
def resolve_assy(query: str) -> list[dict]:
    """Query PN -> [{raw, kategori, units}] dari assy_index (persis dulu, lalu substring)."""
    idx = _load().get("assy_index", {})
    if not idx or not (query or "").strip():
        return []
    qn = _norm(query)
    if qn in idx:
        return [idx[qn]]
    return [v for k, v in idx.items() if qn in k]


def _assy_repr_parts(entry: dict) -> tuple[str, dict[str, str]]:
    """Unit patokan (part terlengkap) + peta part-nya untuk satu assy entry."""
    code = entry["kategori"]
    best_u, best = None, []
    for u in entry["units"]:
        parts = _cat_parts(u, code)
        if best_u is None or len(parts) > len(best):
            best_u, best = u, parts
    return best_u, _pnmap(best)


def compare_assy(query1: str, query2: str, cap: int = 50) -> dict:
    if not available():
        return {"error": "Data katalog BOM belum tersedia di server."}
    h1, h2 = resolve_assy(query1), resolve_assy(query2)
    if not h1:
        return {"error": f"PN assy '{query1}' tidak ditemukan di indeks assembly."}
    if not h2:
        return {"error": f"PN assy '{query2}' tidak ditemukan di indeks assembly."}
    if len(h1) > 1:
        return {"ambigu": 1, "kandidat": [e["raw"] for e in h1],
                "catatan": f"'{query1}' cocok ke beberapa PN assy — sebutkan persis."}
    if len(h2) > 1:
        return {"ambigu": 2, "kandidat": [e["raw"] for e in h2],
                "catatan": f"'{query2}' cocok ke beberapa PN assy — sebutkan persis."}
    e1, e2 = h1[0], h2[0]
    up1, s1 = _assy_repr_parts(e1)
    up2, s2 = _assy_repr_parts(e2)
    res = _diff(s1, s2, cap)
    beda_kategori = e1["kategori"] != e2["kategori"]
    return {
        "mode": "antar_assy",
        "assy_1": {"part_number": e1["raw"], "kategori": e1["kategori"],
                   "kategori_nama": kategori_nama(e1["kategori"]),
                   "unit_patokan": up1, "units": e1["units"]},
        "assy_2": {"part_number": e2["raw"], "kategori": e2["kategori"],
                   "kategori_nama": kategori_nama(e2["kategori"]),
                   "unit_patokan": up2, "units": e2["units"]},
        "beda_kategori": beda_kategori,
        **res,
    }


def assy_detail(query: str, cap: int = 80) -> dict:
    if not available():
        return {"error": "Data katalog BOM belum tersedia di server."}
    hits = resolve_assy(query)
    if not hits:
        return {"error": f"PN assy '{query}' tidak ditemukan di indeks assembly."}
    if len(hits) > 1:
        return {"ambigu": True, "kandidat": [e["raw"] for e in hits],
                "catatan": f"'{query}' cocok ke beberapa PN assy — sebutkan persis."}
    e = hits[0]
    up, pmap = _assy_repr_parts(e)
    parts = _cat_parts(up, e["kategori"])
    return {"part_number": e["raw"], "kategori": e["kategori"],
            "kategori_nama": kategori_nama(e["kategori"]), "unit_patokan": up,
            "units": e["units"], "jumlah_part": len(parts),
            "parts": parts[:cap], "terpotong": max(0, len(parts) - cap)}


# ── 3) Reverse lookup: KOMPONEN → assembly (transmisi/dll) yang memuatnya ─────
_REV: dict = {"mtime": None, "map": {}}


def _reverse() -> dict:
    """Indeks balik {pn_komponen_norm: {'nama':..., 'assy': {assy_pn: kode_kat}}}.
    Di-cache per-mtime (ikut file catalog_bom.json)."""
    data = _load()
    if _REV["mtime"] != _CACHE["mtime"]:
        m: dict[str, dict] = {}
        for u in data.get("units", {}).values():
            for code, c in u.get("kategori", {}).items():
                apn = c.get("assy_pn")
                if not apn:
                    continue
                for p in c.get("parts", []):
                    pn = _norm(p.get("pn", ""))
                    if not pn:
                        continue
                    e = m.setdefault(pn, {"nama": p.get("nama") or "", "assy": {}})
                    if not e["nama"] and p.get("nama"):
                        e["nama"] = p["nama"]
                    e["assy"][apn] = code
        _REV.update(mtime=_CACHE["mtime"], map=m)
    return _REV["map"]


def part_in_assy(pn: str, cap: int = 40) -> dict:
    """Komponen `pn` termasuk di assembly (transmisi/gardan/dll) mana saja."""
    rev = _reverse()
    e = rev.get(_norm(pn))
    if not e:
        return {"found": 0, "part_number": (pn or "").strip().upper()}
    assy = sorted(e["assy"].keys())
    kats = sorted({e["assy"][a] for a in assy})
    return {
        "found": len(assy),
        "part_number": (pn or "").strip().upper(),
        "nama": e["nama"],
        "kategori": [kategori_nama(k) for k in kats],
        "jumlah_assy": len(assy),
        "assy": assy[:cap],
        "terpotong": max(0, len(assy) - cap),
    }


# ── PN sets / peta kategori (untuk rekonsiliasi & kategorisasi hasil EPC) ─────
def unit_parts(unit_name: str) -> dict[str, dict]:
    """Semua part SATU unit (lintas kategori) → {PN_norm: {'nama', 'kategori'}}.
    Kemunculan pertama menang (PN sama di >1 kategori jarang)."""
    u = _load().get("units", {}).get(unit_name, {})
    out: dict[str, dict] = {}
    for code, c in u.get("kategori", {}).items():
        for p in c.get("parts", []):
            pn = _norm(p.get("pn", ""))
            if pn and pn not in out:
                out[pn] = {"nama": p.get("nama", ""), "kategori": code}
    return out


_PNCAT: dict = {"mtime": None, "map": {}}


def pn_category_map() -> dict[str, dict]:
    """Peta global {PN_norm: {'nama', 'kategori', 'poros_ambigu'}} dari SELURUH unit.
    'kategori' = kemunculan pertama (untuk breakdown). 'poros_ambigu'=True bila PN
    SAMA muncul di BOTH kategori 06 (driven/depan) DAN 07 (drive/belakang) di across
    unit — artinya posisi depan/belakang TAK bisa ditentukan dari katalog (lihat
    _axle_posisi). Di-cache ikut mtime catalog_bom.json."""
    _load()  # pastikan _CACHE['mtime'] mutakhir
    if _PNCAT["mtime"] != _CACHE["mtime"]:
        m: dict[str, dict] = {}
        cats_seen: dict[str, set] = {}  # pn -> {kode kategori}
        for u in _CACHE["data"].get("units", {}).values():
            for code, c in u.get("kategori", {}).items():
                for p in c.get("parts", []):
                    pn = _norm(p.get("pn", ""))
                    if not pn:
                        continue
                    cats_seen.setdefault(pn, set()).add(code)
                    if pn not in m:
                        m[pn] = {"nama": p.get("nama", ""), "kategori": code}
        # Tandai PN yg muncul di poros DEPAN(06) DAN BELAKANG(07) → posisi tak pasti.
        for pn, entry in m.items():
            cs = cats_seen.get(pn, set())
            if "06" in cs and "07" in cs:
                entry["poros_ambigu"] = True
        _PNCAT.update(mtime=_CACHE["mtime"], map=m)
    return _PNCAT["map"]


# ─────────────────────────────────────────────────────────────────────────────
#  BUILD / REBUILD catalog_bom.json — dari sheet kategori (01..12) tiap unit.
#  Dipakai oleh endpoint admin (rebuild in-process) & CLI build_catalog_bom.py.
# ─────────────────────────────────────────────────────────────────────────────

# Nama tampil per kode kategori (istilah lapangan Indonesia).
KATEGORI_NAMA = {
    "01": "Kabin (Driver's cab)",
    "02": "Mesin / Powertrain (动力总成)",
    "03": "Aksesori powertrain (动力总成附件)",
    "04": "Kopling (Clutch)",
    "05": "Transmisi / Gearbox (变速箱)",
    "06": "Poros penumpu / Driven axle (从动桥)",
    "07": "Poros penggerak / Drive axle (驱动桥)",
    "08": "Kelistrikan (Electrical system)",
    "09": "Sistem rem (Brake system)",
    "10": "Sasis / Chassis (底盘)",
    "11": "Lainnya (Others)",
    "12": "Bak / Karoseri (上装 Truck loading)",
}
# Kategori "assembly" (baris pertama sheet = PN assy → masuk assy_index).
ASSEMBLY_CODES = {"01", "02", "04", "05", "06", "07"}
_SKIP_DIRS = ("/stok/", "/harga/", "/populasi/")


def _clean(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def _unit_name(path: str) -> str:
    b = os.path.basename(path)
    return _clean(b.split(" - ")[-1].replace(".xlsx", "")) if " - " in b else b


def _sheet_code(sheet: str) -> str | None:
    """Kode kategori = 2 digit di awal nama sheet (01..12), else None."""
    m = re.match(r"\s*(\d{2})", sheet or "")
    return m.group(1) if m else None


def _parse_sheet(ws) -> tuple[str | None, list[dict]]:
    """-> (assy_pn|None, [ {pn,nama,qty}, ... ]) dari satu sheet kategori."""
    assy = None
    parts: dict[str, dict] = {}
    first = True
    for r in ws.iter_rows(values_only=True):
        c0 = r[0] if len(r) > 0 else None
        c1 = r[1] if len(r) > 1 else None
        cn = r[2] if len(r) > 2 else None
        en = r[3] if len(r) > 3 else None
        qty = r[4] if len(r) > 4 else None
        if not (c1 and str(c1).strip() and c0 is not None and str(c0).strip().isdigit()):
            continue
        pn = str(c1).strip().upper()
        nama = _clean(en) or _clean(cn)
        if first:
            assy = pn
            first = False
        if pn not in parts:
            parts[pn] = {"pn": pn, "nama": nama, "qty": _clean(qty)}
        elif not parts[pn]["nama"] and nama:
            parts[pn]["nama"] = nama
    return assy, list(parts.values())


def build_data(data_path: Path) -> tuple[dict, dict]:
    """Pindai semua file katalog di data_path → (output_dict, stats).
    Import openpyxl lokal (hanya saat rebuild, bukan saat import service)."""
    import openpyxl

    files = [f for f in glob.glob(os.path.join(str(data_path), "**", "*.xlsx"), recursive=True)
             if not any(s in f.replace("\\", "/").lower() for s in _SKIP_DIRS)
             and not os.path.basename(f).startswith("~$")]

    units: dict[str, dict] = {}
    kategori_seen: dict[str, str] = {}
    assy_index: dict[str, dict] = {}
    n_files = 0

    for f in sorted(files):
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except Exception:
            continue
        unit = _unit_name(f)
        rel = os.path.relpath(f, str(data_path)).replace("\\", "/")
        ub = units.setdefault(unit, {"file": rel, "kategori": {}})
        had_any = False
        for s in wb.sheetnames:
            code = _sheet_code(s)
            if not code:
                continue
            assy, parts = _parse_sheet(wb[s])
            if not parts:
                continue
            had_any = True
            kategori_seen.setdefault(code, KATEGORI_NAMA.get(code, _clean(s)))
            is_assy = bool(code in ASSEMBLY_CODES and assy)
            ub["kategori"][code] = {"assy_pn": assy if is_assy else None,
                                    "jumlah": len(parts), "parts": parts}
            if is_assy:
                ai = assy_index.setdefault(_norm(assy), {"raw": assy, "kategori": code, "units": []})
                if unit not in ai["units"]:
                    ai["units"].append(unit)
        wb.close()
        if had_any:
            n_files += 1
        elif not ub["kategori"]:
            units.pop(unit, None)

    out = {
        "kategori": dict(sorted(kategori_seen.items())),
        "units": dict(sorted(units.items())),
        "assy_index": dict(sorted(assy_index.items())),
    }
    stats = {
        "file_katalog_dipindai": len(files),
        "unit_berkategori": len(units),
        "kategori": len(kategori_seen),
        "assy_terindeks": len(assy_index),
        "total_baris_part": sum(c["jumlah"] for u in units.values()
                                for c in u["kategori"].values()),
    }
    return out, stats


def rebuild() -> dict:
    """Bangun ulang catalog_bom.json di DATA_DIR (in-process). Cache di-reset →
    fitur kategori langsung pakai data baru tanpa restart. Return stats."""
    dp = get_settings().data_path
    out, stats = build_data(dp)
    p = dp / "catalog_bom.json"
    p.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    _CACHE["mtime"] = None  # paksa reload pada akses berikutnya
    stats["ok"] = True
    stats["ukuran_kb"] = p.stat().st_size // 1024
    return stats
