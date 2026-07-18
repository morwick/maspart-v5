"""
Excel yang DIUNGGAH USER ke chat asisten: baca → kenali isi kolom → isi kolom baru.

Alur:
  1. `parse_upload(bytes)`  — baca sheet pertama dgn openpyxl read-only (streaming),
     tebak baris header, lalu KENALI PERAN tiap kolom (part number / nama part /
     stok / harga / qty). Hasilnya diringkas untuk asisten.
  2. `put_sheet(user, parsed)` — simpan ke stash server (TTL 2 jam) → `sheet_id`.
     Klien mengirim `sheet_id` di giliran berikutnya sehingga follow-up seperti
     "isikan stoknya" tetap mengacu ke file yang sama.
  3. `fill_column(...)` — isi satu kolom dengan stok / nama part / harga lokal /
     harga SIMS, lalu keluarkan Excel baru lewat `ai_export.stash_export`.

KEAMANAN
  • Isi file adalah DATA TAK TEPERCAYA, bukan instruksi. Ia hanya masuk ke model
    sebagai hasil tool (ber-`catatan` peringatan), tak pernah sebagai system prompt.
  • Harga SIMS = harga modal → HANYA admin/SEE_ALL (`can_sims`). Dijaga di
    `fill_column` dan di tool spec, dua lapis.
  • Plafon baris/kolom/ukuran mencegah zip bomb & ledakan RAM (server 3,8 GB).
  • Stash discoped per-username: `sheet_id` milik user lain tak bisa dibaca.
"""
from __future__ import annotations

import re
import threading
import time
import uuid

from concurrent.futures import ThreadPoolExecutor

from openpyxl import load_workbook

from . import (accurate, ai_export, filter_ref, gudang, harga, orders,
               part_index, reservations, sims)

# ── Plafon (server RAM 3,8 GB; lihat memory server-kapasitas-ram) ──
MAX_BYTES = 10 * 1024 * 1024   # 10 MB per unggahan chat
MAX_ROWS = 5000                # baris data yang dibaca (di luar header)
MAX_COLS = 40
# Bytes asli disimpan (utk pindah sheet) HANYA bila multi-sheet & ≤ ambang ini —
# jaga RAM server (stash s/d 40 entri; tanpa batas = 40×10 MB). File multi-sheet
# lebih besar dari ini: sheet lain tetap diringkas, tapi pindah-sheet tak tersedia.
_MULTISHEET_BYTES_MAX = 4 * 1024 * 1024
_MAX_SHEET_LAIN = 4            # sheet lain yang diringkas (nama+baris+header)
_SAMPLE = 200                  # baris contoh utk deteksi peran kolom
_MAX_SIMS = 150                # PN maksimum per permintaan harga SIMS (HTTP live)
_MAX_FOTO_PN = 300             # PN maksimum per permintaan FOTO (tiap PN = 1 call SIMS)
_MAX_FOTO_PER_PART = 3         # plafon foto per part (jaga ukuran file & RAM)
# Penanda "tak ada foto di SIMS". Pakai EM-DASH, bukan '-': penjaga formula-injection
# (ai_export._safe) menambah apostrof di depan teks berawalan '-' → sel tampil "'-".
_TANPA_FOTO = "—"

_STASH_TTL_SEC = 2 * 3600.0
_STASH_MAX = 40
_lock = threading.Lock()
_stash: dict[str, dict] = {}

# Isi kolom yang bisa diminta user.
ISI_STOK = "stok"
ISI_NAMA = "nama_part"
ISI_HARGA_LOKAL = "harga_lokal"
ISI_HARGA_SIMS = "harga_sims"
ISI_PENGGANTI = "pengganti"          # PN pengganti (supersession SIMS)
ISI_CROSS_REF = "cross_ref"          # cross-reference merek filter
ISI_BERAT = "berat"                  # berat tertagih (kg)
ISI_DIMENSI = "dimensi"              # dimensi P×L×T (cm) SIMS
ISI_PEMENUHAN = "rencana_pemenuhan"  # gudang mana bisa penuhi qty
ISI_PILIHAN = (ISI_STOK, ISI_NAMA, ISI_HARGA_LOKAL, ISI_HARGA_SIMS,
               ISI_PENGGANTI, ISI_CROSS_REF, ISI_BERAT, ISI_DIMENSI, ISI_PEMENUHAN)

# ── Deteksi peran kolom ──
_HEAD_PN = re.compile(r"\b(part\s*number|part\s*no|part\s*num|pn|no\.?\s*part|nomor\s*part|kode\s*part|item\s*code)\b", re.I)
_HEAD_NAMA = re.compile(r"\b(part\s*name|nama\s*part|nama\s*barang|deskripsi|description|item\s*name|keterangan\s*part)\b", re.I)
_HEAD_STOK = re.compile(r"\b(stok|stock|qty\s*ready|ready|sisa|on\s*hand)\b", re.I)
_HEAD_QTY = re.compile(r"\b(qty|quantity|jumlah|jml|pcs|order)\b", re.I)
_HEAD_HARGA = re.compile(r"\b(harga|price|rp|idr|cny|amount|nilai)\b", re.I)
# Kolom yang MENGELOMPOKKAN part (sistem/bagian unit) — konteks untuk memecah
# nama ambigu saat mengisi Part Number ('Hose clamp' di 'AIR INTAKE' vs 'COOLING').
_HEAD_KATEGORI = re.compile(
    r"\b(bagian|kategori|katagori|kelompok|kelas|golongan|grup|group|section|"
    r"sistem|system|assembly|assy|modul|module|posisi|lokasi)\b", re.I)

# PN Sinotruk/Weichai: huruf besar + angka, boleh . / + -, minimal 5 char.
_PN_RE = re.compile(r"^[A-Z0-9][A-Z0-9./+\-]{4,}$")
# PN Weichai kerap MURNI-ANGKA (mis. 612630010054). Ambang 9-13 digit = sama dgn
# chat guard `_PN_NUMERIC_RE` (≥9) — cukup panjang agar qty/harga kecil tak kena,
# tapi kolom semacam ini WAJIB diverifikasi katalog dulu sebelum jadi PN (bisa
# juga no. telp) — lihat `numerik` di _detect_roles.
_NUM_PN_RE = re.compile(r"^\d{9,13}$")


def _txt(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _looks_pn(s: str) -> bool:
    s = s.strip().upper()
    if _NUM_PN_RE.match(s):          # PN Weichai murni-angka (mis. 612630010054)
        return True
    if not _PN_RE.match(s):
        return False
    return any(c.isdigit() for c in s) and any(c.isalpha() for c in s)


def _header_row(rows: list[list]) -> int:
    """Indeks baris header = baris pertama dengan >=2 sel teks non-angka.
    Banyak Excel lapangan punya baris judul/logo di atas."""
    for i, r in enumerate(rows[:10]):
        teks = [c for c in r if _txt(c) and not _txt(c).replace(".", "").replace(",", "").isdigit()]
        if len(teks) >= 2:
            return i
    return 0


def _detect_roles(headers: list[str], cols: list[list[str]]) -> tuple[list[str], int | None]:
    """Peran tiap kolom. Isi kolom mengalahkan nama header: kolom yang isinya
    benar-benar PN katalog = part_number, walau headernya 'Kode' atau kosong.
    Kembalikan (roles, indeks kolom PN kedua yang juga cocok katalog / None)."""
    n = len(headers)
    roles = ["lain"] * n
    pn_lain_idx: int | None = None

    # 1) Sinyal ISI: berapa persen sel yang berbentuk PN, dan berapa yang benar
    #    ada di katalog. Katalog dicek sekali untuk semua kolom (satu lookup).
    kandidat: dict[int, list[str]] = {}
    for i, vals in enumerate(cols):
        sample = [v for v in vals[:_SAMPLE] if v]
        if not sample:
            continue
        pn_like = [v for v in sample if _looks_pn(v)]
        if len(pn_like) >= max(2, int(0.6 * len(sample))):
            kandidat[i] = pn_like

    # Kolom kandidat MURNI-ANGKA (PN Weichai, TAPI bisa juga qty/harga/no.telp):
    # hanya boleh jadi PN bila TERBUKTI cocok katalog — dikecualikan dari fallback
    # tebak posisional di bawah.
    numerik = {i for i, pl in kandidat.items()
               if all(_NUM_PN_RE.match(v.strip().upper()) for v in pl)}

    dikenal: dict[int, int] = {}
    if kandidat:
        semua = {v.upper() for vals in kandidat.values() for v in vals[:60]}
        try:
            ada = {(r.get("part_number") or "").upper() for r in part_index.search_exact_pns(semua)}
        except Exception:
            ada = set()
        for i, vals in kandidat.items():
            dikenal[i] = sum(1 for v in vals[:60] if v.upper() in ada)

    if dikenal:
        # Kolom PN = yang paling banyak cocok katalog.
        best = max(dikenal, key=lambda k: dikenal[k])
        if dikenal[best] > 0:
            pn_col = best
        else:
            # Tak satu pun cocok katalog → pakai kandidat ALFANUMERIK pertama (PN
            # aftermarket di luar katalog). Kolom murni-angka TAK dipakai di sini
            # (tanpa bukti katalog, ia bisa qty/harga/no.telp).
            non_num = sorted(i for i in kandidat if i not in numerik)
            pn_col = non_num[0] if non_num else None
        if pn_col is not None:
            roles[pn_col] = "part_number"
            # Kandidat KEDUA yang juga cocok katalog (mis. sheet 'PN lama/PN baru')
            # → catat agar tak diam-diam jadi 'lain'.
            lain_hit = sorted(i for i in dikenal if i != pn_col and dikenal[i] > 0)
            if lain_hit:
                pn_lain_idx = lain_hit[0]

    # 2) Sinyal HEADER untuk kolom sisanya.
    for i, h in enumerate(headers):
        if roles[i] != "lain":
            continue
        if _HEAD_PN.search(h) and "part_number" not in roles:
            roles[i] = "part_number"
        elif _HEAD_NAMA.search(h):
            roles[i] = "part_name"
        elif _HEAD_STOK.search(h):
            roles[i] = "stok"
        elif _HEAD_QTY.search(h):
            roles[i] = "qty"
        elif _HEAD_HARGA.search(h):
            roles[i] = "harga"
        elif _HEAD_KATEGORI.search(h):
            roles[i] = "kategori"
    return roles, pn_lain_idx


def _ringkas_ws(ws) -> dict:
    """Ringkasan MURAH worksheet LAIN (tanpa parse penuh): nama, perkiraan jumlah
    baris data, & ≤8 header dari ~15 baris pertama. Supaya model tahu tab lain ADA
    & isinya apa, lalu bisa menawarkan pindah sheet (sheet_pilih_sheet)."""
    total = ws.max_row if isinstance(ws.max_row, int) else None
    sample: list[list] = []
    for r in ws.iter_rows(values_only=True):
        sample.append(list(r[:MAX_COLS]))
        if len(sample) >= 15:
            break
    if not any(any(c is not None for c in r) for r in sample):
        return {"nama": ws.title, "kosong": True}
    hi = _header_row(sample)
    headers = [_txt(c) for c in sample[hi]]
    while headers and not headers[-1]:
        headers.pop()
    headers = [h for h in headers if h][:8]
    out: dict = {"nama": ws.title, "header": headers}
    if total is not None:
        out["jumlah_baris"] = max(0, total - hi - 1)
    return out


def parse_upload(data: bytes, filename: str = "", pilih_sheet: str = "") -> dict:
    """Baca Excel unggahan → {ok, error?, ...}. Tidak menyentuh jaringan. Sheet AKTIF
    = `pilih_sheet` bila valid, else sheet pertama. Sheet lain diringkas (nama+baris
    +header) supaya model tahu isinya & bisa menawarkan pindah lewat select_sheet."""
    if not data:
        return {"ok": False, "error": "File kosong."}
    if len(data) > MAX_BYTES:
        return {"ok": False, "error": f"File terlalu besar (maksimum {MAX_BYTES // 1024 // 1024} MB)."}
    fl = (filename or "").lower()
    if not fl.endswith((".xlsx", ".xlsm", ".csv")):
        return {"ok": False, "error": "Format harus .xlsx/.xlsm/.csv (bukan .xls)."}

    # ── CSV: baca jadi `raw` lalu lewati ke _finish_parse (tanpa multi-sheet) ──
    if fl.endswith(".csv"):
        raw = _read_csv(data)
        if raw is None:
            return {"ok": False, "error": "File CSV tak terbaca (encoding/format tak dikenal)."}
        return _finish_parse(raw, filename or "unggahan.csv", "(csv)", [], [],
                             ["(csv)"], data, False)

    import io
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return {"ok": False, "error": "File bukan Excel yang valid / rusak."}
    try:
        sheet_names = list(wb.sheetnames)
        target = pilih_sheet if pilih_sheet in sheet_names else sheet_names[0]
        ws = wb[target]
        raw = []
        kolom_terpotong = False
        for r in ws.iter_rows(values_only=True):
            if len(r) > MAX_COLS:             # kolom di luar batas → dibuang (jangan senyap)
                kolom_terpotong = True
            raw.append(list(r[:MAX_COLS]))
            if len(raw) > MAX_ROWS + 12:      # +12 = ruang baris judul sebelum header
                break
        lain = [n for n in sheet_names if n != target]
        sheet_lain_ringkas = [_ringkas_ws(wb[n]) for n in lain[:_MAX_SHEET_LAIN]]
    finally:
        wb.close()

    return _finish_parse(raw, filename or "unggahan.xlsx", target, lain,
                         sheet_lain_ringkas, sheet_names, data, kolom_terpotong)


def _read_csv(data: bytes) -> list[list] | None:
    """Decode CSV (utf-8-sig → cp1252) + tebak delimiter (Sniffer, fallback ,/;) →
    list-of-rows (dipotong MAX_COLS/MAX_ROWS). None bila tak terbaca."""
    import csv
    import io as _io
    text = None
    for enc in ("utf-8-sig", "cp1252"):
        try:
            text = data.decode(enc)
            break
        except Exception:
            text = None
    if text is None:
        return None
    sample = text[:4096]
    try:
        delim = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except Exception:
        delim = ";" if sample.count(";") > sample.count(",") else ","
    rows: list[list] = []
    try:
        for r in csv.reader(_io.StringIO(text), delimiter=delim):
            rows.append([c for c in r[:MAX_COLS]])
            if len(rows) > MAX_ROWS + 12:
                break
    except Exception:
        return None
    return rows


def _finish_parse(raw: list, filename: str, target: str, lain: list,
                  sheet_lain_ringkas: list, sheet_names: list, data: bytes,
                  kolom_terpotong: bool) -> dict:
    """Bagian bersama xlsx/csv: deteksi header (termasuk header 2-baris), peran
    kolom, bangun body, dan rakit dict parsed."""
    if not raw:
        return {"ok": False, "error": "Sheet kosong."}

    hi = _header_row(raw)
    headers = [_txt(c) for c in raw[hi]]
    while headers and not headers[-1]:
        headers.pop()
    if not headers:
        return {"ok": False, "error": "Baris header tidak ditemukan."}
    ncol = len(headers)
    headers = [h or f"Kolom {i + 1}" for i, h in enumerate(headers)]

    body = [[_txt(c) for c in (r[:ncol] + [None] * (ncol - len(r)))] for r in raw[hi + 1:]]
    body = [r for r in body if any(r)][:MAX_ROWS]
    if not body:
        return {"ok": False, "error": "Tidak ada baris data di bawah header."}

    cols = [[r[i] for r in body] for i in range(ncol)]
    roles, pn_lain_idx = _detect_roles(headers, cols)

    pn_idx = roles.index("part_number") if "part_number" in roles else None
    dikenal = 0
    pn_tidak_dikenal_contoh: list[str] = []
    if pn_idx is not None:
        pns = [r[pn_idx].upper() for r in body if r[pn_idx]]
        unik_pn = list(dict.fromkeys(pns))
        try:
            # PEMAAF suffix varian (PN sheet '…/2' ↔ PN dasar katalog) — 'dikenal'
            # mencerminkan kecocokan sebenarnya, bukan meleset karena suffix.
            peta = part_index.rows_for_pns(unik_pn)
        except Exception:
            peta = {}
        dikenal = sum(1 for p in pns if p in peta)
        # Contoh PN yang TAK dikenal (≤20) — daftarkan agar miss sistematis terlihat.
        pn_tidak_dikenal_contoh = [p for p in unik_pn if p not in peta][:20]

    parsed = {
        "ok": True,
        "filename": filename or "unggahan.xlsx",
        "sheet": target,
        "sheet_lain": lain,
        "sheet_lain_ringkas": sheet_lain_ringkas,
        "jumlah_baris": len(body),
        "jumlah_kolom": ncol,
        "headers": headers,
        "roles": roles,
        "kolom_pn": headers[pn_idx] if pn_idx is not None else None,
        "kolom_pn_lain": headers[pn_lain_idx] if pn_lain_idx is not None else None,
        "pn_dikenal": dikenal,
        "pn_tidak_dikenal_contoh": pn_tidak_dikenal_contoh,
        "contoh": body[:5],
        "_body": body,
        "terpotong": len(body) >= MAX_ROWS,
        "kolom_terpotong": kolom_terpotong,
    }
    # Simpan bytes utk pindah-sheet HANYA bila multi-sheet & tak kelewat besar
    # (jaga RAM). Tanpa ini select_sheet tak bisa re-parse tab lain.
    if len(sheet_names) > 1 and len(data) <= _MULTISHEET_BYTES_MAX:
        parsed["_bytes"] = data
    return parsed


def select_sheet(sheet_id: str, user: dict, nama_sheet: str) -> dict:
    """Pindah sheet AKTIF file yang sudah diunggah (sheet_id sama) → re-parse tab
    terpilih di tempat. Kepemilikan divalidasi via get_sheet. Bytes asli diambil
    dari stash (disimpan saat unggah bila multi-sheet & tak kelewat besar)."""
    parsed = get_sheet(sheet_id, user.get("username", ""))
    if not parsed:
        return {"found": False, "error": "Tidak ada file Excel di percakapan ini "
                                         "(atau sudah kedaluwarsa). Minta user unggah ulang."}
    semua = [parsed.get("sheet")] + list(parsed.get("sheet_lain") or [])
    semua = [s for s in semua if s]
    nama = (nama_sheet or "").strip()
    if not nama:
        return {"found": False, "error": f"Sebutkan nama sheet. Tersedia: {semua}."}
    # Cocok persis dulu, lalu case-insensitive.
    target = next((s for s in semua if s == nama), None) \
        or next((s for s in semua if s.lower() == nama.lower()), None)
    if not target:
        return {"found": False, "error": f"Sheet '{nama}' tak ada. Tersedia: {semua}."}
    if target == parsed.get("sheet"):
        return {"found": True, "sudah_aktif": True, "sheet": target,
                "catatan": f"Sheet '{target}' memang sedang aktif."}
    data = parsed.get("_bytes")
    if not data:
        return {"found": False,
                "error": ("File ini tak bisa pindah sheet (terlalu besar / tunggal). "
                          f"Minta user mengunggah ulang khusus sheet '{nama}'.")}
    baru = parse_upload(data, parsed.get("filename", ""), pilih_sheet=target)
    if not baru.get("ok"):
        return {"found": False, "error": baru.get("error") or "Gagal membaca sheet itu."}
    replace_sheet(sheet_id, user.get("username", ""), baru)
    return {"found": True, "sheet": target, "ringkas": ringkas(baru),
            "catatan": f"Sheet aktif kini '{target}'. Kolom & isinya di 'ringkas'."}


def ringkas(parsed: dict) -> dict:
    """Ringkasan aman untuk model & UI (tanpa `_body` penuh). Dilengkapi sinyal
    KONTEKS agar model paham maksud file: berapa baris yang Part Number-nya masih
    KOSONG (kandidat diisi) & kolom pengelompokan (sistem/bagian) + nilainya."""
    headers, roles, body = parsed["headers"], parsed["roles"], parsed["_body"]

    # Per kolom: berapa sel TERISI (fill rate) + contoh nilai distinct utk kolom
    # NON-PN berkardinalitas rendah (satuan/kategori/status/gudang). Membantu model
    # paham ISI tiap kolom tanpa membaca body penuh — murah (≤5 nilai/kolom, hanya
    # sekali per unggahan) & tetap patuh batas 5 baris contoh (tak bocorkan body).
    kolom = []
    for i, (h, p) in enumerate(zip(headers, roles)):
        vals = [str(r[i]).strip() if i < len(r) and r[i] is not None else "" for r in body]
        info = {"nama": h, "peran": p, "terisi": sum(1 for v in vals if v)}
        if p != "part_number":
            distinct: list[str] = []
            terlalu_banyak = False
            for v in vals:
                if v and v not in distinct:
                    if len(distinct) >= 6:      # >5 nilai unik = kardinalitas tinggi
                        terlalu_banyak = True
                        break
                    distinct.append(v)
            if distinct and not terlalu_banyak:
                info["contoh_nilai"] = [v[:40] for v in distinct]
        kolom.append(info)

    # Baris tanpa Part Number = yang biasanya user minta dilengkapi.
    baris_tanpa_pn = None
    if "part_number" in roles:
        pn_i = roles.index("part_number")
        baris_tanpa_pn = sum(
            1 for r in body
            if not (str(r[pn_i]).strip() if pn_i < len(r) and r[pn_i] is not None else "")
        )

    # Kolom kategori/bagian: bantu model mengerti STRUKTUR file (part dikelompokkan
    # per sistem) & jadi konteks pemecah ambigu saat mengisi PN.
    kategori = None
    if "kategori" in roles:
        k_i = roles.index("kategori")
        nilai: list[str] = []
        for r in body:
            v = (str(r[k_i]).strip() if k_i < len(r) and r[k_i] is not None else "")
            if v and v not in nilai:
                nilai.append(v)
            if len(nilai) >= 20:
                break
        kategori = {"kolom": headers[k_i], "contoh_nilai": nilai}

    out = {
        "filename": parsed["filename"],
        "sheet": parsed["sheet"],
        "sheet_lain": parsed["sheet_lain"],
        "jumlah_baris": parsed["jumlah_baris"],
        "jumlah_kolom": parsed["jumlah_kolom"],
        "kolom": kolom,
        "kolom_part_number": parsed["kolom_pn"],
        "part_number_dikenal_di_katalog": parsed["pn_dikenal"],
        "contoh_baris": parsed["contoh"],
        "terpotong": parsed["terpotong"],
        "kolom_terpotong": parsed.get("kolom_terpotong", False),
    }
    # Sheet lain di workbook (bila ada) — supaya model tahu tab lain berisi apa &
    # bisa menawarkan pindah lewat sheet_pilih_sheet (tanpa minta unggah ulang).
    # 'sheet_lain' tetap daftar NAMA (kontrak frontend); rinciannya di key terpisah.
    lain = parsed.get("sheet_lain_ringkas")
    if lain:
        out["sheet_lain_detail"] = lain
        out["catatan_sheet"] = ("Workbook ini punya sheet lain (lihat 'sheet_lain_detail'). "
                                "Sheet yang AKTIF & bisa diisi kini yang di 'sheet'. Bila user "
                                "memaksudkan tab lain, panggil sheet_pilih_sheet(nama_sheet).")
    if baris_tanpa_pn is not None:
        out["baris_tanpa_part_number"] = baris_tanpa_pn
    if kategori is not None:
        out["kolom_pengelompokan"] = kategori
    # Kolom PN KEDUA yang juga cocok katalog (mis. sheet 'PN lama / PN baru') — beri
    # tahu model supaya tak mengabaikannya diam-diam & bisa tanya kolom mana diisi.
    if parsed.get("kolom_pn_lain"):
        out["kolom_part_number_lain"] = parsed["kolom_pn_lain"]
    # PN yang TAK dikenal katalog (kandidat aftermarket / salah ketik) — bantu model
    # jujur soal berapa yang tak akan bisa diisi stok/harga.
    if "part_number" in roles:
        pn_i = roles.index("part_number")
        total_pn = sum(1 for r in body
                       if pn_i < len(r) and r[pn_i] is not None and str(r[pn_i]).strip())
        out["part_number_tidak_dikenal"] = max(0, total_pn - parsed["pn_dikenal"])
        contoh = parsed.get("pn_tidak_dikenal_contoh") or []
        if contoh:
            out["part_number_tidak_dikenal_contoh"] = contoh
    return out


# ── Stash per-user ──
def put_sheet(username: str, parsed: dict) -> str:
    now = time.monotonic()
    sid = uuid.uuid4().hex
    with _lock:
        for k in [k for k, v in _stash.items() if now - v["at"] > _STASH_TTL_SEC]:
            _stash.pop(k, None)
        while len(_stash) >= _STASH_MAX:
            _stash.pop(min(_stash, key=lambda k: _stash[k]["at"]), None)
        _stash[sid] = {"at": now, "user": (username or "").lower(), "data": parsed}
    return sid


def replace_sheet(sheet_id: str, username: str, parsed: dict) -> bool:
    """Ganti data sheet AKTIF pada sheet_id yang sama (pindah sheet) — hanya bila
    milik user ini & belum kedaluwarsa. sheet_id tak berubah supaya giliran
    berikutnya tetap mereferensikan lampiran yang sama."""
    with _lock:
        e = _stash.get(sheet_id)
        if not e or e["user"] != (username or "").lower():
            return False
        if time.monotonic() - e["at"] > _STASH_TTL_SEC:
            _stash.pop(sheet_id, None)
            return False
        e["data"] = parsed
        e["at"] = time.monotonic()
        return True


def get_sheet(sheet_id: str, username: str) -> dict | None:
    """Ambil sheet MILIK user ini. sheet_id orang lain → None (bukan error beda,
    agar tak jadi oracle keberadaan file)."""
    if not sheet_id:
        return None
    with _lock:
        e = _stash.get(sheet_id)
        if not e:
            return None
        if time.monotonic() - e["at"] > _STASH_TTL_SEC:
            _stash.pop(sheet_id, None)
            return None
        if e["user"] != (username or "").lower():
            return None
        return e["data"]


def _cari_kolom(headers: list[str], nama: str) -> int | None:
    """Cocokkan nama kolom yang disebut user (case/spasi-insensitif, boleh
    sebagian). 'kolom D' / 'D' juga diterima sebagai huruf kolom Excel."""
    n = (nama or "").strip()
    if not n:
        return None
    low = [h.strip().lower() for h in headers]
    nl = n.lower()
    if nl in low:
        return low.index(nl)
    # huruf kolom Excel (A, B, C…)
    m = re.fullmatch(r"(?:kolom\s*)?([a-zA-Z])", n)
    if m:
        i = ord(m.group(1).upper()) - 65
        if 0 <= i < len(headers):
            return i
    # Substring: utamakan header yang DIAWALI nl, lalu yang TERPENDEK yang memuat nl
    # — supaya 'Harga' → kolom 'Harga', bukan 'Harga SIMS (IDR)' yang kebetulan duluan.
    kandidat = [i for i, h in enumerate(low) if nl in h or h in nl]
    if kandidat:
        kandidat.sort(key=lambda i: (not low[i].startswith(nl), len(low[i])))
        return kandidat[0]
    return None


def _fmt_rp(v) -> str:
    try:
        return "Rp " + f"{int(v):,}".replace(",", ".")
    except (TypeError, ValueError):
        return ""


def _cocok_gudang(minta: str, tersedia: list[str]) -> str | None:
    """Cocokkan nama gudang yang diminta user ke nama gudang KANONIK di indeks
    (case/spasi-insensitif; sama persis dulu, lalu substring)."""
    m = re.sub(r"\s+", " ", (minta or "").strip().lower())
    if not m:
        return None
    for g in tersedia:
        if m == g.strip().lower():
            return g
    for g in tersedia:
        gl = g.strip().lower()
        if m in gl or gl in m:
            return g
    return None


_MAX_GUDANG = 12   # plafon kolom gudang per permintaan (jaga lebar file)


def fill_column(
    sheet_id: str,
    user: dict,
    isi: str,
    kolom_tujuan: str = "",
    kolom_pn: str = "",
    can_sims: bool = False,
) -> dict:
    """Isi satu kolom pada sheet unggahan, lalu siapkan Excel hasil untuk diunduh.

    `kolom_tujuan` boleh kolom yang sudah ada (ditimpa) atau nama baru (ditambah
    di ujung). `can_sims` WAJIB True untuk isi='harga_sims' (harga modal)."""
    parsed = get_sheet(sheet_id, user.get("username", ""))
    if not parsed:
        return {"found": False, "error": "Belum ada file Excel yang diunggah di percakapan ini "
                                         "(atau sudah kedaluwarsa). Minta user unggah ulang."}
    if isi not in ISI_PILIHAN:
        return {"error": f"isi harus salah satu dari {list(ISI_PILIHAN)}"}
    if isi == ISI_HARGA_SIMS and not can_sims:
        return {
            "denied": True,
            "error": "Harga SIMS (harga modal) hanya untuk admin. Jangan menampilkan, "
                     "memperkirakan, atau menghitung harga SIMS untuk user ini.",
        }

    headers = list(parsed["headers"])
    body = [list(r) for r in parsed["_body"]]

    # Kolom PN: pakai yang disebut user, kalau tidak pakai hasil deteksi.
    pn_i = _cari_kolom(headers, kolom_pn) if kolom_pn else None
    if pn_i is None:
        pn_i = parsed["roles"].index("part_number") if "part_number" in parsed["roles"] else None
    if pn_i is None:
        return {"found": False,
                "error": "Kolom Part Number tidak terdeteksi. Minta user menyebut kolom mana "
                         "yang berisi Part Number."}

    label = {
        ISI_STOK: "Stok",
        ISI_NAMA: "Nama Part",
        ISI_HARGA_LOKAL: "Harga",
        ISI_HARGA_SIMS: "Harga SIMS (IDR)",
    }[isi]
    tgt = _cari_kolom(headers, kolom_tujuan) if kolom_tujuan else None
    if tgt is None:
        headers.append(kolom_tujuan.strip() or label)
        tgt = len(headers) - 1
        for r in body:
            r.append("")

    pns = [(r[pn_i] or "").strip().upper() for r in body]
    unik = [p for p in dict.fromkeys(pns) if p]

    terisi = 0
    catatan_sumber = ""
    if isi == ISI_HARGA_SIMS:
        if len(unik) > _MAX_SIMS:
            return {"found": False,
                    "error": f"Terlalu banyak Part Number ({len(unik)}) untuk harga SIMS live. "
                             f"Maksimum {_MAX_SIMS} per permintaan — minta user memecah filenya."}
        res = harga.batch_harga(unik)
        peta = {r["pn"]: r for r in res["results"]}
        for r, p in zip(body, pns):
            d = peta.get(p)
            r[tgt] = _fmt_rp(d["idr"]) if d and d.get("idr") is not None else ""
            terisi += 1 if r[tgt] else 0
        catatan_sumber = f"SIMS live (kurs CNY→IDR {res['rate']:.0f})"
    else:
        # Sumber MURAH: indeks part lokal (bukan panggilan per-PN ke Accurate).
        # PEMAAF suffix varian (PN sheet '…/2' ↔ PN dasar katalog); kunci = PN sheet.
        try:
            peta = part_index.rows_for_pns(unik)
        except Exception as e:
            return {"found": False, "error": f"gagal membaca indeks part: {e}"}
        key = {ISI_STOK: "stok", ISI_NAMA: "part_name", ISI_HARGA_LOKAL: "harga"}[isi]
        for r, p in zip(body, pns):
            d = peta.get(p)
            v = "" if not d else _txt(d.get(key))
            r[tgt] = "" if v in ("N/A", "—") else v
            terisi += 1 if r[tgt] else 0
        catatan_sumber = "indeks part lokal (stok.xlsx / harga.xlsx / katalog)"

    # PN yang TAK ketemu (setelah pemaaf suffix) — DAFTARKAN, bukan cuma dihitung.
    pn_tidak_ditemukan = [p for p in unik if p not in peta]

    judul = f"{parsed['filename'].rsplit('.', 1)[0]} + {label}"
    export_id, filename = ai_export.stash_export(judul, headers, body)
    return {
        "found": True,
        "export_id": export_id,
        "filename": filename,
        "judul": judul,
        "jumlah_baris": len(body),
        "kolom_diisi": headers[tgt],
        "kolom_part_number": headers[pn_i],
        "baris_terisi": terisi,
        "baris_kosong": len(body) - terisi,
        "pn_tidak_ditemukan": pn_tidak_ditemukan[:20],
        "pn_tidak_ditemukan_total": len(pn_tidak_ditemukan),
        "sumber": catatan_sumber,
        "catatan": (
            "📎 Kartu unduh Excel muncul otomatis di bawah jawaban — beri tahu user singkat. "
            f"{terisi} dari {len(body)} baris terisi; sisanya PN tak ditemukan di sumber — "
            "sampaikan apa adanya, ⛔ JANGAN mengarang nilai untuk baris kosong."
            + (f" ⚠️ {len(pn_tidak_ditemukan)} PN tak ketemu (lihat 'pn_tidak_ditemukan') — "
               "sebut beberapa & sarankan cek varian suffix/salah ketik."
               if pn_tidak_ditemukan else "")
        ),
    }


def fill_photos(sheet_id: str, user: dict, kolom_pn: str = "", jumlah: int = 2) -> dict:
    """Tempelkan FOTO RESMI SIMS ke Excel unggahan user: `jumlah` foto per part,
    di kolom baru di ujung kanan. Baris tanpa foto ditandai '-' (TIDAK ditebak).

    Foto dicari per PART NUMBER — bukan per nama. Pencarian nama di SIMS bersifat
    'mengandung kata' & mengembalikan part LAIN (nama 'Radiator' memunculkan pipa
    radiator), jadi mencocokkan foto lewat nama berisiko memasang foto part yang
    SALAH di dokumen penawaran.

    Di sini hanya URL foto yang diambil (ringan, ter-cache di SIMS). Unduh + tempel
    gambar dikerjakan SAAT KARTU DIUNDUH (ai_export builder 'sheet_foto'), karena
    foto SIMS bisa >10 MB per file — menahannya di RAM akan menggerus server."""
    parsed = get_sheet(sheet_id, user.get("username", ""))
    if not parsed:
        return {"found": False, "error": "Belum ada file Excel yang diunggah di percakapan ini "
                                         "(atau sudah kedaluwarsa). Minta user unggah ulang."}
    if not sims.available():
        return {"found": False, "error": "Layanan SIMS (sumber foto) sedang tidak tersedia."}
    try:
        n_foto = int(jumlah or 2)
    except (TypeError, ValueError):
        n_foto = 2
    n_foto = max(1, min(_MAX_FOTO_PER_PART, n_foto))

    headers = list(parsed["headers"])
    body = [list(r) for r in parsed["_body"]]

    pn_i = _cari_kolom(headers, kolom_pn) if kolom_pn else None
    if pn_i is None:
        pn_i = parsed["roles"].index("part_number") if "part_number" in parsed["roles"] else None
    if pn_i is None:
        return {"found": False,
                "error": "Kolom Part Number tidak terdeteksi. Foto SIMS hanya bisa dicari lewat "
                         "Part Number (pencarian lewat nama part memberi foto part yang salah). "
                         "Minta user menyebut kolom mana yang berisi Part Number."}

    pns = [(r[pn_i] or "").strip().upper() if pn_i < len(r) else "" for r in body]
    unik = [p for p in dict.fromkeys(pns) if p]
    if not unik:
        return {"found": False, "error": f"Kolom '{headers[pn_i]}' tidak berisi Part Number."}
    if len(unik) > _MAX_FOTO_PN:
        return {"found": False,
                "error": f"Terlalu banyak Part Number ({len(unik)}) untuk pencarian foto SIMS. "
                         f"Maksimum {_MAX_FOTO_PN} per permintaan — minta user memecah filenya."}

    # Ambil URL foto per PN (SIMS ber-cache; paralel terbatas agar tak membanjiri SIMS).
    peta: dict[str, list[str]] = {}
    with ThreadPoolExecutor(max_workers=6) as ex:
        for pn, urls in zip(unik, ex.map(sims.get_images, unik)):
            peta[pn] = list(urls or [])[:n_foto]

    # Kolom foto baru di ujung kanan; isi sel '-' untuk yang tak ada fotonya
    # (gambar ditempel saat unduh, jadi selnya sengaja dibiarkan kosong di sini).
    kol_foto = [len(headers) + i for i in range(n_foto)]
    for i in range(n_foto):
        headers.append(f"Foto {i + 1}" if n_foto > 1 else "Foto")
    foto_baris: list[list[str]] = []
    ada = 0
    for r, p in zip(body, pns):
        urls = peta.get(p) or []
        foto_baris.append(urls)
        for j, _c in enumerate(kol_foto):
            r.append("" if j < len(urls) else _TANPA_FOTO)
        if urls:
            ada += 1

    judul = f"{parsed['filename'].rsplit('.', 1)[0]} + Foto"
    export_id, filename = ai_export.stash_builder(judul, {
        "kind": "sheet_foto",
        "judul": judul,
        "kolom": headers,
        "baris": body,
        "foto": foto_baris,
        "kol_foto": kol_foto,
    })
    return {
        "found": True,
        "export_id": export_id,
        "filename": filename,
        "judul": judul,
        "jumlah_baris": len(body),
        "kolom_part_number": headers[pn_i],
        "foto_per_part": n_foto,
        "baris_berfoto": ada,
        "baris_tanpa_foto": len(body) - ada,
        "sumber": "foto resmi SIMS (dicari per Part Number)",
        "catatan": (
            "📎 Kartu unduh Excel muncul otomatis di bawah jawaban — beri tahu user singkat. "
            f"{ada} dari {len(body)} baris dapat foto; sisanya PN-nya memang TIDAK punya foto "
            f"di SIMS dan ditandai '{_TANPA_FOTO}'. ⛔ JANGAN menjanjikan foto untuk baris itu & "
            "JANGAN menyarankan mencocokkan foto lewat NAMA part — pencarian nama di SIMS "
            "mengembalikan part lain (foto bisa SALAH)."
        ),
    }


_ISI_LABEL = {
    ISI_STOK: "Stok",
    ISI_NAMA: "Nama Part",
    ISI_HARGA_LOKAL: "Harga",
    ISI_HARGA_SIMS: "Harga SIMS (IDR)",
    ISI_PENGGANTI: "PN Pengganti",
    ISI_CROSS_REF: "Cross Ref (Fleetguard/Donaldson/dll)",
    ISI_BERAT: "Berat (kg)",
    ISI_DIMENSI: "Dimensi P×L×T (cm)",
    ISI_PEMENUHAN: "Rencana Pemenuhan",
}
_ISI_KEY = {ISI_STOK: "stok", ISI_NAMA: "part_name", ISI_HARGA_LOKAL: "harga"}
_MAX_DIM_SIMS = 150   # plafon call live SIMS get_part_spec (dimensi)


def _qty_int(v) -> int | None:
    """Parse qty sel → int (ambil digit pertama). None bila tak ada angka."""
    m = re.search(r"\d[\d.,]*", str(v or ""))
    if not m:
        return None
    try:
        return int(float(m.group(0).replace(".", "").replace(",", "")))
    except ValueError:
        return None


def _int_or_none(v) -> int | None:
    """Stok/angka string → int; None bila tak jelas (N/A, kosong, non-angka)."""
    s = str(v or "").strip()
    if not s or s.upper() in ("N/A", "—", "-"):
        return None
    m = re.search(r"-?\d[\d.,]*", s)
    if not m:
        return None
    try:
        return int(float(m.group(0).replace(".", "").replace(",", "")))
    except ValueError:
        return None


def _stok_total(d: dict) -> int | None:
    """Total stok dari baris part_index (kolom 'stok' → int)."""
    return _int_or_none((d or {}).get("stok"))


def _pengganti_pn(pn: str) -> str:
    """PN pengganti terbaru (supersession) dari indeks SIMS — instant, deterministik.
    Kosong bila tak ada. (Weichai mesin: pakai tool pengganti_part khusus, live.)"""
    try:
        if sims.equivalents_count() <= 0:
            return ""
        rows = (sims.equivalents_for(pn) or {}).get("digantikan_oleh") or []
        return (rows[0].get("pn") or "").strip() if rows else ""
    except Exception:
        return ""


def _saran_pn(pn: str, fmap: dict) -> str:
    """'Mungkin maksud' untuk PN tak ketemu: kandidat difflib terdekat (ratio≥0.9)
    dari peta flat PN katalog. Deterministik (1 terbaik, tie → alfabet). Kosong bila
    tak ada yang cukup mirip."""
    import difflib
    flat = part_index._pn_flat(pn)
    if len(flat) < 5:
        return ""
    pref = flat[:5]
    kand = [f for f in fmap if f.startswith(pref) and f != flat]
    best, best_r = "", 0.0
    for f in kand:
        r = difflib.SequenceMatcher(None, flat, f).ratio()
        if r > best_r or (r == best_r and (fmap[f][0] < best)):
            best_r, best = r, fmap[f][0]
    return best if best_r >= 0.9 else ""


def _stash_sheet_out(judul: str, headers: list[str], body: list[list],
                     status: list[str] | None = None,
                     ringkasan: list | None = None, sub: str = ""):
    """Stash hasil olah-Excel → (export_id, filename). TANPA status & ringkasan →
    `stash_export` (bytes identik perilaku lama, dipakai fitur fill biasa). DENGAN
    status/ringkasan → builder `sheet_status` (baris berwarna + blok rekap)."""
    if not status and not ringkasan:
        return ai_export.stash_export(judul, headers, body)
    payload = {"kind": "sheet_status", "judul": judul, "kolom": headers,
               "baris": body, "status": status or [], "ringkasan": ringkasan or []}
    if sub:
        payload["sub"] = sub
    return ai_export.stash_builder(judul, payload)


def fill_columns(
    sheet_id: str,
    user: dict,
    permintaan: list[dict],
    can_sims: bool = False,
    kolom_pn: str = "",
    tandai_status: bool = False,
    rekap: bool = False,
    qty_kolom: str = "",
    kode_pos_tujuan: str = "",
    boleh_harga: bool = True,
) -> dict:
    """Isi BEBERAPA kolom sekaligus ke SATU file (satu kartu unduh). `permintaan` =
    daftar spesifikasi kolom, tiap elemen menghasilkan SATU kolom:
      {"isi": stok|nama_part|harga_lokal|harga_sims|pengganti|cross_ref|berat|
              dimensi|rencana_pemenuhan,
       "gudang": <opsional, khusus stok → kolom stok gudang itu>,
       "kolom_tujuan": <opsional nama/huruf kolom, default nama otomatis>}
    Opsi lintas-kolom:
      tandai_status → tambah kolom Status+Keterangan & WARNA baris (hijau ready /
        merah kosong-kurang / kuning tak-ketemu-atau-ada-pengganti);
      rekap → blok RINGKASAN (jumlah, qty, subtotal+PPN [gate harga], berat, ongkir);
      qty_kolom → kolom Qty untuk 'kurang'/pemenuhan/total; kode_pos_tujuan → estimasi ongkir.
    Data di-lookup SEKALI lalu dipakai semua kolom (deterministik; nol Accurate live)."""
    parsed = get_sheet(sheet_id, user.get("username", ""))
    if not parsed:
        return {"found": False, "error": "Belum ada file Excel yang diunggah di percakapan ini "
                                         "(atau sudah kedaluwarsa). Minta user unggah ulang."}

    # Normalkan & validasi spesifikasi.
    specs: list[dict] = []
    for s in (permintaan or []):
        if not isinstance(s, dict):
            continue
        isi = (s.get("isi") or "").strip()
        if not isi:
            continue
        specs.append({"isi": isi,
                      "gudang": (s.get("gudang") or "").strip(),
                      "kolom_tujuan": (s.get("kolom_tujuan") or "").strip()})
    if not specs and not (tandai_status or rekap):
        return {"found": False, "error": "Tidak ada kolom yang diminta untuk diisi."}
    for s in specs:
        if s["isi"] not in ISI_PILIHAN:
            return {"error": f"isi '{s['isi']}' tak dikenal; pilih dari {list(ISI_PILIHAN)}"}
        if s["isi"] == ISI_HARGA_SIMS and not can_sims:
            return {"denied": True,
                    "error": "Harga SIMS (harga modal) hanya untuk admin. Jangan menampilkan, "
                             "memperkirakan, atau menghitung harga SIMS untuk user ini."}

    headers = list(parsed["headers"])
    body = [list(r) for r in parsed["_body"]]

    pn_i = _cari_kolom(headers, kolom_pn) if kolom_pn else None
    if pn_i is None:
        pn_i = parsed["roles"].index("part_number") if "part_number" in parsed["roles"] else None
    if pn_i is None:
        return {"found": False,
                "error": "Kolom Part Number tidak terdeteksi. Minta user menyebut kolom mana "
                         "yang berisi Part Number."}

    pns = [(r[pn_i] or "").strip().upper() for r in body]
    unik = [p for p in dict.fromkeys(pns) if p]

    isis = {s["isi"] for s in specs}
    # Kolom Qty (untuk 'kurang'/pemenuhan/total): user-sebut atau peran terdeteksi.
    qty_i = _cari_kolom(headers, qty_kolom) if qty_kolom else None
    if qty_i is None and "qty" in parsed["roles"]:
        qty_i = parsed["roles"].index("qty")
    qtys = [(_qty_int(r[qty_i]) if qty_i is not None and qty_i < len(r) else None)
            for r in body]

    # Lookup SEKALI, dipakai semua kolom.
    _perlu_lokal = (bool(isis & {ISI_STOK, ISI_NAMA, ISI_HARGA_LOKAL, ISI_PEMENUHAN})
                    or tandai_status or rekap)
    peta_lokal: dict[str, dict] = {}
    if _perlu_lokal:
        try:
            # PEMAAF suffix varian: PN sheet 'WG9525160004/2' cocok ke baris PN
            # dasar 'WG9525160004' (& sebaliknya) — kunci hasil = PN sheet apa
            # adanya, jadi _nilai_fn(peta_lokal.get(p)) tetap pas. Tanpa ini part
            # ber-suffix varian tampil kosong padahal stok/harga-nya ADA.
            peta_lokal = part_index.rows_for_pns(unik)
        except Exception as e:
            return {"found": False, "error": f"gagal membaca indeks part: {e}"}
    peta_sims: dict[str, dict] = {}
    sumber_sims = ""
    if any(s["isi"] == ISI_HARGA_SIMS for s in specs):
        if len(unik) > _MAX_SIMS:
            return {"found": False,
                    "error": f"Terlalu banyak Part Number ({len(unik)}) untuk harga SIMS live. "
                             f"Maksimum {_MAX_SIMS} per permintaan — minta user memecah filenya."}
        res = harga.batch_harga(unik)
        peta_sims = {r["pn"]: r for r in res["results"]}
        sumber_sims = f"SIMS live (kurs CNY→IDR {res['rate']:.0f})"
    tersedia_gudang = part_index.gudang_names()

    # ── Lookup fitur baru (batch 1× per sheet, deterministik) ──
    peta_pengganti: dict[str, str] = {}
    if ISI_PENGGANTI in isis or tandai_status:
        peta_pengganti = {p: _pengganti_pn(p) for p in unik}
    peta_cross: dict[str, dict | None] = {}
    if ISI_CROSS_REF in isis:
        peta_cross = {p: filter_ref.by_pn(p) for p in unik}
    peta_berat: dict[str, int] = {}
    if ISI_BERAT in isis or rekap:
        peta_berat = {p: (harga.shipping_weight_for(p, allow_remote=False) or 0)
                      for p in unik}
    peta_dim: dict[str, str] = {}
    if ISI_DIMENSI in isis:
        sel = unik[:_MAX_DIM_SIMS]

        def _spec_dim(p):
            try:
                return (sims.get_part_spec(p) or {}).get("dimensi_cm") or ""
            except Exception:
                return ""
        with ThreadPoolExecutor(max_workers=6) as ex:
            for p, dim in zip(sel, ex.map(_spec_dim, sel)):
                peta_dim[p] = dim
    peta_pemenuhan: dict[str, dict] = {}
    if ISI_PEMENUHAN in isis:
        _is_buyer = (user.get("role") or "").lower() in ("buyer", "pembeli")
        all_names = accurate.warehouse_names()
        for p in unik:
            bd = accurate.gudang_breakdown(p) or {}
            avail = {}
            for g, q in bd.items():
                try:
                    avail[g] = max(0, int(q) - int(reservations.reserved_for(p, g) or 0))
                except Exception:
                    avail[g] = 0
            if _is_buyer:
                avail = gudang.shippable(avail)
            avail = gudang.scope_breakdown(avail, user.get("username", ""),
                                           user.get("role", ""), all_names, fallback=True)
            peta_pemenuhan[p] = {"bd": bd, "avail": avail}

    def _pemenuhan_str(p: str, qty: int | None) -> str:
        info = peta_pemenuhan.get(p)
        if not info:
            return ""
        avail = info["avail"]
        n = qty or 1
        penuh = sorted([g for g, q in avail.items() if q >= n],
                       key=lambda g: (-avail[g], g))
        if penuh:
            return f"SEMUA dari {gudang.gudang_label(penuh[0])}"
        tot = sum(avail.values())
        if tot >= n:
            parts, rem = [], n
            for g in sorted(avail, key=lambda g: (-avail[g], g)):
                if rem <= 0:
                    break
                take = min(avail[g], rem)
                if take > 0:
                    parts.append(f"{gudang.gudang_label(g)}({take})")
                    rem -= take
            return "SPLIT: " + " + ".join(parts)
        tot_bd = sum(_int_or_none(q) or 0 for q in info["bd"].values())
        if tot_bd >= n:
            return f"READY SETELAH RESERVASI DILEPAS (tertahan {tot_bd - tot})"
        return f"KURANG (butuh {n}, ada {tot})"

    def _nilai_fn(isi: str, gud_canon: str | None):
        """Fungsi nilai per-PN untuk satu kolom (menutup peta lookup)."""
        if isi == ISI_STOK and gud_canon:
            def f(p):
                d = peta_lokal.get(p)
                return "" if not d else str((d.get("gudang") or {}).get(gud_canon, 0))
            return f
        if isi == ISI_HARGA_SIMS:
            def f(p):
                d = peta_sims.get(p)
                return _fmt_rp(d["idr"]) if d and d.get("idr") is not None else ""
            return f
        if isi == ISI_PENGGANTI:
            return lambda p: peta_pengganti.get(p) or ""
        if isi == ISI_CROSS_REF:
            def f(p):
                cr = peta_cross.get(p)
                return ", ".join(cr.get("cross_reference") or []) if cr else ""
            return f
        if isi == ISI_BERAT:
            def f(p):
                g = peta_berat.get(p) or 0
                return f"{g / 1000:.2f}".replace(".", ",") if g > 0 else ""
            return f
        if isi == ISI_DIMENSI:
            return lambda p: peta_dim.get(p) or ""
        key = _ISI_KEY[isi]

        def f(p):
            d = peta_lokal.get(p)
            if not d:
                return ""
            v = _txt(d.get(key))
            return "" if v in ("N/A", "—") else v
        return f

    # PN yang TAK ketemu di sumber mana pun (setelah pemaaf suffix varian) —
    # DAFTARKAN (bukan cuma dihitung) agar miss SISTEMATIS (mis. semua varian '/2',
    # atau seluruh file salah kolom) terlihat oleh model & user. Cap 20 → hormati
    # budget token (_compact_result/_cap_tool_content).
    _ketemu = set(peta_lokal) | set(peta_sims)
    pn_tidak_ditemukan = [p for p in unik if p not in _ketemu]

    hasil: list[dict] = []
    gudang_tak_dikenal: list[str] = []
    gudang_dipakai = 0
    for s in specs:
        isi, gud = s["isi"], s["gudang"]
        gud_canon = None
        if isi == ISI_STOK and gud:
            if gudang_dipakai >= _MAX_GUDANG:
                continue
            gud_canon = _cocok_gudang(gud, tersedia_gudang)
            if not gud_canon:
                if gud not in gudang_tak_dikenal:
                    gudang_tak_dikenal.append(gud)
                continue
            gudang_dipakai += 1
            label = s["kolom_tujuan"] or f"Stok {gud_canon}"
        else:
            label = s["kolom_tujuan"] or _ISI_LABEL[isi]

        tgt = _cari_kolom(headers, label)
        if tgt is None:
            headers.append(label)
            tgt = len(headers) - 1
            for r in body:
                r.append("")

        terisi = 0
        if isi == ISI_PEMENUHAN:
            for idx, (r, p) in enumerate(zip(body, pns)):
                r[tgt] = _pemenuhan_str(p, qtys[idx]) if p else ""
                if r[tgt] != "":
                    terisi += 1
        else:
            fn = _nilai_fn(isi, gud_canon)
            for r, p in zip(body, pns):
                r[tgt] = fn(p)
                if r[tgt] != "":
                    terisi += 1
        hasil.append({"kolom": headers[tgt], "isi": isi,
                      "gudang": gud_canon, "baris_terisi": terisi})

    if not hasil and not (tandai_status or rekap):
        return {"found": False,
                "error": (f"Tak ada kolom yang bisa diisi. Gudang tak dikenal: {gudang_tak_dikenal}. "
                          f"Pilihan gudang: {tersedia_gudang}.")}

    # ── tandai_status (#3+#4+#5): kolom Status+Keterangan + WARNA baris ──
    row_status: list[str] = []
    mungkin_maksud: list[dict] = []
    pn_diganti: list[dict] = []
    status_ringkas = {"hijau": 0, "merah": 0, "kuning": 0}
    if tandai_status:
        fmap = part_index._pn_flat_map()
        saran = {p: _saran_pn(p, fmap) for p in pn_tidak_ditemukan}
        st_i = _cari_kolom(headers, "Status")
        if st_i is None:
            headers.append("Status"); st_i = len(headers) - 1
            for r in body:
                r.append("")
        ket_i = _cari_kolom(headers, "Keterangan")
        if ket_i is None:
            headers.append("Keterangan"); ket_i = len(headers) - 1
            for r in body:
                r.append("")
        for idx, (r, p) in enumerate(zip(body, pns)):
            warna, stat, ket = "", "", ""
            if p:
                d = peta_lokal.get(p)
                if not d:
                    pg = peta_pengganti.get(p) or ""
                    if pg:
                        warna, stat = "kuning", f"ADA PENGGANTI: {pg}"
                        pn_diganti.append({"pn": p, "pengganti": pg})
                    else:
                        warna, stat = "kuning", "PN TAK DITEMUKAN"
                        sg = saran.get(p) or ""
                        if sg:
                            ket = f"mungkin maksud: {sg} (belum dipastikan)"
                            mungkin_maksud.append({"pn": p, "saran": sg})
                else:
                    stok = _stok_total(d)
                    q = qtys[idx]
                    if stok is not None and stok <= 0:
                        warna, stat = "merah", "STOK KOSONG"
                    elif stok is not None and q and stok < q:
                        warna, stat = "merah", f"KURANG (butuh {q}, ada {stok})"
                    else:
                        warna, stat = "hijau", "READY"
            # jangan timpa keterangan user bila sudah terisi
            if not str(r[ket_i] or "").strip():
                r[ket_i] = ket
            r[st_i] = stat
            row_status.append(warna)
            if warna:
                status_ringkas[warna] += 1

    # ── rekap (#2): blok RINGKASAN ──
    ringkasan: list = []
    if rekap:
        pn_ada = [p for p in pns if p]
        ringkasan.append(("Jumlah item (baris ber-PN)", str(len(pn_ada)), ""))
        tot_qty = sum(q for q in qtys if q)
        if any(q for q in qtys):
            ringkasan.append(("Total qty", str(tot_qty), ""))
        if boleh_harga:
            subtotal = 0
            for p, q in zip(pns, qtys):
                d = peta_lokal.get(p)
                hv = _int_or_none((d or {}).get("harga"))
                if hv and q:
                    subtotal += hv * q
            if subtotal:
                ringkasan.append(("Subtotal (qty × harga)", _fmt_rp(subtotal), ""))
                ringkasan.append(("PPN 12% (sudah termasuk)",
                                  _fmt_rp(orders.ppn_included(subtotal)), "hijau"))
        if peta_berat:
            tot_g = sum((peta_berat.get(p) or 0) * (q or 1)
                        for p, q in zip(pns, qtys) if p)
            tanpa = sum(1 for p in set(pns) if p and not peta_berat.get(p))
            note = f"{tot_g / 1000:.1f} kg".replace(".", ",")
            if tanpa:
                note += f" (± {tanpa} part tanpa data berat)"
            ringkasan.append(("Berat tertagih total", note, ""))
            if kode_pos_tujuan and tot_g > 0:
                try:
                    from . import shipping
                    if shipping.available():
                        rates, _err = shipping.get_rates(
                            user.get("username", ""), tot_g, 0, dest_postal=kode_pos_tujuan)
                        for rt in (rates or [])[:3]:
                            ringkasan.append((
                                f"Ongkir {rt.get('courier_name') or rt.get('courier')} "
                                f"{rt.get('service')} (indikatif)",
                                _fmt_rp(rt.get('price')) + f" · {rt.get('etd') or '?'} hari", ""))
                except Exception:  # pragma: no cover
                    pass

    judul = f"{parsed['filename'].rsplit('.', 1)[0]} + Data"
    export_id, filename = _stash_sheet_out(
        judul, headers, body,
        status=row_status if tandai_status else None,
        ringkasan=ringkasan or None)
    out = {
        "found": True,
        "export_id": export_id,
        "filename": filename,
        "judul": judul,
        "jumlah_baris": len(body),
        "kolom_part_number": headers[pn_i],
        "kolom": hasil,
        "gudang_tak_dikenal": gudang_tak_dikenal,
        "pn_tidak_ditemukan": pn_tidak_ditemukan[:20],
        "pn_tidak_ditemukan_total": len(pn_tidak_ditemukan),
        "sumber": ("indeks part lokal (stok.xlsx/harga.xlsx/katalog)"
                   + (f" + {sumber_sims}" if sumber_sims else "")),
        "catatan": (
            "📎 SATU kartu unduh Excel muncul otomatis di bawah — SEMUA kolom yang diminta ada "
            "di file yang SAMA (jangan pernah membuat banyak file kecuali user eksplisit "
            "memintanya). Untuk stok per-gudang, '0' = terlacak tapi kosong di gudang itu, sel "
            "KOSONG = PN tak ada di sumber. "
            + (f"⚠️ Gudang tak dikenal: {gudang_tak_dikenal}. " if gudang_tak_dikenal else "")
            + (f"⚠️ {len(pn_tidak_ditemukan)} PN tak ketemu di sumber (lihat "
               "'pn_tidak_ditemukan'). " if pn_tidak_ditemukan else "")
            + ("Baris DIWARNAI: hijau=ready, merah=kosong/kurang, kuning=tak-ketemu/ada-pengganti. "
               if tandai_status else "")
            + ("⚠️ 'mungkin_maksud' = saran PN mirip yang BELUM DIPASTIKAN — sampaikan sebagai "
               "kemungkinan, ⛔ JANGAN sajikan sebagai PN pasti. " if mungkin_maksud else "")
            + "⛔ JANGAN mengarang nilai untuk baris kosong."
        ),
    }
    if tandai_status:
        out["status_ringkas"] = status_ringkas
        if mungkin_maksud:
            out["mungkin_maksud"] = mungkin_maksud[:20]
        if pn_diganti:
            out["pn_diganti"] = pn_diganti[:20]
    if rekap:
        out["rekap"] = [{"label": a, "nilai": b} for a, b, _c in ringkasan]
    return out
