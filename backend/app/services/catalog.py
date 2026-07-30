"""
Service: Batch Download — katalog Excel berisi gambar part.

Mirror app.py::build_catalog_excel + make_template_excel, decoupled dari
Streamlit. Untuk tiap part number: cari di index lokal (nama + file cocok),
tempel sampai 2 gambar SIMS berbeda ke Excel.
"""
from __future__ import annotations

import hashlib
import io

import requests

from . import part_index, sims

_MAX_BATCH = 300  # batas PN per batch (jaga waktu/proses)

# Cap KHUSUS bila kolom "exploded" dipilih. Satu PN yang belum pernah dibuka =
# 1 panggilan reverse global (respons bisa belasan ribu baris) + s/d 3 tree/item +
# unduh SVG + render resvg; terukur 94 detik untuk PN yang dipakai 17.185 model.
# Publik (tanpa underscore) karena router memang perlu membacanya.
MAX_BATCH_EXPLODED = 25

# Gambar TEKNIK: batas 200 px milik kolom `foto` membuat nomor balon tak terbaca.
# PNG-nya TIDAK di-resample — hanya ukuran TAMPIL yang diatur, jadi zoom & cetak
# di Excel tetap tajam (pola sama dengan ai_export.katalog_excel).
_EXPLODED_IMG_W = 620      # px lebar tampil
_EXPLODED_IMG_H = 500      # px tinggi tampil maks
_EXPLODED_COL_WIDTH = 88   # satuan lebar openpyxl (±7 px/satuan pada Calibri 11)
_MAX_ROW_PT = 409          # batas KERAS tinggi baris Excel = 409,5 pt (±546 px)

# Cegah Excel/CSV formula injection: nama part dari SIMS/lokal bisa diawali
# = + - @ → Excel menganggapnya FORMULA/DDE. Escape dgn prefiks ' → jadi teks.
_FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r", "\n")


def _xlsafe(v):
    if isinstance(v, str):
        s = v.lstrip("﻿")
        if s[:1] in _FORMULA_LEAD:
            return "'" + v
    return v


def parse_part_numbers(text: str) -> tuple[list[str], list[str]]:
    """Dari teks (1 PN per baris) → (unik urut, duplikat). Header dilewati."""
    raw = [ln.strip() for ln in (text or "").splitlines() if ln.strip()]
    if raw and raw[0].lower() in ("part number", "part_number", "partnumber", "no part", "kode"):
        raw = raw[1:]
    seen, uniq, dups = set(), [], []
    for pn in raw:
        up = pn.upper()
        if up in seen:
            dups.append(pn)
        else:
            seen.add(up)
            uniq.append(pn)
    return uniq, dups


def parse_part_numbers_from_file(filename: str, data: bytes) -> list[str]:
    """Ambil PN dari kolom A file Excel/CSV (mirror Streamlit)."""
    import pandas as pd

    bio = io.BytesIO(data)
    if filename.lower().endswith(".csv"):
        df = pd.read_csv(bio, header=None, dtype=str)
    else:
        df = pd.read_excel(bio, header=None, dtype=str)
    col_a = df.iloc[:, 0].dropna().astype(str).str.strip()
    if len(col_a) and col_a.iloc[0].lower() in (
        "part number", "part_number", "partnumber", "no part", "kode"
    ):
        col_a = col_a.iloc[1:]
    return col_a[col_a.str.len() > 0].tolist()


def _lookup_part(pn: str) -> dict:
    """Cari PN di index lokal → {part_name, kecocokan, found}."""
    found = part_index.search_part_number(pn)
    # exact match diutamakan (search PN = substring)
    exact = [r for r in found if r["part_number"].upper() == pn.upper()]
    hits = exact or found
    if hits:
        files = ", ".join(h["file"] for h in hits)
        return {"part_name": hits[0]["part_name"], "kecocokan": files, "found": True}
    return {"part_name": "", "kecocokan": "", "found": False}


def _fetch_image_bytes(url: str) -> bytes | None:
    try:
        r = requests.get(url, timeout=20)
        if r.status_code == 200 and r.content:
            return r.content
    except Exception:
        pass
    return None


# ── Kolom katalog yang bisa DIPILIH user ─────────────────────────────────────
# Part Number SELALU disertakan. Kunci di bawah = opsi yang bisa dicentang di
# frontend. Urutan _COLUMN_ORDER = urutan kolom di Excel (teks dulu, gambar kanan).
_COLUMN_DEFS: dict[str, tuple[str, int]] = {
    "nama": ("Part Name", 30),
    "kecocokan": ("Kecocokan", 45),
    "stok": ("Stok", 16),
    "harga_sims": ("Harga SIMS", 18),
    "harga_accurate": ("Harga Jual Accurate", 22),
    "harga_daftar": ("Harga (Daftar)", 18),
    # Kolom TEKS pendamping gambar exploded (nama figure, nomor balon, peringatan
    # lintas-model). Bukan chip tersendiri di UI — dinyalakan otomatis oleh
    # 'exploded' lewat _TURUNAN, jadi header/lebar/perataannya ikut mesin yang ada.
    "exploded_info": ("Info Exploded View", 46),
}
_COLUMN_ORDER = [
    "nama", "kecocokan", "stok", "harga_sims", "harga_accurate", "harga_daftar",
    "exploded_info",        # teks dulu…
    "foto", "exploded",     # …lalu gambar (foto SIMS, baru exploded EPC)
]
# Satu chip di UI → beberapa kolom di Excel.
_TURUNAN = {"exploded": ("exploded_info",)}
# Kolom gambar: bukan kolom teks, jadi TIDAK ada di _COLUMN_DEFS (pola `foto`).
_IMG_COLUMNS = ("foto", "exploded")
# Teks panjang/multi-baris → rata kiri.
_ALIGN_LEFT = ("nama", "kecocokan", "exploded_info")
# Default bila user tidak memilih apa pun (perilaku ramah: identitas + foto + stok).
_DEFAULT_COLUMNS = ("nama", "foto", "stok")

# Kolom harga — hanya boleh untuk admin & akun SEE_ALL (mis. 'mas'). Di-strip di
# endpoint untuk akun lain (lihat routers/parts.py::batch_catalog).
PRICE_COLUMNS = frozenset({"harga_sims", "harga_accurate", "harga_daftar"})


def clean_columns(cols) -> list[str]:
    """Saring pilihan kolom user → hanya kunci sah, urut sesuai _COLUMN_ORDER.
    Kosong/tak sah → _DEFAULT_COLUMNS. Kolom TURUNAN dinyalakan otomatis
    (mis. 'exploded' → ikut membawa kolom teks 'exploded_info')."""
    if not cols:
        return list(_DEFAULT_COLUMNS)
    want = {str(c).strip().lower() for c in cols}
    for induk, anak in _TURUNAN.items():
        if induk in want:
            want.update(anak)
    out = [c for c in _COLUMN_ORDER if c in want]
    return out or list(_DEFAULT_COLUMNS)


def _teks_info_exploded(d: dict) -> str:
    """Isi sel "Info Exploded View".

    Tak ditemukan / gagal render / kehabisan anggaran waktu → tulis ALASANNYA apa
    adanya. ⛔ Jangan pernah mengisi nama figure atau nomor balon karangan."""
    if not d:
        return "—"
    if not d.get("found"):
        return d.get("alasan") or "Figure exploded view tidak ditemukan di EPC."
    baris = []
    if d.get("figure_nama"):
        baris.append(f"Figure: {d['figure_nama']}"
                     + (f" ({d['figure_pn']})" if d.get("figure_pn") else ""))
    bal = d.get("balon")
    baris.append(f"Balon: {bal}" if bal not in (None, "")
                 else "Balon: tidak terdeteksi di figure ini")
    if d.get("nama_item"):
        baris.append(f"Nama di EPC: {d['nama_item']}")
    if d.get("jumlah_item"):
        baris.append(f"{d['jumlah_item']} part di gambar")
    if d.get("catatan"):
        baris.append("⚠️ " + d["catatan"])
    return "\n".join(b for b in baris if str(b).strip())


def _fmt_qty(v) -> str:
    try:
        f = float(v)
    except (TypeError, ValueError):
        return "0"
    return str(int(f)) if f == int(f) else f"{f:g}"


def build_catalog_excel(part_numbers: list[str], columns=None, on_progress=None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage

    from . import accurate, harga

    cols = clean_columns(columns)
    text_cols = [c for c in cols if c not in _IMG_COLUMNS]  # kolom teks, terurut
    has_foto = "foto" in cols
    has_exploded = "exploded" in cols

    # Prefetch PARALEL semua gambar exploded dulu (pola ai_export.katalog_excel):
    # loop baris di bawah tetap deterministik & cepat. Cache dibagi dengan halaman
    # detail part, jadi PN yang sudah pernah dibuka praktis gratis.
    exp_map: dict[str, dict] = {}
    if has_exploded:
        from . import exploded_view          # impor lokal: konsisten dgn modul lain
        exp_map = exploded_view.png_batch(part_numbers)
    need_accurate = "stok" in cols or "harga_accurate" in cols
    need_sims_price = "harga_sims" in cols

    rate = 0.0
    if need_sims_price:
        try:
            rate, _ = harga.get_rate()
        except Exception:
            rate = 0.0

    wb = Workbook()
    ws = wb.active
    ws.title = "Catalog"

    header_fill = PatternFill("solid", fgColor="1565C0")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Susun header dinamis: Part Number + kolom teks terpilih + (Gambar 1, Gambar 2).
    headers = ["Part Number"] + [_COLUMN_DEFS[c][0] for c in text_cols]
    widths = [20] + [_COLUMN_DEFS[c][1] for c in text_cols]
    img_col_1 = exp_col = None
    if has_foto:
        img_col_1 = len(headers) + 1            # indeks 1-based kolom "Gambar 1"
        headers += ["Gambar 1", "Gambar 2"]
        widths += [38, 38]
    if has_exploded:
        exp_col = len(headers) + 1
        headers += ["Exploded View (EPC · lintas model)"]
        widths += [_EXPLODED_COL_WIDTH]
    for ci, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    fill_even = PatternFill("solid", fgColor="E3F2FD")
    fill_odd = PatternFill("solid", fgColor="FAFAFA")
    fill_nf = PatternFill("solid", fgColor="FFEBEE")

    def _xl_image(img_bytes: bytes, max_h: int = 200):
        pil = PILImage.open(io.BytesIO(img_bytes)).convert("RGB")
        w_px, h_px = pil.size
        if h_px > max_h:
            ratio = max_h / h_px
            w_px, h_px = int(w_px * ratio), max_h
            pil = pil.resize((w_px, h_px), PILImage.LANCZOS)
        bio = io.BytesIO()
        pil.save(bio, format="PNG")
        bio.seek(0)
        xl = XLImage(bio)
        xl.width, xl.height = w_px, h_px
        return xl, h_px

    def _xl_image_exploded(png: bytes):
        """Gambar exploded: TANPA resample — hanya ukuran tampil yang diatur supaya
        zoom/cetak tetap tajam. Di-clamp DUA sisi: tinggi baris Excel maksimum
        409,5 pt (±546 px), jadi figure portrait harus ikut menyusut."""
        xl = XLImage(io.BytesIO(png))
        w, h = int(xl.width or 1), int(xl.height or 1)
        skala = min(_EXPLODED_IMG_W / w, _EXPLODED_IMG_H / h, 1.0)
        xl.width, xl.height = max(1, int(w * skala)), max(1, int(h * skala))
        return xl, xl.height

    row_idx = 2
    total = len(part_numbers)
    for i, pn in enumerate(part_numbers):
        if on_progress:
            on_progress(i, total, pn)

        info = _lookup_part(pn)
        part_name = info["part_name"]
        is_found = info["found"]
        fill = (fill_even if i % 2 == 0 else fill_odd) if is_found else fill_nf

        stok_entry = accurate.stock_for(pn) if need_accurate else None

        # Nilai tiap kolom teks terpilih.
        vals: dict[str, str] = {}
        if "kecocokan" in cols:
            vals["kecocokan"] = info["kecocokan"] or "—"
        if "stok" in cols:
            vals["stok"] = (
                f'{_fmt_qty(stok_entry["available_to_sell"])} {stok_entry.get("unit", "")}'.strip()
                if stok_entry else "—"
            )
        if "harga_accurate" in cols:
            p = stok_entry.get("price") if stok_entry else 0
            vals["harga_accurate"] = harga.fmt_rp(p) if p else "—"
        if "harga_sims" in cols:
            try:
                cny, _note = sims.get_price(pn)
            except Exception:
                cny = None
            vals["harga_sims"] = harga.fmt_rp(round(cny * rate)) if (cny is not None and rate) else "—"
        if "harga_daftar" in cols:
            dp, dname = harga.price_for(pn)
            vals["harga_daftar"] = harga.fmt_rp(dp) if dp else "—"
            if not part_name and dname:
                part_name = dname

        # Fallback nama part: lokal → Accurate → daftar harga → SIMS (paling akhir, jaringan).
        if "nama" in cols and not part_name and stok_entry and stok_entry.get("name"):
            part_name = stok_entry["name"]
        if "nama" in cols and not part_name:
            _dp, dname = harga.price_for(pn)
            if dname:
                part_name = dname

        # Gambar (hanya bila kolom foto dipilih — melewati ini mempercepat batch).
        img_d = img_e = img_exp = None
        row_height = 80 if (has_foto or has_exploded) else 22
        if has_foto:
            urls = sims.get_images(pn)
            if urls:
                b1 = _fetch_image_bytes(urls[0])
                if b1:
                    try:
                        xl, h = _xl_image(b1)
                        img_d = xl
                        row_height = max(int(h * 0.75) + 10, row_height)
                        hash1 = hashlib.md5(b1).hexdigest()
                        for u2 in urls[1:]:
                            b2 = _fetch_image_bytes(u2)
                            if b2 and hashlib.md5(b2).hexdigest() != hash1:
                                xl2, h2 = _xl_image(b2)
                                img_e = xl2
                                row_height = max(int(h2 * 0.75) + 10, row_height)
                                break
                    except Exception:
                        pass

        # Gambar exploded (sudah diambil paralel di atas — di sini cuma menempel).
        if has_exploded:
            d_exp = exp_map.get(exploded_view.kunci_pn(pn)) or {}
            if d_exp.get("found") and d_exp.get("png"):
                try:
                    img_exp, h_exp = _xl_image_exploded(d_exp["png"])
                    row_height = max(row_height,
                                     min(_MAX_ROW_PT, int(h_exp * 0.75) + 10))
                except Exception:
                    img_exp = None
            # Kolom teks: APA ADANYA — PN tanpa figure tidak dikarang.
            vals["exploded_info"] = _teks_info_exploded(d_exp)

        # Nama dari SIMS kalau masih kosong (upaya terakhir, jaringan).
        if "nama" in cols and not part_name:
            try:
                info_sims = sims.get_part_info(pn)
                if info_sims.get("partName"):
                    part_name = info_sims["partName"]
            except Exception:
                pass
        vals["nama"] = part_name or ""

        # Tulis Part Number.
        c = ws.cell(row=row_idx, column=1, value=_xlsafe(pn))
        c.fill, c.border, c.alignment, c.font = fill, border, center, Font(name="Arial", size=10)
        # Tulis kolom teks terpilih (nama/kecocokan rata kiri, sisanya tengah).
        for offset, key in enumerate(text_cols, start=2):
            aln = left if key in _ALIGN_LEFT else center
            c = ws.cell(row=row_idx, column=offset, value=_xlsafe(vals.get(key, "")))
            c.fill, c.border, c.alignment, c.font = fill, border, aln, Font(name="Arial", size=10)
        # Kolom gambar: sel kosong bergaya + tempel gambar.
        if has_foto:
            for ci in (img_col_1, img_col_1 + 1):
                c = ws.cell(row=row_idx, column=ci, value="")
                c.fill, c.border, c.alignment = fill, border, center
            if img_d:
                ws.add_image(img_d, f"{get_column_letter(img_col_1)}{row_idx}")
            if img_e:
                ws.add_image(img_e, f"{get_column_letter(img_col_1 + 1)}{row_idx}")
        if exp_col:
            c = ws.cell(row=row_idx, column=exp_col, value="")
            c.fill, c.border, c.alignment = fill, border, center
            if img_exp:
                ws.add_image(img_exp, f"{get_column_letter(exp_col)}{row_idx}")

        ws.row_dimensions[row_idx].height = row_height
        row_idx += 1

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_template_excel() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Part Number List"
    ws["A1"] = "Part Number"
    ws["A1"].font = Font(bold=True, name="Arial", size=11, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1565C0")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    for i, ex in enumerate(
        ["WG1642821034", "WG9925520270", "AZ9100443082", "WG9718820030"], start=2
    ):
        ws.cell(row=i, column=1, value=ex).font = Font(name="Arial", size=10)
    ws.column_dimensions["A"].width = 28
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
