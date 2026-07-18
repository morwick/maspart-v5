"""Test builder ai_export.sheet_status_excel + _save_stable + tool template.

Fokus: (1) warna baris jatuh di baris yang benar (hijau/merah/kuning via status),
(2) kolom teks 'Status' tetap ada (dwi-encode), (3) blok RINGKASAN muncul,
(4) BYTE-STABLE (build 2× → bytes identik), (5) _t_template_excel hasilkan export
yang bisa dibangun generic_excel.
"""
import io

from openpyxl import load_workbook

from app.services import ai_export


def _payload():
    return {
        "kind": "sheet_status",
        "judul": "Uji Status",
        "kolom": ["No", "Part Number", "Stok", "Status"],
        "baris": [
            ["1", "WG111", "5", "READY"],
            ["2", "WG222", "0", "STOK KOSONG"],
            ["3", "WG333", "", "PN TAK DITEMUKAN"],
        ],
        "status": ["hijau", "merah", "kuning"],
        "ringkasan": [("Jumlah item", "3", ""), ("Total qty", "7", "hijau")],
    }


def _fill_hex(cell):
    fg = cell.fill.fgColor
    return (fg.rgb or "")[-6:] if fg and fg.rgb else ""


def test_warna_dan_status_kolom():
    data, err = ai_export.sheet_status_excel(_payload())
    assert data and not err
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Data"]
    # _title pakai baris 1-2, header di baris 4, data mulai baris 5,6,7
    assert _fill_hex(ws.cell(row=5, column=2)) == "EAF6EC"   # hijau
    assert _fill_hex(ws.cell(row=6, column=2)) == "FDECEA"   # merah
    assert _fill_hex(ws.cell(row=7, column=2)) == "FFF6DC"   # kuning
    # kolom Status berisi teks (bukan hanya warna)
    assert ws.cell(row=6, column=4).value == "STOK KOSONG"
    assert ws.cell(row=7, column=4).value == "PN TAK DITEMUKAN"


def test_blok_ringkasan_muncul():
    data, _ = ai_export.sheet_status_excel(_payload())
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Data"]
    teks = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "RINGKASAN" in teks
    assert "Jumlah item" in teks and "Total qty" in teks


def test_byte_stable():
    a, _ = ai_export.sheet_status_excel(_payload())
    b, _ = ai_export.sheet_status_excel(_payload())
    assert a == b and a is not None


def test_payload_kosong():
    data, err = ai_export.sheet_status_excel({"kind": "sheet_status", "kolom": []})
    assert data is None and err


def test_template_tool_bisa_dibangun():
    from app.services import ai_assistant as ai
    r = ai._t_template_excel({"dengan_contoh": True}, {"username": "x", "role": "buyer"})
    assert r["found"] and r["export_id"]
    data, fn = ai_export.generic_excel(r["export_id"])
    assert data and fn.endswith(".xlsx")
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Data"]
    hdr = [ws.cell(row=4, column=j).value for j in range(1, 6)]
    assert hdr == ["No", "Part Number", "Nama Part", "Qty", "Keterangan"]
