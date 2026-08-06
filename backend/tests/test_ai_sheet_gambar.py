"""Tool asisten `sheet_isi_gambar`: FOTO fisik (SIMS) dan/atau GAMBAR TEKNIS
(exploded view EPC) ke Excel unggahan user — SATU panggilan = SATU file.

Latar (keluhan pemilik 2026-08-06): minta "foto part + gambar exploded"
menghasilkan DUA kartu unduh & dua file, karena dulu ada dua tool terpisah yang
masing-masing menstash filenya sendiri. Aturan pemilik: satu permintaan = SATU
file Excel, kecuali user memang minta dipisah.

Yang dijaga di sini:
  • dua jenis dalam satu panggilan → SATU export_id, SATU builder, kolom foto &
    kolom gambar teknis berdampingan di file yang sama;
  • gerbang "tanya VIN dulu" hanya berlaku untuk jenis 'exploded' — permintaan
    FOTO saja tak boleh mendadak ditanyai nomor rangka;
  • plafon PN memakai batas TERSEMPIT dari jenis yang diminta;
  • tool lama (sheet_isi_foto / sheet_isi_exploded) tak lagi ditawarkan ke model,
    tapi tetap SAH dieksekusi (alias legacy) & tetap memberi file satu-jenis.
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from app.services import ai_export, ai_sheet, epc_bom, exploded_view
from app.services import ai_assistant as A

USER = {"username": "budi", "role": "staf"}
FRAME = "RJ326978"
PN_ADA = "WG9725520278"
PN_TAK = "AZ9100440002"

FOTO = {PN_ADA: ["http://sims/a-1.jpg", "http://sims/a-2.jpg"], PN_TAK: []}


def _png(w: int = 1400, h: int = 900) -> bytes:
    from PIL import Image
    b = io.BytesIO()
    Image.new("RGB", (w, h), "white").save(b, format="PNG")
    return b.getvalue()


PNG = _png()

_ROOT = {"data": {"orderNo": "ORD-1", "partRoots": [{"id": 1}]}}
_REVERSE_ADA = {"data": [{"partCode": "ASSY-1", "partName": "Front axle",
                          "rootId": 1, "partId": 22, "partListId": 222}]}
_ITEMS = {"data": {"partListName": "Front axle assembly", "d2s": ["FIG-1.svg"],
                   "items": [{"code": PN_ADA, "name": "Wheel hub",
                              "ballNum": 7, "amount": 2}]}}


@pytest.fixture(autouse=True)
def bersih(monkeypatch):
    exploded_view.bersihkan_cache()
    epc_bom._root_cache.clear()
    monkeypatch.setattr(ai_sheet.sims, "available", lambda: True)
    monkeypatch.setattr(ai_sheet.sims, "get_images", lambda pn: list(FOTO.get(pn, [])))
    yield
    exploded_view.bersihkan_cache()
    epc_bom._root_cache.clear()


@pytest.fixture
def epc_stub(monkeypatch):
    def fake_get_auto(url, params, **kw):
        if url == epc_bom._ATLAS_ROOT_URL:
            return _ROOT
        if url == epc_bom._REVERSE_URL:
            return _REVERSE_ADA if (params.get("k") or "").upper() == PN_ADA else {"data": []}
        if url == epc_bom._ATLAS_ITEM_URL:
            return _ITEMS
        return {"_err": "api"}

    monkeypatch.setattr(epc_bom, "_get_auto", fake_get_auto)
    monkeypatch.setattr(ai_export, "exploded_png", lambda svg, ball=None: PNG)


def _sheet(headers, roles, body, filename="recom tambahan CPM.xlsx") -> str:
    return ai_sheet.put_sheet("budi", {
        "filename": filename, "headers": list(headers), "roles": list(roles),
        "_body": [list(r) for r in body],
        "contoh": [], "sheet": "Sheet1", "sheet_lain": [], "pn_dikenal": len(body),
        "jumlah_baris": len(body), "jumlah_kolom": len(headers),
        "kolom_pn": headers[roles.index("part_number")] if "part_number" in roles else None,
        "terpotong": False,
    })


@pytest.fixture
def sheet():
    return _sheet(["Part No.", "Nama", "Qty"], ["part_number", "part_name", "qty"],
                  [[PN_ADA, "Wheel hub", "2"], [PN_TAK, "Bracket", "1"]])


# ── Inti keluhan: dua jenis → SATU file ─────────────────────────────────────

def test_foto_dan_exploded_jadi_satu_file(sheet):
    out = ai_sheet.fill_gambar(sheet, USER, ["foto", "exploded"], rangka=FRAME)
    assert out["found"] is True and out["satu_file"] is True
    assert out["jenis"] == ["foto", "exploded"]
    # SATU kartu unduh: satu export_id, satu builder.
    b = ai_export._stash[out["export_id"]]["builder"]
    assert b["kind"] == "sheet_gambar"
    # Urutan kolom: teks info → foto → gambar teknis (sama dgn Batch Download).
    assert b["kolom"][3:] == ["Info Gambar Teknis", "Foto 1", "Foto 2",
                              "Gambar Teknis (Exploded View)"]
    assert b["kol_info"] == 3 and b["kol_foto"] == [4, 5] and b["kol_gambar"] == 6
    assert b["rangka"] == FRAME and b["pns"] == [PN_ADA, PN_TAK]
    # Laporan kedua jenis ada di SATU hasil tool.
    assert out["baris_berfoto"] == 1 and out["baris_tanpa_foto"] == 1
    assert out["jumlah_part"] == 2 and out["estimasi_durasi_unduhan"]
    assert "SATU kartu unduh" in out["catatan"] and "JANGAN membuat file kedua" in out["catatan"]


def test_builder_menempel_foto_dan_gambar_teknis_di_satu_workbook(sheet, epc_stub, monkeypatch):
    from PIL import Image as PILImage
    buf = io.BytesIO()
    PILImage.new("RGB", (3000, 2000), "red").save(buf, format="JPEG")

    class _Resp:
        status_code = 200
        content = buf.getvalue()

    import requests
    monkeypatch.setattr(requests, "get", lambda url, timeout=0: _Resp())

    out = ai_sheet.fill_gambar(sheet, USER, ["foto", "exploded"], rangka=FRAME)
    data, err = ai_export.sheet_gambar_excel(ai_export._stash[out["export_id"]]["builder"])
    assert err == "" and data
    ws = openpyxl.load_workbook(io.BytesIO(data))["Data"]
    # 2 foto + 1 gambar teknis, semua di file yang SAMA.
    tinggi = sorted(im.height for im in ws._images)
    assert len(tinggi) == 3
    assert tinggi[:2] == [ai_export._FOTO_H_PX, ai_export._FOTO_H_PX]   # foto diciutkan
    assert tinggi[2] > ai_export._FOTO_H_PX                             # figure jauh lebih besar
    # Tinggi baris mengikuti gambar TERTINGGI (kalau ikut foto, figure terpotong).
    assert ws.row_dimensions[5].height > ai_export._FOTO_H_PX * 0.78
    kol = {c.value: c.column for c in ws[4]}
    info = ws.cell(row=5, column=kol["Info Gambar Teknis"]).value
    assert "per-VIN" in info and "Balon: 7" in info


# ── Gerbang tanya VIN: hanya untuk exploded ─────────────────────────────────

def test_foto_saja_tak_ditanyai_vin(sheet):
    out = ai_sheet.fill_gambar(sheet, USER, ["foto"])
    assert out["found"] is True and "perlu_jawaban_user" not in out
    b = ai_export._stash[out["export_id"]]["builder"]
    assert "rangka" not in b and b["kol_foto"] == [3, 4]


def test_exploded_ikut_diminta_tetap_memicu_tanya_vin(sheet):
    out = ai_sheet.fill_gambar(sheet, USER, ["foto", "exploded"])
    assert out["perlu_jawaban_user"] is True
    assert out["jenis_diminta"] == ["foto", "exploded"]
    assert "export_id" not in out                      # tak ada file dibuat dulu
    assert A._tool_fail_kind(out) == ""                # bertanya ≠ lookup gagal
    # Model harus memanggil ULANG dengan KEDUA jenis → tetap satu file.
    assert "jenis_diminta" in out["jawaban_wajib"] and "SATU file" in out["jawaban_wajib"]


def test_jenis_kosong_atau_tak_dikenal_ditolak_dengan_penjelasan(sheet):
    for j in ([], ["entah"], ""):
        out = ai_sheet.fill_gambar(sheet, USER, j)
        assert out["found"] is False and "jenis" in out["error"].lower()


def test_jenis_diterima_dalam_bentuk_string_majemuk(sheet):
    """Model tak selalu patuh skema array — 'foto, exploded' harus tetap dipahami."""
    assert ai_sheet.jenis_norm("foto, exploded") == ["foto", "exploded"]
    assert ai_sheet.jenis_norm("gambar_teknis") == ["exploded"]
    assert ai_sheet.jenis_norm(["photo"]) == ["foto"]
    assert ai_sheet.jenis_norm(["exploded", "foto"]) == ["foto", "exploded"]   # urut tetap


# ── Plafon: batas tersempit yang menang ─────────────────────────────────────

def test_plafon_memakai_batas_tersempit(monkeypatch):
    n = ai_sheet._MAX_EXPLODED_PN_GLOBAL + 1        # 26 PN: muat foto, tak muat exploded
    sid = _sheet(["Part No."], ["part_number"], [[f"WG97255202{i:02d}"] for i in range(n)])
    monkeypatch.setattr(ai_sheet.sims, "get_images", lambda pn: [])
    assert ai_sheet.fill_gambar(sid, USER, ["foto"])["found"] is True
    out = ai_sheet.fill_gambar(sid, USER, ["foto", "exploded"], lintas_model=True)
    assert out["found"] is False and "terlalu banyak" in out["error"].lower()
    assert "per-VIN" in out["error"]                # tawarkan jalur yang muat 60 PN
    assert ai_sheet.fill_gambar(sid, USER, ["foto", "exploded"], rangka=FRAME)["found"] is True


# ── Tool lama: tak ditawarkan, tapi tetap jalan ─────────────────────────────

def test_tool_lama_jadi_shim_satu_jenis(sheet):
    dengan = {f["function"]["name"] for f in A._tool_specs(USER, sheet)}
    assert "sheet_isi_gambar" in dengan
    assert not {"sheet_isi_foto", "sheet_isi_exploded"} & dengan
    izin = A._allowed_tool_names(USER, sheet)
    assert {"sheet_isi_foto", "sheet_isi_exploded"} <= izin      # alias legacy

    r = A._run_tool("sheet_isi_foto", {}, USER, sheet_id=sheet)
    assert r["found"] is True and r["jenis"] == ["foto"]
    r2 = A._run_tool("sheet_isi_exploded", {"rangka": FRAME}, USER, sheet_id=sheet)
    assert r2["found"] is True and r2["jenis"] == ["exploded"]


def test_dispatch_dan_kartu_unduh_terpasang():
    assert A._DISPATCH["sheet_isi_gambar"] is A._t_sheet_isi_gambar
    import inspect
    assert '"sheet_isi_gambar"' in inspect.getsource(A.chat)   # kartu unduh ditangkap
