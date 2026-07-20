"""Isi kolom Excel unggahan → SEL NUMERIK, dari ujung ke ujung.

Membuktikan keluhan pemilik (2026-07-20) benar-benar tertutup: sebelumnya kolom
Harga berisi teks "Rp 1.500.000" sehingga SUM() = 0. Test ini membangun file
sungguhan lalu memeriksa TIPE sel lewat openpyxl.
"""
import io

import pytest
from openpyxl import Workbook, load_workbook

from app.services import ai_export, ai_sheet

USER = {"username": "admin", "role": "admin"}


@pytest.fixture
def dunia(monkeypatch):
    """Indeks part palsu: field string tampilan + field *_num mentah."""
    baris = {
        "WG9925520270": {
            "part_number": "WG9925520270", "part_name": "Spring bracket",
            "stok": "12", "harga": "Rp 1.500.000",
            "stok_num": 12, "harga_num": 1_500_000,
            "gudang": {"01.Jakarta": 8, "02.Pekanbaru": 4},
        },
        "AZ9925520271": {
            "part_number": "AZ9925520271", "part_name": "Leaf spring",
            "stok": "0", "harga": "—",
            "stok_num": 0, "harga_num": None,
            "gudang": {},
        },
    }
    monkeypatch.setattr(ai_sheet.part_index, "rows_for_pns", lambda pns: {
        p: baris[p] for p in pns if p in baris})
    return baris


def _sid(rows):
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    parsed = ai_sheet.parse_upload(buf.getvalue(), "uji.xlsx")
    assert parsed.get("ok"), parsed
    return ai_sheet.put_sheet(USER["username"], parsed)


def _ws(export_id):
    data, _ = ai_export.generic_excel(export_id)
    return load_workbook(io.BytesIO(data))["Data"]


def _kolom(ws, header):
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value == header:
                return r, c
    raise AssertionError(f"header {header!r} tak ada")


# ── inti keluhan ─────────────────────────────────────────────────────
def test_isi_kolom_harga_sel_int(dunia):
    sid = _sid([["Part Number"], ["WG9925520270"]])
    r = ai_sheet.fill_columns(sid, USER, permintaan=[{"isi": "harga_lokal"}])
    ws = _ws(r["export_id"])
    hr, hc = _kolom(ws, "Harga")
    sel = ws.cell(row=hr + 1, column=hc)
    assert isinstance(sel.value, int) and sel.value == 1_500_000
    assert sel.data_type == "n"          # numerik di mata Excel, bukan teks


def test_isi_kolom_stok_sel_int_termasuk_nol(dunia):
    sid = _sid([["Part Number"], ["WG9925520270"], ["AZ9925520271"]])
    r = ai_sheet.fill_columns(sid, USER, permintaan=[{"isi": "stok"}])
    ws = _ws(r["export_id"])
    hr, hc = _kolom(ws, "Stok")
    assert ws.cell(row=hr + 1, column=hc).value == 12
    # stok 0 harus tertulis 0, bukan kosong — SUM & pembacaan manusia beda arti
    assert ws.cell(row=hr + 2, column=hc).value == 0


def test_part_number_tetap_teks(dunia):
    sid = _sid([["Part Number"], ["WG9925520270"]])
    r = ai_sheet.fill_columns(sid, USER, permintaan=[{"isi": "harga_lokal"}])
    ws = _ws(r["export_id"])
    pr, pc = _kolom(ws, "Part Number")
    assert isinstance(ws.cell(row=pr + 1, column=pc).value, str)


def test_kolom_nama_tetap_teks(dunia):
    sid = _sid([["Part Number"], ["WG9925520270"]])
    r = ai_sheet.fill_columns(sid, USER, permintaan=[{"isi": "nama_part"}])
    ws = _ws(r["export_id"])
    hr, hc = _kolom(ws, "Nama Part")
    assert ws.cell(row=hr + 1, column=hc).value == "Spring bracket"


def test_stok_per_gudang_int(dunia):
    sid = _sid([["Part Number"], ["WG9925520270"]])
    r = ai_sheet.fill_columns(sid, USER,
                              permintaan=[{"isi": "stok", "gudang": "Jakarta"}])
    ws = _ws(r["export_id"])
    hdr = [ws.cell(row=4, column=j).value for j in range(1, ws.max_column + 1)]
    kol = [h for h in hdr if h and h.startswith("Stok")][0]
    hr, hc = _kolom(ws, kol)
    assert ws.cell(row=hr + 1, column=hc).value == 8


def test_qty_milik_user_ikut_jadi_angka(dunia):
    """Kolom Qty diketik user sendiri; parse unggahan menstringkannya. Koersi di
    titik tulis mengembalikannya jadi angka supaya =Qty*Harga & SUM jalan."""
    sid = _sid([["Part Number", "Qty"], ["WG9925520270", 3]])
    r = ai_sheet.fill_columns(sid, USER, permintaan=[{"isi": "harga_lokal"}])
    ws = _ws(r["export_id"])
    qr, qc = _kolom(ws, "Qty")
    assert ws.cell(row=qr + 1, column=qc).value == 3
    assert isinstance(ws.cell(row=qr + 1, column=qc).value, int)


def test_seluruh_kolom_harga_bertipe_angka_sehingga_sum_jalan(dunia):
    """Pembuktian langsung keluhan: SUM() atas kolom Harga tak akan 0 karena
    SEMUA selnya bertipe numerik."""
    sid = _sid([["Part Number"], ["WG9925520270"], ["WG9925520270"]])
    r = ai_sheet.fill_columns(sid, USER, permintaan=[{"isi": "harga_lokal"}])
    ws = _ws(r["export_id"])
    hr, hc = _kolom(ws, "Harga")
    nilai = [ws.cell(row=hr + 1 + i, column=hc) for i in range(2)]
    assert all(c.data_type == "n" for c in nilai)
    assert sum(c.value for c in nilai) == 3_000_000


def test_fallback_tanpa_field_num(monkeypatch):
    """Pemanggil/fixture lama yang hanya punya string tampilan tetap
    menghasilkan sel numerik lewat jalur fallback."""
    monkeypatch.setattr(ai_sheet.part_index, "rows_for_pns", lambda pns: {
        "WG9925520270": {"part_number": "WG9925520270", "part_name": "X",
                         "stok": "12", "harga": "Rp 1.500.000", "gudang": {}}})
    sid = _sid([["Part Number"], ["WG9925520270"]])
    r = ai_sheet.fill_columns(sid, USER, permintaan=[{"isi": "harga_lokal"}])
    ws = _ws(r["export_id"])
    hr, hc = _kolom(ws, "Harga")
    assert ws.cell(row=hr + 1, column=hc).value == 1_500_000


def test_harga_tak_ada_tetap_kosong_bukan_nol(dunia):
    """PN tanpa harga di Accurate jangan jadi 0 — SUM akan berbohong."""
    sid = _sid([["Part Number"], ["AZ9925520271"]])
    r = ai_sheet.fill_columns(sid, USER, permintaan=[{"isi": "harga_lokal"}])
    ws = _ws(r["export_id"])
    hr, hc = _kolom(ws, "Harga")
    assert ws.cell(row=hr + 1, column=hc).value in (None, "")


# ── pagar: tampilan tetap "Rp …" ─────────────────────────────────────
def test_rp_tampilan_masih_berformat():
    assert ai_sheet._rp_tampilan(1_500_000) == "Rp 1.500.000"


def test_int_or_none_float_tidak_dikali_sepuluh():
    """Regresi: jalur string membuang titik → 1500000.0 dulu jadi 15000000,
    membuat Subtotal user 10x lipat tanpa gejala."""
    assert ai_sheet._int_or_none(1_500_000.0) == 1_500_000
    assert ai_sheet._int_or_none(1_500_000) == 1_500_000
    assert ai_sheet._int_or_none("Rp 1.500.000") == 1_500_000
    assert ai_sheet._int_or_none(True) is None
