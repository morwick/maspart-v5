"""Excel unggahan user di chat asisten: deteksi kolom, isi kolom, scoping harga SIMS."""
import io

import pytest
from openpyxl import Workbook

from app.services import ai_assistant as ai
from app.services import ai_sheet

ADMIN = {"username": "admin", "role": "admin"}
USER = {"username": "budi", "role": "user"}
PEMBELI = {"username": "toko", "role": "pembeli"}


def _xlsx(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_multi(sheets: dict) -> bytes:
    """sheets = {nama_sheet: [rows]}. Sheet pertama jadi aktif default."""
    wb = Workbook()
    wb.remove(wb.active)
    for nama, rows in sheets.items():
        ws = wb.create_sheet(title=nama)
        for r in rows:
            ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def katalog(monkeypatch):
    """Katalog palsu: 2 PN dikenal, lengkap dgn stok/harga/nama."""
    rows = [
        {"part_number": "WG9925520270", "part_name": "Spring bracket", "stok": "12", "harga": "Rp 1.500.000"},
        {"part_number": "AZ9925520271", "part_name": "Leaf spring", "stok": "0", "harga": "—"},
    ]
    monkeypatch.setattr(ai_sheet.part_index, "search_exact_pns",
                        lambda pns: [r for r in rows if (r["part_number"] or "").upper() in {p.upper() for p in pns}])
    # rows_for_pns (pemaaf suffix varian) melangkah ke _pn_flat_map utk PN tak
    # ketemu → jangan sentuh indeks nyata di test.
    monkeypatch.setattr(ai_sheet.part_index, "_pn_flat_map", lambda: {})
    return rows


# ── Parsing & deteksi peran kolom ────────────────────────────────────────────

def test_deteksi_kolom_pn_nama_qty(katalog):
    data = _xlsx([
        ["DAFTAR PERMINTAAN PART"],                 # baris judul, harus dilewati
        ["No", "Kode", "Deskripsi", "Jumlah"],      # header sebenarnya
        [1, "WG9925520270", "Spring bracket", 4],
        [2, "AZ9925520271", "Leaf spring", 2],
    ])
    p = ai_sheet.parse_upload(data, "permintaan.xlsx")
    assert p["ok"]
    assert p["jumlah_baris"] == 2
    # 'Kode' bukan kata kunci PN, tapi ISI-nya cocok katalog → tetap terdeteksi.
    assert p["kolom_pn"] == "Kode"
    assert p["pn_dikenal"] == 2
    roles = dict(zip(p["headers"], p["roles"]))
    assert roles["Deskripsi"] == "part_name"
    assert roles["Jumlah"] == "qty"


# ── P5: PN Weichai murni-angka ───────────────────────────────────────────────

def test_deteksi_pn_weichai_murni_angka(monkeypatch):
    """Kolom PN Weichai murni-angka (612630010054) TERDETEKSI bila cocok katalog;
    kolom qty & harga murni-angka TIDAK salah jadi PN."""
    kat = [{"part_number": "612630010054", "part_name": "Oil cooler"},
           {"part_number": "612600115000", "part_name": "Filter"}]
    monkeypatch.setattr(ai_sheet.part_index, "search_exact_pns",
                        lambda pns: [r for r in kat if r["part_number"] in {str(p) for p in pns}])
    monkeypatch.setattr(ai_sheet.part_index, "_pn_flat_map", lambda: {})
    data = _xlsx([
        ["Kode", "Nama", "Qty", "Harga"],
        [612630010054, "Oil cooler", 2, 1500000],
        [612600115000, "Filter", 5, 250000],
    ])
    p = ai_sheet.parse_upload(data, "weichai.xlsx")
    assert p["ok"]
    assert p["kolom_pn"] == "Kode"                 # kolom numerik terverifikasi → PN
    assert p["pn_dikenal"] == 2
    roles = dict(zip(p["headers"], p["roles"]))
    assert roles["Qty"] != "part_number"           # qty tak salah jadi PN
    assert roles["Harga"] != "part_number"


def test_kolom_murni_angka_tanpa_bukti_katalog_bukan_pn(monkeypatch):
    """Guard: kolom murni-angka TANPA bukti katalog (mis. no. telp) TAK dipaksa PN."""
    monkeypatch.setattr(ai_sheet.part_index, "search_exact_pns", lambda pns: [])
    monkeypatch.setattr(ai_sheet.part_index, "_pn_flat_map", lambda: {})
    data = _xlsx([
        ["Nama", "No HP"],
        ["Budi", 81234567890],
        ["Andi", 81298765432],
    ])
    p = ai_sheet.parse_upload(data, "kontak.xlsx")
    assert p["ok"]
    assert p["kolom_pn"] is None                    # angka tak terverifikasi ≠ PN


def test_dua_kolom_pn_kolom_kedua_dicatat(monkeypatch):
    """Dua kolom PN sama-sama cocok katalog → kolom kedua dicatat, tak jadi 'lain'."""
    kat = [{"part_number": p, "part_name": "x"} for p in
           ("WG9925520270", "AZ9925520271", "VG1560080012", "WG9725580006")]
    monkeypatch.setattr(ai_sheet.part_index, "search_exact_pns",
                        lambda pns: [r for r in kat if r["part_number"] in {str(p).upper() for p in pns}])
    monkeypatch.setattr(ai_sheet.part_index, "_pn_flat_map", lambda: {})
    data = _xlsx([
        ["PN Lama", "PN Baru", "Nama"],
        ["WG9925520270", "VG1560080012", "A"],
        ["AZ9925520271", "WG9725580006", "B"],
    ])
    p = ai_sheet.parse_upload(data, "supersession.xlsx")
    assert p["ok"]
    assert p["kolom_pn"] in ("PN Lama", "PN Baru")
    assert p["kolom_pn_lain"] in ("PN Lama", "PN Baru")
    assert p["kolom_pn"] != p["kolom_pn_lain"]
    r = ai_sheet.ringkas(p)
    assert r["kolom_part_number_lain"] == p["kolom_pn_lain"]


# ── P9.4/P9.5: poles _cari_kolom & truncation kolom ─────────────────────────

def test_cari_kolom_utamakan_exact_lalu_terpendek():
    # exact menang walau ada header lain yang MEMUAT kata & posisinya duluan
    assert ai_sheet._cari_kolom(["Harga SIMS (IDR)", "Harga", "Nama"], "Harga") == 1
    # tanpa exact: prefix + TERPENDEK menang (bukan yang kebetulan duluan)
    assert ai_sheet._cari_kolom(["Harga SIMS (IDR)", "Harga Jual"], "harga") == 1


def test_kolom_terpotong_ditandai():
    """Kolom melewati MAX_COLS dibuang → DITANDAI (tak senyap)."""
    header = [f"C{i}" for i in range(ai_sheet.MAX_COLS + 5)]
    row = list(range(ai_sheet.MAX_COLS + 5))
    p = ai_sheet.parse_upload(_xlsx([header, row]), "lebar.xlsx")
    assert p["ok"] and p["kolom_terpotong"] is True
    assert ai_sheet.ringkas(p)["kolom_terpotong"] is True


def test_kolom_tidak_terpotong_bila_dalam_batas(katalog):
    p = ai_sheet.parse_upload(_xlsx([["Part Number", "Qty"], ["WG9925520270", 1]]), "kecil.xlsx")
    assert p["kolom_terpotong"] is False


def test_tolak_format_dan_file_rusak():
    assert ai_sheet.parse_upload(b"", "a.xlsx")["error"] == "File kosong."
    assert "Format" in ai_sheet.parse_upload(b"x", "a.xls")["error"]   # .xls tetap ditolak
    assert "valid" in ai_sheet.parse_upload(b"bukan-excel", "a.xlsx")["error"]


def test_tolak_file_kelewat_besar():
    r = ai_sheet.parse_upload(b"x" * (ai_sheet.MAX_BYTES + 1), "a.xlsx")
    assert not r["ok"] and "besar" in r["error"]


# ── Stash discoped per-user ──────────────────────────────────────────────────

def test_sheet_id_user_lain_tidak_terbaca(katalog):
    p = ai_sheet.parse_upload(_xlsx([["PN"], ["WG9925520270"]]), "a.xlsx")
    sid = ai_sheet.put_sheet("budi", p)
    assert ai_sheet.get_sheet(sid, "budi") is not None
    assert ai_sheet.get_sheet(sid, "orang-lain") is None       # milik user lain
    assert ai_sheet.get_sheet("ngawur", "budi") is None


# ── Multi-sheet ──────────────────────────────────────────────────────────────

def test_parse_ringkas_sheet_lain(katalog):
    """Sheet pertama di-parse penuh; tab lain diringkas (nama+header+baris)."""
    data = _xlsx_multi({
        "Order": [["Part Number", "Qty"], ["WG9925520270", 4]],
        "Retur": [["Kode", "Alasan"], ["AZ9925520271", "rusak"], ["X", "salah"]],
    })
    p = ai_sheet.parse_upload(data, "wb.xlsx")
    assert p["ok"] and p["sheet"] == "Order"
    assert p["sheet_lain"] == ["Retur"]
    out = ai_sheet.ringkas(p)
    detail = {d["nama"]: d for d in out["sheet_lain_detail"]}
    assert "Retur" in detail
    assert detail["Retur"]["header"][:2] == ["Kode", "Alasan"]
    assert detail["Retur"]["jumlah_baris"] == 2
    assert "catatan_sheet" in out


def test_pilih_sheet_ganti_aktif_dan_isi_di_tab_itu(gudang_stok):
    data = _xlsx_multi({
        "Kosong": [["Catatan"], ["abaikan"]],
        "Barang": [["Part Number"], ["WG9925520270"], ["AZ9925520271"]],
    })
    p = ai_sheet.parse_upload(data, "wb.xlsx")
    sid = ai_sheet.put_sheet(USER["username"], p)
    assert p["sheet"] == "Kosong"                   # default tab pertama

    r = ai_sheet.select_sheet(sid, USER, "Barang")
    assert r["found"] and r["sheet"] == "Barang"
    # sheet_id SAMA kini menunjuk tab 'Barang' → fill jalan di situ.
    f = ai_sheet.fill_column(sid, USER, isi="stok")
    assert f["found"] and f["baris_terisi"] == 2


def test_pilih_sheet_nama_tak_ada(gudang_stok):
    data = _xlsx_multi({"A": [["Part Number"], ["WG9925520270"]],
                        "B": [["Part Number"], ["AZ9925520271"]]})
    p = ai_sheet.parse_upload(data, "wb.xlsx")
    sid = ai_sheet.put_sheet(USER["username"], p)
    r = ai_sheet.select_sheet(sid, USER, "Zzz")
    assert r["found"] is False and "tak ada" in r["error"].lower()


def test_pilih_sheet_milik_user_lain_ditolak(gudang_stok):
    data = _xlsx_multi({"A": [["Part Number"], ["WG9925520270"]],
                        "B": [["Part Number"], ["AZ9925520271"]]})
    p = ai_sheet.parse_upload(data, "wb.xlsx")
    sid = ai_sheet.put_sheet("budi", p)
    r = ai_sheet.select_sheet(sid, {"username": "orang-lain"}, "B")
    assert r["found"] is False


def test_tool_pilih_sheet_hanya_ada_bila_ada_lampiran():
    tanpa = {s["function"]["name"] for s in ai._tool_specs(ADMIN)}
    dengan = {s["function"]["name"] for s in ai._tool_specs(ADMIN, sheet_id="x")}
    assert "sheet_pilih_sheet" not in tanpa
    assert "sheet_pilih_sheet" in dengan
    assert ai._DISPATCH["sheet_pilih_sheet"] is ai._t_sheet_pilih_sheet


# ── Isi kolom ────────────────────────────────────────────────────────────────

def _sheet_untuk(user, katalog):
    p = ai_sheet.parse_upload(_xlsx([
        ["Part Number", "Qty"],
        ["WG9925520270", 4],
        ["ZZZ0000000", 1],        # PN tak dikenal → dibiarkan kosong
    ]), "order.xlsx")
    return ai_sheet.put_sheet(user["username"], p)


def test_isi_kolom_stok_tambah_kolom_baru(katalog):
    sid = _sheet_untuk(USER, katalog)
    r = ai_sheet.fill_column(sid, USER, isi="stok")
    assert r["found"]
    assert r["kolom_diisi"] == "Stok" and r["kolom_part_number"] == "Part Number"
    # PN dikenal terisi, PN asing dibiarkan kosong (tidak dikarang).
    assert r["baris_terisi"] == 1 and r["baris_kosong"] == 1
    assert r["export_id"]


def test_isi_kolom_pemaaf_suffix_varian(katalog):
    """PN sheet ber-suffix varian ('WG9925520270/2') harus tetap terisi dari baris
    PN dasar katalog ('WG9925520270') — pencocokan lewat rows_for_pns yang pemaaf.
    Sebaliknya juga (sheet base, katalog di tes ini base) — cukup uji arah suffix."""
    p = ai_sheet.parse_upload(_xlsx([
        ["Part Number", "Qty"],
        ["WG9925520270/2", 4],       # suffix varian — dulu meleset (kosong)
    ]), "order.xlsx")
    sid = ai_sheet.put_sheet(USER["username"], p)
    r = ai_sheet.fill_column(sid, USER, isi="stok")
    assert r["found"] and r["baris_terisi"] == 1     # terisi lewat PN dasar
    # parse_upload juga menghitung 'dikenal' dgn pemaaf.
    assert p["pn_dikenal"] == 1


# ── P6: daftar PN gagal-cocok ────────────────────────────────────────────────

def test_pn_tidak_ditemukan_didaftar(katalog):
    """PN yang tak ketemu DIDAFTAR (bukan cuma dihitung) di fill_column & ringkas."""
    p = ai_sheet.parse_upload(_xlsx([
        ["Part Number", "Qty"],
        ["WG9925520270", 1],       # dikenal
        ["ZZZ111AA", 2],           # tak dikenal
        ["ZZZ222BB", 3],           # tak dikenal
    ]), "order.xlsx")
    sid = ai_sheet.put_sheet(USER["username"], p)
    r = ai_sheet.fill_column(sid, USER, isi="stok")
    assert r["found"]
    assert set(r["pn_tidak_ditemukan"]) == {"ZZZ111AA", "ZZZ222BB"}
    assert r["pn_tidak_ditemukan_total"] == 2
    # sheet_ringkasan juga mendaftarkan contoh PN tak dikenal.
    rr = ai_sheet.ringkas(p)
    assert "ZZZ111AA" in rr["part_number_tidak_dikenal_contoh"]


def test_pn_tidak_ditemukan_tercap_20(katalog):
    baris = [["Part Number"]] + [[f"NOPE{i:04d}X"] for i in range(30)]
    p = ai_sheet.parse_upload(_xlsx(baris), "besar.xlsx")
    sid = ai_sheet.put_sheet(USER["username"], p)
    r = ai_sheet.fill_column(sid, USER, isi="stok")
    assert len(r["pn_tidak_ditemukan"]) == 20        # ter-cap
    assert r["pn_tidak_ditemukan_total"] == 30       # total jujur


def test_pn_tidak_ditemukan_absen_bila_semua_cocok(katalog):
    p = ai_sheet.parse_upload(_xlsx([
        ["Part Number", "Qty"],
        ["WG9925520270", 1],
        ["AZ9925520271", 2],
    ]), "order.xlsx")
    sid = ai_sheet.put_sheet(USER["username"], p)
    r = ai_sheet.fill_column(sid, USER, isi="stok")
    assert r["pn_tidak_ditemukan"] == []
    assert r["pn_tidak_ditemukan_total"] == 0


def test_isi_kolom_menimpa_kolom_yang_disebut_huruf(katalog):
    sid = _sheet_untuk(USER, katalog)
    r = ai_sheet.fill_column(sid, USER, isi="nama_part", kolom_tujuan="B")  # kolom Qty
    assert r["found"] and r["kolom_diisi"] == "Qty"


def test_isi_kolom_sheet_kedaluwarsa():
    r = ai_sheet.fill_column("tidak-ada", USER, isi="stok")
    assert r["found"] is False and "unggah" in r["error"].lower()


# ── HARGA SIMS: hanya admin ──────────────────────────────────────────────────

def test_harga_sims_ditolak_untuk_user_biasa(katalog):
    sid = _sheet_untuk(USER, katalog)
    r = ai_sheet.fill_column(sid, USER, isi="harga_sims", can_sims=False)
    assert r["denied"] is True and "admin" in r["error"]


def test_harga_sims_ditolak_untuk_pembeli(katalog):
    sid = _sheet_untuk(PEMBELI, katalog)
    r = ai_sheet.fill_column(sid, PEMBELI, isi="harga_sims", can_sims=False)
    assert r["denied"] is True


def test_harga_sims_jalan_untuk_admin(monkeypatch, katalog):
    monkeypatch.setattr(ai_sheet.harga, "batch_harga", lambda pns, **k: {
        "rate": 2200.0, "count": len(pns), "found": 1,
        "results": [{"pn": "WG9925520270", "cny": 700, "idr": 1540000, "status": "ok"},
                    {"pn": "ZZZ0000000", "cny": None, "idr": None, "status": "not_found"}],
    })
    sid = _sheet_untuk(ADMIN, katalog)
    r = ai_sheet.fill_column(sid, ADMIN, isi="harga_sims", can_sims=True)
    assert r["found"] and r["baris_terisi"] == 1 and "SIMS" in r["sumber"]


# ── HARGA SIMS = harga MODAL ber-CNY; jangan dikonversi kecuali diminta ──────

@pytest.fixture
def sims_palsu(monkeypatch):
    """1 CNY = Rp 2.200 → 700 CNY = Rp 1.540.000. Dua angka yang jelas beda."""
    monkeypatch.setattr(ai_sheet.harga, "batch_harga", lambda pns, **k: {
        "rate": 2200.0, "count": len(pns), "found": 1,
        "results": [{"pn": "WG9925520270", "cny": 700, "idr": 1540000, "status": "ok"},
                    {"pn": "ZZZ0000000", "cny": None, "idr": None, "status": "not_found"}],
    })


def _sel_kolom(export_id: str, nama_kolom: str):
    """Nilai terisi PERTAMA di bawah header `nama_kolom` pada Excel hasil."""
    from openpyxl import load_workbook

    from app.services import ai_export
    data, _fn = ai_export.generic_excel(export_id)
    rows = list(load_workbook(io.BytesIO(data)).active.iter_rows(values_only=True))
    for i, row in enumerate(rows):
        teks = [str(c) if c is not None else "" for c in (row or ())]
        if nama_kolom in teks:
            j = teks.index(nama_kolom)
            for bawah in rows[i + 1:]:
                v = bawah[j] if bawah and j < len(bawah) else None
                if v not in (None, ""):
                    return v
            return None
    raise AssertionError(f"kolom '{nama_kolom}' tak ada di Excel hasil")


def test_harga_sims_default_cny_tanpa_konversi(sims_palsu, katalog):
    """Harga SIMS = harga MODAL, mata uang aslinya CNY. Default TIDAK dikonversi:
    kalau diam-diam jadi rupiah, kolom modal tampak seperti harga jual dan ikut
    bergoyang mengikuti kurs harian."""
    sid = _sheet_untuk(ADMIN, katalog)
    r = ai_sheet.fill_column(sid, ADMIN, isi="harga_sims", can_sims=True)
    assert r["kolom_diisi"] == "Harga SIMS (CNY)"
    assert r["mata_uang_harga_sims"] == "CNY"
    assert "tanpa konversi" in r["sumber"]
    assert _sel_kolom(r["export_id"], "Harga SIMS (CNY)") == 700     # ⛔ bukan 1_540_000


def test_harga_sims_dikonversi_bila_user_minta(sims_palsu, katalog):
    sid = _sheet_untuk(ADMIN, katalog)
    r = ai_sheet.fill_column(sid, ADMIN, isi="harga_sims", can_sims=True, konversi_idr=True)
    assert r["kolom_diisi"] == "Harga SIMS (IDR)"
    assert r["mata_uang_harga_sims"] == "IDR"
    assert _sel_kolom(r["export_id"], "Harga SIMS (IDR)") == 1_540_000


def test_harga_sims_banyak_kolom_juga_cny_secara_default(sims_palsu, katalog):
    sid = _sheet_untuk(ADMIN, katalog)
    r = ai_sheet.fill_columns(sid, ADMIN, [{"isi": "harga_sims"}], can_sims=True)
    assert r["kolom"][0]["kolom"] == "Harga SIMS (CNY)"
    assert r["mata_uang_harga_sims"] == "CNY"
    assert _sel_kolom(r["export_id"], "Harga SIMS (CNY)") == 700


def test_harga_sims_banyak_kolom_konversi_bila_diminta(sims_palsu, katalog):
    sid = _sheet_untuk(ADMIN, katalog)
    r = ai_sheet.fill_columns(sid, ADMIN, [{"isi": "harga_sims"}], can_sims=True,
                              konversi_idr=True)
    assert r["kolom"][0]["kolom"] == "Harga SIMS (IDR)"
    assert _sel_kolom(r["export_id"], "Harga SIMS (IDR)") == 1_540_000


def test_tool_harga_sims_tak_kirim_rupiah_kecuali_diminta(monkeypatch):
    """Tool chat: bila IDR selalu ikut dikirim, model hampir selalu menyajikan
    yang rupiah — jadi nilai IDR hanya disertakan saat user memintanya."""
    monkeypatch.setattr(ai.harga, "cari_harga",
                        lambda pn, **k: {"pn": pn, "cny": 700, "idr": 1540000,
                                         "rate": 2200.0, "note": ""})
    polos = ai._t_harga_sims({"part_number": "WG9925520270"}, ADMIN)
    assert polos["harga_cny"] == 700 and polos["mata_uang"] == "CNY"
    assert "harga_idr" not in polos and "kurs_cny_idr" not in polos
    assert "jangan dikonversi" in polos["catatan"].lower()

    diminta = ai._t_harga_sims({"part_number": "WG9925520270", "konversi_idr": True}, ADMIN)
    assert diminta["harga_idr"] == 1540000 and diminta["kurs_cny_idr"] == 2200.0


def test_harga_sims_batasi_jumlah_pn(monkeypatch, katalog):
    baris = [["Part Number"]] + [[f"WG99255{i:05d}"] for i in range(ai_sheet._MAX_SIMS + 5)]
    p = ai_sheet.parse_upload(_xlsx(baris), "besar.xlsx")
    sid = ai_sheet.put_sheet("admin", p)
    r = ai_sheet.fill_column(sid, ADMIN, isi="harga_sims", can_sims=True)
    assert r["found"] is False and "Maksimum" in r["error"]


# ── Scoping di lapisan tool (allow-list) ─────────────────────────────────────

def test_tool_sheet_hanya_ada_bila_ada_lampiran():
    assert "sheet_ringkasan" not in ai._allowed_tool_names(USER, "")
    assert "sheet_isi_kolom" not in ai._allowed_tool_names(ADMIN, "")
    assert "sheet_ringkasan" in ai._allowed_tool_names(USER, "sid-apapun")


def test_pilihan_harga_sims_hanya_ditawarkan_ke_admin():
    def _enum(user):
        for f in ai._tool_specs(user, "sid"):
            if f["function"]["name"] == "sheet_isi_kolom":
                props = f["function"]["parameters"]["properties"]
                return props["kolom"]["items"]["properties"]["isi"]["enum"]
        return []

    assert "harga_sims" in _enum(ADMIN)
    assert "harga_sims" not in _enum(USER)
    assert "harga_sims" not in _enum(PEMBELI)


def test_harga_sims_tool_tak_pernah_untuk_non_admin():
    """Aturan pemilik: SEMUA akses harga SIMS di asisten hanya admin."""
    for u in (USER, PEMBELI):
        assert "harga_sims" not in ai._allowed_tool_names(u, "sid")
    assert "harga_sims" in ai._allowed_tool_names(ADMIN, "sid")


def test_run_tool_menolak_sheet_tool_tanpa_lampiran():
    r = ai._run_tool("sheet_isi_kolom", {"isi": "stok"}, ADMIN, sheet_id="")
    assert r["denied"] is True


def test_chat_lampiran_excel_sampai_kartu_unduh(monkeypatch, katalog):
    """Jalur PENUH: sheet_id → tool ditawarkan → model panggil sheet_isi_kolom →
    kartu unduh Excel muncul di metadata. DeepSeek di-mock (tanpa network)."""
    monkeypatch.setattr(ai, "_system_prompt", lambda user: "system uji")
    monkeypatch.setattr(ai, "_unit_name_tokens", lambda: set())
    sid = _sheet_untuk(USER, katalog)

    dilihat_tools: list[list[str]] = []
    seq = [
        {"choices": [{"message": {"content": "", "tool_calls": [
            {"id": "t1", "function": {"name": "sheet_isi_kolom",
                                      "arguments": '{"isi":"stok","kolom_tujuan":"Stok"}'}},
        ]}, "finish_reason": "tool_calls"}]},
        {"choices": [{"message": {"content": "Sudah saya isikan stoknya."}, "finish_reason": "stop"}]},
    ]
    n = {"i": 0}

    def fake_post(messages, tools, max_tokens=6000):
        dilihat_tools.append([t["function"]["name"] for t in (tools or [])])
        c = seq[min(n["i"], len(seq) - 1)]
        n["i"] += 1
        return c

    monkeypatch.setattr(ai, "_post_chat", fake_post)
    out = ai.chat(USER, [{"role": "user", "content": "isikan stoknya di kolom Stok"}], sheet_id=sid)

    assert "sheet_isi_kolom" in dilihat_tools[0]      # tool ditawarkan krn ada lampiran
    assert "sheet_isi_kolom" in out["tools_used"]
    assert out["excel_exports"] and out["excel_exports"][0]["id"]   # kartu unduh muncul
    assert "stok" in out["reply"].lower()


def test_chat_sheet_id_milik_orang_lain_diabaikan(monkeypatch, katalog):
    """sheet_id curian → lampiran dianggap tidak ada, tool sheet_* tak ditawarkan."""
    monkeypatch.setattr(ai, "_system_prompt", lambda user: "system uji")
    monkeypatch.setattr(ai, "_unit_name_tokens", lambda: set())
    sid = _sheet_untuk(USER, katalog)          # milik 'budi'
    dilihat: list[list[str]] = []

    def fake_post(messages, tools, max_tokens=6000):
        dilihat.append([t["function"]["name"] for t in (tools or [])])
        return {"choices": [{"message": {"content": "Tidak ada lampiran."}, "finish_reason": "stop"}]}

    monkeypatch.setattr(ai, "_post_chat", fake_post)
    penyusup = {"username": "penyusup", "role": "admin"}
    ai.chat(penyusup, [{"role": "user", "content": "isi stok"}], sheet_id=sid)
    assert not any(t.startswith("sheet_") for t in dilihat[0])


def test_model_tak_bisa_memilih_sheet_id_lewat_argumen(monkeypatch, katalog):
    """_sheet_id selalu ditimpa server; argumen dari model diabaikan."""
    sid = _sheet_untuk(USER, katalog)
    dilihat = {}

    def _spy(args, user):
        dilihat.update(args)
        return {"found": True}

    monkeypatch.setitem(ai._DISPATCH, "sheet_ringkasan", _spy)
    ai._run_tool("sheet_ringkasan", {"_sheet_id": "sheet-orang-lain"}, USER, sheet_id=sid)
    assert dilihat["_sheet_id"] == sid


# ── Isi BANYAK kolom / stok MULTI-GUDANG ke SATU file (dinamis) ───────────────

@pytest.fixture
def gudang_stok(monkeypatch):
    rows = {
        "WG9925520270": {"part_number": "WG9925520270", "part_name": "Spring bracket",
                         "stok": "12", "harga": "Rp 1.500.000",
                         "gudang": {"JAKARTA": 8, "PEKANBARU": 4}},
        "AZ9925520271": {"part_number": "AZ9925520271", "part_name": "Leaf spring",
                         "stok": "3", "harga": "—", "gudang": {"JAKARTA": 3}},
    }
    monkeypatch.setattr(ai_sheet.part_index, "search_exact_pns",
                        lambda pns: [rows[p] for p in {x.upper() for x in pns} if p in rows])
    monkeypatch.setattr(ai_sheet.part_index, "gudang_names",
                        lambda: ["JAKARTA", "PEKANBARU", "MEDAN"])
    monkeypatch.setattr(ai_sheet.part_index, "_pn_flat_map", lambda: {})
    return rows


def _sheet_pn(user) -> str:
    p = ai_sheet.parse_upload(_xlsx([
        ["Part Number"], ["WG9925520270"], ["AZ9925520271"], ["ZZZ0000000"],
    ]), "order.xlsx")
    return ai_sheet.put_sheet(user["username"], p)


def test_isi_multi_gudang_dan_harga_satu_file(gudang_stok):
    sid = _sheet_pn(USER)
    r = ai_sheet.fill_columns(sid, USER, [
        {"isi": "stok", "gudang": "jakarta"},
        {"isi": "stok", "gudang": "Pekanbaru"},
        {"isi": "harga_lokal"},
    ])
    assert r["found"] and r["export_id"]                 # SATU file
    per = {c["kolom"]: c["baris_terisi"] for c in r["kolom"]}
    assert list(per) == ["Stok JAKARTA", "Stok PEKANBARU", "Harga"]
    assert per["Stok JAKARTA"] == 2                       # WG=8, AZ=3
    assert per["Stok PEKANBARU"] == 2                     # WG=4, AZ='0' (terlacak)
    assert per["Harga"] == 1                              # AZ '—' → kosong


def test_fill_columns_daftar_pn_tidak_ditemukan(gudang_stok):
    """P6 (jalur produksi fill_columns): PN tak ketemu didaftar, bukan cuma dihitung."""
    sid = _sheet_pn(USER)   # berisi ZZZ0000000 (di luar katalog)
    r = ai_sheet.fill_columns(sid, USER, [{"isi": "stok", "gudang": "Jakarta"}])
    assert r["found"]
    assert "ZZZ0000000" in r["pn_tidak_ditemukan"]
    assert r["pn_tidak_ditemukan_total"] == 1


def test_isi_gudang_campur_dikenal_dan_tidak(gudang_stok):
    sid = _sheet_pn(USER)
    r = ai_sheet.fill_columns(sid, USER, [
        {"isi": "stok", "gudang": "Jakarta"},
        {"isi": "stok", "gudang": "Surabaya"},   # tak ada
    ])
    assert r["found"] and len(r["kolom"]) == 1 and r["gudang_tak_dikenal"] == ["Surabaya"]


def test_isi_gudang_semua_tak_dikenal(gudang_stok):
    sid = _sheet_pn(USER)
    r = ai_sheet.fill_columns(sid, USER, [{"isi": "stok", "gudang": "Surabaya"}])
    assert r["found"] is False and "Surabaya" in r["error"]


def test_fill_columns_harga_sims_admin_only(gudang_stok):
    sid = _sheet_pn(USER)
    r = ai_sheet.fill_columns(sid, USER, [{"isi": "harga_sims"}], can_sims=False)
    assert r["denied"] is True and "admin" in r["error"]


def test_tool_isi_kolom_multi_lewat_run_tool(gudang_stok):
    sid = _sheet_pn(USER)
    r = ai._run_tool("sheet_isi_kolom", {"kolom": [
        {"isi": "stok", "gudang": "Jakarta"},
        {"isi": "stok", "gudang": "Pekanbaru"},
    ]}, USER, sheet_id=sid)
    assert r["found"] and len(r["kolom"]) == 2       # dua gudang, SATU file


def test_tool_isi_kolom_backcompat_single(gudang_stok):
    """Model lama kirim isi tunggal (bukan 'kolom') → tetap jalan (1 kolom)."""
    sid = _sheet_pn(USER)
    r = ai._run_tool("sheet_isi_kolom", {"isi": "stok"}, USER, sheet_id=sid)
    assert r["found"] and len(r["kolom"]) == 1 and r["kolom"][0]["isi"] == "stok"


# ── Isi PART NUMBER dari nama part (kebalikan sheet_isi_kolom), lingkup 1 unit ──

# BOM unit palsu (per nomor rangka) + nama lokal katalog.
_BOM_PARTS = [
    {"pn": "WG9925520270", "nama_cn": "钢板弹簧支架", "qty": 4},
    {"pn": "AZ9925520271", "nama_cn": "钢板弹簧", "qty": 2},
    {"pn": "VG1560080126", "nama_cn": "机油滤清器", "qty": 1},
]
_BOM_NAMA_LOKAL = {
    "WG9925520270": "Spring bracket",
    "AZ9925520271": "Leaf spring",
    "VG1560080126": "Oil filter",
}


@pytest.fixture
def bom_unit(monkeypatch):
    """Kunci EPC & katalog lokal untuk lingkup BOM satu unit (RJ-TEST)."""
    monkeypatch.setattr(ai.epc_bom, "loading_list",
                        lambda rangka: {"found": True, "frame_number": rangka, "parts": _BOM_PARTS})
    monkeypatch.setattr(ai.epc_weichai, "engine_bom", lambda rangka: {"found": False})
    monkeypatch.setattr(ai.part_index, "search_exact_pns", lambda pns: [
        {"part_number": pn, "part_name": _BOM_NAMA_LOKAL[pn]}
        for pn in {p.upper() for p in pns} if pn in _BOM_NAMA_LOKAL])


def _sheet_nama(user, rows) -> str:
    p = ai_sheet.parse_upload(_xlsx([["Nama Part"]] + [[r] for r in rows]), "daftar.xlsx")
    return ai_sheet.put_sheet(user["username"], p)


def test_isi_pn_cocok_persis_dan_kosongkan_asing(bom_unit):
    sid = _sheet_nama(USER, ["Spring bracket", "Oil filter", "Barang asing xyz"])
    r = ai._t_sheet_isi_part_number({"_sheet_id": sid, "rangka": "RJ-TEST"}, USER)
    assert r["found"]
    assert r["kolom_nama"] == "Nama Part" and r["kolom_diisi"] == "Part Number (EPC)"
    assert r["baris_terisi"] == 2 and r["baris_kosong"] == 1
    body = ai_sheet.get_sheet(sid, USER["username"])["_body"]
    # (catatan: fill_column bekerja pada salinan; verifikasi lewat hasil ringkas)
    assert r["frame_number"] == "RJ-TEST" and r["export_id"]


def test_isi_pn_hanya_sel_kosong_jaga_yang_terisi(bom_unit):
    """File sudah punya kolom Part Number sebagian terisi → hanya baris KOSONG
    yang diisi; PN yang sudah ada TIDAK ditimpa (walau namanya tak cocok BOM)."""
    p = ai_sheet.parse_upload(_xlsx([
        ["Nama Part", "Part Number"],
        ["Spring bracket", ""],            # kosong → diisi WG9925520270
        ["Oil filter", ""],               # kosong → diisi VG1560080126
        ["Barang asing xyz", "PN-MANUAL"],  # sudah terisi & tak cocok → JANGAN diubah
    ]), "campur.xlsx")
    sid = ai_sheet.put_sheet(USER["username"], p)
    r = ai._t_sheet_isi_part_number({"_sheet_id": sid, "rangka": "RJ-TEST"}, USER)
    assert r["kolom_diisi"] == "Part Number"   # pakai kolom yang ADA, bukan bikin baru
    assert r["baris_terisi"] == 2 and r["baris_sudah_terisi"] == 1
    assert r["baris_kosong"] == 0


def test_isi_pn_nama_ambigu_dikosongkan(bom_unit):
    # 'spring' cocok ke DUA PN (Spring bracket & Leaf spring) → ambigu → kosong.
    # 'leaf spring' hanya cocok satu → terisi.
    sid = _sheet_nama(USER, ["spring", "leaf spring"])
    r = ai._t_sheet_isi_part_number({"_sheet_id": sid, "rangka": "RJ-TEST"}, USER)
    assert r["baris_terisi"] == 1 and r["baris_ambigu"] == 1


def test_kategori_pecah_ambigu(monkeypatch, bom_unit):
    """Nama sama ('Clamp') di dua sistem → kolom kategori memilih PN yang benar."""
    # Tambah dua 'Clamp' ke BOM: satu di intake, satu di cooling (via nama EN).
    monkeypatch.setattr(ai.epc_bom, "loading_list", lambda rangka: {
        "found": True, "frame_number": rangka, "parts": [
            {"pn": "P-INTAKE-CLAMP", "nama_cn": "", "qty": 1},
            {"pn": "P-COOLING-CLAMP", "nama_cn": "", "qty": 1},
        ]})
    monkeypatch.setattr(ai.part_index, "search_exact_pns", lambda pns: [
        {"part_number": "P-INTAKE-CLAMP", "part_name": "Air intake clamp"},
        {"part_number": "P-COOLING-CLAMP", "part_name": "Cooling water clamp"},
    ])
    p = ai_sheet.parse_upload(_xlsx([
        ["Bagian", "Nama Part", "Part Number"],
        ["AIR INTAKE", "Clamp", ""],       # → P-INTAKE-CLAMP (via konteks)
        ["COOLING", "Clamp", ""],          # → P-COOLING-CLAMP
    ]), "sistem.xlsx")
    assert "kategori" in p["roles"]        # kolom 'Bagian' dikenali sbg kategori
    sid = ai_sheet.put_sheet(USER["username"], p)
    r = ai._t_sheet_isi_part_number({"_sheet_id": sid, "rangka": "RJ-TEST"}, USER)
    assert r["baris_terisi"] == 2 and r["baris_ambigu"] == 0


def test_ringkas_sinyal_konteks(bom_unit):
    """ringkas() melaporkan baris tanpa PN + kolom pengelompokan."""
    p = ai_sheet.parse_upload(_xlsx([
        ["Bagian", "Nama Part", "Part Number"],
        ["AIR INTAKE", "Oil filter", ""],
        ["AIR INTAKE", "Spring bracket", "WG9925520270"],
    ]), "x.xlsx")
    out = ai_sheet.ringkas(p)
    assert out["baris_tanpa_part_number"] == 1
    assert out["kolom_pengelompokan"]["kolom"] == "Bagian"
    assert "AIR INTAKE" in out["kolom_pengelompokan"]["contoh_nilai"]


def test_ringkas_fill_rate_dan_contoh_nilai(katalog):
    """Tiap kolom bawa 'terisi' (fill rate); kolom non-PN kardinalitas rendah
    bawa 'contoh_nilai'; kolom PN tidak (kardinalitas tinggi & untuk anti-bocor)."""
    p = ai_sheet.parse_upload(_xlsx([
        ["Part Number", "Satuan", "Qty"],
        ["WG9925520270", "PCS", 4],
        ["AZ9925520271", "SET", ""],          # Qty kosong 1 baris
        ["WG9925520270/2", "PCS", 2],         # suffix varian → dikenal via base
    ]), "x.xlsx")
    out = ai_sheet.ringkas(p)
    per = {k["nama"]: k for k in out["kolom"]}
    assert per["Qty"]["terisi"] == 2                       # 1 sel kosong
    assert set(per["Satuan"]["contoh_nilai"]) == {"PCS", "SET"}
    assert "contoh_nilai" not in per["Part Number"]        # kolom PN tak diekspos
    # 3 PN, semua dikenal (WG…/2 lewat base) → tak ada yang asing.
    assert out["part_number_tidak_dikenal"] == 0


def test_ringkas_pn_tidak_dikenal_dihitung(katalog):
    p = ai_sheet.parse_upload(_xlsx([
        ["Part Number"], ["WG9925520270"], ["ZZZ0000000"], ["QQQ1111111"],
    ]), "x.xlsx")
    out = ai_sheet.ringkas(p)
    assert out["part_number_dikenal_di_katalog"] == 1
    assert out["part_number_tidak_dikenal"] == 2


def test_isi_pn_lewat_sinonim(monkeypatch, bom_unit):
    monkeypatch.setattr(ai, "_expand_query",
                        lambda q: ([q] + {"saringan": ["filter"], "oli": ["oil"]}.get(q.lower(), []), []))
    sid = _sheet_nama(USER, ["saringan oli"])
    r = ai._t_sheet_isi_part_number({"_sheet_id": sid, "rangka": "RJ-TEST"}, USER)
    assert r["baris_terisi"] == 1     # 'saringan oli' → Oil filter (VG1560080126)


def test_isi_pn_frasa_sinonim_filter_solar(monkeypatch, bom_unit):
    """Istilah lapangan multi-kata 'filter solar' → nama BOM Inggris 'Fuel filter'
    (sinonim tingkat-frasa; kata 'solar' sendiri tak punya sinonim)."""
    monkeypatch.setattr(ai.epc_bom, "loading_list", lambda rangka: {
        "found": True, "frame_number": rangka,
        "parts": [{"pn": "612600080933", "nama_cn": "", "qty": 1}]})
    monkeypatch.setattr(ai.part_index, "search_exact_pns", lambda pns: [
        {"part_number": "612600080933", "part_name": "Fuel filter"}])
    monkeypatch.setattr(ai, "_expand_query",
                        lambda q: ([q, "fuel filter", "diesel filter"], ["filter solar"])
                        if q.lower() == "filter solar" else ([q], []))
    sid = _sheet_nama(USER, ["filter solar"])
    r = ai._t_sheet_isi_part_number({"_sheet_id": sid, "rangka": "RJ-TEST"}, USER)
    assert r["baris_terisi"] == 1


def test_isi_pn_gabung_part_mesin_weichai(monkeypatch, bom_unit):
    """Part internal mesin (tak ada di Loading List) ikut dari engine_bom."""
    monkeypatch.setattr(ai.epc_weichai, "engine_bom", lambda rangka: {
        "found": True, "groups": [
            {"pn": "612600030201", "nama": "Piston", "parts": [
                {"pn": "612600030001", "nama": "Piston ring"}]},
        ]})
    sid = _sheet_nama(USER, ["Piston ring"])
    r = ai._t_sheet_isi_part_number({"_sheet_id": sid, "rangka": "RJ-TEST"}, USER)
    assert r["baris_terisi"] == 1


def test_isi_pn_wajib_rangka(bom_unit):
    sid = _sheet_nama(USER, ["Oil filter"])
    r = ai._t_sheet_isi_part_number({"_sheet_id": sid}, USER)
    assert r["found"] is False and "rangka" in r["error"].lower()


def test_isi_pn_token_epc_kedaluwarsa(monkeypatch, bom_unit):
    monkeypatch.setattr(ai.epc_bom, "loading_list",
                        lambda rangka: {"found": False, "_err": "token_expired", "frame_number": rangka})
    sid = _sheet_nama(USER, ["Oil filter"])
    r = ai._t_sheet_isi_part_number({"_sheet_id": sid, "rangka": "RJ-TEST"}, USER)
    assert r["found"] is False and r.get("_token_issue")


def test_isi_pn_kolom_nama_tak_terdeteksi(bom_unit):
    # Sheet tanpa kolom nama (hanya angka) → minta user sebut kolomnya.
    p = ai_sheet.parse_upload(_xlsx([["Qty"], [1], [2]]), "x.xlsx")
    sid = ai_sheet.put_sheet(USER["username"], p)
    r = ai._t_sheet_isi_part_number({"_sheet_id": sid, "rangka": "RJ-TEST"}, USER)
    assert r["found"] is False and "nama part" in r["error"].lower()


def test_tool_isi_pn_hanya_ada_bila_ada_lampiran():
    assert "sheet_isi_part_number" not in ai._allowed_tool_names(USER, "")
    assert "sheet_isi_part_number" in ai._allowed_tool_names(USER, "sid-apapun")


def test_isi_pn_lewat_run_tool_menimpa_sheet_id(bom_unit):
    """Jalur _run_tool: _sheet_id dari model diabaikan, diganti sheet_id server."""
    sid = _sheet_nama(USER, ["Oil filter"])
    r = ai._run_tool("sheet_isi_part_number",
                     {"_sheet_id": "curian", "rangka": "RJ-TEST"}, USER, sheet_id=sid)
    assert r["found"] and r["baris_terisi"] == 1


# ── Cek/isi Qty dari BOM unit ─────────────────────────────────────────────────

def test_cek_qty_isi_kosong_dan_tandai_selisih(bom_unit):
    p = ai_sheet.parse_upload(_xlsx([
        ["Part Number", "Qty"],
        ["WG9925520270", ""],       # kosong → diisi dari BOM (4)
        ["AZ9925520271", 2],        # cocok BOM (2)
        ["VG1560080126", 5],        # beda BOM (1) → ditandai, TAK ditimpa
        ["ZZZ0000000", 3],          # PN tak di BOM → tanpa referensi
    ]), "qty.xlsx")
    sid = ai_sheet.put_sheet(USER["username"], p)
    r = ai._t_sheet_cek_qty({"_sheet_id": sid, "rangka": "RJ-TEST"}, USER)
    assert r["found"]
    assert r["qty_diisi_dari_bom"] == 1
    assert r["qty_cocok"] == 1
    assert r["qty_selisih"] == 1
    assert r["tanpa_referensi_bom"] == 1


def test_cek_qty_wajib_rangka(bom_unit):
    p = ai_sheet.parse_upload(_xlsx([["Part Number", "Qty"], ["WG9925520270", 1]]), "q.xlsx")
    sid = ai_sheet.put_sheet(USER["username"], p)
    r = ai._t_sheet_cek_qty({"_sheet_id": sid}, USER)
    assert r["found"] is False and "rangka" in r["error"].lower()


def test_tool_cek_qty_hanya_ada_bila_lampiran():
    assert "sheet_cek_qty" not in ai._allowed_tool_names(USER, "")
    assert "sheet_cek_qty" in ai._allowed_tool_names(USER, "sid-apapun")


# ── Pemahaman proaktif ────────────────────────────────────────────────────────

def test_semua_tool_sheet_ditawarkan_saat_lampiran():
    names = ai._allowed_tool_names(USER, "sid")
    for t in ("sheet_ringkasan", "sheet_isi_kolom", "sheet_isi_part_number", "sheet_cek_qty"):
        assert t in names


def test_chat_lampiran_proaktif_panggil_ringkasan(monkeypatch, katalog):
    """User hanya melampirkan file + minta samar → model boleh proaktif membaca
    (sheet_ringkasan) lalu menawarkan aksi. Menguji jalur & ketersediaan tool."""
    monkeypatch.setattr(ai, "_system_prompt", lambda user: "system uji")
    monkeypatch.setattr(ai, "_unit_name_tokens", lambda: set())
    sid = _sheet_untuk(USER, katalog)
    seq = [
        {"choices": [{"message": {"content": "", "tool_calls": [
            {"id": "t1", "function": {"name": "sheet_ringkasan", "arguments": "{}"}}]},
            "finish_reason": "tool_calls"}]},
        {"choices": [{"message": {"content": "File berisi 2 baris part. Mau saya lengkapi "
                                             "Part Number / stok?"}, "finish_reason": "stop"}]},
    ]
    n = {"i": 0}

    def fake_post(messages, tools, max_tokens=6000):
        c = seq[min(n["i"], len(seq) - 1)]
        n["i"] += 1
        return c

    monkeypatch.setattr(ai, "_post_chat", fake_post)
    out = ai.chat(USER, [{"role": "user", "content": "tolong bantu file ini"}], sheet_id=sid)
    assert "sheet_ringkasan" in out["tools_used"]
