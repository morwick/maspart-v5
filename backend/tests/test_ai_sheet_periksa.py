"""Fase 2 olah-Excel: tandai_status (warna+flag+mungkin-maksud), kolom pengganti/
cross_ref/berat, rencana_pemenuhan, rekap (gate harga), determinisme."""
import io

import pytest
from openpyxl import Workbook, load_workbook

from app.services import ai_export, ai_sheet

ADMIN = {"username": "admin", "role": "admin"}
USER = {"username": "budi", "role": "user"}


def _xlsx(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def dunia(monkeypatch):
    rows = [
        {"part_number": "WG9925520270", "part_name": "Spring bracket", "stok": "12", "harga": "Rp 1.500.000"},
        {"part_number": "AZ9925520271", "part_name": "Leaf spring", "stok": "0", "harga": "Rp 500.000"},
    ]
    monkeypatch.setattr(ai_sheet.part_index, "search_exact_pns",
                        lambda pns: [r for r in rows if r["part_number"].upper() in {p.upper() for p in pns}])
    monkeypatch.setattr(ai_sheet.part_index, "_pn_flat_map",
                        lambda: {"WG9925520270": ["WG9925520270"], "AZ9925520271": ["AZ9925520271"]})
    monkeypatch.setattr(ai_sheet.part_index, "gudang_names", lambda: [])
    # pengganti (SIMS index)
    monkeypatch.setattr(ai_sheet.sims, "equivalents_count", lambda: 9)
    monkeypatch.setattr(ai_sheet.sims, "equivalents_for",
                        lambda pn: ({"digantikan_oleh": [{"pn": "NEW777", "nama": "baru"}]}
                                    if pn == "GANTIME1" else {}))
    # cross-ref & berat
    monkeypatch.setattr(ai_sheet.filter_ref, "by_pn",
                        lambda pn: ({"cross_reference": ["FF5052", "P550008"]}
                                    if pn == "WG9925520270" else None))
    monkeypatch.setattr(ai_sheet.harga, "shipping_weight_for",
                        lambda pn, allow_remote=False: 2500 if pn == "WG9925520270" else 0)
    return rows


def _sid(rows, user=USER):
    p = ai_sheet.parse_upload(_xlsx(rows), "order.xlsx")
    return ai_sheet.put_sheet(user["username"], p)


def test_status_warna_dan_flag(dunia):
    sid = _sid([["Part Number", "Qty"], ["WG9925520270", 3], ["AZ9925520271", 1],
                ["GANTIME1", 1], ["WG9925520271", 1]])
    r = ai_sheet.fill_columns(sid, USER, permintaan=[], tandai_status=True)
    assert r["found"]
    data, _ = ai_export.generic_excel(r["export_id"])
    ws = load_workbook(io.BytesIO(data))["Data"]
    # kolom Status ada; baca teks per baris data (mulai baris 5)
    hdr = [ws.cell(row=4, column=j).value for j in range(1, ws.max_column + 1)]
    si = hdr.index("Status") + 1
    ki = hdr.index("Keterangan") + 1
    assert ws.cell(row=5, column=si).value == "READY"          # stok 12
    assert ws.cell(row=6, column=si).value == "STOK KOSONG"    # stok 0
    assert ws.cell(row=7, column=si).value.startswith("ADA PENGGANTI: NEW777")
    assert ws.cell(row=8, column=si).value == "PN TAK DITEMUKAN"
    # 'mungkin maksud' hanya di Keterangan, TAK menimpa kolom PN
    assert "mungkin maksud" in (ws.cell(row=8, column=ki).value or "")
    assert ws.cell(row=8, column=1).value == "WG9925520271"    # PN user utuh
    assert {"pn": "WG9925520271", "saran": "WG9925520270"} in r["mungkin_maksud"]
    assert r["pn_diganti"] == [{"pn": "GANTIME1", "pengganti": "NEW777"}]
    assert r["status_ringkas"]["merah"] >= 1 and r["status_ringkas"]["kuning"] == 2


def test_status_kurang_pakai_qty(dunia):
    sid = _sid([["Part Number", "Qty"], ["WG9925520270", 99]])  # butuh 99, stok 12
    r = ai_sheet.fill_columns(sid, USER, permintaan=[], tandai_status=True)
    data, _ = ai_export.generic_excel(r["export_id"])
    ws = load_workbook(io.BytesIO(data))["Data"]
    hdr = [ws.cell(row=4, column=j).value for j in range(1, ws.max_column + 1)]
    assert ws.cell(row=5, column=hdr.index("Status") + 1).value.startswith("KURANG")


def test_kolom_pengganti_crossref_berat(dunia):
    sid = _sid([["Part Number"], ["WG9925520270"], ["GANTIME1"]])
    r = ai_sheet.fill_columns(sid, USER, permintaan=[
        {"isi": "pengganti"}, {"isi": "cross_ref"}, {"isi": "berat"}])
    data, _ = ai_export.generic_excel(r["export_id"])
    ws = load_workbook(io.BytesIO(data))["Data"]
    hdr = [ws.cell(row=4, column=j).value for j in range(1, ws.max_column + 1)]
    pg = hdr.index("PN Pengganti") + 1
    cr = [h for h in hdr if h.startswith("Cross Ref")][0]
    cri = hdr.index(cr) + 1
    br = hdr.index("Berat (kg)") + 1
    assert ws.cell(row=6, column=pg).value == "NEW777"          # GANTIME1 → NEW777
    assert ws.cell(row=5, column=cri).value == "FF5052, P550008"
    assert ws.cell(row=5, column=br).value == "2,50"


def test_rekap_gate_harga(dunia):
    rows = [["Part Number", "Qty"], ["WG9925520270", 2]]
    # admin (boleh_harga=True) → ada Subtotal
    r = ai_sheet.fill_columns(_sid(rows, ADMIN), ADMIN, permintaan=[], rekap=True, boleh_harga=True)
    labels = [x["label"] for x in r["rekap"]]
    assert any("Subtotal" in l for l in labels) and any("PPN" in l for l in labels)
    # non-privileged (boleh_harga=False) → TANPA Subtotal/PPN
    r2 = ai_sheet.fill_columns(_sid(rows), USER, permintaan=[], rekap=True, boleh_harga=False)
    labels2 = [x["label"] for x in r2["rekap"]]
    assert not any("Subtotal" in l or "PPN" in l for l in labels2)
    assert any("Jumlah item" in l for l in labels2)


def test_determinisme(dunia):
    rows = [["Part Number", "Qty"], ["WG9925520270", 2], ["GANTIME1", 1]]
    a = ai_sheet.fill_columns(_sid(rows), USER, permintaan=[{"isi": "pengganti"}],
                              tandai_status=True)
    b = ai_sheet.fill_columns(_sid(rows), USER, permintaan=[{"isi": "pengganti"}],
                              tandai_status=True)
    da, _ = ai_export.generic_excel(a["export_id"])
    db, _ = ai_export.generic_excel(b["export_id"])
    assert da == db
