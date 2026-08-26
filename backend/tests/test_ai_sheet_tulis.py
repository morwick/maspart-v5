"""Tool `sheet_tulis` — asisten menulis nilai/catatan/RUMUS yang DIDIKTE USER ke
Excel lampirannya, di tempat.

Sebelum ini asisten hanya bisa mengisi menu tetap data MASPART (`sheet_isi_kolom`):
permintaan sesederhana "tulis 'kirim batch 2' di kolom Keterangan untuk 3 PN ini"
atau "kolom Total = Qty × Harga" tak ada alatnya. Test ini memakai file "punya
user" yang kaya format (kop, baris kosong, header di baris 3, rumus, baris TOTAL,
sheet kedua) dan memeriksa hasilnya SEL PER SEL — termasuk pagar yang menjaga
milik user: sel berisi tak ditimpa, rumus user tak dirusak, baris TOTAL tak
kena "isi semua baris", dan teks berawalan '=' tetap dijinakkan (anti injeksi).
"""
import io

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.services import ai_assistant as ai
from app.services import ai_export, ai_sheet

USER = {"username": "budi", "role": "admin"}

_JUDUL = "PT CONTOH JAYA — Daftar Permintaan Part"


def _file_user() -> bytes:
    """Excel gaya lapangan: kop bergaya, baris kosong, header di baris 3, rumus
    milik user, kolom Keterangan yang sebagian sudah diisi, dan baris TOTAL."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Permintaan"
    ws.merge_cells("A1:D1")
    ws["A1"] = _JUDUL
    ws["A1"].font = Font(bold=True, size=16, color="FFFF0000")
    ws["A1"].fill = PatternFill("solid", fgColor="FFFFFF00")
    ws["A2"] = "Tanggal: 26 Agustus 2026"
    for j, h in enumerate(["No", "Part Number", "Nama", "Qty", "Harga",
                           "Stok", "Keterangan"], 1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor="FF1F4E78")
    # baris 4
    ws.cell(row=4, column=1, value=1)
    ws.cell(row=4, column=2, value="WG9925520270")
    ws.cell(row=4, column=3, value="Filter oli")
    ws.cell(row=4, column=4, value=2)
    ws.cell(row=4, column=5, value=150000).number_format = '"Rp"#,##0'
    ws.cell(row=4, column=6, value=5)
    ws.cell(row=4, column=7, value="punya user")        # Keterangan SUDAH berisi
    # baris 5
    ws.cell(row=5, column=1, value=2)
    ws.cell(row=5, column=2, value="AZ9925520271")
    ws.cell(row=5, column=3, value="Filter solar")
    ws.cell(row=5, column=4, value=3)
    ws.cell(row=5, column=5, value=90000)
    ws.cell(row=5, column=6, value=1)                    # stok < qty
    # baris 6 sengaja KOSONG (pemisah); baris 7 = baris TOTAL milik user
    ws.cell(row=7, column=1, value="TOTAL")
    ws.cell(row=7, column=5, value="=SUM(E4:E5)")
    ws.column_dimensions["C"].width = 42
    ws.freeze_panes = "A4"
    wb.create_sheet("Catatan")["A1"] = "jangan dihapus"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sid(data: bytes = None, nama: str = "permintaan.xlsx") -> str:
    p = ai_sheet.parse_upload(data if data is not None else _file_user(), nama)
    assert p.get("ok"), p
    return ai_sheet.put_sheet(USER["username"], p)


def _ws(export_id, nama: str = "Permintaan"):
    data, _ = ai_export.generic_excel(export_id)
    assert data
    return load_workbook(io.BytesIO(data))[nama]


@pytest.fixture(autouse=True)
def _tanpa_katalog(monkeypatch):
    """Deteksi peran kolom memanggil indeks part; test ini tak menguji katalog."""
    monkeypatch.setattr(ai_sheet.part_index, "rows_for_pns", lambda pns: {})
    monkeypatch.setattr(ai_sheet.part_index, "search_exact_pns", lambda pns: set(pns))


# ── sasaran per PART NUMBER ──────────────────────────────────────────────────
def test_per_pn_mendarat_di_baris_yang_benar_kolom_baru():
    r = ai_sheet.tulis(_sid(), USER, kolom="Catatan Kirim",
                       nilai=[{"pn": "AZ9925520271", "nilai": "kirim batch 2"}])
    assert r["found"] and r["sel_ditulis"] == 1
    assert r["kolom_baru"] is True and r["format_asli_dipertahankan"] is True

    ws = _ws(r["export_id"])
    kol = ws[r["kolom_excel"] + "3"].column      # kolom baru, huruf dilaporkan
    assert ws.cell(row=3, column=kol).value == "Catatan Kirim"
    assert ws.cell(row=5, column=kol).value == "kirim batch 2"   # baris PN itu saja
    assert ws.cell(row=4, column=kol).value is None
    # file user tetap utuh
    assert ws["A1"].value == _JUDUL and ws["A1"].font.size == 16
    assert ws["E4"].number_format == '"Rp"#,##0'
    assert ws["E7"].value == "=SUM(E4:E5)"


def test_pn_tak_ada_di_file_dilaporkan_bukan_ditulis_asal():
    r = ai_sheet.tulis(_sid(), USER, kolom="Catatan",
                       nilai=[{"pn": "WG9925520270", "nilai": "ok"},
                              {"pn": "TIDAK-ADA-123", "nilai": "ok"}])
    assert r["found"] and r["sel_ditulis"] == 1
    assert r["pn_tidak_ada_di_file"] == ["TIDAK-ADA-123"]
    assert "TIDAK-ADA-123" in r["catatan"]


def test_semua_pn_meleset_gagal_terang_terangan():
    r = ai_sheet.tulis(_sid(), USER, kolom="Catatan",
                       nilai=[{"pn": "TIDAK-ADA-123", "nilai": "ok"}])
    assert r["found"] is False
    assert "TIDAK-ADA-123" in r["error"]


# ── pagar: sel user yang SUDAH BERISI ────────────────────────────────────────
def test_sel_user_sudah_berisi_tak_ditimpa_tanpa_izin():
    r = ai_sheet.tulis(_sid(), USER, kolom="Keterangan",
                       nilai=[{"pn": "WG9925520270", "nilai": "ganti"}])
    assert r["found"] is False and r["sel_dilewati_sudah_terisi"] == 1
    assert "timpa=true" in r["error"]


def test_timpa_true_baru_mengganti_isi_user():
    r = ai_sheet.tulis(_sid(), USER, kolom="Keterangan", timpa=True,
                       nilai=[{"pn": "WG9925520270", "nilai": "ganti"}])
    assert r["found"] and r["sel_ditulis"] == 1
    assert r["kolom_baru"] is False and r["kolom_excel"] == "G"
    assert _ws(r["export_id"])["G4"].value == "ganti"


def test_rumus_milik_user_tak_dirusak_walau_timpa():
    """E7 milik user berisi =SUM(E4:E5). Mesin isi-di-tempat menolak menimpanya —
    dan laporan tool ikut JUJUR: 0 sel ditulis, bukan '1 sel terisi'."""
    r = ai_sheet.tulis(_sid(), USER, kolom="Harga", timpa=True,
                       nilai=[{"baris": 7, "nilai": 999}])
    assert r["found"] is False and r["sel_dilewati_rumus"] == 1
    assert "RUMUS" in r["error"]


# ── sasaran per NOMOR BARIS Excel ────────────────────────────────────────────
def test_per_nomor_baris_memakai_baris_asli_bukan_urutan_data():
    r = ai_sheet.tulis(_sid(), USER, kolom="Qty", timpa=True,
                       nilai=[{"baris": 5, "nilai": 7}])
    assert r["found"] and r["sel_ditulis"] == 1
    ws = _ws(r["export_id"])
    assert ws["D5"].value == 7 and ws["D4"].value == 2


def test_nomor_baris_di_luar_data_dilaporkan():
    r = ai_sheet.tulis(_sid(), USER, kolom="Catatan",
                       nilai=[{"baris": 4, "nilai": "a"}, {"baris": 99, "nilai": "b"}])
    assert r["found"] and r["baris_tidak_ada"] == [99]


# ── satu nilai untuk SEMUA baris ─────────────────────────────────────────────
def test_nilai_semua_tak_menodai_baris_total_milik_user():
    r = ai_sheet.tulis(_sid(), USER, kolom="Supplier", nilai_semua="MAS")
    assert r["found"] and r["sel_ditulis"] == 2          # baris 4 & 5, BUKAN baris 7
    assert r["baris_tanpa_part_number_dilewati"] == 1
    ws = _ws(r["export_id"])
    kol = ws[r["kolom_excel"] + "3"].column
    assert [ws.cell(row=b, column=kol).value for b in (4, 5, 7)] == ["MAS", "MAS", None]


# ── syarat `bila` (dihitung server, bukan ditebak model) ─────────────────────
def test_bila_kosong_hanya_mengisi_sel_yang_masih_kosong():
    r = ai_sheet.tulis(_sid(), USER, kolom="Keterangan", nilai_semua="perlu dicek",
                       bila={"kolom": "Keterangan", "operator": "kosong"})
    assert r["found"] and r["sel_ditulis"] == 1
    ws = _ws(r["export_id"])
    assert ws["G4"].value == "punya user" and ws["G5"].value == "perlu dicek"


def test_bila_banding_kolom_lain_stok_di_bawah_qty():
    r = ai_sheet.tulis(_sid(), USER, kolom="Status Stok", nilai_semua="KURANG",
                       bila={"kolom": "Stok", "operator": "lebih_kecil", "nilai": "{Qty}"})
    assert r["found"] and r["sel_ditulis"] == 1          # hanya baris 5 (stok 1 < qty 3)
    ws = _ws(r["export_id"])
    kol = ws[r["kolom_excel"] + "3"].column
    assert ws.cell(row=5, column=kol).value == "KURANG"
    assert ws.cell(row=4, column=kol).value is None


def test_bila_kolom_tak_ada_gagal_dengan_daftar_kolom():
    r = ai_sheet.tulis(_sid(), USER, kolom="Catatan", nilai_semua="y",
                       bila={"kolom": "Gudang Asal", "operator": "terisi"})
    assert r["found"] is False and "Gudang Asal" in r["error"]


# ── RUMUS HIDUP ──────────────────────────────────────────────────────────────
def test_rumus_ditulis_sebagai_rumus_sungguhan_per_baris():
    r = ai_sheet.tulis(_sid(), USER, kolom="Total", rumus="={Qty}*{Harga}")
    assert r["found"] and r["sel_ditulis"] == 2
    assert r["rumus_contoh"] == "=D4*E4"
    ws = _ws(r["export_id"])
    kol = ws[r["kolom_excel"] + "3"].column
    assert ws.cell(row=4, column=kol).value == "=D4*E4"   # bukan "'=D4*E4"
    assert ws.cell(row=5, column=kol).value == "=D5*E5"


def test_rumus_boleh_disasar_ke_pn_tertentu_saja():
    r = ai_sheet.tulis(_sid(), USER, kolom="Total", rumus="{Qty}*{Harga}",
                       nilai=[{"pn": "AZ9925520271"}])
    assert r["found"] and r["sel_ditulis"] == 1 and r["rumus_contoh"] == "=D5*E5"


def test_rumus_di_kolom_baru_dapat_format_angka_kolomnya():
    """'=D4*E4' di kolom "Total Harga" harus tampil sebagai Rupiah, bukan angka
    telanjang — number_format kolom baru ikut berlaku untuk rumus."""
    r = ai_sheet.tulis(_sid(), USER, kolom="Total Harga", rumus="={Qty}*{Harga}")
    assert r["found"]
    ws = _ws(r["export_id"])
    sel = ws.cell(row=4, column=ws[r["kolom_excel"] + "3"].column)
    assert sel.value == "=D4*E4"
    assert sel.number_format == ai_export.num_format("Total Harga") != "General"


def test_rumus_merujuk_kolom_tak_ada_ditolak():
    r = ai_sheet.tulis(_sid(), USER, kolom="Total", rumus="={Diskon}*{Harga}")
    assert r["found"] is False and "Diskon" in r["error"]


def test_rumus_melingkar_ditolak():
    r = ai_sheet.tulis(_sid(), USER, kolom="Harga", rumus="={Harga}*2", timpa=True)
    assert r["found"] is False and "melingkar" in r["error"]


def test_rumus_butuh_peta_baris_asli_csv_ditolak():
    sid = _sid(b"Part Number,Qty,Harga\nWG9925520270,2,1000\n", "daftar.csv")
    r = ai_sheet.tulis(sid, USER, kolom="Total", rumus="={Qty}*{Harga}")
    assert r["found"] is False and "csv" in r["error"].lower()


# ── keamanan & koersi nilai ──────────────────────────────────────────────────
def test_teks_berawalan_sama_dengan_tetap_dijinakkan():
    """Hanya `rumus` yang boleh jadi rumus hidup; nilai biasa tetap TEKS."""
    r = ai_sheet.tulis(_sid(), USER, kolom="Catatan",
                       nilai=[{"pn": "WG9925520270", "nilai": "=1+1"}])
    assert r["found"]
    ws = _ws(r["export_id"])
    assert ws.cell(row=4, column=ws[r["kolom_excel"] + "3"].column).value == "'=1+1"


@pytest.mark.parametrize("masuk,keluar", [
    (5, 5), ("12", 12), (1500.5, 1500.5), ("1500,5", 1500.5),
    ("0812345678", "0812345678"),      # nomor telepon → TEKS
    ("612630010054", "612630010054"),  # PN numerik Weichai → TEKS
    ("kirim batch 2", "kirim batch 2"),
    (True, "YA"),
])
def test_koersi_nilai_sel(masuk, keluar):
    assert ai_sheet._nilai_sel(masuk) == keluar


def test_angka_ditulis_sebagai_angka_supaya_sum_user_jalan():
    r = ai_sheet.tulis(_sid(), USER, kolom="Qty Kirim",
                       nilai=[{"pn": "WG9925520270", "nilai": "4"}])
    ws = _ws(r["export_id"])
    assert ws.cell(row=4, column=ws[r["kolom_excel"] + "3"].column).value == 4


# ── kontrak masukan ──────────────────────────────────────────────────────────
def test_tanpa_nilai_apapun_ditolak():
    r = ai_sheet.tulis(_sid(), USER, kolom="Catatan")
    assert r["found"] is False and "nilai" in r["error"]


def test_tanpa_kolom_ditolak():
    r = ai_sheet.tulis(_sid(), USER, kolom="", nilai_semua="x")
    assert r["found"] is False


def test_sheet_id_milik_user_lain_tak_bisa_ditulisi():
    sid = _sid()
    r = ai_sheet.tulis(sid, {"username": "orang_lain"}, kolom="Catatan", nilai_semua="x")
    assert r["found"] is False and "kedaluwarsa" in r["error"]


# ── ringkasan: sasaran tulis harus bisa disebut model ────────────────────────
def test_ringkasan_memberi_huruf_kolom_dan_nomor_baris_asli():
    p = ai_sheet.parse_upload(_file_user(), "permintaan.xlsx")
    ring = ai_sheet.ringkas(p)
    assert ring["baris_header_excel"] == 3
    assert ring["contoh_baris_nomor"] == [4, 5, 7]
    assert [k["kolom_excel"] for k in ring["kolom"]][:3] == ["A", "B", "C"]


def test_huruf_kolom_di_luar_jangkauan_ditolak_bukan_bikin_kolom_aneh():
    """'tulis di kolom J' padahal file cuma sampai G: jangan diam-diam membuat
    kolom bernama 'J' — sebutkan kolom yang memang ada."""
    r = ai_sheet.tulis(_sid(), USER, kolom="J", nilai_semua="x")
    assert r["found"] is False and "B=Part Number" in r["error"]


def test_huruf_kolom_yang_ada_dipakai_apa_adanya():
    r = ai_sheet.tulis(_sid(), USER, kolom="G", nilai_semua="cek", timpa=True)
    assert r["found"] and r["kolom_excel"] == "G" and r["kolom_tujuan"] == "Keterangan"


# ── lapisan tool: ketersediaan & pemaaf bentuk argumen ───────────────────────
def test_tool_hanya_ditawarkan_saat_ada_lampiran():
    assert "sheet_tulis" not in ai._allowed_tool_names(USER, "")
    assert "sheet_tulis" in ai._allowed_tool_names(USER, "sid-apapun")


def test_handler_terima_nilai_dict_tunggal():
    r = ai._t_sheet_tulis({"_sheet_id": _sid(), "kolom": "Catatan",
                           "nilai": {"pn": "WG9925520270", "nilai": "ok"}}, USER)
    assert r["found"] and r["sel_ditulis"] == 1


def test_handler_terima_peta_pn_ke_nilai():
    r = ai._t_sheet_tulis({"_sheet_id": _sid(), "kolom": "Catatan",
                           "nilai": {"WG9925520270": "a", "AZ9925520271": "b"}}, USER)
    assert r["found"] and r["sel_ditulis"] == 2


def test_handler_nilai_skalar_dibaca_sebagai_semua_baris():
    r = ai._t_sheet_tulis({"_sheet_id": _sid(), "kolom": "Supplier",
                           "nilai": "MAS"}, USER)
    assert r["found"] and r["sel_ditulis"] == 2


def test_model_tak_bisa_memilih_file_lewat_argumen():
    """`_sheet_id` selalu ditimpa server (pola semua tool sheet_*)."""
    sid = _sid()
    args = {"sheet_id": "punya-orang-lain", "kolom": "Catatan", "nilai_semua": "x"}
    r = ai._run_tool("sheet_tulis", dict(args), USER, sid)
    assert r["found"] is True


def test_chat_penuh_sampai_kartu_unduh(monkeypatch):
    """Jalur PENUH: lampiran → tool ditawarkan → model panggil sheet_tulis → KARTU
    UNDUH muncul. Tanpa nama tool di daftar kartu (`chat`), file terbuat tapi user
    tak pernah bisa mengunduhnya — kegagalan yang tak terlihat dari test unit."""
    monkeypatch.setattr(ai, "_system_prompt", lambda user: "system uji")
    monkeypatch.setattr(ai, "_unit_name_tokens", lambda: set())
    sid = _sid()
    dilihat: list[list[str]] = []
    seq = [
        {"choices": [{"message": {"content": "", "tool_calls": [
            {"id": "t1", "function": {"name": "sheet_tulis", "arguments":
                '{"kolom":"Keterangan","nilai":[{"pn":"AZ9925520271",'
                '"nilai":"kirim batch 2"}]}'}}]},
            "finish_reason": "tool_calls"}]},
        {"choices": [{"message": {"content": "Sudah saya tulis di kolom Keterangan."},
                      "finish_reason": "stop"}]},
    ]
    n = {"i": 0}

    def fake_post(messages, tools, max_tokens=6000):
        dilihat.append([t["function"]["name"] for t in (tools or [])])
        c = seq[min(n["i"], len(seq) - 1)]
        n["i"] += 1
        return c

    monkeypatch.setattr(ai, "_post_chat", fake_post)
    out = ai.chat(USER, [{"role": "user",
                          "content": "tulis 'kirim batch 2' di keterangan PN AZ9925520271"}],
                  sheet_id=sid)
    assert "sheet_tulis" in dilihat[0] and "sheet_tulis" in out["tools_used"]
    assert out["excel_exports"] and out["excel_exports"][0]["id"]
