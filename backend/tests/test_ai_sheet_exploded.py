"""Tool asisten `sheet_isi_exploded`: tempel GAMBAR TEKNIS (exploded view EPC)
ke Excel unggahan user.

Aturan yang dijaga di sini:
  • ⛔ TANYA DULU soal nomor rangka. Tanpa `rangka` DAN tanpa `lintas_model`,
    tool TIDAK menebak: ia mengembalikan perintah bertanya & TIDAK menyentuh EPC.
    Bertanya juga BUKAN "lookup gagal" (tak ber-found:False) — kalau tidak,
    giliran tercatat tool_failed dan model disuntik nota "jangan mengarang".
  • Ada rangka  → figure PER-VIN (gambar unit itu sendiri).
    Tak ada rangka → figure LINTAS MODEL + peringatannya wajib ikut.
  • Gambar dicocokkan lewat PART NUMBER (nama part tak cukup).
  • PN tanpa figure: sel gambar KOSONG + ALASANNYA ditulis apa adanya.
  • Berat → hanya daftar kerja yang disimpan saat tool jalan; unduh & render
    gambar dikerjakan SAAT KARTU DIUNDUH (builder 'sheet_exploded'), di bawah
    gerbang satu-batch (server 1 vCPU).

Tanpa jaringan: `epc_bom._get_auto` & `ai_export.exploded_png` di-stub.
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from app.services import ai_export, ai_sheet, epc_bom, exploded_view
from app.services import ai_assistant as A

USER = {"username": "budi", "role": "staf"}
FRAME = "RJ326978"
PN_ADA = "WG9725520278"
PN_TAK = "AZ9100440002"


def _png(w: int = 1400, h: int = 900) -> bytes:
    """PNG SAH — byte palsu ditolak openpyxl, gambarnya tak akan tertempel."""
    from PIL import Image
    b = io.BytesIO()
    Image.new("RGB", (w, h), "white").save(b, format="PNG")
    return b.getvalue()


PNG = _png()

_ROOT = {"data": {"orderNo": "ORD-1", "partRoots": [{"id": 1}]}}
_REVERSE_ADA = {"data": [{"partCode": "ASSY-1", "partName": "Front axle",
                          "rootId": 1, "partId": 22, "partListId": 222}]}
_ITEMS = {"data": {"partListName": "Front axle assembly", "d2s": ["FIG-1.svg"],
                   "items": [{"code": PN_ADA, "name": "Wheel hub",
                              "ballNum": 7, "amount": 2}]}}


@pytest.fixture(autouse=True)
def bersih():
    exploded_view.bersihkan_cache()
    epc_bom._root_cache.clear()
    yield
    exploded_view.bersihkan_cache()
    epc_bom._root_cache.clear()


@pytest.fixture
def epc_stub(monkeypatch):
    """EPC per-VIN tiruan: PN_ADA punya figure, PN_TAK tidak. Mengembalikan
    daftar panggilan supaya test bisa membuktikan EPC TAK disentuh saat bertanya."""
    panggilan: list[tuple] = []

    def fake_get_auto(url, params, **kw):
        panggilan.append((url, dict(params)))
        if url == epc_bom._ATLAS_ROOT_URL:
            return _ROOT
        if url == epc_bom._REVERSE_URL:
            return _REVERSE_ADA if (params.get("k") or "").upper() == PN_ADA else {"data": []}
        if url == epc_bom._ATLAS_ITEM_URL:
            return _ITEMS
        return {"_err": "api"}

    monkeypatch.setattr(epc_bom, "_get_auto", fake_get_auto)
    monkeypatch.setattr(ai_export, "exploded_png", lambda svg, ball=None: PNG)
    return panggilan


def _sheet(headers, roles, body, filename="PERMINTAAN.xlsx") -> str:
    return ai_sheet.put_sheet("budi", {
        "filename": filename, "headers": list(headers), "roles": list(roles),
        "_body": [list(r) for r in body],
        "contoh": [], "sheet": "Sheet1", "sheet_lain": [], "pn_dikenal": len(body),
        "jumlah_baris": len(body), "jumlah_kolom": len(headers),
        "kolom_pn": headers[roles.index("part_number")] if "part_number" in roles else None,
        "terpotong": False,
    })


@pytest.fixture
def sheet():
    return _sheet(["Part No.", "Nama", "Qty"], ["part_number", "part_name", "qty"],
                  [[PN_ADA, "Wheel hub", "2"], [PN_TAK, "Bracket", "1"]])


# ── Gerbang TANYA: inti permintaan pemilik ──────────────────────────────────

def test_tanpa_vin_tool_bertanya_dan_tak_menyentuh_epc(sheet, epc_stub):
    out = ai_sheet.fill_exploded(sheet, USER)
    assert out["perlu_jawaban_user"] is True
    assert "export_id" not in out                      # tak ada file dibuat
    assert epc_stub == [], "EPC tak boleh disentuh sebelum user menjawab"
    teks = out["jawaban_wajib"].lower()
    assert "rangka" in teks and "lintas_model=true" in out["jawaban_wajib"]
    assert len(out["pilihan"]) == 2                    # ADA VIN / TIDAK ADA VIN
    assert out["jumlah_part"] == 2


def test_bertanya_bukan_lookup_gagal(sheet):
    """found:False akan membuat giliran tercatat tool_failed & model disuntik
    nota 'lookup gagal, jangan mengarang' — padahal tak ada yang gagal."""
    out = ai_sheet.fill_exploded(sheet, USER)
    assert "found" not in out and "error" not in out
    assert A._tool_fail_kind(out) == ""


def test_vin_di_file_ditawarkan_tapi_tak_dipakai_diam_diam(epc_stub):
    sid = _sheet(["No Rangka", "Part No."], ["lain", "part_number"],
                 [[FRAME, PN_ADA], [FRAME, PN_TAK]])
    out = ai_sheet.fill_exploded(sid, USER)
    assert out["perlu_jawaban_user"] is True
    assert out["rangka_di_file"] == {"kolom": "No Rangka", "nilai": [FRAME]}
    assert FRAME in out["jawaban_wajib"]
    assert epc_stub == []                              # tetap menunggu jawaban user


# ── Dua jalur: per-VIN & lintas-model ───────────────────────────────────────

def test_jalur_per_vin_menyiapkan_kartu_dan_kolom(sheet):
    out = ai_sheet.fill_exploded(sheet, USER, rangka=FRAME)
    assert out["found"] is True and out["sumber_gambar_teknis"] == "per-VIN"
    assert out["rangka"] == FRAME and out["jumlah_part"] == 2
    b = ai_export._stash[out["export_id"]]["builder"]
    assert b["kind"] == "sheet_gambar" and b["rangka"] == FRAME
    assert b["kolom"][-2:] == [ai_sheet._KOL_EXPLODED_INFO, ai_sheet._KOL_EXPLODED_GAMBAR]
    assert b["pns"] == [PN_ADA, PN_TAK]
    assert b["kol_info"] == 3 and b["kol_gambar"] == 4


def test_jalur_lintas_model_membawa_peringatan(sheet):
    out = ai_sheet.fill_exploded(sheet, USER, lintas_model=True)
    assert out["found"] is True and out["sumber_gambar_teknis"] == "lintas-model"
    assert ai_export._stash[out["export_id"]]["builder"]["rangka"] == ""
    assert "LINTAS MODEL" in out["catatan"]            # model wajib menyampaikannya
    assert "estimasi_durasi_unduhan" in out


def test_plafon_pn_beda_per_jalur():
    n = ai_sheet._MAX_EXPLODED_PN_GLOBAL + 1
    sid = _sheet(["Part No."], ["part_number"], [[f"WG97255202{i:02d}"] for i in range(n)])
    lintas = ai_sheet.fill_exploded(sid, USER, lintas_model=True)
    assert lintas["found"] is False and "terlalu banyak" in lintas["error"].lower()
    # Jalur per-VIN jauh lebih murah → plafonnya lebih longgar, bukan ikut ditolak.
    assert ai_sheet.fill_exploded(sid, USER, rangka=FRAME)["found"] is True


def test_tanpa_kolom_pn_ditolak_bukan_ditebak_dari_nama():
    sid = _sheet(["Nama", "Qty"], ["part_name", "qty"], [["Wheel hub", "2"]])
    out = ai_sheet.fill_exploded(sid, USER, lintas_model=True)
    assert out["found"] is False
    assert "part number" in out["error"].lower() and "nama" in out["error"].lower()


def test_sheet_orang_lain_tak_bisa_dibaca(sheet):
    out = ai_sheet.fill_exploded(sheet, {"username": "orang_lain", "role": "staf"},
                                 rangka=FRAME)
    assert out["found"] is False


# ── Builder: gambar & alasan jujur saat kartu diunduh ───────────────────────

def _bangun(sheet_id: str, **kw) -> tuple:
    out = ai_sheet.fill_exploded(sheet_id, USER, **kw)
    return ai_export.sheet_exploded_excel(ai_export._stash[out["export_id"]]["builder"])


def test_builder_per_vin_menempel_gambar_dan_menulis_balon(sheet, epc_stub):
    data, err = _bangun(sheet, rangka=FRAME)
    assert err == "" and data
    ws = openpyxl.load_workbook(io.BytesIO(data))["Data"]
    assert len(ws._images) == 1                        # hanya PN yang punya figure
    kol = {c.value: c.column for c in ws[4]}           # baris header (judul 3 baris)
    info = [ws.cell(row=r, column=kol[ai_sheet._KOL_EXPLODED_INFO]).value
            for r in (5, 6)]
    assert "per-VIN" in info[0] and FRAME in info[0]
    assert "Balon: 7" in info[0] and "Front axle assembly" in info[0]
    # PN tanpa figure: ALASAN apa adanya, bukan figure karangan.
    assert "tidak ditemukan" in info[1].lower()
    assert "Balon" not in info[1]


def test_builder_lintas_model_menyertakan_peringatan_di_sel(sheet, monkeypatch):
    monkeypatch.setattr(exploded_view, "figure_untuk_pn", lambda pn, **kw: (
        {"found": True, "part_number": pn, "figure_nama": "Axle", "balon": 7,
         "catatan": exploded_view.catatan_lintas_model("HOWO T7H"), "png": PNG}
        if pn == PN_ADA else
        {"found": False, "part_number": pn, "alasan": "Figure tidak ditemukan di EPC."}))
    data, err = _bangun(sheet, lintas_model=True)
    assert err == "" and data
    ws = openpyxl.load_workbook(io.BytesIO(data))["Data"]
    kol = {c.value: c.column for c in ws[4]}
    sel = ws.cell(row=5, column=kol[ai_sheet._KOL_EXPLODED_INFO]).value
    assert "lintas-model" in sel and "HOWO T7H" in sel
    assert "per-VIN" not in sel


def test_builder_menolak_saat_batch_lain_berjalan(sheet, epc_stub):
    out = ai_sheet.fill_exploded(sheet, USER, rangka=FRAME)
    b = ai_export._stash[out["export_id"]]["builder"]
    assert exploded_view._gerbang.acquire(blocking=False)
    try:
        data, err = ai_export.sheet_exploded_excel(b)
    finally:
        exploded_view._gerbang.release()
    assert data is None and "satu sekaligus" in err     # bukan file setengah jadi


def test_cache_figure_hemat_panggilan_tapi_kegagalan_dicoba_lagi(sheet, epc_stub):
    """Figure yang SUDAH didapat dilayani dari cache (dibagi dengan halaman detail
    part & Batch Download). Kegagalan sengaja TIDAK di-cache: bisa cuma EPC ngadat."""
    b = ai_export._stash[ai_sheet.fill_exploded(sheet, USER, rangka=FRAME)["export_id"]]["builder"]
    ai_export.sheet_exploded_excel(b)
    n_item = sum(1 for u, _p in epc_stub if u == epc_bom._ATLAS_ITEM_URL)
    n_rev_gagal = sum(1 for u, p in epc_stub
                      if u == epc_bom._REVERSE_URL and p.get("k") == PN_TAK)
    ai_export.sheet_exploded_excel(b)
    assert sum(1 for u, _p in epc_stub if u == epc_bom._ATLAS_ITEM_URL) == n_item
    assert sum(1 for u, p in epc_stub
               if u == epc_bom._REVERSE_URL and p.get("k") == PN_TAK) == n_rev_gagal + 1
    assert exploded_view.kunci(PN_ADA, FRAME) in exploded_view.CACHE
    assert exploded_view.kunci(PN_ADA) not in exploded_view.CACHE   # kunci tak bertabrakan


def test_kartu_unduh_membangun_lewat_generic_excel(sheet, epc_stub):
    """Tanpa cabang 'sheet_exploded' di generic_excel, klik kartu unduh hanya
    memberi 'Jenis export tidak dikenal'."""
    out = ai_sheet.fill_exploded(sheet, USER, rangka=FRAME)
    data, fname = ai_export.generic_excel(out["export_id"])
    assert data and fname == out["filename"]
    assert openpyxl.load_workbook(io.BytesIO(data))["Data"]._images


# ── Terdaftar di asisten (kalau lupa, tool/kartu unduh tak pernah muncul) ────

def test_tool_terdaftar_dan_hanya_muncul_saat_ada_lampiran(sheet):
    assert A._DISPATCH["sheet_isi_exploded"] is A._t_sheet_isi_exploded
    tanpa = {f["function"]["name"] for f in A._tool_specs(USER, "")}
    dengan = {f["function"]["name"] for f in A._tool_specs(USER, sheet)}
    assert "sheet_isi_kolom" not in tanpa
    assert "sheet_isi_kolom" in dengan
    # Nama lama: tak ditawarkan lagi, tapi tetap SAH dieksekusi (alias legacy).
    assert "sheet_isi_exploded" not in dengan
    assert "sheet_isi_exploded" in A._allowed_tool_names(USER, sheet)


def test_kartu_unduh_ditangkap_chat_loop():
    """Tanpa nama ini di daftar _capture_meta, file jadi tapi kartunya tak muncul."""
    import inspect
    src = inspect.getsource(A.chat)
    assert '"sheet_isi_kolom"' in src
