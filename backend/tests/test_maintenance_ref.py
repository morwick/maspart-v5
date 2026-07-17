"""Test jadwal perawatan berkala Shantui — parser 3-template (kini di BUILDER
tools/build_maintenance.py, rombakan 2026-07-17) + service maintenance_ref
(loader tipis atas store kanonik jadwal_perawatan.json.gz).

Fokus: (1) ingest multi-file, (2) 3 TEMPLATE sheet berbeda (Spanyol X / Inggris
√ dgn jam di baris berikutnya / Excavator kolom bergeser + 'Part Name' +
'2,250h'), (3) kolom terdeteksi DINAMIS (jangkar N/P SHANTUI), (4) `_is_mark` —
sel catatan panjang TAK dihitung tanda ganti, (5) jenis alat + varian dipisah,
(6) dedup baris & sheet '中英文', (7) filter model/jenis/jam/kata-kunci +
sinonim lintas-bahasa & batas kata, (8) store hilang → kosong tanpa error.
"""
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.services import maintenance_ref
from app.services.knowledge_util import write_json_gz

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "tools"))
import build_maintenance  # noqa: E402


def _book_dozer_loader(path):
    """Template A (Spanyol, tanda X) + sheet varian 国三 + sheet bilingual 中英文."""
    wb = Workbook()
    # -- SD99 国二 (Euro II) --
    ws = wb.active
    ws.title = "SD99"
    ws.append(["HISTORIAL DE MANTENIMIENTO PREVENTIVO"])
    ws.append(["UNIDAD:Bulldozer SD99国二排放（重发）"])
    ws.append(["MARCA : SHANTUI"])
    ws.append(["MODELO : SD99国二"])
    ws.append([]); ws.append([]); ws.append([])
    ws.append(["PLAN DE MANTENIMIENTO PREVENTIVO 服务指南"])
    ws.append([" SERVICIOS 服务", "N/P SHANTUI", "N/P EQUIVALENTE",
               "N/P FLEETGUARD", "CANTIDAD", "Intervalo en horas 更换时间",
               "50", "250", "500", "1000"])
    ws.append(["MOTOR 发动机"])
    ws.append(["Filtro de aceite de motor 机油滤", "OF-1", "", "", "1", "",
               "", "X", "X", "X"])
    ws.append(["Filtro de aceite de motor 机油滤", "OF-1", "", "", "1", "",
               "", "X", "X", "X"])                                  # DUPLIKAT persis
    ws.append(["Filtro de Diesel 柴滤", "FF-1", "", "", "2", "",
               "", "", "X", "X"])
    ws.append(["SISTEMA HIDRAULICO 液压系统"])
    ws.append(["Hydraulic fluids 液压油", "HY-1", "", "", "100L", "",
               "", "", "", "X"])

    # -- SD99 国三 (Euro III) — varian TERPISAH, model dasar sama --
    ws3 = wb.create_sheet("SD99G3")
    ws3.append(["HISTORIAL DE MANTENIMIENTO PREVENTIVO"])
    ws3.append(["UNIDAD:Bulldozer SD99国三排放（重发）"])
    ws3.append(["MARCA : SHANTUI"])
    ws3.append(["MODELO : SD99国三"])
    ws3.append([]); ws3.append([]); ws3.append([])
    ws3.append(["PLAN DE MANTENIMIENTO PREVENTIVO 服务指南"])
    ws3.append([" SERVICIOS 服务", "N/P SHANTUI", "N/P EQUIVALENTE",
                "N/P FLEETGUARD", "CANTIDAD", "Intervalo en horas 更换时间",
                "50", "250", "500", "1000"])
    ws3.append(["MOTOR 发动机"])
    ws3.append(["Filtro de aceite de motor 机油滤", "OF-3", "", "", "1", "",
                "", "X", "X", "X"])

    # -- SD99 中英文 (terjemahan → HARUS di-skip krn SD99 sudah tercakup) --
    wb2 = wb.create_sheet("SD99 中英文")
    wb2.append(["Bulldozer SD99 Cummins Engine"])
    wb2.append(["Brand : SHANTUI"])
    wb2.append(["Model : SD99E/SD99S"])
    wb2.append([]); wb2.append([]); wb2.append([])
    wb2.append(["Maintenance Service Guide 服务指南"])
    wb2.append(["Service 服务", "N/P SHANTUI", "N/P EQUIVALENTE",
                "N/P FLEETGUARD", "CANTIDAD", "Intervalo en horas 更换时间",
                "50", "250", "500", "1000"])
    wb2.append(["Engine 发动机"])
    wb2.append(["Engine Oil Filter 机油滤", "OF-1", "", "", "1", "",
                "", "X", "X", "X"])
    wb.save(path)


def _book_excavator(path):
    """Template Excavator — kolom bergeser (Part Name@1, PN@2, qty@5) + '2,250h'."""
    wb = Workbook()
    ws = wb.active
    ws.title = "EX99"
    ws.append(["HISTORIAL DE MANTENIMIENTO PREVENTIVO"])
    ws.append(["UNIDAD:Excavator SE99W国三排放（WP6H）"])
    ws.append(["MARCA : SHANTUI"])
    ws.append(["MODELO : SE99W国三"])
    ws.append([]); ws.append([]); ws.append([])
    ws.append(["PLAN DE MANTENIMIENTO PREVENTIVO 服务指南"])
    ws.append([" SERVICIOS 服务", "Part Name 零件名称", "N/P SHANTUI",
               "N/P EQUIVALENTE", "N/P FLEETGUARD", "CANTIDAD",
               "Intervalo en horas 更换时间", "50", "250", "2,250h"])
    ws.append(["MOTOR 发动机"])
    ws.append(["燃油滤芯", "Fuel filter element", "EPN-1", "", "", "1", "",
               "X", "X", "X"])
    ws.append(["机油滤芯", "Oil filter element", "EPN-2", "", "", "1", "",
               "", "X", "X"])
    wb.save(path)


def _book_bulldozer_en(path):
    """Template B (Inggris) — jam di baris SETELAH header, tanda √, qty@2, +
    sel CATATAN panjang di kolom interval (harus TIDAK terhitung)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "DB99"
    ws.append(["Maintenance schedule 保养周期表"])
    ws.append(["Machine type: Bulldozer "])
    ws.append(["Make : SHANTUI"])
    ws.append(["Model : DB99"])
    ws.append([]); ws.append([]); ws.append([])
    ws.append(["Maintenance details 保养明细"])
    ws.append(["   Item  项目", "N/P SHANTUI", "Quantity"])
    ws.append(["Engine System 发动机系统", "", "", "50", "250", "500"])
    ws.append(["fuel filter 燃油滤", "BPN-1", "1", "", "√", "√"])
    ws.append(["air filter 空滤", "BPN-2", "1",
               "定期检查报警指示，风枪清理次数不超过6次 Check alarm regularly, blow max 6 times"])
    wb.save(path)


@pytest.fixture()
def dunia(tmp_path, monkeypatch):
    manuals = tmp_path / "manuals"
    manuals.mkdir()
    _book_dozer_loader(manuals / "test loader maintenance periodic table.xlsx")
    _book_excavator(manuals / "test excavator maintenance periodic table.xlsx")
    _book_bulldozer_en(manuals / "test bulldozer maintenance periodic table.xlsx")
    # File LAIN yang harus DIABAIKAN (bukan maintenance).
    (manuals / "PART FILTER SHANTUI.xlsx").write_bytes(b"dummy")

    # Alur produksi: BUILDER parse xlsx → store kanonik; service = loader tipis.
    rows = build_maintenance.build_rows(manuals)
    store = tmp_path / "jadwal_perawatan.json.gz"
    write_json_gz(store, rows)
    monkeypatch.setattr(maintenance_ref, "_DATA", store)
    return tmp_path


def test_ingest_multifile_dan_jenis(dunia):
    assert maintenance_ref.available() is True
    assert set(maintenance_ref.jenis_list()) == {"dozer", "excavator"}
    mbj = maintenance_ref.models_by_jenis()
    assert "SD99" in mbj["dozer"] and "DB99" in mbj["dozer"]
    assert mbj["excavator"] == ["SE99W"]


def test_varian_dipisah(dunia):
    varian = {(r["model"], r["varian"]) for r in maintenance_ref.search("SD99")}
    assert ("SD99", "Euro II") in varian
    assert ("SD99", "Euro III") in varian


def test_bilingual_dan_duplikat_dibuang(dunia):
    # Sheet '中英文' (model dasar SD99 sudah tercakup) → dibuang; nama Inggris
    # 'Engine Oil Filter' TAK muncul.
    names = [r["nama"] for r in maintenance_ref.search("SD99", jam=250)]
    assert not any("Engine Oil Filter" in n for n in names)
    # Baris OF-1 DUPLIKAT persis → tinggal satu.
    of1 = [r for r in maintenance_ref.search("SD99", query="oli", jam=250)
           if r["part_number"] == "OF-1" and r["varian"] == "Euro II"]
    assert len(of1) == 1


def test_excavator_kolom_dinamis_dan_koma_jam(dunia):
    # PN@2, qty@5, label '2,250h' → jam 2250; Part Name disambung ke nama.
    r = maintenance_ref.search("SE99W", jenis="excavator", jam=2250)
    pns = {x["part_number"] for x in r}
    assert pns == {"EPN-1", "EPN-2"}
    epn1 = next(x for x in r if x["part_number"] == "EPN-1")
    assert "Fuel filter element" in epn1["nama"]
    assert epn1["qty"] == "1"
    assert epn1["varian"] == "Euro III / WP6H"


def test_template_inggris_marks_dan_catatan(dunia):
    # Tanda √ terbaca; jam ada di baris setelah header.
    bpn1 = maintenance_ref.search("DB99", query="solar")
    assert bpn1 and bpn1[0]["part_number"] == "BPN-1"
    assert bpn1[0]["ganti_jam"] == [250, 500]
    # air filter: sel interval berisi CATATAN panjang → BUKAN tanda ganti.
    air = [r for r in maintenance_ref.search("DB99") if r["part_number"] == "BPN-2"]
    assert air and air[0]["ganti_jam"] == []


def test_filter_jenis_dan_jam(dunia):
    exc = maintenance_ref.search(jenis="excavator")
    assert exc and all(r["jenis"] == "excavator" for r in exc)
    pn500 = {r["part_number"] for r in maintenance_ref.search("SD99", jam=500)}
    assert pn500 == {"OF-1", "OF-3", "FF-1"}


def test_sinonim_lintas_bahasa_dan_batas_kata(dunia):
    # 'solar' → 柴滤/fuel: cocok FF-1 (dozer) & BPN-1 (loader-EN).
    assert "FF-1" in {r["part_number"] for r in maintenance_ref.search("SD99", query="solar")}
    # 'hidrolik' mengandung substring 'oli' — TIDAK boleh menyeret item oli mesin.
    hid = {r["part_number"] for r in maintenance_ref.search("SD99", query="hidrolik")}
    assert hid == {"HY-1"}


def test_part_filter_diabaikan(dunia):
    # File 'PART FILTER SHANTUI.xlsx' bukan sumber jadwal perawatan.
    assert all("FILTER" not in (r.get("nama") or "").upper() or r["part_number"]
               for r in maintenance_ref.search())  # tak crash; tak ada baris dummy
    assert not any(r["part_number"] == "dummy" for r in maintenance_ref.search())


def test_file_hilang_kosong(tmp_path, monkeypatch):
    # Store belum dibangun/hilang → service kosong tanpa error.
    monkeypatch.setattr(maintenance_ref, "_DATA", tmp_path / "tidak_ada.json.gz")
    assert maintenance_ref.available() is False
    assert maintenance_ref.models() == []
    assert maintenance_ref.search("SD99") == []


def test_builder_fail_loud_bila_anchor_hilang(tmp_path):
    # File maintenance TANPA sel 'N/P SHANTUI' → build_rows harus SystemExit
    # (kesalahan parsing ketahuan saat build, bukan diam-diam di produksi).
    manuals = tmp_path / "manuals"
    manuals.mkdir()
    wb = Workbook()
    ws = wb.active
    ws.append(["Tabel tanpa anchor"])
    wb.save(manuals / "rusak maintenance periodic table.xlsx")
    with pytest.raises(SystemExit):
        build_maintenance.build_rows(manuals)
