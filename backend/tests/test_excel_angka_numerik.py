"""Sel Excel HARUS numerik agar rumus user jalan (keluhan pemilik 2026-07-20).

Sebelumnya kolom Harga berisi teks "Rp 1.500.000" sehingga SUM() = 0. Test ini
MEMBUKA workbook hasil dengan openpyxl dan memeriksa TIPE selnya — test lama
hanya memeriksa payload, itu sebabnya regresi ini tak pernah tertangkap.
"""
import io

import pytest
from openpyxl import load_workbook

from app.services import ai_export as X


def _bangun(kolom, baris, judul="Uji"):
    """payload → bytes .xlsx → worksheet, lewat jalur produksi sungguhan."""
    eid, _ = X.stash_export(judul, kolom, baris)
    data, _fname = X.generic_excel(eid)
    assert data, "generic_excel gagal membangun file"
    return load_workbook(io.BytesIO(data)).active


# ── Fase 0: crash len() ──────────────────────────────────────────────
def test_lebar_kolom_dengan_sel_int_tidak_crash():
    """len(int) melempar TypeError — dulu ini menjatuhkan seluruh unduhan."""
    ws = _bangun(["No", "Harga"], [["1", 1500000]])
    assert ws is not None


# ── Fase 1: koersi sadar-kolom ───────────────────────────────────────
def _sel(ws, header, baris_ke=0):
    """Cari kolom lewat headernya lalu ambil sel data ke-n."""
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value == header:
                return ws.cell(row=r + 1 + baris_ke, column=c)
    raise AssertionError(f"header {header!r} tak ditemukan")


def test_kolom_harga_jadi_int():
    ws = _bangun(["Part Number", "Harga"], [["WG9925520270", "Rp 1.500.000"]])
    c = _sel(ws, "Harga")
    assert isinstance(c.value, int) and c.value == 1500000
    assert c.data_type == "n"


def test_sel_angka_pakai_format_pemisah_ribuan():
    """Tampilan bertitik (pemisah ribuan) TANPA mengubah nilai — SUM tetap jalan
    karena selnya tetap numerik. Kolom NON-uang → format polos '#,##0'."""
    ws = _bangun(["Qty"], [["1.500"]])
    c = _sel(ws, "Qty")
    assert c.value == 1500 and c.data_type == "n"      # nilai tetap angka murni
    assert c.number_format == "#,##0"                  # tampilan bertitik polos
    # Kolom uang kini bersimbol Rp (nilai tetap int → SUM jalan).
    ws2 = _bangun(["Harga"], [["Rp 1.500.000"]])
    ch = _sel(ws2, "Harga")
    assert ch.value == 1500000 and ch.number_format == '"Rp"#,##0'


def test_sel_teks_tidak_kena_number_format():
    ws = _bangun(["Nama Part"], [["Spring bracket 12"]])
    assert _sel(ws, "Nama Part").number_format == "General"


def test_pn_murni_angka_tetap_teks():
    """PN 10 digit tak boleh jadi angka — hilang makna & tampil notasi ilmiah."""
    ws = _bangun(["Part Number", "Qty"], [["1013133963", "2"]])
    pn = _sel(ws, "Part Number")
    assert isinstance(pn.value, str) and pn.value == "1013133963"
    assert isinstance(_sel(ws, "Qty").value, int)


def test_angka_panjang_di_kolom_harga_tetap_dikoersi():
    ws = _bangun(["Harga"], [["1200000000"]])
    assert _sel(ws, "Harga").value == 1200000000


def test_kolom_status_dan_nama_tidak_dikoersi():
    ws = _bangun(["Status", "Nama Part"], [["READY", "Spring bracket 12"]])
    assert isinstance(_sel(ws, "Status").value, str)
    assert isinstance(_sel(ws, "Nama Part").value, str)


def test_stok_per_gudang_tetap_teks():
    """Headernya mengandung 'stok' tapi isinya naratif — deny harus menang."""
    ws = _bangun(["Stok Total", "Stok per Gudang"],
                 [["3", "01.Jakarta: 2 · 02.Pekanbaru: 1"]])
    assert _sel(ws, "Stok Total").value == 3
    assert isinstance(_sel(ws, "Stok per Gudang").value, str)


def test_penanda_kosong_tetap_teks():
    """'—' jangan jadi 0 — SUM akan berbohong."""
    ws = _bangun(["Harga", "Stok"], [["—", "N/A"]])
    assert _sel(ws, "Harga").value == "—"
    assert _sel(ws, "Stok").value == "N/A"


def test_desimal_koma_jadi_float():
    ws = _bangun(["Berat (kg)"], [["1,25"]])
    v = _sel(ws, "Berat (kg)").value
    assert isinstance(v, float) and abs(v - 1.25) < 1e-9


def test_ribuan_plus_desimal():
    assert X.ke_angka("1.500,25") == 1500.25


def test_nilai_ambigu_tetap_teks():
    for v in ("1.50", "12 pcs", "85%", "(1.500)", "3 - 5", "0012"):
        assert isinstance(X.ke_angka(v, uang=True), str), v


def test_koersi_idempoten_dan_bool_aman():
    assert X.ke_angka(1500000) == 1500000
    assert X.ke_angka(1.25) == 1.25
    assert X.ke_angka(True) is True          # bool jangan jadi 1
    assert X.ke_angka(None) is None


def test_negatif_dikoersi():
    assert X.ke_angka("-500") == -500


def test_formula_injection_tetap_diescape():
    """Koersi tak boleh melubangi guard anti formula-injection."""
    ws = _bangun(["Harga"], [["=SUM(A1:A9)"]])
    v = _sel(ws, "Harga").value
    assert isinstance(v, str) and v.startswith("'")


def test_kolom_angka_deny_allow():
    peta = X.kolom_angka(["Part Number", "Nama Part", "Harga", "Stok Total",
                          "Stok per Gudang", "Qty", "Status", "No"])
    assert 0 not in peta and 1 not in peta      # PN, Nama
    assert peta.get(2) is True                  # Harga = kolom uang
    assert peta.get(3) is False                 # Stok = angka non-uang
    assert 4 not in peta and 6 not in peta      # per gudang, status
    assert peta.get(5) is False                 # Qty
    assert 7 not in peta                        # "No" bukan kolom angka


def test_sheet_status_excel_harga_int_ringkasan_tetap_rp():
    """Sel tabel numerik, TAPI blok RINGKASAN tetap 'Rp …' (dibaca manusia)."""
    data, _f = X.sheet_status_excel({
        "judul": "Uji", "kolom": ["Part Number", "Harga"],
        "baris": [["WG9925520270", "Rp 1.500.000"]],
        "status": [""],
        "ringkasan": [["Subtotal", "Rp 1.500.000"]],
    })
    ws = load_workbook(io.BytesIO(data)).active
    assert _sel(ws, "Harga").value == 1500000
    teks = [ws.cell(row=r, column=c).value
            for r in range(1, ws.max_row + 1) for c in range(1, ws.max_column + 1)]
    assert any(isinstance(v, str) and v.startswith("Rp ") for v in teks), \
        "blok RINGKASAN harus tetap 'Rp …'"


# ── Format angka PINTAR per-kolom (2026-07-22) ───────────────────────
def test_num_format_per_jenis_kolom():
    """Format tampilan mengikuti jenis kolom dari nama header."""
    assert X.num_format("% Terpakai") == '0.0"%"'
    assert X.num_format("persen_terpakai") == '0.0"%"'
    assert X.num_format("Harga") == '"Rp"#,##0'
    assert X.num_format("Total Harga") == '"Rp"#,##0'
    assert X.num_format("Harga (CNY)") == '#,##0.00" CNY"'
    assert X.num_format("total_cny") == '#,##0.00" CNY"'
    assert X.num_format("Nilai USD") == '"$"#,##0.00'
    assert X.num_format("Berat") == '#,##0" kg"'
    assert X.num_format("Berat (kg)") == '#,##0" kg"'
    # non-uang/non-persen → default polos
    assert X.num_format("Qty") == "#,##0"
    assert X.num_format("Stok Total") == "#,##0"
    assert X.num_format("KM") == "#,##0"
    assert X.num_format("Sisa Hari") == "#,##0"


def test_generic_excel_number_format_terpasang_per_kolom():
    """Workbook nyata: number_format tiap kolom sesuai jenis, nilai tetap angka."""
    ws = _bangun(["No", "Harga", "% Terpakai", "Berat", "Qty"],
                 [["1", "Rp 720.000", "87,5", "1.301", "5"]])
    assert _sel(ws, "Harga").number_format == '"Rp"#,##0'
    assert _sel(ws, "Harga").value == 720000
    p = _sel(ws, "% Terpakai")
    assert p.number_format == '0.0"%"'
    assert p.value == 87.5                     # literal 87.5 (BUKAN 0.875/8750)
    assert _sel(ws, "Berat").number_format == '#,##0" kg"'
    assert _sel(ws, "Qty").number_format == "#,##0"


def test_persen_tidak_dikali_seratus():
    """Regresi: format '0.0%' Excel mengali 100 — kita pakai literal '%'."""
    ws = _bangun(["% Terpakai"], [[87.5]])
    c = _sel(ws, "% Terpakai")
    assert c.value == 87.5 and '%' in c.number_format and c.number_format != "0.0%"


def test_sheet_status_excel_pakai_format_pintar():
    data, _f = X.sheet_status_excel({
        "judul": "Uji", "kolom": ["Part Number", "Harga", "% Terpakai"],
        "baris": [["WG9925520270", "Rp 1.500.000", "50"]],
        "status": [""], "ringkasan": [],
    })
    ws = load_workbook(io.BytesIO(data)).active
    assert _sel(ws, "Harga").number_format == '"Rp"#,##0'
    assert _sel(ws, "% Terpakai").number_format == '0.0"%"'
