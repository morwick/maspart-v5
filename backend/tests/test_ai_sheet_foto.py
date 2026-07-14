"""Tool asisten `sheet_isi_foto`: tempel foto SIMS ke Excel unggahan user.

Aturan yang dijaga di sini:
  • Foto dicocokkan lewat PART NUMBER, TIDAK PERNAH lewat nama part. Pencarian nama
    di SIMS bersifat 'mengandung kata' → mengembalikan part LAIN (nama 'Radiator'
    memunculkan PIPA radiator). Tanpa kolom PN, tool HARUS menolak — bukan menebak.
  • PN tanpa foto di SIMS ditandai '-' (tidak dikarang).
  • Foto berat → hanya URL yang disimpan saat tool jalan; gambar diunduh & ditempel
    SAAT KARTU DIUNDUH (builder 'sheet_foto'), supaya RAM server tak jebol.
"""
import io

import openpyxl
import pytest

from app.services import ai_export, ai_sheet
from app.services import ai_assistant as A

USER = {"username": "budi", "role": "pembeli"}

FOTO = {
    "FG9804532170": ["http://sims/radiator-1.jpg", "http://sims/radiator-2.jpg", "http://sims/radiator-3.jpg"],
    "FG9806130202": ["http://sims/ac-1.jpg"],
    "SPHG0000000111": [],                       # ada di file, tak punya foto di SIMS
}


@pytest.fixture
def sheet(monkeypatch):
    """Sheet unggahan tiruan: kolom Part No. + Nama."""
    monkeypatch.setattr(ai_sheet.sims, "available", lambda: True)
    monkeypatch.setattr(ai_sheet.sims, "get_images", lambda pn: list(FOTO.get(pn, [])))
    parsed = {
        "filename": "QUOTATION.xlsx",
        "headers": ["Part No.", "Nama"],
        "roles": ["part_number", "part_name"],
        "_body": [["FG9804532170", "Radiator"],
                  ["FG9806130202", "AC compressor"],
                  ["SPHG0000000111", "Wheel bolt"]],
        "contoh": [], "sheet": "Sheet1", "sheet_lain": [], "pn_dikenal": 3,
        "jumlah_baris": 3, "jumlah_kolom": 2, "kolom_pn": "Part No.", "terpotong": False,
    }
    return ai_sheet.put_sheet("budi", parsed)


def test_isi_foto_2_per_part_dan_tandai_yang_kosong(sheet):
    out = ai_sheet.fill_photos(sheet, USER, jumlah=2)
    assert out["found"] is True
    assert out["baris_berfoto"] == 2 and out["baris_tanpa_foto"] == 1
    assert out["foto_per_part"] == 2
    assert out["export_id"] and out["filename"].endswith(".xlsx")

    b = ai_export._stash[out["export_id"]]["builder"]
    assert b["kind"] == "sheet_foto"
    assert b["kolom"][-2:] == ["Foto 1", "Foto 2"]          # 2 kolom foto di ujung
    assert b["foto"][0] == FOTO["FG9804532170"][:2]         # dipotong ke 2, bukan 3
    assert b["foto"][2] == []                               # tak ada foto
    # Baris tanpa foto ditandai em-dash; sel yang akan diisi gambar dibiarkan kosong.
    D = ai_sheet._TANPA_FOTO
    assert b["baris"][2][-2:] == [D, D]
    assert b["baris"][0][-2:] == ["", ""]
    assert b["baris"][1][-2:] == ["", D]                    # cuma 1 foto → slot ke-2 ditandai


def test_tanpa_kolom_pn_ditolak_bukan_ditebak_dari_nama(monkeypatch):
    """⛔ Inti keamanan fitur: file tanpa Part Number TIDAK boleh dicocokkan lewat
    nama part (foto bisa milik part lain). Tool wajib menolak & minta kolom PN."""
    monkeypatch.setattr(ai_sheet.sims, "available", lambda: True)
    monkeypatch.setattr(ai_sheet.sims, "get_images",
                        lambda pn: pytest.fail("SIMS tak boleh dipanggil tanpa kolom PN"))
    sid = ai_sheet.put_sheet("budi", {
        "filename": "tanpa-pn.xlsx", "headers": ["Nama", "Qty"],
        "roles": ["part_name", "qty"],
        "_body": [["Radiator", "1"], ["Oil filter", "2"]],
        "contoh": [], "sheet": "S", "sheet_lain": [], "pn_dikenal": 0,
        "jumlah_baris": 2, "jumlah_kolom": 2, "kolom_pn": None, "terpotong": False,
    })
    out = ai_sheet.fill_photos(sid, USER)
    assert out["found"] is False
    assert "part number" in out["error"].lower()
    assert "nama" in out["error"].lower()          # alasannya dijelaskan ke model


def test_sheet_orang_lain_tak_bisa_dibaca(sheet):
    out = ai_sheet.fill_photos(sheet, {"username": "orang_lain", "role": "pembeli"})
    assert out["found"] is False


def test_plafon_pn_menolak_file_raksasa(monkeypatch):
    monkeypatch.setattr(ai_sheet.sims, "available", lambda: True)
    monkeypatch.setattr(ai_sheet.sims, "get_images", lambda pn: [])
    n = ai_sheet._MAX_FOTO_PN + 1
    sid = ai_sheet.put_sheet("budi", {
        "filename": "besar.xlsx", "headers": ["Part No."], "roles": ["part_number"],
        "_body": [[f"PN{i:06d}"] for i in range(n)],
        "contoh": [], "sheet": "S", "sheet_lain": [], "pn_dikenal": 0,
        "jumlah_baris": n, "jumlah_kolom": 1, "kolom_pn": "Part No.", "terpotong": False,
    })
    out = ai_sheet.fill_photos(sid, USER)
    assert out["found"] is False and "terlalu banyak" in out["error"].lower()


# ── Builder: gambar benar-benar tertanam saat kartu diunduh ──────────────────
def test_builder_menempel_gambar_dan_menciutkannya(monkeypatch):
    """Foto SIMS bisa 6000 px/17 MB → builder wajib memperkecil sebelum menempel.
    Jaringan diganti: `requests.get` dibuat mengembalikan JPEG raksasa, sehingga
    _foto_thumb yang ASLI (termasuk logika resize-nya) yang diuji."""
    from PIL import Image as PILImage

    buf = io.BytesIO()
    PILImage.new("RGB", (3000, 2000), "red").save(buf, format="JPEG")
    besar = buf.getvalue()

    class _Resp:
        status_code = 200
        content = besar

    import requests
    monkeypatch.setattr(requests, "get", lambda url, timeout=0: _Resp())

    D = ai_sheet._TANPA_FOTO
    data, err = ai_export.sheet_foto_excel({
        "judul": "QUOTATION + Foto",
        "kolom": ["Part No.", "Nama", "Foto 1", "Foto 2"],
        "baris": [["FG9804532170", "Radiator", "", ""],
                  ["SPHG0000000111", "Wheel bolt", D, D]],
        "foto": [["http://sims/a.jpg", "http://sims/b.jpg"], []],
        "kol_foto": [2, 3],
    })
    assert err == "" and data
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Data"]
    assert len(ws._images) == 2                       # 2 foto baris pertama
    for im in ws._images:
        assert im.height <= ai_export._FOTO_H_PX      # diciutkan, bukan 2000 px
    # Penanda "tak ada foto" ditulis APA ADANYA (tak di-escape jadi "'-").
    vals = [c.value for c in ws[ws.max_row]]
    assert vals[-2:] == [D, D]


# ── Terdaftar di asisten (kalau lupa, kartu unduh tak muncul) ────────────────
def test_tool_terdaftar_dan_hanya_muncul_saat_ada_lampiran():
    assert "sheet_isi_foto" in A._DISPATCH
    tanpa = {f["function"]["name"] for f in A._tool_specs(USER, "")}
    dengan = {f["function"]["name"] for f in A._tool_specs(USER, "sid-123")}
    assert "sheet_isi_foto" not in tanpa       # tanpa lampiran: model tak melihatnya
    assert "sheet_isi_foto" in dengan


def test_pembeli_boleh_pakai(sheet):
    """Foto bukan data sensitif (beda dari harga/stok gudang) → pembeli boleh."""
    assert "sheet_isi_foto" in A._allowed_tool_names(USER, sheet)
