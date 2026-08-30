"""Katalog LENGKAP = SATU Excel dengan SATU SHEET PER KATEGORI (+ sheet Daftar
Isi), dan `katalog_kategori` menerima BANYAK kategori dalam satu panggilan.

Kenapa ada (audit ai_chat_log 2026-08-29, sesi user 'mas'): "pisahkan katalog
per bagian" → 9 panggilan katalog_kategori (plafon tool berat 4 menolak sisanya,
dua giliran "lanjut" tetap 0 file baru), lalu TIGA kali minta "1 Excel, sheet per
kategori" — model dua kali menjawab "sudah jadi" (file = 1 sheet 350 seksi) dan
ketiga kali MENGKLAIM susunan sheet yang tidak ada.

Sifat yang dijaga:
 1. Kategori tunggal → satu sheet "Katalog" persis bentuk lama.
 2. Lengkap → sheet "Daftar Isi" ber-hyperlink + satu sheet per kategori; tak ada
    part yang hilang; nama sheet sah (≤31 char, tanpa []:*?/\\) dan unik.
 3. Hasil tool menyebut susunan file ('isi_file') agar model tak mengarang.
 4. Array kategori → kartu unduh PER kategori tetap sampai ke klien
    (ringkasan_per_item dibaca _capture_meta), termasuk bendera sedang_disusun.
"""
import io
import json

import openpyxl
import pytest

from app.services import ai_assistant as ai, ai_export, epc_bom, part_index

ADMIN = {"username": "admin", "role": "admin"}
VIN = "LZZ5EXSF9RJ380449"


def _fig(kat, nama, kode, pns):
    return {"kategori": kat, "nama": nama, "kode": kode, "kode_kategori": "",
            "svg": "", "svg_lain": [],
            "items": [{"balon": i + 1, "pn": p, "nama": "part " + p, "nama_cn": "",
                       "qty": 1, "pengganti": []} for i, p in enumerate(pns)]}


FIGS = [
    _fig("Cab", "Door", "ZZ01", ["A1", "A2"]),
    _fig("Cab", "Seat", "ZZ02", ["A3"]),
    _fig("Engine", "Piston", "ZZ10", ["B1", "B2", "B3"]),
    _fig("Brake/Axle: Front", "Drum", "ZZ20", ["C1"]),
]


def _walk(lengkap):
    return lambda rangka, kategori: {"found": True, "frame_number": "RJ380449",
                                     "lengkap": lengkap, "figures": FIGS, "incomplete": False}


@pytest.fixture(autouse=True)
def _tanpa_render(monkeypatch):
    monkeypatch.setattr(ai_export, "_svg_to_png", lambda svg, width=0: None)
    monkeypatch.setattr(part_index, "search_exact_pns", lambda pns: [])


def _pn_di_sheet(ws):
    return {c.value for row in ws.iter_rows(min_col=2, max_col=2) for c in row
            if isinstance(c.value, str) and len(c.value) == 2 and c.value[1].isdigit()}


# ── 1. bentuk lama utk kategori tunggal ─────────────────────────────────────

def test_kategori_tunggal_satu_sheet_katalog(monkeypatch):
    monkeypatch.setattr(epc_bom, "catalog_walk", _walk(False))
    data, fn = ai_export.katalog_excel(VIN, "kabin")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Katalog"]
    ws = wb["Katalog"]
    assert _pn_di_sheet(ws) == {"A1", "A2", "A3", "B1", "B2", "B3", "C1"}
    # link "↑ Daftar Isi" menunjuk sheet ini sendiri (nama dikutip)
    links = [c.hyperlink.location for row in ws.iter_rows(min_col=7, max_col=7)
             for c in row if c.hyperlink]
    assert links and all(l == "'Katalog'!A5" for l in links)
    assert fn == "Katalog_kabin_RJ380449.xlsx"


# ── 2. lengkap → sheet per kategori ─────────────────────────────────────────

def test_lengkap_sheet_per_kategori(monkeypatch):
    monkeypatch.setattr(epc_bom, "catalog_walk", _walk(True))
    data, _ = ai_export.katalog_excel(VIN, "semua")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Daftar Isi", "Cab", "Engine", "Brake Axle Front"]
    # tak ada part hilang; tiap kategori hanya memuat part-nya
    assert _pn_di_sheet(wb["Cab"]) == {"A1", "A2", "A3"}
    assert _pn_di_sheet(wb["Engine"]) == {"B1", "B2", "B3"}
    assert _pn_di_sheet(wb["Brake Axle Front"]) == {"C1"}
    # Daftar Isi: hyperlink ke tiap sheet (nama dikutip), jumlah figure/part benar
    di = wb["Daftar Isi"]
    assert di["B5"].hyperlink.location == "'Cab'!A1" and di["B5"].value == "Cab"
    assert di["B7"].hyperlink.location == "'Brake Axle Front'!A1"
    assert (di["C5"].value, di["D5"].value) == (2, 3)
    # sheet kategori: link kembali ke Daftar Isi + link internal pakai nama sheet-nya
    eng = wb["Engine"]
    assert eng["G3"].hyperlink.location == "'Daftar Isi'!A1"
    links = [c.hyperlink.location for row in eng.iter_rows(min_col=7, max_col=7)
             for c in row if c.hyperlink and c.row > 3]
    assert links and all(l == "'Engine'!A5" for l in links)


def test_lengkap_satu_kelompok_tetap_bentuk_lama(monkeypatch):
    figs = [f for f in FIGS if f["kategori"] == "Cab"]
    monkeypatch.setattr(epc_bom, "catalog_walk",
                        lambda r, k: {"found": True, "frame_number": "RJ380449",
                                      "lengkap": True, "figures": figs, "incomplete": False})
    data, _ = ai_export.katalog_excel(VIN, "lengkap")
    assert openpyxl.load_workbook(io.BytesIO(data)).sheetnames == ["Katalog"]


def test_nama_sheet_sah_dan_unik():
    m = ai_export._nama_sheet_unik(["Cab", "Cab", "x" * 40, "Daftar Isi", "A[1]:B*?"],
                                   terpakai={"Daftar Isi"})
    assert m["Cab"] == "Cab" and len(m) == 4
    assert m["x" * 40] == "x" * 31
    assert m["Daftar Isi"] == "Daftar Isi (2)"
    assert m["A[1]:B*?"] == "A 1 B"
    # tabrakan: dua kunci sama → dict cuma satu; uji lewat daftar berbeda huruf besar
    m2 = ai_export._nama_sheet_unik(["Rem", "REM"])
    assert m2 == {"Rem": "Rem", "REM": "REM (2)"}


# ── 3. hasil tool menyebut susunan file ─────────────────────────────────────

def test_isi_file_disebut_di_hasil_tool(monkeypatch):
    monkeypatch.setattr(epc_bom, "catalog_walk", lambda r, k: {
        "found": True, "frame_number": "RJ380449", "kategori_kode": None, "lengkap": True,
        "jumlah_figure": 4, "jumlah_part": 7, "kategori_cocok": ["Cab"], "figures": FIGS,
        "incomplete": False})
    monkeypatch.setattr(ai_export, "stash_builder", lambda j, b, ext="xlsx": ("EID", "f." + ext))
    r = ai._t_katalog_kategori({"rangka": VIN, "kategori": "semua", "format": "excel"}, ADMIN)
    assert "SATU SHEET PER KATEGORI" in r["isi_file"] and "isi_file" in r["catatan"]
    r = ai._t_katalog_kategori({"rangka": VIN, "kategori": "kabin", "format": "pdf"}, ADMIN)
    assert r["isi_file"].startswith("PDF")


# ── 4. banyak kategori satu panggilan → kartu per kategori ──────────────────

def test_array_kategori_satu_panggilan(monkeypatch):
    monkeypatch.setattr(epc_bom, "catalog_walk", lambda r, k: {
        "found": True, "frame_number": "RJ380449", "kategori_kode": "01", "lengkap": False,
        "jumlah_figure": 1, "jumlah_part": 2, "kategori_cocok": [k], "figures": FIGS[:1],
        "incomplete": False})
    n = {"i": 0}

    def stash(judul, builder, ext="xlsx"):
        n["i"] += 1
        return f"E{n['i']}", f"{judul}.{ext}"

    monkeypatch.setattr(ai_export, "stash_builder", stash)
    r = ai._t_katalog_kategori({"rangka": VIN, "format": "excel",
                                "kategori": ["kabin", "rem", "sasis", "mesin", "ac"]}, ADMIN)
    assert r["jumlah_item"] == 4                       # plafon 4, sisanya dilaporkan
    ids = sorted(it["export_id"] for it in r["ringkasan_per_item"])
    assert ids == ["E1", "E2", "E3", "E4"]
    assert "ac" in r["catatan_batch"]


def test_kartu_unduh_dari_hasil_batch_sampai_ke_klien(monkeypatch):
    """_capture_meta membaca ringkasan_per_item + meneruskan sedang_disusun."""
    monkeypatch.setattr(ai, "_system_prompt", lambda user: "system uji")
    monkeypatch.setattr(ai, "_tool_specs", lambda user, sheet_id="": [{"x": 1}])
    monkeypatch.setattr(ai, "_unit_name_tokens", lambda: set())
    batch = {"jumlah_item": 2, "kategori_cocok": [], "ringkasan_per_item": [
        {"kategori": "kabin", "found": True, "export_id": "K1", "filename": "Katalog_Kabin.xlsx",
         "judul": "Katalog Kabin RJ380449", "jumlah_baris": 5},
        {"kategori": "rem", "found": True, "export_id": "K2", "filename": "Katalog_Rem.xlsx",
         "judul": "Katalog Rem RJ380449", "sedang_disusun": True},
    ]}
    monkeypatch.setattr(ai, "_run_tool", lambda name, args, u, sid="": batch)
    seq = iter([
        {"choices": [{"message": {"content": None, "tool_calls": [
            {"id": "1", "type": "function", "function": {
                "name": "katalog_kategori",
                "arguments": json.dumps({"rangka": VIN, "kategori": ["kabin", "rem"],
                                         "format": "excel"})}}]},
            "finish_reason": "tool_calls"}]},
        {"choices": [{"message": {"content": "Dua katalog siap."}, "finish_reason": "stop"}]},
    ])
    monkeypatch.setattr(ai, "_post_chat", lambda m, t, max_tokens=6000: next(seq))
    out = ai.chat(ADMIN, [{"role": "user", "content": f"katalog kabin & rem {VIN} excel"}])
    kartu = out["excel_exports"]
    assert [k["id"] for k in kartu] == ["K1", "K2"]
    assert "sedang_disusun" not in kartu[0] and kartu[1]["sedang_disusun"] is True
