"""⛔ Aturan pemilik 2026-08-25: Excel yang dikirim user JANGAN diubah formatnya —
asisten hanya MENGISI data (atau membacanya).

Sebelum ini asisten membangun ULANG file dari `headers`+`baris`, jadi yang diunduh
user bukan filenya lagi: kop/judul di atas header hilang, baris kosong hilang,
rumus jadi angka mati, warna & lebar kolom diganti tema MASPART, sheet lain lenyap,
dan kolom di luar 40 kolom pertama terbuang. Test ini membangun file "punya user"
yang kaya format, menjalankan pengisian sungguhan, lalu memeriksa file hasilnya
sel per sel.
"""
import io

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.services import ai_export, ai_sheet

USER = {"username": "budi", "role": "admin"}

_JUDUL = "PT CONTOH JAYA — Daftar Permintaan Part"


def _file_user(harga_baris1=150000, rumus_harga=False) -> bytes:
    """Excel gaya lapangan: kop bergaya, baris kosong, header di baris 3, rumus,
    format angka sendiri, kolom TANPA header, dan sheet kedua."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Permintaan"
    ws.merge_cells("A1:D1")
    ws["A1"] = _JUDUL
    ws["A1"].font = Font(bold=True, size=16, color="FFFF0000")
    ws["A1"].fill = PatternFill("solid", fgColor="FFFFFF00")
    ws["A2"] = "Tanggal: 25 Agustus 2026"
    for j, h in enumerate(["No", "Part Number", "Nama", "Qty", "Harga", "Total"], 1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor="FF1F4E78")
    ws.cell(row=4, column=1, value=1)
    ws.cell(row=4, column=2, value="WG9925520270")
    ws.cell(row=4, column=3, value="Filter oli")
    ws.cell(row=4, column=4, value=2)
    if rumus_harga:
        ws.cell(row=4, column=5, value="=100*3")
    else:
        ws.cell(row=4, column=5, value=harga_baris1).number_format = '"Rp"#,##0'
    ws.cell(row=4, column=6, value="=D4*E4")
    ws.cell(row=4, column=7, value="catatan tangan")      # kolom G: TANPA header
    ws.cell(row=5, column=1, value=2)
    ws.cell(row=5, column=2, value="AZ9925520271")
    ws.cell(row=5, column=3, value="Filter solar")
    ws.cell(row=5, column=4, value=1)
    ws.cell(row=5, column=6, value="=D5*E5")
    # baris 6 sengaja KOSONG (pemisah), baris 7 baris TOTAL milik user
    ws.cell(row=7, column=1, value="TOTAL")
    ws.cell(row=7, column=6, value="=SUM(F4:F5)")
    ws.column_dimensions["C"].width = 42
    ws.freeze_panes = "A4"
    ws2 = wb.create_sheet("Catatan")
    ws2["A1"] = "jangan dihapus"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def dunia(monkeypatch):
    baris = {
        "WG9925520270": {"part_number": "WG9925520270", "part_name": "Spring bracket",
                         "stok": "12", "harga": "Rp 1.500.000",
                         "stok_num": 12, "harga_num": 1_500_000, "gudang": {}},
    }
    monkeypatch.setattr(ai_sheet.part_index, "rows_for_pns",
                        lambda pns: {p: baris[p] for p in pns if p in baris})
    monkeypatch.setattr(ai_sheet.part_index, "gudang_names", lambda: [])
    return baris


def _sid(data: bytes, nama: str = "permintaan.xlsx") -> str:
    p = ai_sheet.parse_upload(data, nama)
    assert p.get("ok"), p
    return ai_sheet.put_sheet(USER["username"], p)


def _hasil(export_id):
    data, _ = ai_export.generic_excel(export_id)
    assert data
    return load_workbook(io.BytesIO(data))


# ── inti aturan: file user pulang UTUH ───────────────────────────────────────
def test_isian_masuk_tanpa_mengubah_format_file_user(dunia):
    sid = _sid(_file_user())
    r = ai_sheet.fill_columns(sid, USER, permintaan=[{"isi": "stok"}])
    assert r["found"] and r["format_asli_dipertahankan"] is True

    wb = _hasil(r["export_id"])
    ws = wb["Permintaan"]
    assert wb.sheetnames == ["Permintaan", "Catatan"]        # sheet lain tetap ada
    assert wb["Catatan"]["A1"].value == "jangan dihapus"

    # kop & gayanya
    assert ws["A1"].value == _JUDUL
    assert ws["A1"].font.bold and ws["A1"].font.size == 16
    assert (ws["A1"].fill.fgColor.rgb or "").endswith("FFFF00")
    assert "A1:D1" in [str(m) for m in ws.merged_cells.ranges]
    assert ws["A2"].value == "Tanggal: 25 Agustus 2026"

    # header user apa adanya (warna & posisi baris 3)
    assert [ws.cell(row=3, column=j).value for j in range(1, 7)] == \
        ["No", "Part Number", "Nama", "Qty", "Harga", "Total"]
    assert (ws["A3"].fill.fgColor.rgb or "").endswith("1F4E78")

    # rumus, format angka, lebar kolom, freeze pane
    assert ws["F4"].value == "=D4*E4"
    assert ws["E4"].number_format == '"Rp"#,##0'
    assert ws["C"][0].parent.column_dimensions["C"].width == 42
    assert ws.freeze_panes == "A4"

    # baris kosong & baris TOTAL user tetap di tempatnya
    assert all(ws.cell(row=6, column=j).value in (None, "") for j in range(1, 7))
    assert ws["A7"].value == "TOTAL" and ws["F7"].value == "=SUM(F4:F5)"

    # kolom G (tanpa header, tak terbaca parser) tak tertimpa kolom baru
    assert ws["G4"].value == "catatan tangan"
    hdr = {ws.cell(row=3, column=j).value: j for j in range(1, ws.max_column + 1)}
    assert hdr["Stok"] >= 8                       # kolom baru DI KANAN kolom G
    assert ws.cell(row=4, column=hdr["Stok"]).value == 12
    assert ws.cell(row=5, column=hdr["Stok"]).value in (None, "")   # PN tak ketemu


def test_kolom_milik_user_diisi_di_selnya_sendiri(dunia):
    """User punya kolom 'Harga' → diisi DI SITU (bukan bikin kolom baru), dan
    number_format miliknya tak diganti."""
    sid = _sid(_file_user())
    r = ai_sheet.fill_columns(sid, USER, permintaan=[{"isi": "harga_lokal"}])
    ws = _hasil(r["export_id"])["Permintaan"]
    assert ws["E3"].value == "Harga"                 # tetap kolom E milik user
    assert ws["E4"].value == 1_500_000
    assert ws["E4"].number_format == '"Rp"#,##0'     # format user menang


def test_nilai_lama_user_tak_dikosongkan_saat_data_tak_ketemu(dunia):
    """PN tak ada di indeks → sel yang SUDAH BERISI angka user dibiarkan, bukan
    dikosongkan. Dulu jalur bangun-ulang menimpanya dengan sel kosong."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Part Number", "Harga"])
    ws.append(["TIDAKADA123", 777])
    buf = io.BytesIO()
    wb.save(buf)
    sid = _sid(buf.getvalue())
    r = ai_sheet.fill_columns(sid, USER, permintaan=[{"isi": "harga_lokal"}])
    assert r["sel_dilewati_sudah_terisi"] == 1
    assert _hasil(r["export_id"]).active["B2"].value == 777


def test_rumus_milik_user_tidak_dirusak(dunia):
    """Sel target yang berisi RUMUS dilewati — hitungan user tak boleh dihancurkan."""
    sid = _sid(_file_user(rumus_harga=True))
    r = ai_sheet.fill_columns(sid, USER, permintaan=[{"isi": "harga_lokal"}])
    assert r["sel_dilewati_rumus"] == 1
    assert _hasil(r["export_id"])["Permintaan"]["E4"].value == "=100*3"


def test_rekap_di_sheet_terpisah_bukan_di_bawah_data_user(dunia):
    """Blok RINGKASAN tak disisipkan di bawah tabel user (di sana sudah ada baris
    TOTAL miliknya) — ditaruh di sheet sendiri."""
    sid = _sid(_file_user())
    r = ai_sheet.fill_columns(sid, USER, permintaan=[{"isi": "stok"}], rekap=True)
    wb = _hasil(r["export_id"])
    assert ai_export._SHEET_REKAP in wb.sheetnames
    assert wb["Permintaan"]["A7"].value == "TOTAL"          # baris user tak tergeser
    assert wb[ai_export._SHEET_REKAP]["A1"].value == "RINGKASAN"


def test_status_hanya_mewarnai_kolom_yang_kita_tambah(dunia):
    """Warna status tak boleh mengecat ulang sel milik user; statusnya tetap
    terbaca karena kolom teks 'Status' ikut ditulis (dwi-encode)."""
    sid = _sid(_file_user())
    r = ai_sheet.fill_columns(sid, USER, permintaan=[], tandai_status=True)
    ws = _hasil(r["export_id"])["Permintaan"]
    hdr = {ws.cell(row=3, column=j).value: j for j in range(1, ws.max_column + 1)}
    assert ws.cell(row=4, column=hdr["Status"]).value == "READY"
    assert (ws.cell(row=4, column=hdr["Status"]).fill.fgColor.rgb or "").endswith("EAF6EC")
    # sel milik user (Nama part) tak dicat
    assert ws.cell(row=4, column=3).fill.fgColor.rgb in (None, "00000000")


def test_determinisme_byte_sama(dunia):
    """Dua pengisian identik → bytes identik (aman untuk cache & diff)."""
    src = _file_user()
    a = ai_sheet.fill_columns(_sid(src), USER, permintaan=[{"isi": "stok"}])
    b = ai_sheet.fill_columns(_sid(src), USER, permintaan=[{"isi": "stok"}])
    da, _ = ai_export.generic_excel(a["export_id"])
    db, _ = ai_export.generic_excel(b["export_id"])
    assert da == db


def test_properti_dokumen_user_tak_diganti(dunia):
    """`dcterms:modified` milik file user dipertahankan (dulu dipin ke 2000-01-01)."""
    src = _file_user()
    asli = ai_export._modified_asli(src)
    r = ai_sheet.fill_columns(_sid(src), USER, permintaan=[{"isi": "stok"}])
    data, _ = ai_export.generic_excel(r["export_id"])
    assert ai_export._modified_asli(data) == asli


# ── jalur cadangan tetap ada & JUJUR ─────────────────────────────────────────
def test_csv_jatuh_ke_jalur_bangun_ulang_dan_mengakui(dunia):
    """CSV tak punya format untuk dipertahankan → file dibangun seperti dulu, dan
    tool mengatakannya apa adanya (bukan diam-diam)."""
    sid = _sid(b"Part Number,Qty\nWG9925520270,2\n", "order.csv")
    r = ai_sheet.fill_columns(sid, USER, permintaan=[{"isi": "stok"}])
    assert r["found"] and r["format_asli_dipertahankan"] is False
    assert "csv" in r["alasan_format"].lower()
    assert _hasil(r["export_id"])["Data"]         # workbook bikinan builder lama


def test_header_berubah_maka_menolak_menulis_di_tempat(dunia, monkeypatch):
    """Kalau peta baris/kolom tak lagi cocok dengan file, JANGAN menulis ke sel
    yang salah — lebih baik jatuh ke jalur bangun-ulang & mengaku."""
    sid = _sid(_file_user())
    parsed = ai_sheet.get_sheet(sid, USER["username"])
    parsed["headers"] = ["No", "SALAH", "Nama", "Qty", "Harga", "Total"]
    r = ai_sheet.fill_columns(sid, USER, permintaan=[{"isi": "stok"}],
                              kolom_pn="Part Number")
    assert r["found"] and r["format_asli_dipertahankan"] is False
    assert "header" in r["alasan_format"].lower()


# ── gambar ikut ke file ASLI, bukan ke workbook baru ─────────────────────────
def test_foto_ditempel_ke_file_asli(dunia, monkeypatch):
    from PIL import Image as PILImage

    buf = io.BytesIO()
    PILImage.new("RGB", (400, 300), "red").save(buf, format="JPEG")
    isi = buf.getvalue()

    class _Resp:
        status_code = 200
        content = isi

    import requests
    monkeypatch.setattr(requests, "get", lambda url, timeout=0: _Resp())
    monkeypatch.setattr(ai_sheet.sims, "available", lambda: True)
    monkeypatch.setattr(ai_sheet.sims, "get_images",
                        lambda pn: ["http://sims/a.jpg"] if pn == "WG9925520270" else [])

    sid = _sid(_file_user())
    r = ai_sheet.fill_columns(sid, USER, permintaan=[{"isi": "stok"}],
                              gambar=["foto"], jumlah_foto=1)
    assert r["found"] and r["format_asli_dipertahankan"] is True
    ws = _hasil(r["export_id"])["Permintaan"]
    assert ws["A1"].value == _JUDUL                       # kop user tetap
    assert ws["F4"].value == "=D4*E4"                     # rumus user tetap
    assert len(ws._images) == 1                           # foto tertanam
