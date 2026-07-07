"""Tool buat_excel (export Excel generik/dinamis) — handler + builder + pagar
anti-karangan. Alur: user "buatkan excelnya" → model panggil buat_excel(judul,
kolom, baris dari hasil tool) → payload di-stash → kartu unduh di frontend →
GET /api/ai/excel/{id} membangun xlsx via ai_export.generic_excel.
"""
import io

from openpyxl import load_workbook

from app.services import ai_assistant as ai
from app.services import ai_export


def test_buat_excel_sukses_dan_file_terbangun():
    args = {"judul": "Part Air Compressor RJ345233",
            "kolom": ["No", "Part Number", "Nama Part"],
            "baris": [["1", "1013133963", "Air Compressor Assembly"],
                      ["2", "1013133966", "Air Compressor"]],
            "_grounded": {"1013133963", "1013133966"}}
    res = ai._t_buat_excel(args, {})
    assert res.get("found") and res.get("export_id") and res["jumlah_baris"] == 2
    assert res["filename"].endswith(".xlsx")

    data, fname = ai_export.generic_excel(res["export_id"])
    assert data and fname == res["filename"]
    ws = load_workbook(io.BytesIO(data)).active
    assert ws.cell(row=4, column=2).value == "Part Number"   # header (judul di baris 1)
    assert ws.cell(row=5, column=2).value == "1013133963"


def test_buat_excel_tolak_pn_karangan():
    # PN yang tak pernah muncul dari tool/riwayat → ditolak, bukan diloloskan.
    args = {"judul": "X", "kolom": ["PN"],
            "baris": [["ZZ9999999888777"]], "_grounded": set()}
    res = ai._t_buat_excel(args, {})
    assert "error" in res and "ZZ9999999888777" in res["error"]


def test_buat_excel_baris_dict_dan_padding():
    # Model kadang kirim dict per baris / baris kepanjangan — dinormalkan, tak crash.
    args = {"judul": "Y", "kolom": ["A", "B"],
            "baris": [{"A": "1"}, ["x", "y", "z"]]}
    res = ai._t_buat_excel(args, {})
    assert res["found"] and res["jumlah_baris"] == 2


def test_buat_excel_input_kosong():
    assert "error" in ai._t_buat_excel({"judul": "Z", "kolom": [], "baris": []}, {})


def test_generic_excel_id_tak_dikenal():
    data, msg = ai_export.generic_excel("tidak-ada")
    assert data is None and "kedaluwarsa" in msg.lower()


# ── stash_builder (katalog bergambar): dibangun saat unduh + cache bytes ─────

def test_stash_builder_katalog_dibangun_saat_unduh(monkeypatch):
    calls = []

    def _fake_katalog(rangka, kategori, source="sinotruk"):
        calls.append((rangka, kategori, source))
        return b"XLSX-BYTES", "x.xlsx"

    monkeypatch.setattr(ai_export, "katalog_excel", _fake_katalog)
    eid, fname = ai_export.stash_builder(
        "Katalog Kabin SJ346500", {"kind": "katalog", "rangka": "SJ346500", "kategori": "kabin"})
    assert fname == "Katalog_Kabin_SJ346500.xlsx"

    data, out_name = ai_export.generic_excel(eid)
    assert data == b"XLSX-BYTES" and out_name == fname
    assert calls == [("SJ346500", "kabin", "sinotruk")]   # kind 'katalog' → sumber Sinotruk
    # klik kedua: dari cache bytes, builder TIDAK dipanggil lagi
    data2, _ = ai_export.generic_excel(eid)
    assert data2 == data and len(calls) == 1


def test_svg_to_png_buang_ukuran_mm():
    # SVG gaya Creo (width/height ber-mm seperti file EPC asli) harus terkonversi.
    svg = (b'<?xml version="1.0"?>\r<svg width="100mm" height="50mm" '
           b'viewBox="0 0 100 50" xmlns="http://www.w3.org/2000/svg">'
           b'<rect x="10" y="10" width="30" height="20" fill="black"/></svg>')
    png = ai_export._svg_to_png(svg, width=200)
    assert png and png[:8] == b"\x89PNG\r\n\x1a\n"
