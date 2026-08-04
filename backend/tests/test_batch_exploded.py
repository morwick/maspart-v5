"""Batch Download kolom "Exploded View" + cache bersama `services/exploded_view`.

Latar: exploded view per-PN terukur **94 detik** panggilan pertama di produksi
(PN dipakai 17.185 model) di server 1 vCPU. Karena itu kolom ini datang dengan
pagar: cap 25 PN, anggaran waktu, paralel 3, gerbang satu-batch, dan build
dipindah ke luar event loop. Yang dijaga di sini adalah pagar-pagar itu — plus
KEJUJURAN sel: PN tanpa figure menulis alasannya, tidak pernah dikarang.

Tanpa jaringan: `epc_bom._get_auto` & `ai_export.exploded_png` di-stub.
"""
from __future__ import annotations

import asyncio
import io
import threading
import time

import openpyxl
import pytest
from fastapi import HTTPException

from app.routers import parts as parts_router
from app.services import catalog, epc_bom, exploded_view

ADMIN = {"username": "agus", "role": "admin"}
PN = "WG9000361402"
PN2 = "WG9000361009"

# PNG 1×1 palsu cukup untuk uji alur; uji ukuran memakai PNG asli bikinan PIL.
def _png_asli(w: int, h: int) -> bytes:
    """PNG SAH — byte palsu ditolak openpyxl, gambarnya tak akan tertempel."""
    from PIL import Image
    b = io.BytesIO()
    Image.new("RGB", (w, h), "white").save(b, format="PNG")
    return b.getvalue()


PNG_PALSU = _png_asli(1400, 900)


def _px(img) -> tuple[float, float]:
    """Ukuran TAMPIL (px) dari anchor. Setelah round-trip, `img.width/height`
    openpyxl melaporkan ukuran ASLI gambar — ukuran tampil hidup di
    `anchor.ext` (EMU; 1 px = 9525 EMU)."""
    ext = img.anchor.ext
    return ext.cx / 9525, ext.cy / 9525

_REVERSE = {"data": [
    {"partCode": "FIG-ADA-PN", "rootId": 2, "partId": 22, "partListId": 222,
     "partName": "Auxiliary gas energy storage device", "model": "Model B"},
]}
_ITEMS = {"data": {"d2s": ["I00050726_C.3.svg"], "items": [
    {"code": "WG9100360108", "name": "Air reservoir bracket", "ballNum": 1},
    {"code": PN + "/1", "name": "Drain valve", "ballNum": 9},
    {"code": PN2, "name": "Nut seat NG12", "ballNum": 12},
]}}


@pytest.fixture(autouse=True)
def bersih():
    exploded_view.bersihkan_cache()
    yield
    exploded_view.bersihkan_cache()


@pytest.fixture
def epc_stub(monkeypatch):
    """Stub EPC + render. Mengembalikan daftar panggilan untuk dihitung test."""
    panggilan: list[tuple] = []

    def fake_get_auto(url, params, **kw):     # **kw: timeout/retries jalur GLOBAL
        panggilan.append((url, dict(params)))
        if url == epc_bom._REVERSE_URL:
            return _REVERSE
        if url == epc_bom._ATLAS_ITEM_URL:
            return _ITEMS
        return {"_err": "api"}

    render: list[str] = []
    monkeypatch.setattr(epc_bom, "_get_auto", fake_get_auto)
    monkeypatch.setattr(exploded_view.ai_export, "exploded_png",
                        lambda svg, ball=None: render.append(svg) or PNG_PALSU)
    monkeypatch.setattr(catalog.part_index, "search_part_number", lambda q: [])
    monkeypatch.setattr(catalog.sims, "get_part_info", lambda pn: {})
    return {"epc": panggilan, "render": render}


def _wb(xls: bytes):
    return openpyxl.load_workbook(io.BytesIO(xls)).active


def _header(ws):
    return [c.value for c in ws[1]]


# ── Bentuk Excel ────────────────────────────────────────────────────────────

def test_chip_exploded_menghasilkan_dua_kolom(epc_stub):
    ws = _wb(catalog.build_catalog_excel([PN], columns=["nama", "exploded"]))
    assert _header(ws) == ["Part Number", "Part Name", "Info Exploded View",
                           "Exploded View (EPC · lintas model)"]
    assert len(ws._images) == 1


def test_urutan_kolom_foto_lalu_exploded(epc_stub, monkeypatch):
    monkeypatch.setattr(catalog.sims, "get_images", lambda pn: [])
    ws = _wb(catalog.build_catalog_excel([PN], columns=["nama", "foto", "exploded"]))
    h = _header(ws)
    assert h.index("Info Exploded View") < h.index("Gambar 1"), "teks sebelum gambar"
    assert h.index("Gambar 2") < h.index("Exploded View (EPC · lintas model)")


def test_clean_columns_menyalakan_kolom_turunan():
    assert catalog.clean_columns(["exploded"]) == ["exploded_info", "exploded"]
    assert catalog.clean_columns(["exploded", "nama"]) == ["nama", "exploded_info", "exploded"]
    # default lama TIDAK boleh berubah
    assert catalog.clean_columns([]) == ["nama", "foto", "stok"]


def test_exploded_tanpa_foto_tak_membuat_kolom_gambar_sims(epc_stub):
    ws = _wb(catalog.build_catalog_excel([PN], columns=["exploded"]))
    assert "Gambar 1" not in _header(ws)


# ── Kejujuran sel (paling penting) ──────────────────────────────────────────

def _sel_info(ws) -> str:
    kol = _header(ws).index("Info Exploded View") + 1
    return str(ws.cell(row=2, column=kol).value or "")


def test_pn_tanpa_figure_menulis_alasan_apa_adanya(monkeypatch):
    monkeypatch.setattr(exploded_view, "png_batch",
                        lambda pns, **kw: {exploded_view.kunci_pn(PN): {
                            "found": False, "part_number": PN,
                            "alasan": "Figure exploded view untuk PN ini tidak ditemukan di EPC."}})
    monkeypatch.setattr(catalog.part_index, "search_part_number", lambda q: [])
    monkeypatch.setattr(catalog.sims, "get_part_info", lambda pn: {})
    ws = _wb(catalog.build_catalog_excel([PN], columns=["exploded"]))
    assert "tidak ditemukan" in _sel_info(ws)
    assert "Figure:" not in _sel_info(ws), "jangan menulis nama figure karangan"
    assert len(ws._images) == 0


def test_epc_ngadat_dibedakan_dari_tidak_ada(monkeypatch):
    """Gangguan EPC != figure tak ada. Menyamakannya membuat user menyimpulkan
    part-nya memang tak punya gambar padahal cuma perlu dicoba lagi."""
    monkeypatch.setattr(epc_bom, "figure_global",
                        lambda pn, max_figures=3: {"found": False, "_err": "api"})
    d = exploded_view.figure_untuk_pn(PN)
    assert d["found"] is False
    assert "tak menjawab" in d["alasan"] and "coba lagi" in d["alasan"].lower()
    assert "tidak ditemukan" not in d["alasan"]

    monkeypatch.setattr(epc_bom, "figure_global",
                        lambda pn, max_figures=3: {"found": False})
    assert "tidak ditemukan" in exploded_view.figure_untuk_pn(PN)["alasan"]


def test_peringatan_lintas_model_ada_di_baris_berhasil(epc_stub):
    ws = _wb(catalog.build_catalog_excel([PN], columns=["exploded"]))
    info = _sel_info(ws)
    assert "lintas-model" in info and "nomor rangka" in info
    assert "Balon: 9" in info, "suffix varian PN harus dikenali → balon ketemu"


def test_kehabisan_anggaran_ditandai_bukan_dikarang(epc_stub):
    hasil = exploded_view.png_batch([PN, PN2], anggaran=0)
    assert set(hasil) == {exploded_view.kunci_pn(PN), exploded_view.kunci_pn(PN2)}
    assert all(not d["found"] and "waktu" in d["alasan"] for d in hasil.values())
    assert exploded_view.CACHE == {}, "yang tak terambil jangan mengotori cache"


def test_gagal_per_pn_tak_menggagalkan_excel(epc_stub, monkeypatch):
    def kadang_meledak(pn, **kw):
        if exploded_view.kunci_pn(pn) == exploded_view.kunci_pn(PN2):
            raise RuntimeError("EPC ngadat")
        return {"found": True, "part_number": pn, "png": PNG_PALSU, "balon": 9,
                "figure_nama": "F", "catatan": exploded_view.catatan_lintas_model()}
    monkeypatch.setattr(exploded_view, "figure_untuk_pn", kadang_meledak)
    hasil = exploded_view.png_batch([PN, PN2])
    assert hasil[exploded_view.kunci_pn(PN)]["found"] is True
    assert hasil[exploded_view.kunci_pn(PN2)]["found"] is False


# ── Ukuran gambar ───────────────────────────────────────────────────────────

def _stub_png(monkeypatch, png: bytes):
    monkeypatch.setattr(exploded_view, "png_batch",
                        lambda pns, **kw: {exploded_view.kunci_pn(pns[0]): {
                            "found": True, "part_number": pns[0], "png": png,
                            "balon": 9, "figure_nama": "F",
                            "catatan": exploded_view.catatan_lintas_model()}})
    monkeypatch.setattr(catalog.part_index, "search_part_number", lambda q: [])
    monkeypatch.setattr(catalog.sims, "get_part_info", lambda pn: {})


def test_gambar_exploded_jauh_lebih_besar_dari_foto(monkeypatch):
    _stub_png(monkeypatch, _png_asli(1400, 900))
    ws = _wb(catalog.build_catalog_excel([PN], columns=["exploded"]))
    w, _h = _px(ws._images[-1])
    assert w == catalog._EXPLODED_IMG_W and w > 200


def test_png_tidak_diresample(monkeypatch):
    png = _png_asli(1400, 900)
    _stub_png(monkeypatch, png)
    ws = _wb(catalog.build_catalog_excel([PN], columns=["exploded"]))
    data = ws._images[-1]._data()
    assert len(data) == len(png), "byte PNG harus utuh — hanya ukuran tampil diubah"


def test_tinggi_baris_tak_melampaui_batas_excel(monkeypatch):
    _stub_png(monkeypatch, _png_asli(600, 2400))     # portrait ekstrem
    ws = _wb(catalog.build_catalog_excel([PN], columns=["exploded"]))
    assert ws.row_dimensions[2].height <= catalog._MAX_ROW_PT
    _w, h = _px(ws._images[-1])
    assert h <= catalog._EXPLODED_IMG_H


# ── Router ──────────────────────────────────────────────────────────────────

def _panggil_router(pns, columns, monkeypatch, build=None):
    monkeypatch.setattr(catalog, "build_catalog_excel",
                        build or (lambda *a, **kw: b"XLSX"))
    monkeypatch.setattr(parts_router.gudang, "can_see_price", lambda u, r: True)
    monkeypatch.setattr(parts_router.permissions, "boleh_stok", lambda u: True)
    return asyncio.run(parts_router.batch_catalog(
        text="\n".join(pns), file=None, columns=columns, user=ADMIN))


def test_router_tolak_lebih_dari_25_pn_saat_exploded(monkeypatch):
    pns = [f"WG900036{i:04d}" for i in range(26)]
    with pytest.raises(HTTPException) as ex:
        _panggil_router(pns, "nama,exploded", monkeypatch)
    assert ex.value.status_code == 400
    assert "25" in ex.value.detail and "Exploded View" in ex.value.detail


def test_router_26_pn_tanpa_exploded_tetap_lolos(monkeypatch):
    pns = [f"WG900036{i:04d}" for i in range(26)]
    r = _panggil_router(pns, "nama,stok", monkeypatch)
    assert r.status_code == 200


def test_router_gerbang_satu_batch_429(monkeypatch):
    assert exploded_view._gerbang.acquire(blocking=False)
    try:
        with pytest.raises(HTTPException) as ex:
            _panggil_router([PN], "nama,exploded", monkeypatch)
        assert ex.value.status_code == 429
        assert "satu" in ex.value.detail
        assert (ex.value.headers or {}).get("Retry-After") == "300"
    finally:
        exploded_view._gerbang.release()


def test_router_bangun_di_luar_main_thread(monkeypatch):
    """Anti-regresi cacat lama: build sinkron di `async def` membekukan server."""
    dicatat = {}

    def build(*a, **kw):
        dicatat["main"] = threading.current_thread() is threading.main_thread()
        return b"XLSX"

    _panggil_router([PN], "nama", monkeypatch, build=build)
    assert dicatat["main"] is False


def test_router_meneruskan_on_progress(monkeypatch):
    """`on_progress` builder dulu ADA tapi tak pernah dipasang router."""
    dicatat = {}

    def build(pns, columns=None, on_progress=None):
        dicatat["cb"] = on_progress
        return b"XLSX"

    _panggil_router([PN], "nama", monkeypatch, build=build)
    assert callable(dicatat["cb"])


# ── Cache bersama ───────────────────────────────────────────────────────────

def test_cache_dibagi_detail_part_dan_batch(epc_stub):
    """Buka di halaman detail part → batch untuk PN itu tak menembak EPC lagi."""
    parts_router.exploded_figure_global(pn=PN, _user=ADMIN)
    n_epc, n_render = len(epc_stub["epc"]), len(epc_stub["render"])
    hasil = exploded_view.png_batch([PN])
    assert hasil[exploded_view.kunci_pn(PN)]["found"] is True
    assert len(epc_stub["epc"]) == n_epc and len(epc_stub["render"]) == n_render


def test_cache_dibagi_arah_sebaliknya(epc_stub):
    exploded_view.png_batch([PN])
    n_render = len(epc_stub["render"])
    out = parts_router.exploded_figure_global(pn=PN.lower(), _user=ADMIN)
    assert out["found"] is True and out["png_base64"]
    assert len(epc_stub["render"]) == n_render


def test_endpoint_tidak_merusak_entri_cache(epc_stub):
    """Jaga bug `pop('png')` tanpa salinan: gambar hilang dari cache."""
    parts_router.exploded_figure_global(pn=PN, _user=ADMIN)
    parts_router.exploded_figure_global(pn=PN, _user=ADMIN)
    d = exploded_view.ambil_cache(PN)
    assert d and d.get("png"), "PNG harus masih ada di cache"
    assert exploded_view.png_batch([PN])[exploded_view.kunci_pn(PN)]["png"]


def test_bentuk_json_endpoint_tidak_berubah(epc_stub):
    """Parser web (api.ts) & mobile (PartExplodedFigure) bergantung pada ini."""
    out = parts_router.exploded_figure_global(pn=PN, _user=ADMIN)
    assert set(out) == {
        "found", "part_number", "svg", "figure_pn", "figure_nama", "nama_item",
        "balon", "jumlah_item", "sumber_model", "jumlah_model_pemakai",
        "png_base64", "catatan",
    }
    assert "png" not in out, "bytes mentah jangan bocor ke JSON"


def test_alias_cache_lama_menunjuk_objek_sama():
    assert parts_router._exploded_cache is exploded_view.CACHE


def test_kegagalan_tidak_dicache(monkeypatch):
    monkeypatch.setattr(epc_bom, "_get_auto", lambda u, p, **kw: {"data": []})
    out = parts_router.exploded_figure_global(pn=PN, _user=ADMIN)
    assert out["found"] is False and "alasan" in out
    assert exploded_view.CACHE == {}


def test_cache_dibatasi_maks_entri(epc_stub):
    for i in range(exploded_view.MAKS_ENTRI + 5):
        exploded_view._simpan_cache(f"PN{i:05d}", {"found": True})
    assert len(exploded_view.CACHE) <= exploded_view.MAKS_ENTRI


# ── Paralelisme & izin ──────────────────────────────────────────────────────

def test_paralel_maksimal_tiga(monkeypatch):
    kunci = threading.Lock()
    aktif = {"n": 0, "puncak": 0}

    def lambat(pn, **kw):
        with kunci:
            aktif["n"] += 1
            aktif["puncak"] = max(aktif["puncak"], aktif["n"])
        time.sleep(0.05)
        with kunci:
            aktif["n"] -= 1
        return {"found": True, "part_number": pn, "png": PNG_PALSU}

    monkeypatch.setattr(exploded_view, "figure_untuk_pn", lambat)
    exploded_view.png_batch([f"PN{i}" for i in range(9)])
    assert aktif["puncak"] <= exploded_view.PEKERJA_BATCH


def test_kolom_exploded_tak_tergerus_gate_harga_stok():
    """exploded bukan data harga/stok → tak boleh ikut di-strip."""
    col = ["nama", "exploded", "harga_sims", "stok"]
    tanpa_harga = [c for c in col if c not in catalog.PRICE_COLUMNS]
    tanpa_stok = [c for c in tanpa_harga if c != "stok"]
    assert "exploded" in tanpa_stok and "harga_sims" not in tanpa_stok
